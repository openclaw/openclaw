#!/usr/bin/env python3
"""Build the final table-first MoClaw operating dashboard web page."""
from __future__ import annotations

import csv
import datetime as dt
import html
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
SRC = ROOT / "web_restore_sources" / "JJHV_MoClaw_Dashboard_v31.xlsx"
OUT = ROOT / "MoClaw_Operating_Dashboard_Web.html"
MISSING_DEFS_OUT = ROOT / "MoClaw_Operating_Dashboard_missing_definitions.txt"
GENERATED_METRICS = ROOT / "generated_posthog_metrics"
COMMENTS_CSS_PATH = ROOT / "dashboard_comments" / "comment_overlay.css"
COMMENT_SCRIPT_PATHS = [
    ROOT / "dashboard_comments" / "comment_model.js",
    ROOT / "dashboard_comments" / "comment_store.js",
    ROOT / "dashboard_comments" / "comment_overlay.js",
]
PAGE_KEY = "moclaw_operating_dashboard"
PAGE_VERSION = "v2"

ORDER = [
    "Dashboard",
    "用户获取",
    "用户激活与转化",
    "用户活跃与使用分布",
    "用户留存与流失",
    "Agent 质量",
    "工程质量",
    "财务",
    "附录1_口径裁决",
    "附录2 个人看板反馈",
]

SHEET_DISPLAY_NAMES = {
    "用户留存与流失": "留存与流失",
    "Agent 质量": "Agent 使用与交付",
    "财务": "业务经营",
}

LEADING_SECTION_TITLES = {
    "Agent 质量": "1. 用户调用",
}

LABEL_RENAMES = {
    "绑卡成功": "新增试用用户",
    "试用订阅": "freetrial",
    "新增试用用户": "新增试用用户",
    "累计付费用户": "累计付费用户",
    "付费用户数": "付费用户数",
    "净 MRR": "MRR",
    "现金收入": "Stripe 现金收入",
    "净现金收入": "Stripe 净现金收入",
    "退款": "退款",
    "注册 CPA": "注册 CPA",
    "付费 CAC": "付费 CAC",
    "新增 UV 占比": "新增 UV 占比",
    "CTA CTR": "CTA CTR",
    "对话时长 P50 (分钟)": "对话时长 P50 (分钟)",
    "对话时长 P90 (分钟)": "对话时长 P90 (分钟)",
}

TEXT_RENAMES = {
    "绑卡成功用户总数": "freetrial 用户总数",
    "绑卡漏斗": "freetrial 漏斗",
    "发起绑卡→绑卡成功": "发起绑卡→新增 freetrial",
    "访问→绑卡成功": "访问→新增 freetrial",
    "注册→绑卡成功": "注册→新增 freetrial",
    "绑卡成功": "新增 freetrial",
    "逾期订阅 past_due": "扣款失败订阅",
    "past_due": "扣款失败",
    "Trialing": "freetrial 中",
    "trialing": "freetrial 中",
    "LCP P75": "页面加载体验 P75",
    "INP P75": "输入响应 P75",
    "CLS P75": "页面稳定 P75",
    "LLM — 流式完成率": "LLM 流式完成率",
    "TTFB P50": "首字节时间 P50",
    "TTFB P95": "首字节时间 P95",
    "环境初始化失败率": "环境启动失败率",
}

DISPLAY_LABEL_RENAMES = {
    "新增注册用户": "新增注册",
    "新增试用用户": "新增freetrial",
    "新增试订阅用户": "新增freetrial",
    "试订阅率": "freetrial 率",
    "试用订阅": "freetrial",
    "试用用户总数": "freetrial 用户总数",
}

FIELD_DEFINITIONS = {
    "指标": "表格第一列，展示业务指标、分区或下钻维度。",
    "UV": "官网或获客入口的去重访问用户数；不是 DAU，也不是全域活跃。",
    "新增 UV": "当日首次访问官网或获客入口的去重用户数。",
    "新增 UV 占比": "新增 UV / UV。",
    "付费渠道新访客 UV": "来自付费投放渠道的新访问去重用户数。",
    "新增注册用户": "当日创建账户的去重用户数。",
    "新增注册": "当日创建账户的去重用户数。",
    "注册开始": "开始注册流程的去重用户数。",
    "注册完成": "完成注册流程的去重用户数。",
    "注册完成率": "注册完成 / 注册开始。",
    "发起绑卡": "开始绑定支付方式的去重用户数。",
    "绑卡开始": "开始绑定支付方式的去重用户数。",
    "新增试用用户": "Stripe 创建 freetrial subscription 的去重用户数。",
    "新增试订阅用户": "Stripe 创建 trial subscription 的去重用户数。",
    "新增 freetrial": "Stripe 创建 freetrial subscription 的去重用户数。",
    "新增freetrial": "Stripe 创建 freetrial subscription 的去重用户数。",
    "freetrial 率": "完成注册后进入 freetrial 的比例，新增 freetrial / 注册完成。",
    "新增付费用户": "当日首次成功扣款的去重 Stripe customer。",
    "累计付费用户": "历史首次成功扣款过的去重 customer 累计数。",
    "付费用户数": "当前有效且成功扣款过的去重付费 customer。",
    "试用用户总数": "历史或当前口径下创建 freetrial subscription 的去重用户总数，按所在表定义窗口读取。",
    "freetrial 用户总数": "历史或当前口径下创建 freetrial subscription 的去重用户总数，按所在表定义窗口读取。",
    "广告花费": "广告平台当日花费；当前主要来自 Google Ads。",
    "广告花费 (美元)": "广告平台当日花费，单位美元。",
    "付费曝光": "广告平台展示次数。",
    "付费点击": "广告平台点击次数。",
    "KOC / KOL": "KOC / KOL 渠道带来的投放或流量，成本和映射接入前显示 —。",
    "campaign": "广告平台 campaign 维度下钻。",
    "CPM": "广告花费 / 曝光数 * 1000。",
    "CPC": "广告花费 / 点击数。",
    "CTR": "广告点击数 / 广告曝光数。",
    "注册 CPA": "广告花费 / 新增注册用户。",
    "付费 CAC": "广告花费 / 新增付费用户；当前只在成本和归因完整时展示。",
    "DAU": "当日发过至少 1 条消息的去重用户数。",
    "WAU": "近 7 日发过至少 1 条消息的去重用户数。",
    "MAU": "近 30 日发过至少 1 条消息的去重用户数。",
    "新增 DAU": "当日首次发消息的去重用户数。",
    "新增 DAU 人均对话量": "新增 DAU 当日平均发起的对话数。",
    "新增 DAU 前三日人均对话量": "新增 DAU 队列在前三日内平均发起的对话数。",
    "首日人均对话量": "新增 DAU 当日平均发起的对话数。",
    "前三日人均对话量": "新增 DAU 队列在前三日内平均发起的对话数。",
    "人均消息数": "消息数 / DAU。",
    "活跃任务数": "待真实 task 事件接入后统计；当前不使用 command:selected 或 tool 事件代理。",
    "人均任务数": "活跃任务数 / DAU；待真实 task 事件接入后计算。",
    "AI Gateway 请求": "AI Gateway 当日处理的模型请求总数；一条用户消息可能触发多次模型请求。",
    "模型供应商请求": "AI Gateway 按模型供应商拆分的请求数。",
    "模型请求": "AI Gateway 按模型名称拆分的请求数。",
    "AI Gateway 错误": "AI Gateway 当日模型请求错误总数。",
    "入口服务 5xx 率": "入口 ALB 返回 5xx 的比例，(Target 5xx + ELB 5xx) / ALB RequestCount。",
    "入口请求": "入口 ALB 当日收到的请求数。",
    "入口 5xx": "入口 ALB 当日 5xx 总数，包含 Target 5xx 和 ELB 5xx。",
    "Target 5xx": "应用目标服务返回给 ALB 的 5xx 次数。",
    "ELB 5xx": "ALB 自身返回的 5xx 次数。",
    "流式中断": "AI Gateway 返回流式响应过程中断的错误次数。",
    "超时": "AI Gateway 调用模型上游或返回响应超时的错误次数。",
    "上游错误": "模型 provider 返回错误或不可用导致的 AI Gateway 错误次数。",
    "Bedrock 错误": "AWS Bedrock 代理或上游返回错误的次数。",
    "LLM 首 token P95": "AI Gateway 从收到请求到返回首个 token 的 P95 秒数。",
    "沙盒不可达率": "Sandbox Fleet 监控中 unreachable / checked 的比例。",
    "沙盒检查次数": "Sandbox Fleet 监控检查次数。",
    "沙盒不可达": "Sandbox Fleet 监控发现 sandbox 不可达的次数。",
    "结账失败率": "结账发起、验证或履约失败 / 发起结账用户。",
    "发起结账用户": "进入 Stripe Checkout 或发起结账流程的去重用户数。",
    "结账失败": "结账发起、验证或履约失败的总次数。",
    "推理成本 credits": "AI Gateway 当日 credits_consumed 汇总；未接入美元单价前作为推理成本代理。",
    "新增任务完成用户": "待真实 task completion 事件接入后统计。",
    "任务完成率": "新增任务完成用户 / 新增 DAU；待真实 task 事件接入后计算。",
    "首次任务开始": "待真实首次 task start 事件接入后统计。",
    "首次任务完成": "待真实首次 task completion 事件接入后统计。",
    "首次发起任务率": "首次任务开始 / 新增 DAU；待真实 task 事件接入后计算。",
    "首次任务完成率": "首次任务完成 / 首次任务开始；待真实 task 事件接入后计算。",
        "次日回访用户数": "新增 DAU cohort 次日发生页面访问、对话、任务发起或响应任一行为的去重用户数。",
        "次日发起会话用户数": "新增 DAU cohort 次日发生 chat:session_start 或 chat:message_sent 的去重用户数。",
        "次日发起任务用户数": "待真实 task start 事件接入后统计。",
        "次日完成任务用户数": "待真实 task completion 事件接入后统计。",
    "回访 DAU": "当日发消息但不是首次发消息的去重用户数。",
    "回访 UV": "当日访问官网或获客入口、但不是首次访问的去重用户数。",
    "持续活跃 DAU": "当日发消息，且上一个周期也发过消息的去重用户数。",
    "复活 DAU": "当日发消息、上一个周期未发消息、但历史曾发过消息的去重用户数。",
    "新增流失用户": "本周期未活跃、上一周期仍活跃，因此新进入流失状态的用户数。",
    "DAU / WAU": "DAU / WAU，用于观察短周期使用粘性。",
    "DAU / MAU": "DAU / MAU，用于观察月度使用粘性。",
    "WAU / MAU": "WAU / MAU，用于观察周活跃用户在月活中的覆盖度。",
    "新增活跃用户": "近 7 日活跃用户中，本周首次发消息的去重用户数。",
    "持续活跃用户": "本周活跃、且上周也活跃的去重用户数。",
    "复活用户": "本周活跃、上周未活跃、但历史曾活跃的去重用户数。",
    "新增活跃用户占 WAU": "新增活跃用户 / WAU。",
    "持续活跃用户占 WAU": "持续活跃用户 / WAU。",
    "复活用户占 WAU": "复活用户 / WAU。",
    "对话数": "按 30 分钟无对话活动切分后的 chat session 数。",
    "会话数": "按 30 分钟无对话活动切分后的 chat session 数。",
    "人均会话数": "会话数 / DAU。",
    "单对话消息数": "chat session 内平均消息数。",
    "单消息对话数": "只有 1 条消息的 chat session 数。",
    "对话时长": "chat session 内最后一条活动与第一条活动的时间差。",
    "对话时长 P50": "对话时长的中位数。",
    "对话时长 P90": "对话时长的 P90 分位。",
    "总对话段": "按 30 分钟无对话活动切分后的 chat session 总数。",
    "总对话段 (托付量)": "按 30 分钟无对话活动切分后的 chat session 总数。",
    "次日留存": "首次发消息用户在次日再次发消息的比例。",
    "次日回访用户": "新增 DAU 在次日再次访问或再次使用的去重用户数。",
    "次日回访用户数": "新增 DAU 在次日再次访问或再次使用的去重用户数。",
    "次日回访率": "次日回访用户数 / 新增 DAU。",
    "次日发起对话用户": "新增 DAU 在次日再次发起对话的去重用户数。",
    "次日发起会话用户数": "次日回访用户中发起会话的去重用户数。",
    "次日回访发起会话率": "次日发起会话用户数 / 次日回访用户数。",
    "次日发起任务用户": "新增 DAU 在次日发起任务的去重用户数。",
    "次日发起任务用户数": "待真实 task start 事件接入后统计。",
    "次日回访发起任务率": "次日发起任务用户数 / 次日回访用户数。",
    "次日完成任务用户": "新增 DAU 在次日完成任务的去重用户数。",
    "次日完成任务用户数": "待真实 task completion 事件接入后统计。",
    "次日任务完成率": "次日完成任务用户数 / 次日发起任务用户数；待真实 task 事件接入后计算。",
    "7 日留存": "首次发消息用户在第 7 日或 7 日窗口内再次发消息的比例，以口径表为准。",
    "D1 留存": "首次发消息队列在 D1 再次发消息的比例。",
    "D1 留存率": "首次发消息队列在 D1 再次发消息的比例。",
    "首次发消息队列人数": "当日首次发消息的去重用户数，作为 cohort 留存分母。",
    "D1 留存用户": "首次发消息队列在 D1 留存的用户数。",
    "D7 留存用户": "首次发消息队列在 D7 留存的用户数。",
    "D30 留存用户": "首次发消息队列在 D30 留存的用户数。",
    "D3 留存": "首次发消息队列在 D3 再次发消息的比例。",
    "D3 留存率": "首次发消息队列在 D3 再次发消息的比例。",
    "D7 留存": "首次发消息队列在 D7 再次发消息的比例。",
    "D7 留存率": "首次发消息队列在 D7 再次发消息的比例。",
    "D14 留存": "首次发消息队列在 D14 再次发消息的比例。",
    "D14 留存率": "首次发消息队列在 D14 再次发消息的比例。",
    "D30 留存": "首次发消息队列在 D30 再次发消息的比例。",
    "D30 留存率": "首次发消息队列在 D30 再次发消息的比例。",
    "当日次日留存": "当日观测到的次日留存率；不是 cohort D1 用户数。",
    "当日7日留存": "当日观测到的 7 日窗口留存率；不是 cohort D7 用户数。",
    "CTA 点击": "获客或激活路径中 CTA 被点击的次数或人数，按所在行口径读取。",
    "新增 UV CTA 点击": "新增 UV 用户触发 CTA 点击的总次数；这是事件次数，不是去重点击用户数。",
    "新增 UV CTA 点击次数": "新增 UV 用户触发 CTA 点击的总次数；这是事件次数，不是去重点击用户数。",
    "新增 UV 人均 CTA 点击": "新增 UV CTA 点击次数 / 新增 UV。",
    "新增 UV CTA 点击率": "新增 UV 中至少点击一次 CTA 的用户数 / 新增 UV；缺少去重点用户数时不展示。",
    "登录 CTA 点击次数": "跳转到登录/注册入口的 CTA 点击次数；不包含其他目的的 CTA。",
    "人均登录 CTA 点击": "登录 CTA 点击次数 / UV。",
    "官网主页 CTA 点击": "官网主页 CTA 被点击的次数。",
    "引导页内 CTA 点击": "引导页内 CTA 被点击的次数。",
    "CTA 点击次数 - 回访": "回访用户触发 CTA 点击的次数。",
    "CTA 点击次数 - 新访": "新访用户触发 CTA 点击的次数。",
    "CTA CTR": "CTA 点击数 / 对应访问或曝光分母。",
    "CTA CTR - 回访": "回访用户 CTA 点击数 / 回访 UV。",
    "CTA CTR - 新访": "新增访问用户 CTA 点击数 / 新增 UV。",
    "CTA 点击用户 → 登录完成": "CTA 点击用户中完成登录/注册流程的比例。",
    "访问→注册": "新增注册用户 / UV。",
    "注册率": "新增 UV 中完成注册的比例，新增注册 / 新增 UV。",
    "试订阅率": "完成注册后进入 freetrial 的比例，新增 freetrial / 注册完成。",
    "付费率": "新增付费用户 / 新增 UV。",
    "注册→发起绑卡": "发起绑卡用户 / 新增注册用户。",
    "访问→发起绑卡": "发起绑卡用户 / UV。",
    "注册→发起绑卡 P50 时长": "从注册完成到发起绑卡的 P50 用时。",
    "注册→发起绑卡 P95 时长": "从注册完成到发起绑卡的 P95 用时。",
    "发起绑卡→新增试用": "新增 freetrial / 发起绑卡用户。",
    "发起绑卡→新增 freetrial": "新增 freetrial / 发起绑卡用户。",
    "绑卡完成率": "新增 freetrial / 绑卡开始。",
    "开口率": "新增 UV 中首次发消息的比例，新增 DAU / 新增 UV。",
    "首次任务开始": "待真实首次 task start 事件接入后统计。",
    "首次发起任务率": "首次任务开始 / 新增 DAU；待真实 task 事件接入后计算。",
    "首次任务完成": "待真实首次 task completion 事件接入后统计。",
    "首次任务完成率": "首次任务完成 / 首次任务开始；待真实 task 事件接入后计算。",
    "新增 DAU 次日回访率": "次日回访用户 / 新增 DAU。",
    "访问→新增试用": "新增 freetrial / UV。",
    "注册→新增试用": "新增 freetrial / 新增注册。",
    "访问→新增 freetrial": "新增 freetrial / UV。",
    "注册→新增 freetrial": "新增 freetrial / 新增注册。",
    "Stripe 现金收入": "Stripe 当日成功扣款金额，未扣除手续费和税。",
    "Stripe 净现金收入": "Stripe 当日成功扣款 - 退款；未扣除手续费和税。",
    "退款": "Stripe 当日退款金额。",
    "D3 应扣款用户": "3 天前新增 freetrial、理论上当日应进入扣款窗口的用户数。",
    "D3 付费用户": "D3 扣款窗口内首次成功扣款用户数。",
    "D3 付费率": "D3 付费用户 / D3 应扣款用户。",
    "D+3 应扣款用户": "3 天前新增 freetrial、理论上当日应进入扣款窗口的用户数。",
    "D+3 付费用户": "D+3 扣款窗口内首次成功扣款用户数。",
    "D+3 付费率": "D+3 付费用户 / D+3 应扣款用户。",
    "D3 扣款成功率": "D3 成功扣款用户 / D3 应扣款用户。",
    "到期扣款失败用户": "freetrial 到期后扣款失败的用户数。",
    "试用期取消用户": "freetrial 期内取消或预约取消的用户数。",
    "MRR": "当前有效付费订阅对应的月化经常性收入。",
    "ARPU": "月经常性收入 / 付费用户数。",
    "总成本": "经营成本汇总；成本源未接通时显示 —。",
    "LLM 成本": "模型推理成本。",
    "云服务成本": "云服务与基础设施成本。",
    "毛利": "收入 - 成本。",
    "毛利率": "毛利 / 收入。",
    "任务成功率": "任务完成数 / 任务发起数；不得用 API 成功率冒充。",
    "活跃用户国家 / 地区分布": "活跃用户按国家和地区的分布。",
    "活跃用户端类型分布": "活跃用户按客户端类型的分布。",
    "文件上传用户数": "当日上传文件的去重用户数。",
    "文件上传次数": "当日文件上传次数。",
    "文件上传成功率": "文件上传完成次数 / 文件上传开始次数。",
    "文件下载用户数": "当日下载文件的去重用户数。",
    "文件下载次数": "当日文件下载次数。",
    "定时任务创建用户数": "当日创建定时任务的去重用户数。",
    "定时任务创建数": "当日创建定时任务次数。",
    "定时任务执行用户数": "当日触发定时任务执行的去重用户数。",
    "定时任务执行数": "当日定时任务执行次数。",
    "skill 使用用户数": "当日使用 skill 的去重用户数。",
    "skill 使用比例": "skill 使用用户数 / DAU。",
    "connector 使用用户数": "当日使用 connector 的去重用户数。",
    "connector 使用比例": "connector 使用用户数 / DAU。",
    "流失用户数": "到达应回访窗口但未回访的用户数。",
    "流失率": "流失用户数 / 应回访用户数。",
    "新增流失用户 / WAU": "新增流失用户 / WAU。",
    "付费流失用户数": "真付费有效用户中进入流失状态的人数。",
    "付费流失率": "付费流失用户数 / 真付费有效用户。",
    "回流用户数": "流失后再次活跃的用户数。",
    "回流后活跃用户数": "回流后达到活跃标准的用户数。",
    "回流后活跃率": "回流后活跃用户数 / 回流用户数。",
    "首个任务发起用户": "首次进入有效 Agent 任务流的用户数。",
    "首个任务完成用户": "首次完成有效 Agent 任务的用户数。",
    "首条消息→首个任务完成": "首次发消息用户中完成首个有效 Agent 任务的比例。",
    "端到端成功率": "用户请求从入口到最终完成的成功比例。",
    "消息错误率": "消息链路中出现错误的比例。",
    "WS 连接错误率": "WebSocket 连接错误数 / WebSocket 连接尝试数。",
    "API 调用量": "对应 API 的请求数。",
    "API P50 延迟": "API 响应时间 P50。",
    "API P95 延迟": "API 响应时间 P95。",
    "API — 整体成功率": "API 2xx 成功请求 / 全部请求。",
    "LLM — 流式完成率": "LLM 流式响应正常完成数 / LLM 流式请求数。",
    "LLM 流式完成率": "LLM 流式响应正常完成数 / LLM 流式请求数。",
    "LLM 全失败": "LLM 请求完全失败的次数。",
    "沙盒失败": "沙盒创建、执行或健康检查失败次数。",
    "沙盒失败率": "沙盒失败数 / 沙盒请求数。",
    "沙盒池饱和度": "沙盒资源占用 / 可用容量。",
    "用户主动中止率": "用户主动中断响应或任务的比例。",
    "流式错误": "流式响应过程中发生错误的次数。",
    "中位中止时点": "用户或系统中止发生时，在响应生命周期中的中位位置。",
    "中位循环深度": "Agent 任务执行过程中的中位循环轮次。",
    "中位消息长度": "用户消息长度的中位数。",
    "中位时长": "对应对象时长的中位数。",
    "有附件对话段占比": "包含附件的 chat session / 全部 chat session。",
    "24 小时续聊率": "首次对话后 24 小时内再次发消息的用户比例。",
    "授权通过率": "发起授权后完成授权的比例。",
    "缓存命中率": "命中缓存的请求 / 可缓存请求。",
    "页面加载体验 P75": "页面最大内容完成渲染的 P75 用时。",
    "输入响应 P75": "页面输入交互响应的 P75 用时。",
    "页面稳定 P75": "页面布局稳定性的 P75 指标。",
    "首字节时间 P50": "请求发出到收到首个响应字节的 P50 用时。",
    "首字节时间 P95": "请求发出到收到首个响应字节的 P95 用时。",
    "环境启动失败率": "运行环境初始化失败数 / 初始化请求数。",
    "不可回滚副作用": "执行后不能自动撤销的高风险动作数量。",
    "数量": "对应分组下的计数。",
    "占比": "对应分组数量 / 所在分组总量。",
    "打开账单门户": "打开 Stripe billing portal 的去重用户数。",
    "撞额度墙": "触发额度不足或用量限制提示的去重用户数。",
    "当周回流": "长期未活跃后本周再次发消息的用户数。",
    "回流后 7 天再激活": "回流用户在 7 天内再次达到活跃标准的人数。",
    "回流后又流失": "回流后短期内再次停止活跃的人数。",
    "重度日用": "近 7 天高频发消息的付费用户分桶。",
    "中度": "近 7 天中频发消息的付费用户分桶。",
    "轻度": "近 7 天低频发消息的付费用户分桶。",
    "低频": "近 7 天仅少量使用的付费用户分桶。",
    "冷却": "近 7 天未活跃、但更早窗口内仍有活动的付费用户分桶。",
    "流失中": "较长时间未发消息、但未达到深度流失阈值的付费用户分桶。",
    "深度流失": "长时间未发消息或从未发消息的付费用户分桶。",
    "中断异常": "个人 chat:response_aborted 次数高于自身历史均值显著区间的用户或事件。",
    "之后 30 天真取消": "触发前兆事件后 30 天内真实取消订阅的人数。",
    "7-13 天前撞墙队列后 7 天升级": "7-13 天前撞额度墙的队列，在随后 7 天内升级的人数。",
    "7-13 天前撞墙队列后 7 天未活跃": "7-13 天前撞额度墙的队列，在随后 7 天内未再发消息的人数。",
    "直接访问": "没有可识别来源参数或推荐来源的访问。",
    "Direct": "没有可识别来源参数或推荐来源的访问。",
    "SEO / 自然搜索": "来自自然搜索的访问。",
    "SEO / Organic Search": "来自自然搜索的访问。",
    "自然社交": "来自非付费社交渠道的访问。",
    "Organic Social": "来自非付费社交渠道的访问。",
    "外链推荐": "来自第三方链接推荐的访问。",
    "Referral": "来自第三方链接推荐的访问。",
    "Google Ads": "来自 Google Ads 的广告流量或广告平台数据。",
    "Meta Ads": "来自 Meta Ads 的广告流量或广告平台数据。",
    "Twitter / X": "来自 Twitter / X 的访问。",
    "Reddit": "来自 Reddit 的访问。",
    "Product Hunt": "来自 Product Hunt 的访问。",
    "TikTok": "来自 TikTok 的访问。",
    "YouTube": "来自 YouTube 的访问。",
    "未识别 UTM": "有 UTM 但无法归入已知渠道的访问。",
    "KOC / KOL": "KOC / KOL 渠道；成本和归因映射接入前显示 —。",
    "KOC / KOL（待映射）": "KOC / KOL 渠道；成本和归因映射接入前显示 —。",
    "onboarding 页面": "经 onboarding 路径进入注册或激活流程的用户。",
    "/start": "经 /start 页面进入注册或激活流程的用户。",
    "/usecase": "经 /usecase 页面进入注册或激活流程的用户。",
    "首页": "经首页进入注册或激活流程的用户。",
    "campaign（待数据）": "广告 campaign 维度；接入 campaign 映射前显示 —。",
        "绑卡失败": "绑定支付方式过程中发生发起、验证或履约失败的流程数。",
    "结账页连续点击": "结账页同一用户短时间内连续点击同一操作的次数。",
    "登录失败": "登录或注册流程失败的次数或人数，按所在行口径读取。",
    "登录页连续点击": "登录页同一用户短时间内连续点击同一操作的次数。",
    "绑卡成本": "广告花费 / 新增试用用户。",
    "ROAS": "广告归因收入 / 广告花费。",
    "WebSocket — 连接健康": "WebSocket 连接稳定性分区。",
    "沙盒 — 健康": "沙盒可用性与稳定性分区。",
    "浏览器/前端 — Web 核心指标健康": "浏览器与前端体验指标分区。",
    "端到端 P50 延迟": "用户请求从入口到最终响应完成的 P50 用时。",
    "端到端 P95 延迟": "用户请求从入口到最终响应完成的 P95 用时。",
    "完成率": "完成数 / 发起数。",
    "总请求": "对应服务或路径的请求总数。",
    "回调调用量": "回调接口被调用的次数。",
    "受影响用户": "被错误、故障或性能问题影响的去重用户数。",
    "今日发布": "当日生产环境发布次数。",
    "回滚": "当日生产环境回滚次数。",
    "发版前后 1 小时错误率 Δ": "发布后 1 小时错误率 - 发布前 1 小时错误率。",
    "重启恢复率": "重启后恢复正常的实例或任务 / 需要恢复的总数。",
    "卡顿率": "前端交互或页面渲染发生明显卡顿的比例。",
    "客户端异常": "客户端捕获到的异常次数。",
    "客户端连续点击": "同一控件短时间连续点击导致的异常交互次数。",
    "前端版本切换": "用户会话中发生前端版本切换的次数。",
    "断连率 / 小时": "每小时 WebSocket 或连接断开的比例。",
    "AI / 网络 / 工具三分": "把失败或问题按 AI、网络、工具三类归因的分布。",
    "跨桶 — 副作用治理": "跨任务桶统计不可回滚副作用与治理情况的分区。",
}

MISSING_DEFINITIONS: list[tuple[str, str]] = []

FIELD_DEFINITIONS.update(
    {
        "对话数": "按 30 分钟无对话活动切分后的 chat session 数。",
        "人均对话数": "对话数 / DAU。",
        "Agent 响应数": "chat:response_received 事件数。",
        "响应率": "Agent 响应数 / 消息数。",
        "消息失败数": "error:message_failed 事件数。",
        "流式错误数": "chat.stream.errored 事件数。",
        "LLM 请求": "llm:request_completed 事件数。",
        "LLM 2xx": "status_code 为 2xx 的 LLM 请求数。",
        "LLM 4xx / 5xx": "status_code 为 4xx 或 5xx 的 LLM 请求数。",
        "LLM 2xx 率": "LLM 2xx / LLM 请求。",
        "LLM 耗时": "llm:request_completed 的 duration_ms 分位数。",
        "credits": "llm:request_completed 的 credits_consumed 汇总。",
        "人均 credits": "credits / DAU。",
        "tokens": "LLM 请求的 token 消耗汇总。",
        "input tokens": "llm:request_completed 的 input_tokens 汇总。",
        "output tokens": "llm:request_completed 的 output_tokens 汇总。",
        "cache read tokens": "llm:request_completed 的 cache_read_tokens 汇总。",
        "connector 连接完成": "connector:*_connected 事件数汇总。",
        "Google Workspace 设置完成": "connector:google_workspace_folder_set 事件数。",
        "Google Workspace picker unavailable": "connector:google_workspace_picker_unavailable 事件数。",
        "Google Workspace 连接率": "Google Workspace 设置完成 / (Google Workspace 设置完成 + Google Workspace picker unavailable)。",
        "API 请求": "server:api_called 事件数。",
        "API 成功": "server:api_called 中 status_code 为 2xx/3xx 的请求数。",
        "API 失败": "server:api_called 中 status_code 为 4xx/5xx 的请求数。",
        "API 成功率": "API 成功 / API 请求。",
        "API 耗时": "server:api_called 的 latency_ms 分位数。",
        "实时连接错误": "WebSocket 连接错误事件数。",
        "消息发送失败": "消息发送失败事件数。",
        "对话流中断": "对话流返回过程中断或报错的事件数。",
        "沙盒启动次数": "需要执行环境的任务触发沙盒环境初始化的次数。",
        "沙盒启动失败": "沙盒环境初始化失败次数。",
        "沙盒启动失败率": "沙盒启动失败 / 沙盒启动次数。",
        "沙盒启动耗时": "沙盒环境初始化耗时分位数。",
        "沙盒重启失败": "已有沙盒热恢复或重启失败次数。",
        "附件上传开始": "用户开始上传附件的次数。",
        "附件上传成功": "附件上传成功完成的次数。",
        "附件上传成功率": "附件上传成功 / 附件上传开始。",
        "附件上传失败": "附件上传失败次数。",
        "附件上传失败率": "附件上传失败 / 附件上传开始。",
        "结账发起失败": "支付或订阅结账发起失败次数。",
        "结账验证失败": "支付或订阅结账验证失败次数。",
        "结账履约失败": "支付或订阅订单履约失败次数。",
        "结账失败率": "结账失败数 / 创建结账会话。",
        "pricing opened 用户": "触发 pricing:dialog_opened 的去重用户数。",
        "checkout started 用户": "触发 payment:checkout_started 的去重用户数。",
        "pricing → checkout started": "checkout started 用户 / pricing opened 用户。",
        "查看价格用户": "打开价格或套餐入口的去重用户数。",
        "发起结账用户": "开始结账流程的去重用户数。",
        "价格页发起结账率": "发起结账用户 / 查看价格用户。",
        "checkout session created flows": "按 checkout_flow_id 去重的 payment:checkout_session_created flow 数。",
        "checkout verified flows": "按 checkout_flow_id 去重的 payment:checkout_verified flow 数。",
        "session created → verified": "checkout verified flows / checkout session created flows。",
        "checkout fulfilled flows": "按 checkout_flow_id 去重的 payment:checkout_fulfilled flow 数。",
        "session created → fulfilled": "checkout fulfilled flows / checkout session created flows。",
        "checkout failed": "checkout start / verify / fulfillment failed 汇总。",
        "subscription flows": "订阅类结账流程数。",
        "credit pack flows": "点数包类结账流程数。",
        "创建结账会话": "已创建 Stripe 结账会话的去重结账流程数。",
        "支付验证完成": "完成支付验证的去重结账流程数。",
        "结账验证率": "支付验证完成 / 创建结账会话。",
        "订单完成": "已完成履约的去重结账流程数。",
        "订单完成率": "订单完成 / 创建结账会话。",
        "结账失败": "结账发起、支付验证、订单履约失败的合计。",
        "发起失败": "结账发起阶段失败数。",
        "验证失败": "支付验证阶段失败数。",
        "履约失败": "订单履约阶段失败数。",
        "订阅订单": "订阅类结账流程数。",
        "点数包订单": "点数包类结账流程数。",
        "到期 freetrial 用户": "3 天前进入 freetrial、在当日进入 D+3 扣款观察窗口的去重 Stripe customer。",
        "D-3 绑卡成功客户": "3 天前进入 freetrial、在当日进入 D+3 扣款观察窗口的去重 Stripe customer。",
        "D+3 成功扣款客户": "D+3 观察日首次成功扣款的 Stripe customer。",
        "D+3 新增付费用户": "D+3 观察日首次成功扣款的 Stripe customer。",
        "扣款成功率": "新增付费用户 / 到期 freetrial 用户。",
        "D+3 past_due 客户": "D+3 观察日扣款失败并进入 past_due 的 Stripe customer。",
        "past_due 比例": "D+3 past_due 客户 / 到期 freetrial 用户。",
        "D+3 canceled 客户": "D+3 观察日已取消的 Stripe customer。",
        "canceled 比例": "D+3 canceled 客户 / 到期 freetrial 用户。",
        "D+3 扣款失败客户": "freetrial 到期扣款失败的客户数。",
        "扣款失败用户": "freetrial 到期扣款失败的客户数。",
        "扣款失败率": "扣款失败用户 / 到期 freetrial 用户。",
        "D+3 取消客户": "freetrial 到期前后取消的客户数。",
        "取消用户": "freetrial 到期前后取消的客户数。",
        "取消率": "取消用户 / 到期 freetrial 用户。",
        "active 订阅": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "付费有效订阅": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "付费有效订阅数": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "freetrial 中订阅": "当前 Stripe status=trialing 的订阅数。",
        "trialing 订阅": "当前 Stripe status=trialing 的订阅数。",
        "past_due 订阅": "当前扣款失败的订阅数。",
        "扣款失败订阅": "当前扣款失败的订阅数。",
        "月经常性收入": "当前有效付费订阅折算的月经常性收入。",
        "付费用户平均收入": "月经常性收入 / 付费有效订阅数。",
        "付费订阅流失": "当日取消的付费订阅数。",
    }
)


MISSING = {"", "—", "-", "–", "None", "none", "null"}
DERIVED_METRIC_TOKENS = [
    "率",
    "占比",
    "CTR",
    "CPC",
    "CPM",
    "CPA",
    "CAC",
    "ARPU",
    "MRR",
    "P50",
    "P90",
    "P95",
    "P99",
]
PAID_EFFICIENCY_NO_DRILLDOWN = {"CPM", "CTR", "CPC"}
PAID_PLATFORM_LABELS = {"Google Ads", "Meta Ads"}
CHANNEL_LABELS = {
    "直接访问",
    "SEO / 自然搜索",
    "自然社交",
    "外链推荐",
    "Google Ads",
    "Meta Ads",
    "Twitter / X",
    "Reddit",
    "Product Hunt",
    "TikTok",
    "YouTube",
    "未识别 UTM",
}


def is_missing_text(value: str) -> bool:
    return value.strip() in MISSING


def metric_kind(label: str) -> str:
    if is_percent_label(label):
        return "percent"
    if any(k in label for k in ["美元", "现金", "收入", "退款", "MRR", "ARPU", "CAC", "CPA", "CPC", "CPM", "成本", "毛利", "广告花费"]):
        return "currency"
    if any(k in label for k in ["时长", "时点", "延迟", "P50", "P90", "P95", "P99"]):
        return "decimal"
    return "number"


def is_derived_metric_label(label: str) -> bool:
    return any(token in label for token in DERIVED_METRIC_TOKENS)


def is_paid_efficiency_label(label: str) -> bool:
    normalized = normalize_label_for_definition(label)
    return normalized in PAID_EFFICIENCY_NO_DRILLDOWN


def derived_child_label(label: str, parent_label: str) -> str:
    normalized_label = normalize_label_for_definition(label)
    if normalized_label not in CHANNEL_LABELS:
        return label
    normalized_parent = normalize_label_for_definition(parent_label)
    suffix = ""
    if "占比" in normalized_parent:
        suffix = "占比"
    elif "CTR" in normalized_parent:
        suffix = "CTR"
    elif "CPM" in normalized_parent:
        suffix = "CPM"
    elif "CPC" in normalized_parent:
        suffix = "CPC"
    elif "CPA" in normalized_parent:
        suffix = "CPA"
    elif "CAC" in normalized_parent:
        suffix = "CAC"
    elif "率" in normalized_parent:
        suffix = "率"
    if not suffix or suffix in label:
        return label
    return f"{label} {suffix}"


def fmt_number(value: float, *, decimals: int = 1, currency: bool = False) -> str:
    sign = "-" if value < 0 else ""
    abs_value = abs(value)
    if decimals == 0 or abs_value == int(abs_value):
        body = f"{int(abs_value):,}"
    else:
        body = f"{abs_value:,.{decimals}f}".rstrip("0").rstrip(".")
    return f"{sign}${body}" if currency else f"{sign}{body}"


def fmt_value(value, number_format: str = "", *, kind: str = "number") -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        nf = (number_format or "").lower()
        if "%" in nf or kind == "percent":
            return f"{value * 100:.1f}%"
        if kind == "currency":
            return fmt_number(float(value), currency=True)
        if kind == "decimal":
            return fmt_number(float(value), decimals=1)
        return fmt_number(float(value), decimals=1)
    return str(value)


def parse_display_number(value: str) -> float | None:
    cleaned = value.strip()
    if not cleaned or is_missing_text(cleaned) or "%" in cleaned:
        if "(" not in cleaned:
            return None
        cleaned = cleaned.split("(", 1)[0].strip()
        if not cleaned or is_missing_text(cleaned) or "%" in cleaned:
            return None
    cleaned = cleaned.replace("$", "").replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def append_child_share(value: str, child: float | None, parent: float | None) -> str:
    if child is None or parent in (None, 0):
        return value
    return f"{value} ({child / parent * 100:.0f}%)"


SHARE_ANNOTATION_RE = re.compile(r"^(.+?)\s+\(-?\d+(?:\.\d+)?%\)$")


def strip_share_annotation(value: str) -> str:
    match = SHARE_ANNOTATION_RE.match(value.strip())
    return match.group(1) if match else value


def has_present_value(value: str) -> bool:
    cleaned = value.strip()
    return bool(cleaned) and not is_missing_text(cleaned)


def has_quantitative_value(value: str) -> bool:
    cleaned = value.strip()
    if not cleaned or is_missing_text(cleaned):
        return False
    if "%" in cleaned:
        numeric = cleaned.split("%", 1)[0].split("(")[-1].strip()
        try:
            float(numeric.replace(",", "").replace("$", ""))
            return True
        except ValueError:
            return False
    return parse_display_number(cleaned) is not None


def has_nonzero_value(value: str) -> bool:
    if not has_present_value(value):
        return False
    number = parse_display_number(value)
    if number is None:
        return True
    return number != 0


def filter_empty_rows(rows: list[dict]) -> list[dict]:
    keep_ids: set[str] = set()
    row_by_id = {row["id"]: row for row in rows}

    for row in rows:
        if row["kind"] in {"header", "section"}:
            continue
        values = row["values"][1:]
        if row.get("keep_empty"):
            keep = True
        elif row["kind"] == "child":
            keep = any(has_nonzero_value(value) for value in values)
        else:
            keep = any(has_present_value(value) for value in values)
        if keep:
            keep_ids.add(row["id"])

    changed = True
    while changed:
        changed = False
        for row_id in list(keep_ids):
            parent_id = row_by_id.get(row_id, {}).get("parent")
            if parent_id and parent_id not in keep_ids:
                keep_ids.add(parent_id)
                changed = True

    section_has_rows: dict[str, bool] = {}
    current_section = ""
    for row in rows:
        if row["kind"] == "section":
            current_section = row["id"]
            section_has_rows[current_section] = False
            continue
        if current_section and row["id"] in keep_ids:
            section_has_rows[current_section] = True

    filtered = []
    for row in rows:
        if row["kind"] == "header":
            filtered.append(row)
        elif row["kind"] == "section":
            if section_has_rows.get(row["id"]):
                filtered.append(row)
        elif row["id"] in keep_ids:
            filtered.append(row)

    child_ids = {row["parent"] for row in filtered if row.get("parent")}
    for row in filtered:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return filtered


def filter_non_quantitative_metric_rows(rows: list[dict]) -> list[dict]:
    removed = {
        row["id"]
        for row in rows
        if row["kind"] not in {"header", "section"}
        and any(has_present_value(value) for value in row["values"][1:])
        and not any(has_quantitative_value(value) for value in row["values"][1:])
    }
    if not removed:
        return rows
    for row in rows:
        while row.get("parent") in removed:
            parent = next((candidate for candidate in rows if candidate["id"] == row["parent"]), None)
            if not parent:
                row["parent"] = ""
                break
            row["parent"] = parent.get("parent", "")
            row["level"] = max(0, row.get("level", 0) - 1)
    kept = [row for row in rows if row["id"] not in removed]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


SHARE_BREAKDOWN_EXACT_LABELS = {
    "付费渠道新访客 UV",
    "新增 UV",
    "回访 UV",
    "新增 DAU",
    "回访 DAU",
    "持续活跃 DAU",
    "复活 DAU",
    "新增流失用户",
    "单消息对话数",
    "新增",
    "持续活跃",
    "复活",
    "流失转入",
    "重度日用",
    "中度",
    "轻度",
    "低频",
    "冷却",
    "流失中",
    "深度流失",
}

SHARE_BREAKDOWN_DENY_TOKENS = [
    "平均",
    "人均",
    "单对话",
    "P50",
    "P75",
    "P90",
    "P95",
    "P99",
    "时长",
    "延迟",
    "CPA",
    "CAC",
    "CPM",
    "CPC",
    "CTR",
    "ARPU",
    "MRR",
    "率",
    "占比",
]


def is_share_breakdown_child(child_label: str, parent_label: str) -> bool:
    child = normalize_label_for_definition(child_label)
    parent = normalize_label_for_definition(parent_label)
    if any(token in child for token in SHARE_BREAKDOWN_DENY_TOKENS):
        return False
    if child.startswith("失败："):
        return True
    if child in CHANNEL_LABELS:
        return True
    if child in SHARE_BREAKDOWN_EXACT_LABELS:
        return True
    if child in {"新增试用用户", "新增付费用户"} and "注册" in parent:
        return False
    return False


FLOW_OR_STAT_TOKENS = [
    "→",
    "P50",
    "P75",
    "P90",
    "P95",
    "P99",
    "平均",
    "人均",
    "单对话",
    "单消息",
    "中位",
    "时长",
    "延迟",
    "率",
    "占比",
    "CTR",
    "CPC",
    "CPM",
    "CPA",
    "CAC",
    "ARPU",
    "MRR",
]


def is_valid_drilldown(parent_label: str, child_label: str) -> bool:
    parent = normalize_label_for_definition(parent_label)
    child = normalize_label_for_definition(child_label)
    if not parent or not child:
        return False
    channel_base = next((channel for channel in CHANNEL_LABELS if child == channel or child.startswith(f"{channel} ")), "")
    if channel_base:
        return parent in {
            "UV",
            "新增 UV",
            "CTA 点击",
            "新增注册用户",
            "新增试用用户",
            "广告花费",
            "付费曝光",
            "付费点击",
            "新增 UV 占比",
            "CTR",
            "CPC",
            "CPM",
            "注册 CPA",
            "付费 CAC",
        }
    if any(token in child for token in FLOW_OR_STAT_TOKENS):
        return False
    if child in PAID_PLATFORM_LABELS:
        return parent in {"广告花费", "付费曝光", "付费点击"}
    if parent == "UV" and child in {"新增 UV", "回访 UV"}:
        return True
    if parent == "DAU" and child in {"新增 DAU", "回访 DAU"}:
        return True
    if parent in {"登录失败", "绑卡失败"} and child.startswith("失败："):
        return True
    if parent == "试用用户总数" and child in {"重度日用", "中度", "轻度", "低频", "冷却", "流失中", "深度流失"}:
        return True
    if parent == "Stripe 净现金收入" and child in {"新增付费用户", "退款"}:
        return False
    return False


def normalize_drilldown_relationships(rows: list[dict]) -> list[dict]:
    by_id = {row["id"]: row for row in rows}
    for row in rows:
        parent_id = row.get("parent")
        if not parent_id:
            continue
        parent = by_id.get(parent_id)
        if not parent or is_valid_drilldown(parent["values"][0], row["values"][0]):
            continue
        row["parent"] = ""
        row["hidden"] = False
        if row["kind"] == "child":
            row["kind"] = "derived" if row.get("ratio") or is_ratio_label(row["values"][0]) else "row"
        row["level"] = 0

    child_ids = {row["parent"] for row in rows if row.get("parent")}
    for row in rows:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return rows


DASHBOARD_EXCLUDE_LABELS = {
    "付费渠道新访客 UV",
    "单对话消息数",
    "单消息对话数",
}


def prune_dashboard_summary(rows: list[dict]) -> list[dict]:
    kept = [
        row
        for row in rows
        if row["kind"] in {"header", "section"}
        or normalize_label_for_definition(row["values"][0]) not in DASHBOARD_EXCLUDE_LABELS
    ]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


def remove_single_child_share_annotations(rows: list[dict]) -> list[dict]:
    children_by_parent: dict[str, list[dict]] = {}
    for row in rows:
        if row.get("parent") and row["kind"] == "child":
            children_by_parent.setdefault(row["parent"], []).append(row)

    for children in children_by_parent.values():
        if len(children) != 1:
            continue
        child = children[0]
        child["values"] = [child["values"][0]] + [strip_share_annotation(value) for value in child["values"][1:]]
    return rows


def strip_marker(text: str) -> tuple[str, bool]:
    value = text.strip()
    had_marker = bool(re.match(r"^(└|┄|·)\s*", value))
    value = re.sub(r"^(└|┄|·)\s*", "", value)
    for src, dst in TEXT_RENAMES.items():
        value = value.replace(src, dst)
    return LABEL_RENAMES.get(value, value), had_marker


def clean_text(text: str) -> str:
    value = text
    for src, dst in TEXT_RENAMES.items():
        value = value.replace(src, dst)
    return value


def raw_cell_text(cell) -> str:
    return clean_text(fmt_value(cell.value, cell.number_format))


def used_bounds(ws) -> tuple[int, int]:
    max_row = max_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def is_section(values: list[str]) -> bool:
    first = values[0].strip() if values else ""
    return bool(re.match(r"^\d+[\.\、]\s*\S+", first)) and all(not v.strip() for v in values[1:])


def is_header_row(row_index: int, values: list[str]) -> bool:
    return row_index == 1 and values and values[0] in {"指标", "反馈对象", "口径项", "指标名称", "业务指标名"}


def is_ratio_label(label: str) -> bool:
    return any(k in label for k in ["率", "占比", "CTR", "CPC", "CPM", "CPA", "CAC", "ARPU", "MRR"])


def is_percent_label(label: str) -> bool:
    return any(k in label for k in ["率", "占比", "CTR", "留存", "转化", "粘性", "DAU / WAU", "DAU / MAU"])


UNIT_SUFFIX_RE = re.compile(r"\s*[（(](个|美元|百分比|毫秒|秒|分钟|人|次|天|月|%|\$)[）)]\s*$")
DISPLAY_UNIT_RE = re.compile(r"\s*[（(](个|美元|百分比|毫秒|秒|分钟|人|次|天|月|%|\$)[）)]")
SECTION_PREFIX_RE = re.compile(r"^\d+[\.\、]\s*")


def slug_key(value) -> str:
    text = html.unescape(str(value or "")).strip().lower()
    text = SECTION_PREFIX_RE.sub("", text)
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "-", text)
    text = text.strip("-")
    return text or "blank"


def column_key(label, index: int) -> str:
    label_key = slug_key(label)
    return f"col-{index}-{label_key}" if label_key != "blank" else f"col-{index}"


def row_anchor_key(row: dict) -> str:
    return slug_key(row.get("id") or row.get("values", [""])[0])


def section_anchor_key(section_row: dict | None) -> str:
    if not section_row:
        return ""
    return row_anchor_key(section_row)


def anchor_id_from_parts(*, sheet: str, section: str, row: str, column: str, anchor_type: str) -> str:
    return ":".join(
        [
            PAGE_KEY,
            PAGE_VERSION,
            sheet,
            section or "_",
            row or "_",
            column or "_",
            anchor_type,
        ]
    )


def anchor_label(value) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(value or "")).strip())


def anchor_attrs(
    anchor_type: str,
    *,
    sheet_title: str,
    row: dict,
    column: str = "",
    section_row: dict | None = None,
    row_key: str | None = None,
    section_key: str | None = None,
    row_label: str = "",
    section_label: str = "",
    column_label: str = "",
) -> str:
    sheet = slug_key(sheet_title)
    section = section_key if section_key is not None else section_anchor_key(section_row)
    anchor_row = row_key if row_key is not None else row_anchor_key(row)
    if anchor_type == "section":
        section = section or row_anchor_key(row)
        anchor_row = ""
        column = ""
    elif anchor_type == "row":
        column = ""
    anchor_id = anchor_id_from_parts(
        sheet=sheet,
        section=section,
        row=anchor_row,
        column=column,
        anchor_type=anchor_type,
    )
    readable_row = anchor_label(row_label or (row.get("values") or [""])[0])
    readable_section = anchor_label(section_label or ((section_row or row).get("values") or [""])[0])
    attrs = {
        "data-anchor-id": anchor_id,
        "data-anchor-type": anchor_type,
        "data-anchor-page": PAGE_KEY,
        "data-anchor-version": PAGE_VERSION,
        "data-anchor-sheet": sheet,
        "data-anchor-sheet-title": display_sheet_title(sheet_title),
        "data-anchor-section": section,
        "data-anchor-section-title": readable_section,
        "data-anchor-row": anchor_row,
        "data-anchor-row-label": readable_row,
        "data-anchor-column": column,
        "data-anchor-column-label": anchor_label(column_label),
    }
    return "".join(f' {name}="{html.escape(value, quote=True)}"' for name, value in attrs.items())


def tooltip_attrs(text: str | None) -> str:
    if not text:
        return ""
    return f' data-tooltip="{html.escape(text, quote=True)}" tabindex="0"'


def normalize_label_for_definition(label: str) -> str:
    value = html.unescape(label).strip()
    value = SECTION_PREFIX_RE.sub("", value)
    value = re.sub(r"\s+", " ", value)
    value = UNIT_SUFFIX_RE.sub("", value).strip()
    value = value.replace("SEO / Organic Search", "SEO / 自然搜索")
    value = value.replace("Direct", "直接访问")
    value = value.replace("Organic Social", "自然社交")
    value = value.replace("Referral", "外链推荐")
    value = value.replace("Unrecognized UTM", "未识别 UTM")
    return value


def display_label_text(label: str) -> str:
    value = html.unescape(label).strip()
    value = DISPLAY_UNIT_RE.sub("", value)
    value = re.sub(r"\s+", " ", value)
    value = DISPLAY_LABEL_RENAMES.get(value, value)
    value = (
        value.replace("试订阅", "freetrial")
        .replace("试用订阅", "freetrial")
        .replace("新增试用用户", "新增freetrial")
        .replace("新增试用", "新增 freetrial")
        .replace("试用用户", "freetrial 用户")
        .replace("试用中", "freetrial 中")
        .replace("试用期", "freetrial 期")
        .replace("试用", "freetrial")
    )
    return value.strip()


def definition_for(label: str, sheet_title: str, row_kind: str) -> str | None:
    normalized = normalize_label_for_definition(label)
    if not normalized:
        return None
    if row_kind == "header":
        return FIELD_DEFINITIONS.get(normalized, "表头字段，用于说明后续数据列。")
    if row_kind == "section":
        return f"分区标题：{normalized}。"

    direct = FIELD_DEFINITIONS.get(normalized)
    if direct:
        return direct

    if sheet_title == "附录2 个人看板反馈":
        return "反馈对象、反馈主题或反馈明细。"
    if sheet_title == "附录1_口径裁决":
        return "口径裁决表字段，用于记录指标定义和最终取舍。"

    if normalized.startswith("总对话段"):
        return "按 30 分钟无对话活动切分后的 chat session 总数。"

    if normalized.startswith("/api/") or normalized.startswith("/auth") or normalized.startswith("/internal"):
        if "P95" in normalized:
            return f"{normalized.replace(' P95', '')} 的 P95 响应时间。"
        if "P99" in normalized:
            return f"{normalized.replace(' P99', '')} 的 P99 响应时间。"
        if "成功率" in normalized:
            return f"{normalized.replace(' 成功率', '')} 的成功请求 / 全部请求。"
        if "调用量" in normalized:
            return f"{normalized.replace(' 调用量', '')} 的请求量。"

    if normalized.startswith("失败："):
        return f"失败原因「{normalized.removeprefix('失败：')}」对应的次数。"
    if re.match(r"^(断连原因|履约失败原因|慢请求路径)第\s*\d+", normalized):
        return "按发生次数或影响排序的原因/路径占位，具体名称来自对应数据源。"
    if re.search(r"P(50|75|90|95|99)$", normalized):
        return f"{normalized} 对应分位数。"
    if re.search(r"P(50|75|90|95|99)\s*时长$", normalized):
        return f"{normalized}，按对应起止事件计算的分位用时。"
    if "CTA CTR" in normalized:
        return f"{normalized}，CTA 点击数 / 对应访问分母。"
    if "CTA 点击" in normalized:
        return f"{normalized}，按对应页面或人群统计 CTA 点击。"
    if "扣款" in normalized and normalized.endswith("用户"):
        return f"{normalized}，按 Stripe 扣款状态统计的去重用户数。"
    if normalized.endswith("占比"):
        return f"{normalized}，对应分组数量 / 所在分组总量。"
    if normalized.endswith("成功率"):
        return f"{normalized.removesuffix('成功率')}成功数 / 对应总数。"
    if normalized.endswith("失败率"):
        return f"{normalized.removesuffix('失败率')}失败数 / 对应总数。"
    if normalized.endswith("数"):
        return f"{normalized}，按所在页面口径统计的数量。"

    return None


def sheet_model(ws) -> dict:
    max_row, max_col = used_bounds(ws)
    max_col = min(max_col, 13)
    rows = []
    stack: dict[int, str] = {}
    stack_metric_kind: dict[int, str] = {}
    active_child_parent: dict[int, str] = {}
    active_child_kind: dict[int, str] = {}
    child_counts: dict[str, int] = {}
    row_labels: dict[str, str] = {}

    for r in range(1, max_row + 1):
        cells = [ws.cell(r, c) for c in range(1, max_col + 1)]
        raw_label = raw_cell_text(cells[0])
        raw_rest_probe = [raw_cell_text(c) for c in cells[1:]]
        if not raw_label:
            continue
        if not any(not is_missing_text(v) for v in [raw_label] + raw_rest_probe):
            continue

        label, had_marker = strip_marker(raw_label)
        outline = ws.row_dimensions[r].outlineLevel or 0
        level = outline if outline else (1 if had_marker else 0)
        probe_values = [label] + raw_rest_probe
        kind = "header" if is_header_row(r, probe_values) else ("section" if is_section(probe_values) else ("child" if level else "row"))
        if kind == "child" and is_derived_metric_label(label):
            kind = "derived"
        row_id = f"{ws.title}-{r}".replace(" ", "_")
        parent = ""
        own_kind = metric_kind(label)
        display_level = level

        if kind == "derived":
            parent = ""
            inherited_kind = own_kind
            display_level = max(level, 1)
        elif kind == "child":
            if own_kind == "number" and active_child_parent.get(level):
                parent = active_child_parent[level]
                inherited_kind = active_child_kind.get(level, "number")
                display_level = level + 1
            else:
                parent = stack.get(max(level - 1, 0), "")
                inherited_kind = stack_metric_kind.get(max(level - 1, 0), "number")
        else:
            inherited_kind = "number"

        parent_label = row_labels.get(parent, "")
        display_label = derived_child_label(label, parent_label) if kind == "child" and parent_label and is_derived_metric_label(parent_label) else label
        kind_for_row = inherited_kind if kind == "child" and own_kind == "number" else own_kind
        rest_values = [clean_text(fmt_value(c.value, c.number_format, kind=kind_for_row)) for c in cells[1:]]
        values = [display_label] + rest_values

        if parent:
            child_counts[parent] = child_counts.get(parent, 0) + 1

        if kind in {"child", "derived"}:
            stack[level] = row_id
            stack_metric_kind[level] = kind_for_row
            if own_kind != "number":
                active_child_parent[level] = row_id
                active_child_kind[level] = kind_for_row
        else:
            stack[0] = row_id
            stack_metric_kind[0] = kind_for_row
            for k in list(stack):
                if k > 0:
                    del stack[k]
                    stack_metric_kind.pop(k, None)
            active_child_parent.clear()
            active_child_kind.clear()

        comments = [c.comment.text.strip() for c in cells if c.comment and c.comment.text]
        definition = definition_for(label, ws.title, kind)
        if not definition:
            MISSING_DEFINITIONS.append((ws.title, label))
            definition = "待定义：请补充这个字段的业务定义和分母。"
        rows.append(
            {
                "id": row_id,
                "parent": parent,
                "level": display_level,
                "kind": kind,
                "ratio": kind in {"child", "derived"} and is_ratio_label(label),
                "value_kind": kind_for_row,
                "values": values,
                "comments": comments,
                "definition": definition,
            }
        )
        row_labels[row_id] = label

    for row in rows:
        row["has_children"] = child_counts.get(row["id"], 0) > 0
        row["hidden"] = row["kind"] == "child"
    row_by_id = {row["id"]: row for row in rows}
    for row in rows:
        parent_row = row_by_id.get(row.get("parent", ""))
        if not parent_row:
            continue
        if row["kind"] != "child":
            continue
        if parent_row["kind"] == "derived":
            continue
        if row.get("ratio") or row.get("value_kind") not in {"number", "currency"}:
            continue
        if parent_row.get("value_kind") not in {"number", "currency"}:
            continue
        if not is_share_breakdown_child(row["values"][0], parent_row["values"][0]):
            continue
        values = row["values"][:]
        for idx in range(1, len(values)):
            child_number = parse_display_number(values[idx])
            parent_number = parse_display_number(parent_row["values"][idx]) if idx < len(parent_row["values"]) else None
            values[idx] = append_child_share(values[idx], child_number, parent_number)
        row["values"] = values
    if not ws.title.startswith("附录"):
        rows = filter_non_quantitative_metric_rows(rows)
        rows = filter_empty_rows(rows)
        rows = remove_or_promote_empty_rows(rows)
    else:
        # Keep single-column marker rows (e.g., "高频诉求") so appendix segmentation works.
        pass
    rows = normalize_sections(ws.title, rows)
    if not ws.title.startswith("附录"):
        rows = normalize_drilldown_relationships(rows)
    if ws.title == "Dashboard":
        rows = prune_dashboard_summary(rows)
    if ws.title == "用户激活与转化":
        rows = restructure_activation(rows)
    if ws.title == "工程质量":
        rows = restructure_engineering(rows)
    if not ws.title.startswith("附录"):
        rows = remove_single_child_share_annotations(rows)
    return {"title": ws.title, "rows": rows, "max_col": max_col}


def workbook_model() -> list[dict]:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sheets = [sheet_model(wb[name]) for name in ORDER if name in wb.sheetnames]
    return apply_operating_schema(sheets)


def display_sheet_title(title: str) -> str:
    return SHEET_DISPLAY_NAMES.get(title, title)


def blank_values() -> list[str]:
    return ["—"] * METRIC_VALUE_COUNT


METRIC_DAYS = [
    "2026-05-15",
    "2026-05-16",
    "2026-05-17",
    "2026-05-18",
    "2026-05-19",
    "2026-05-20",
    "2026-05-21",
    "2026-05-22",
    "2026-05-23",
    "2026-05-24",
    "2026-05-25",
]
SOURCE_METRIC_DAYS = [
    "2026-05-15",
    "2026-05-16",
    "2026-05-17",
    "2026-05-18",
    "2026-05-19",
    "2026-05-20",
    "2026-05-21",
    "2026-05-22",
]
METRIC_WOW_REF = "2026-05-18"
METRIC_MOM_REF = "2026-04-25"
METRIC_VALUE_COUNT = len(METRIC_DAYS) + 4
DATE_COL_WIDTH = 76
TODAY_COL_WIDTH = 108
COMP_COL_WIDTH = 128
WOW_COL_WIDTH = 78
MOM_COL_WIDTH = 78
LEFT_COL_WIDTH = 240
TODAY_GRID_NTH = len(METRIC_DAYS)
TODAY_TABLE_NTH = len(METRIC_DAYS) + 1
DEFAULT_FIRST_VISIBLE_DATE = (dt.date.fromisoformat(METRIC_DAYS[-1]) - dt.timedelta(days=7)).isoformat()
DEFAULT_FIRST_VISIBLE_INDEX = (
    METRIC_DAYS.index(DEFAULT_FIRST_VISIBLE_DATE)
    if DEFAULT_FIRST_VISIBLE_DATE in METRIC_DAYS
    else max(0, len(METRIC_DAYS) - 8)
)
DEFAULT_DATE_SCROLL_LEFT = DEFAULT_FIRST_VISIBLE_INDEX * DATE_COL_WIDTH
RIGHT_GRID_TEMPLATE = (
    f"repeat({len(METRIC_DAYS) - 1}, {DATE_COL_WIDTH}px) "
    f"{TODAY_COL_WIDTH}px {COMP_COL_WIDTH}px {WOW_COL_WIDTH}px {COMP_COL_WIDTH}px {MOM_COL_WIDTH}px"
)
RIGHT_INNER_WIDTH = (
    (len(METRIC_DAYS) - 1) * DATE_COL_WIDTH
    + TODAY_COL_WIDTH
    + COMP_COL_WIDTH
    + WOW_COL_WIDTH
    + COMP_COL_WIDTH
    + MOM_COL_WIDTH
)
MAIN_TABLE_WIDTH = LEFT_COL_WIDTH + RIGHT_INNER_WIDTH
RIGHT_VIEWPORT_WIDTH = RIGHT_INNER_WIDTH - DEFAULT_DATE_SCROLL_LEFT
MAIN_TABLE_VIEWPORT_WIDTH = LEFT_COL_WIDTH + RIGHT_VIEWPORT_WIDTH
_EXTRA_CSV_CACHE: dict[str, list[dict[str, str]]] = {}


def normalize_metric_values(values: list[str] | None) -> list[str]:
    raw = list(values or [])
    if len(raw) == METRIC_VALUE_COUNT:
        return raw
    if len(raw) == len(SOURCE_METRIC_DAYS) + 4 and len(METRIC_DAYS) >= len(SOURCE_METRIC_DAYS):
        date_values = raw[: len(SOURCE_METRIC_DAYS)]
        return date_values + ["—"] * (len(METRIC_DAYS) - len(SOURCE_METRIC_DAYS)) + ["—"] * 4
    if len(raw) < METRIC_VALUE_COUNT:
        return raw + ["—"] * (METRIC_VALUE_COUNT - len(raw))
    return raw[:METRIC_VALUE_COUNT]


def schema_header_row(sheet_title: str) -> dict:
    labels = [
        "指标",
        *[day[5:] for day in METRIC_DAYS[:-1]],
        f"当日 {METRIC_DAYS[-1][5:]}",
        f"上周同日 {METRIC_WOW_REF[5:]}",
        "WoW %",
        f"上月同日 {METRIC_MOM_REF[5:]}",
        "MoM %",
    ]
    return {
        "id": f"{sheet_title}-schema-header".replace(" ", "_"),
        "parent": "",
        "level": 0,
        "kind": "header",
        "ratio": False,
        "value_kind": "number",
        "values": labels,
        "comments": [],
        "definition": "",
        "has_children": False,
        "hidden": False,
    }


def extra_csv_rows(filename: str) -> list[dict[str, str]]:
    if filename not in _EXTRA_CSV_CACHE:
        path = GENERATED_METRICS / filename
        if not path.exists():
            match = re.match(r"^(.+)_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.csv$", filename)
            candidates = sorted(GENERATED_METRICS.glob(f"{match.group(1)}_*.csv")) if match else []
            path = candidates[-1] if candidates else path
        if not path.exists():
            _EXTRA_CSV_CACHE[filename] = []
        else:
            with path.open(encoding="utf-8-sig", newline="") as handle:
                _EXTRA_CSV_CACHE[filename] = list(csv.DictReader(handle))
    return _EXTRA_CSV_CACHE[filename]


def numeric_raw(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or is_missing_text(text):
        return None
    try:
        return float(text.replace(",", "").replace("$", "").replace("%", ""))
    except ValueError:
        return None


def fmt_extra_value(value: float | None, *, kind: str = "number", decimals: int = 1) -> str:
    if value is None:
        return "—"
    if kind == "percent":
        return f"{value * 100:.1f}%"
    if kind == "currency":
        return fmt_number(value, decimals=decimals, currency=True)
    return fmt_number(value, decimals=decimals)


def extra_series_from_daily(daily: dict[str, float | None], *, kind: str = "number", decimals: int = 1) -> list[str]:
    values = [fmt_extra_value(daily.get(day), kind=kind, decimals=decimals) for day in METRIC_DAYS]
    current = daily.get(METRIC_DAYS[-1])
    wow_ref = daily.get(METRIC_WOW_REF)
    mom_ref = daily.get(METRIC_MOM_REF)
    values.append(fmt_extra_value(wow_ref, kind=kind, decimals=decimals))
    if current is not None and wow_ref not in (None, 0):
        values.append(f"{(current / wow_ref - 1) * 100:.1f}%")
    else:
        values.append("—")
    values.append(fmt_extra_value(mom_ref, kind=kind, decimals=decimals))
    if current is not None and mom_ref not in (None, 0):
        values.append(f"{(current / mom_ref - 1) * 100:.1f}%")
    else:
        values.append("—")
    return values


def extra_series(filename: str, column: str, *, kind: str = "number", decimals: int = 1, date_column: str = "date") -> list[str]:
    daily = {
        str(row.get(date_column, ""))[:10]: numeric_raw(row.get(column))
        for row in extra_csv_rows(filename)
        if row.get(date_column)
    }
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def extra_event_series(filename: str, event: str, column: str = "events", *, kind: str = "number", decimals: int = 0) -> list[str]:
    daily: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        if row.get("event") != event:
            continue
        day = str(row.get("day", ""))[:10]
        value = numeric_raw(row.get(column))
        if day and value is not None:
            daily[day] += value
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def extra_group_series(
    filename: str,
    group_column: str,
    group_value: str,
    value_column: str = "value",
    *,
    kind: str = "number",
    decimals: int = 0,
) -> list[str]:
    daily: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        if row.get(group_column) != group_value:
            continue
        day = str(row.get("date", ""))[:10]
        value = numeric_raw(row.get(value_column))
        if day and value is not None:
            daily[day] += value
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def nonzero_groups(filename: str, group_column: str, *, value_column: str = "value") -> list[str]:
    totals: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        key = row.get(group_column)
        value = numeric_raw(row.get(value_column))
        if key and value is not None:
            totals[key] += value
    return [key for key, value in sorted(totals.items()) if value > 0]


def sum_value_series(*series: list[str], kind: str = "number", decimals: int = 0) -> list[str]:
    values: list[str] = []
    for columns in zip(*series):
        numbers = [parse_display_number(value) for value in columns]
        if any(number is not None for number in numbers):
            values.append(fmt_extra_value(sum(number or 0 for number in numbers), kind=kind, decimals=decimals))
        else:
            values.append("—")
    return values


def one_day_series(day: str, value: float | int | None, *, kind: str = "number", decimals: int = 0) -> list[str]:
    return extra_series_from_daily({day: None if value is None else float(value)}, kind=kind, decimals=decimals)


def row_lookup(sheets: list[dict]) -> list[tuple[str, dict, str]]:
    output: list[tuple[str, dict, str]] = []
    for sheet in sheets:
        by_id = {row["id"]: row for row in sheet["rows"]}
        for row in sheet["rows"]:
            if row["kind"] in {"header", "section"}:
                continue
            parent = by_id.get(row.get("parent", ""))
            parent_label = normalize_label_for_definition(parent["values"][0]) if parent else ""
            output.append((sheet["title"], row, parent_label))
    return output


def find_source_values(
    lookup: list[tuple[str, dict, str]],
    labels: list[str] | str,
    *,
    titles: list[str] | None = None,
    parent: str | None = None,
) -> list[str]:
    wanted = [labels] if isinstance(labels, str) else labels
    wanted_norm = {normalize_label_for_definition(label) for label in wanted}
    title_set = set(titles or [])
    parent_norm = normalize_label_for_definition(parent) if parent else None
    for title, row, parent_label in lookup:
        if title_set and title not in title_set:
            continue
        if parent_norm is not None and parent_label != parent_norm:
            continue
        if normalize_label_for_definition(row["values"][0]) in wanted_norm:
            return normalize_metric_values(row["values"][1:])
    for title, row, parent_label in lookup:
        if parent_norm is not None and parent_label != parent_norm:
            continue
        if normalize_label_for_definition(row["values"][0]) in wanted_norm:
            return normalize_metric_values(row["values"][1:])
    return blank_values()


def ratio_values(numerator: list[str], denominator: list[str]) -> list[str]:
    values: list[str] = []
    for n_value, d_value in zip(numerator, denominator):
        n = parse_display_number(n_value)
        d = parse_display_number(d_value)
        values.append(f"{n / d * 100:.1f}%" if n is not None and d not in (None, 0) else "—")
    return values


def zero_missing_by_denominator(values: list[str], denominator: list[str]) -> list[str]:
    output: list[str] = []
    for value, denominator_value in zip(values, denominator):
        d = parse_display_number(denominator_value)
        output.append("0" if parse_display_number(value) is None and d not in (None, 0) else value)
    return output


def division_values(numerator: list[str], denominator: list[str]) -> list[str]:
    values: list[str] = []
    for n_value, d_value in zip(numerator, denominator):
        n = parse_display_number(n_value)
        d = parse_display_number(d_value)
        values.append(fmt_number(n / d, decimals=1) if n is not None and d not in (None, 0) else "—")
    return values


def strip_share_values(values: list[str]) -> list[str]:
    return [strip_share_annotation(value) for value in values]


def metric_row(
    sheet_title: str,
    seq: int,
    label: str,
    values: list[str] | None,
    *,
    kind: str = "row",
    parent: str = "",
    level: int = 0,
    keep_empty: bool = False,
    ratio: bool | None = None,
) -> dict:
    if kind == "derived" and not parent and level == 0:
        level = 1
    definition = definition_for(label, sheet_title, kind) or f"{normalize_label_for_definition(label)}。"
    return {
        "id": f"{sheet_title}-schema-{seq}".replace(" ", "_"),
        "parent": parent,
        "level": level,
        "kind": kind,
        "ratio": (kind == "derived" or is_ratio_label(label)) if ratio is None else ratio,
        "value_kind": metric_kind(label),
        "values": [label] + normalize_metric_values(values),
        "comments": [],
        "definition": definition,
        "has_children": False,
        "hidden": kind == "child",
        "keep_empty": keep_empty,
    }


def finalize_schema_rows(rows: list[dict]) -> list[dict]:
    child_ids = {row["parent"] for row in rows if row.get("parent")}
    for row in rows:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return remove_single_child_share_annotations(rows)


def build_schema_sheet(source_sheet: dict, lookup: list[tuple[str, dict, str]], sections: list[dict]) -> dict:
    rows = [schema_header_row(source_sheet["title"])]
    seq = 0
    for section_idx, section in enumerate(sections, start=1):
        rows.append(make_section_row(source_sheet["title"], section_idx, section["title"]))
        for item in section["items"]:
            seq += 1
            label = item["label"]
            values = item.get("values")
            if values is None:
                values = find_source_values(
                    lookup,
                    item.get("source", label),
                    titles=item.get("titles", [source_sheet["title"]]),
                    parent=item.get("source_parent"),
                )
            row = metric_row(
                source_sheet["title"],
                seq,
                label,
                values,
                kind=item.get("kind", "row"),
                parent="",
                level=item.get("level", 0),
                keep_empty=item.get("keep_empty", False),
                ratio=item.get("ratio"),
            )
            rows.append(row)
            parent_id = row["id"]
            for child in item.get("children", []):
                seq += 1
                child_label = child["label"]
                child_values = child.get("values")
                if child_values is None:
                    child_values = find_source_values(
                        lookup,
                        child.get("source", child_label),
                        titles=child.get("titles", [source_sheet["title"]]),
                        parent=child.get("source_parent", label),
                    )
                child_level = child.get("level", row.get("level", 0) + 1)
                child_row = metric_row(
                    source_sheet["title"],
                    seq,
                    child_label,
                    child_values,
                    kind="child",
                    parent=parent_id,
                    level=child_level,
                    keep_empty=child.get("keep_empty", item.get("keep_empty_children", False)),
                    ratio=child.get("ratio"),
                )
                rows.append(child_row)
                child_parent_id = child_row["id"]
                for grandchild in child.get("children", []):
                    seq += 1
                    grand_values = grandchild.get("values")
                    if grand_values is None:
                        grand_values = find_source_values(
                            lookup,
                            grandchild.get("source", grandchild["label"]),
                            titles=grandchild.get("titles", [source_sheet["title"]]),
                            parent=grandchild.get("source_parent", child_label),
                        )
                    rows.append(
                        metric_row(
                            source_sheet["title"],
                            seq,
                            grandchild["label"],
                            grand_values,
                            kind="child",
                            parent=child_parent_id,
                            level=grandchild.get("level", child_row.get("level", 0) + 1),
                            keep_empty=grandchild.get("keep_empty", child.get("keep_empty_children", item.get("keep_empty_children", False))),
                            ratio=grandchild.get("ratio"),
                        )
                    )
    rows = finalize_schema_rows(rows)
    rows = filter_empty_rows(rows)
    return {**source_sheet, "rows": rows}


def apply_operating_schema(sheets: list[dict]) -> list[dict]:
    lookup = row_lookup(sheets)
    by_title = {sheet["title"]: sheet for sheet in sheets}
    if "用户获取" in by_title:
        by_title["用户获取"] = build_acquisition_schema(by_title["用户获取"], lookup)
    if "用户激活与转化" in by_title:
        by_title["用户激活与转化"] = build_activation_schema(by_title["用户激活与转化"], lookup)
    if "用户留存与流失" in by_title:
        activity_sheet = {**by_title["用户留存与流失"], "title": "用户活跃与使用分布"}
        by_title["用户活跃与使用分布"] = build_activity_usage_schema(activity_sheet, lookup)
        by_title["用户留存与流失"] = build_retention_schema(by_title["用户留存与流失"], lookup)
    if "Agent 质量" in by_title:
        by_title["Agent 质量"] = build_agent_schema(by_title["Agent 质量"], lookup)
    if "工程质量" in by_title:
        by_title["工程质量"] = build_engineering_schema(by_title["工程质量"], lookup)
    if "财务" in by_title:
        by_title["财务"] = build_finance_schema(by_title["财务"], lookup)
    if "Dashboard" in by_title:
        schema_lookup = row_lookup([by_title[title] for title in ORDER if title in by_title])
        by_title["Dashboard"] = build_dashboard_home_schema(by_title["Dashboard"], schema_lookup)
    return [by_title[title] for title in ORDER if title in by_title]


def source_values(lookup: list[tuple[str, dict, str]], label: list[str] | str, *, titles: list[str] | None = None, parent: str | None = None) -> list[str]:
    return find_source_values(lookup, label, titles=titles, parent=parent)


def channel_children(parent_label: str, *, include_campaign: bool = False) -> list[dict]:
    children = [
        {"label": "Direct", "source": "直接访问", "source_parent": parent_label},
        {"label": "SEO / Organic Search", "source": "SEO / 自然搜索", "source_parent": parent_label},
        {"label": "Organic Social", "source": "自然社交", "source_parent": parent_label},
        {"label": "Referral", "source": "外链推荐", "source_parent": parent_label},
        {
            "label": "Google Ads",
            "source": "Google Ads",
            "source_parent": parent_label,
            "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 2}] if include_campaign else [],
        },
        {"label": "Meta Ads", "source": "Meta Ads", "source_parent": parent_label},
        {"label": "KOC / KOL（待映射）", "values": blank_values()},
    ]
    return children


def value_by_channel(lookup: list[tuple[str, dict, str]], parent_label: str, channel_label: str) -> list[str]:
    source_label = {
        "Direct": "直接访问",
        "SEO / Organic Search": "SEO / 自然搜索",
        "Organic Social": "自然社交",
        "Referral": "外链推荐",
    }.get(channel_label, channel_label)
    return source_values(lookup, source_label, titles=["用户获取"], parent=parent_label)


def channel_ratio_children(
    lookup: list[tuple[str, dict, str]],
    numerator_parent: str,
    denominator_parent: str,
    *,
    include_campaign: bool = True,
) -> list[dict]:
    # Do not divide separate attribution rows. Channel-level funnel rates need
    # same-user cohort joins; until that exists, keep the drilldown structure
    # but leave values empty.
    return blank_channel_children(include_campaign=include_campaign, keep_empty=True)


def blank_channel_children(*, include_campaign: bool = True, keep_empty: bool = False) -> list[dict]:
    output: list[dict] = []
    for label in ["Direct", "SEO / Organic Search", "Organic Social", "Referral", "Google Ads", "Meta Ads", "KOC / KOL（待映射）"]:
        child = {"label": label, "values": blank_values(), "keep_empty": keep_empty}
        if label == "Google Ads" and include_campaign:
            child["children"] = [{"label": "campaign（待数据）", "values": blank_values(), "level": 3, "keep_empty": keep_empty}]
        output.append(child)
    return output


def onboarding_path_children() -> list[dict]:
    return [
        {
            "label": "onboarding 页面",
            "values": blank_values(),
            "keep_empty": True,
            "keep_empty_children": True,
            "children": [
                {"label": "/start", "values": blank_values(), "level": 3, "keep_empty": True},
                {"label": "/usecase", "values": blank_values(), "level": 3, "keep_empty": True},
            ],
        },
        {"label": "首页", "values": blank_values(), "keep_empty": True},
    ]


def paid_efficiency_children(label: str) -> list[dict]:
    return [
        {
            "label": "Google Ads",
            "source": f"Google Ads {normalize_label_for_definition(label)}",
            "source_parent": normalize_label_for_definition(label),
            "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 3}],
        },
        {"label": "Meta Ads（待数据）", "values": blank_values()},
        {"label": "KOC / KOL（待数据）", "values": blank_values()},
    ]


def add_series(*series: list[str], kind: str = "number", decimals: int = 0) -> list[str]:
    return sum_value_series(*series, kind=kind, decimals=decimals)


def subtract_series(left: list[str], right: list[str], *, kind: str = "number", decimals: int = 0) -> list[str]:
    output: list[str] = []
    for left_value, right_value in zip(left, right):
        left_number = parse_display_number(left_value)
        right_number = parse_display_number(right_value)
        if left_number is None and right_number is None:
            output.append("—")
            continue
        output.append(fmt_extra_value((left_number or 0) - (right_number or 0), kind=kind, decimals=decimals))
    return output


def currency_division(numerator: list[str], denominator: list[str], *, factor: float = 1) -> list[str]:
    output: list[str] = []
    for numerator_value, denominator_value in zip(numerator, denominator):
        numerator_number = parse_display_number(numerator_value)
        denominator_number = parse_display_number(denominator_value)
        if numerator_number is None or denominator_number in (None, 0):
            output.append("—")
            continue
        output.append(fmt_extra_value(numerator_number / denominator_number * factor, kind="currency", decimals=1))
    return output


def values_with_share(child: list[str], parent: list[str]) -> list[str]:
    output: list[str] = []
    for child_value, parent_value in zip(child, parent):
        child_number = parse_display_number(child_value)
        parent_number = parse_display_number(parent_value)
        output.append(append_child_share(strip_share_annotation(child_value), child_number, parent_number))
    return output


def build_dashboard_home_schema(source_sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    task_csv = "posthog_task_proxy_2026-05-15_2026-05-26.csv"
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    entry_csv = "grafana_entry_security_daily_2026-05-15_2026-05-26.csv"
    sandbox_csv = "grafana_sandbox_fleet_daily_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"
    new_uv = extra_series(web_core_csv, "new_uv", decimals=0)
    google_ads_uv = source_values(lookup, "Google Ads", titles=["用户获取"], parent="新增 UV")
    meta_ads_uv = source_values(lookup, "Meta Ads", titles=["用户获取"], parent="新增 UV")
    koc_uv = source_values(lookup, "KOC / KOL（待映射）", titles=["用户获取"], parent="新增 UV")
    paid_uv = add_series(google_ads_uv, meta_ads_uv, koc_uv)
    organic_uv = subtract_series(new_uv, paid_uv)

    ad_spend = source_values(lookup, "广告花费 (美元)", titles=["用户获取"])
    paid_trial_users = add_series(
        source_values(lookup, "Google Ads", titles=["用户获取"], parent="新增试订阅用户"),
        source_values(lookup, "Meta Ads", titles=["用户获取"], parent="新增试订阅用户"),
        source_values(lookup, "KOC / KOL（待映射）", titles=["用户获取"], parent="新增试订阅用户"),
    )

    new_reg = extra_series(web_core_csv, "login_completed_users", decimals=0)
    new_trial = extra_series(checkout_csv, "fulfilled_subscription", decimals=0)
    new_dau = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    new_paid_users = extra_series(finance_csv, "new_paid_users", decimals=0)

    dau = extra_series(retention_csv, "dau", decimals=0)
    returning_dau = subtract_series(dau, new_dau)
    wau = extra_series(retention_csv, "wau", decimals=0)

    messages = extra_series(agent_csv, "messages", decimals=0)
    api_calls = source_values(lookup, "API 请求", titles=["工程质量"])
    api_fail = source_values(lookup, "API 失败", titles=["工程质量"])
    chat_stream_errored = source_values(lookup, "对话流中断", titles=["工程质量"])
    env_init_count = source_values(lookup, "沙盒启动次数", titles=["工程质量"])
    env_init_failed = source_values(lookup, "沙盒启动失败", titles=["工程质量"])
    upload_started = extra_series(file_csv, "upload_started", decimals=0)
    upload_failed = extra_series(file_csv, "upload_failed", decimals=0)
    ttfb_p95 = extra_series(grafana_csv, "ttfb_p95_s", decimals=1)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    entry_5xx_rate = extra_series(entry_csv, "entry_5xx_rate", kind="percent")
    sandbox_unreachable_rate = extra_series(sandbox_csv, "sandbox_unreachable_rate", kind="percent")
    checkout_started_users = extra_series(checkout_csv, "checkout_started_users", decimals=0)
    checkout_failed = add_series(
        extra_series(checkout_csv, "checkout_start_failed", decimals=0),
        extra_series(checkout_csv, "checkout_verify_failed", decimals=0),
        extra_series(checkout_csv, "checkout_fulfillment_failed", decimals=0),
    )
    inference_credits = extra_series(grafana_csv, "credits_consumed", decimals=0)
    cash_net = extra_series(finance_csv, "cash_net", kind="currency", decimals=0)
    refund = extra_series(finance_csv, "cash_refund", kind="currency", decimals=0)
    mrr = extra_series(finance_csv, "mrr_active", kind="currency", decimals=0)
    active_subs = extra_series(finance_csv, "status_active", decimals=0)
    past_due_subs = extra_series(finance_csv, "status_past_due", decimals=0)

    sections = [
        {
            "title": "新用户链路",
            "items": [
                {
                    "label": "新增 UV",
                    "values": new_uv,
                    "children": [
                        {
                            "label": "付费渠道新增 UV",
                            "values": values_with_share(paid_uv, new_uv),
                            "children": [
                                {"label": "广告花费 (美元)", "values": ad_spend, "level": 2},
                                {"label": "CPA (美元)", "values": currency_division(ad_spend, paid_trial_users), "level": 2},
                                {"label": "CAC（待数据）", "values": blank_values(), "level": 2},
                            ],
                        },
                        {"label": "自然渠道新增 UV", "values": values_with_share(organic_uv, new_uv)},
                    ],
                },
                {"label": "注册完成", "values": new_reg},
                {"label": "注册率 (百分比)", "kind": "derived", "values": ratio_values(new_reg, new_uv)},
                {"label": "新增freetrial", "values": new_trial},
                {"label": "freetrial 率 (百分比)", "kind": "derived", "values": ratio_values(new_trial, new_reg)},
                {"label": "新增 DAU", "values": new_dau},
                {"label": "开口率 (百分比)", "kind": "derived", "values": ratio_values(new_dau, new_uv)},
                {"label": "新增任务完成用户", "values": blank_values(), "keep_empty": True},
                {"label": "任务完成率 (百分比)", "kind": "derived", "values": blank_values(), "keep_empty": True},
                {"label": "新增付费用户", "values": new_paid_users},
            ],
        },
        {
            "title": "活跃与使用",
            "items": [
                {
                    "label": "DAU",
                    "values": dau,
                    "children": [
                        {"label": "新增 DAU", "values": new_dau},
                        {"label": "回访 DAU", "values": returning_dau},
                    ],
                },
                {"label": "WAU", "values": wau},
                {"label": "DAU / WAU", "kind": "derived", "values": ratio_values(dau, wau)},
                {"label": "D1 留存率 (百分比)", "kind": "derived", "values": source_values(lookup, "D1 留存率 (百分比)", titles=["用户留存与流失"])},
                {"label": "D7 留存率 (百分比)", "kind": "derived", "values": source_values(lookup, "D7 留存率 (百分比)", titles=["用户留存与流失"])},
            ],
        },
        {
            "title": "Agent 使用",
            "items": [
                {"label": "消息数", "values": messages},
                {"label": "人均消息数", "kind": "derived", "values": division_values(messages, dau)},
                {"label": "活跃任务数", "values": blank_values(), "keep_empty": True},
                {"label": "人均任务数", "kind": "derived", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "工程质量",
            "items": [
                {"label": "入口服务 5xx 率 (百分比)", "values": entry_5xx_rate, "ratio": False},
                {"label": "AI Gateway 错误率 (百分比)", "values": ratio_values(gateway_errors, gateway_requests), "ratio": False},
                {"label": "LLM 首 token P95", "values": ttfb_p95},
                {"label": "沙盒不可达率 (百分比)", "values": sandbox_unreachable_rate, "ratio": False},
                {"label": "附件上传失败率 (百分比)", "values": ratio_values(upload_failed, upload_started), "ratio": False},
                {"label": "结账失败率 (百分比)", "values": ratio_values(checkout_failed, checkout_started_users), "ratio": False},
            ],
        },
        {
            "title": "业务经营",
            "items": [
                {"label": "月经常性收入", "values": mrr},
                {
                    "label": "净现金收入",
                    "values": cash_net,
                    "children": [
                        {"label": "退款", "values": refund},
                    ],
                },
                {
                    "label": "付费有效订阅数",
                    "values": active_subs,
                    "children": [
                        {"label": "扣款失败订阅", "values": past_due_subs},
                    ],
                },
                {"label": "推理成本 credits", "values": inference_credits},
                {
                    "label": "总成本（待数据）",
                    "values": blank_values(),
                    "keep_empty": True,
                    "children": [
                        {"label": "云服务成本（待数据）", "values": blank_values(), "keep_empty": True},
                    ],
                },
                {"label": "毛利（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "毛利率（待数据）", "kind": "derived", "values": blank_values(), "keep_empty": True},
            ],
        },
    ]
    return build_schema_sheet(source_sheet, lookup, sections)


def build_acquisition_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    titles = ["用户获取"]
    new_uv = extra_series(web_core_csv, "new_uv", decimals=0)
    new_reg = extra_series(web_core_csv, "login_completed_users", decimals=0)
    new_trial = extra_series(checkout_csv, "fulfilled_subscription", decimals=0)
    new_dau = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    new_paid = extra_series(finance_csv, "new_paid_users", decimals=0)
    sections = [
        {
            "title": "投放",
            "items": [
                {
                    "label": "广告花费 (美元)",
                    "children": [
                        {
                            "label": "Google Ads",
                            "source": "Google Ads",
                            "source_parent": "广告花费 (美元)",
                            "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 2}],
                        },
                        {"label": "Meta Ads（待数据）", "values": blank_values()},
                        {"label": "KOC / KOL（待成本与映射）", "values": blank_values()},
                    ],
                },
                {"label": "付费曝光 (个)"},
                {"label": "CPM (美元)", "kind": "derived", "children": paid_efficiency_children("CPM")},
                {"label": "付费点击 (个)"},
                {"label": "CTR (百分比)", "kind": "derived", "children": paid_efficiency_children("CTR")},
                {"label": "CPC (美元)", "kind": "derived", "children": paid_efficiency_children("CPC")},
            ],
        },
        {
            "title": "新访数量",
            "items": [
                {"label": "新增 UV", "values": new_uv, "children": channel_children("新增 UV", include_campaign=True)},
            ],
        },
        {
            "title": "新访质量",
            "items": [
                {"label": "新增注册", "values": new_reg, "children": channel_children("新增注册用户", include_campaign=True)},
                {
                    "label": "注册率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(new_reg, new_uv),
                    "children": channel_ratio_children(lookup, "新增注册用户", "新增 UV"),
                },
                {"label": "新增freetrial", "values": new_trial, "children": channel_children("新增试用用户 (个)", include_campaign=True)},
                {
                    "label": "freetrial 率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(new_trial, new_reg),
                    "children": channel_ratio_children(lookup, "新增试用用户 (个)", "新增注册用户"),
                },
                {"label": "新增 DAU", "values": new_dau, "children": channel_children("新增 DAU", include_campaign=True)},
                {
                    "label": "开口率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(new_dau, new_uv),
                    "children": blank_channel_children(keep_empty=True),
                },
                {"label": "首日人均对话量", "kind": "derived", "values": division_values(extra_series(nextday_csv, "d0_messages", decimals=0), new_dau)},
                {"label": "前三日人均对话量", "kind": "derived", "values": division_values(extra_series(nextday_csv, "d0_d2_messages", decimals=0), new_dau)},
                {"label": "新增付费用户", "values": new_paid},
                {
                    "label": "付费率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(new_paid, new_uv),
                    "children": channel_ratio_children(lookup, "新增付费用户", "新增 UV"),
                },
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_activation_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_flow_cohort_2026-05-15_2026-05-26.csv"
    checkout_funnel_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    titles = ["用户激活与转化", "用户获取"]
    uv = extra_series(web_core_csv, "uv", decimals=0)
    new_uv = extra_series(web_core_csv, "new_uv", decimals=0)
    login_cta = extra_series(web_core_csv, "landing_cta_clicks", decimals=0)
    reg_start = blank_values()
    reg_done = extra_series(web_core_csv, "login_completed_users", decimals=0)
    card_start = extra_series(checkout_funnel_csv, "checkout_started_users", decimals=0)
    trial = extra_series(checkout_funnel_csv, "fulfilled_subscription", decimals=0)
    new_dau = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    first_task_start = extra_series(nextday_csv, "first_task_start_users", decimals=0)
    first_task_done = extra_series(nextday_csv, "first_task_done_users", decimals=0)
    next_return = extra_series(nextday_csv, "next_return_users", decimals=0)
    next_session = extra_series(nextday_csv, "next_chat_users", decimals=0)
    next_task_start = extra_series(nextday_csv, "next_task_start_users", decimals=0)
    next_task_done = extra_series(nextday_csv, "next_task_done_users", decimals=0)
    checkout_start_failed = extra_series(checkout_csv, "has_start_failed", decimals=0)
    checkout_verify_failed = extra_series(checkout_csv, "has_verify_failed", decimals=0)
    checkout_fulfillment_failed = extra_series(checkout_csv, "has_fulfillment_failed", decimals=0)
    bind_failed = sum_value_series(
        checkout_start_failed,
        checkout_verify_failed,
        checkout_fulfillment_failed,
    )
    sections = [
        {
            "title": "访问到注册",
            "items": [
                {"label": "UV", "values": uv},
                {"label": "新增 UV", "values": new_uv},
                {"label": "登录 CTA 点击次数", "values": login_cta},
                {"label": "人均登录 CTA 点击", "kind": "derived", "values": division_values(login_cta, uv), "children": onboarding_path_children()},
                {"label": "注册开始", "values": reg_start},
                {"label": "注册完成", "values": reg_done},
                {"label": "注册完成率 (百分比)", "kind": "derived", "values": ratio_values(reg_done, reg_start), "children": onboarding_path_children()},
                {
                    "label": "注册率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(reg_done, new_uv),
                    "children": onboarding_path_children(),
                },
                {"label": "注册失败", "source": "登录失败 (个)", "children": [{"label": "失败原因（待结构化）", "values": blank_values()}]},
            ],
        },
        {
            "title": "注册到 freetrial",
            "items": [
                {"label": "绑卡开始", "source": "发起绑卡 (个)", "values": card_start},
                {"label": "新增试订阅用户", "source": ["新增试用用户", "新增试用用户 (个)"], "values": trial},
                {
                    "label": "freetrial 率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(trial, reg_done),
                    "children": onboarding_path_children(),
                },
                {
                    "label": "绑卡完成率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(trial, card_start),
                    "children": [
                        {
                            "label": "绑卡失败",
                            "values": bind_failed,
                            "level": 2,
                            "children": [
                                {"label": "结账发起失败", "values": checkout_start_failed, "level": 3},
                                {"label": "结账验证失败", "values": checkout_verify_failed, "level": 3},
                                {"label": "结账履约失败", "values": checkout_fulfillment_failed, "level": 3},
                            ],
                        }
                    ],
                },
            ],
        },
        {
            "title": "freetrial 到首次任务",
            "items": [
                {"label": "新增 DAU", "values": new_dau},
                {"label": "开口率 (百分比)", "kind": "derived", "values": ratio_values(new_dau, new_uv), "children": onboarding_path_children()},
                {"label": "首日人均对话量", "kind": "derived", "values": division_values(extra_series(nextday_csv, "d0_messages", decimals=0), new_dau)},
                {"label": "首次任务开始", "values": blank_values(), "keep_empty": True},
                {"label": "首次发起任务率 (百分比)", "kind": "derived", "values": blank_values(), "children": onboarding_path_children(), "keep_empty": True},
                {"label": "首次任务完成", "values": blank_values(), "keep_empty": True},
                {"label": "首次任务完成率 (百分比)", "kind": "derived", "values": blank_values(), "children": onboarding_path_children(), "keep_empty": True},
            ],
        },
        {
            "title": "次日回访",
            "items": [
                {"label": "次日回访用户数", "values": next_return},
                {"label": "次日回访率 (百分比)", "kind": "derived", "values": ratio_values(next_return, new_dau), "children": onboarding_path_children()},
                {"label": "次日发起会话用户数", "values": next_session},
                {"label": "次日回访发起会话率 (百分比)", "kind": "derived", "values": ratio_values(next_session, next_return), "children": onboarding_path_children()},
                {"label": "次日发起任务用户数", "values": blank_values(), "keep_empty": True},
                {"label": "次日回访发起任务率 (百分比)", "kind": "derived", "values": blank_values(), "children": onboarding_path_children(), "keep_empty": True},
                {"label": "次日完成任务用户数", "values": blank_values(), "keep_empty": True},
                {"label": "次日任务完成率 (百分比)", "kind": "derived", "values": blank_values(), "children": onboarding_path_children(), "keep_empty": True},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_activity_usage_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    connector_csv = "posthog_connector_events_2026-05-15_2026-05-26.csv"
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"

    dau = extra_series(retention_csv, "dau", decimals=0)
    wau = extra_series(retention_csv, "wau", decimals=0)
    mau = extra_series(retention_csv, "mau", decimals=0)
    new_dau = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    returning_dau = subtract_series(dau, new_dau)
    sessions = extra_series(agent_csv, "chat_sessions", decimals=0)
    messages = extra_series(agent_csv, "messages", decimals=0)
    single_message_sessions = source_values(lookup, "单消息对话数", titles=["Dashboard"])
    active_tasks = blank_values()

    upload_users = extra_series(file_csv, "upload_started_users", decimals=0)
    upload_count = extra_series(file_csv, "upload_started", decimals=0)
    upload_completed = extra_series(file_csv, "upload_completed", decimals=0)
    upload_failed = extra_series(file_csv, "upload_failed", decimals=0)
    attach_processed = extra_series(file_csv, "attach_processed", decimals=0)
    file_viewed = extra_series(file_csv, "file_viewed", decimals=0)

    tool_requests = extra_series(agent_csv, "requests_with_tool_use", decimals=0)
    tool_request_rate = extra_series(agent_csv, "tool_use_request_rate", kind="percent")
    tool_call_count = extra_series(agent_csv, "tool_use_count", decimals=0)
    trial_total = source_values(lookup, ["试用用户总数", "绑卡成功用户总数"], titles=["用户留存与流失"])

    connector_events = [
        extra_event_series(connector_csv, "connector:telegram_connected"),
        extra_event_series(connector_csv, "connector:google_connected"),
        extra_event_series(connector_csv, "connector:slack_connected"),
        extra_event_series(connector_csv, "connector:discord_connected"),
        extra_event_series(connector_csv, "connector:lark_connected"),
    ]
    connector_connected = sum_value_series(*connector_events)

    sections = [
        {
            "title": "活跃规模",
            "items": [
                {"label": "DAU", "values": dau},
                {"label": "WAU", "values": wau},
                {"label": "MAU", "values": mau},
                {"label": "DAU / WAU", "kind": "derived", "values": ratio_values(dau, wau)},
                {"label": "DAU / MAU", "kind": "derived", "values": ratio_values(dau, mau)},
                {"label": "WAU / MAU", "kind": "derived", "values": ratio_values(wau, mau)},
            ],
        },
        {
            "title": "活跃用户结构",
            "items": [
                {"label": "新增 DAU", "values": new_dau},
                {"label": "回访 DAU", "values": returning_dau},
            ],
        },
        {
            "title": "使用深度",
            "items": [
                {"label": "对话数", "values": sessions},
                {"label": "人均对话数", "kind": "derived", "values": division_values(sessions, dau)},
                {"label": "消息数", "values": messages},
                {"label": "人均消息数", "kind": "derived", "values": division_values(messages, dau)},
                {"label": "单消息对话数", "values": single_message_sessions},
                {"label": "会话时长 P50", "source": "对话时长 P50 (分钟)", "titles": ["Dashboard"]},
                {"label": "会话时长 P90", "source": "对话时长 P90 (分钟)", "titles": ["Dashboard"]},
            ],
        },
        {
            "title": "功能使用分布",
            "items": [
                {
                    "label": "附件上传用户数",
                    "values": upload_users,
                    "children": [
                        {"label": "附件上传次数", "values": upload_count},
                        {"label": "附件处理次数", "values": attach_processed},
                        {"label": "附件查看次数", "values": file_viewed},
                        {"label": "附件上传成功次数", "values": upload_completed},
                        {"label": "附件上传失败次数", "values": upload_failed},
                    ],
                },
                {
                    "label": "tool 使用请求数",
                    "values": tool_requests,
                    "children": [
                        {"label": "tool 调用次数", "values": tool_call_count},
                    ],
                },
                {"label": "tool 使用请求占比", "kind": "derived", "values": tool_request_rate},
                {"label": "connector 连接完成", "values": connector_connected},
                {"label": "活跃任务数", "values": active_tasks, "keep_empty": True},
                {"label": "人均任务数", "kind": "derived", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "freetrial 使用强度",
            "items": [
                {
                    "label": "freetrial 用户总数",
                    "values": trial_total,
                    "children": [
                        {"label": "重度日用", "source": "重度日用", "values": source_values(lookup, "重度日用", titles=["用户留存与流失"])},
                        {"label": "中度", "source": "中度", "values": source_values(lookup, "中度", titles=["用户留存与流失"])},
                        {"label": "轻度", "source": "轻度", "values": source_values(lookup, "轻度", titles=["用户留存与流失"])},
                        {"label": "低频", "source": "低频", "values": source_values(lookup, "低频", titles=["用户留存与流失"])},
                        {"label": "冷却", "source": "冷却", "values": source_values(lookup, "冷却", titles=["用户留存与流失"])},
                        {"label": "流失中", "source": "流失中", "values": source_values(lookup, "流失中", titles=["用户留存与流失"])},
                        {"label": "深度流失", "source": "深度流失", "values": source_values(lookup, "深度流失", titles=["用户留存与流失"])},
                    ],
                },
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_retention_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    dau = extra_series(retention_csv, "dau", decimals=0)
    wau = extra_series(retention_csv, "wau", decimals=0)
    new_churn = source_values(lookup, "新增流失用户", titles=["用户留存与流失"])
    d1_rate = source_values(lookup, "次日留存", titles=["Dashboard"])
    returning = source_values(lookup, "当周回流 (个)", titles=["用户留存与流失"])
    returning_active = source_values(lookup, "回流后 7 天再激活 (个)", titles=["用户留存与流失"])
    returning_churn = source_values(lookup, "回流后又流失 (个)", titles=["用户留存与流失"])
    sections = [
        {
            "title": "留存",
            "items": [
                {"label": "D1 留存率", "kind": "derived", "values": d1_rate},
            ],
        },
        {
            "title": "流失",
            "items": [
                {"label": "新增流失用户", "values": new_churn},
                {"label": "流失率", "kind": "derived", "values": ratio_values(new_churn, wau)},
            ],
        },
        {
            "title": "回流",
            "items": [
                {"label": "回流用户数", "source": "当周回流 (个)", "values": returning},
                {"label": "回流后活跃用户数", "source": "回流后 7 天再激活 (个)", "values": returning_active},
                {"label": "回流后活跃率 (百分比)", "kind": "derived", "values": ratio_values(returning_active, returning)},
                {"label": "回流后又流失", "source": "回流后又流失 (个)", "values": returning_churn},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_agent_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    connector_csv = "posthog_connector_events_2026-05-15_2026-05-26.csv"
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    grafana_errors_csv = "grafana_ai_gateway_errors_by_type_2026-05-15_2026-05-26.csv"
    grafana_provider_csv = "grafana_ai_gateway_requests_by_provider_2026-05-15_2026-05-26.csv"
    grafana_model_csv = "grafana_ai_gateway_requests_by_model_2026-05-15_2026-05-26.csv"

    dau = extra_series(retention_csv, "dau", decimals=0)
    chat_sessions = extra_series(agent_csv, "chat_sessions", decimals=0)
    messages = extra_series(agent_csv, "messages", decimals=0)
    active_tasks = blank_values()
    message_failed = extra_series(agent_csv, "message_failed", decimals=0)
    stream_errors = extra_series(agent_csv, "stream_errors", decimals=0)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    credits = extra_series(grafana_csv, "credits_consumed", decimals=0)
    input_tokens = extra_series(grafana_csv, "input_tokens", decimals=0)
    output_tokens = extra_series(grafana_csv, "output_tokens", decimals=0)
    cache_read_tokens = extra_series(agent_csv, "cache_read_tokens", decimals=0)
    tool_requests = extra_series(agent_csv, "requests_with_tool_use", decimals=0)
    tool_call_count = extra_series(agent_csv, "tool_use_count", decimals=0)
    tool_request_rate = extra_series(agent_csv, "tool_use_request_rate", kind="percent")
    gateway_error_children = [
        {"label": "流式中断", "values": extra_group_series(grafana_errors_csv, "error_type", "stream_abort", decimals=0)},
        {"label": "超时", "values": extra_group_series(grafana_errors_csv, "error_type", "timeout", decimals=0)},
        {"label": "上游错误", "values": extra_group_series(grafana_errors_csv, "error_type", "upstream_error", decimals=0)},
        {"label": "Bedrock 错误", "values": extra_group_series(grafana_errors_csv, "error_type", "bedrock_error", decimals=0)},
    ]
    provider_children = [
        {
            "label": provider,
            "values": extra_group_series(grafana_provider_csv, "provider", provider, decimals=0),
            "level": 2,
        }
        for provider in nonzero_groups(grafana_provider_csv, "provider")
    ]
    model_children = [
        {
            "label": model,
            "values": extra_group_series(grafana_model_csv, "model", model, decimals=0),
            "level": 2,
        }
        for model in nonzero_groups(grafana_model_csv, "model")
    ]
    telegram_connected = extra_event_series(connector_csv, "connector:telegram_connected")
    google_connected = extra_event_series(connector_csv, "connector:google_connected")
    slack_connected = extra_event_series(connector_csv, "connector:slack_connected")
    discord_connected = extra_event_series(connector_csv, "connector:discord_connected")
    lark_connected = extra_event_series(connector_csv, "connector:lark_connected")
    connector_connected = sum_value_series(
        telegram_connected,
        google_connected,
        slack_connected,
        discord_connected,
        lark_connected,
    )
    gw_set = extra_event_series(connector_csv, "connector:google_workspace_folder_set")
    gw_unavailable = extra_event_series(connector_csv, "connector:google_workspace_picker_unavailable")

    sections = [
        {
            "title": "用户调用",
            "items": [
                {"label": "对话数", "values": chat_sessions},
                {"label": "人均对话数", "kind": "derived", "values": division_values(chat_sessions, dau)},
                {"label": "消息数", "values": messages},
                {"label": "人均消息数", "kind": "derived", "values": division_values(messages, dau)},
                {"label": "单对话消息数", "kind": "derived", "values": division_values(messages, chat_sessions)},
                {"label": "活跃任务数", "values": active_tasks, "keep_empty": True},
                {"label": "人均任务数", "kind": "derived", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "Agent 响应",
            "items": [
                {
                    "label": "LLM 首 token",
                    "values": blank_values(),
                    "children": [
                        {"label": "首 token P50", "values": extra_series(grafana_csv, "ttfb_p50_s", decimals=1)},
                        {"label": "首 token P95", "values": extra_series(grafana_csv, "ttfb_p95_s", decimals=1)},
                    ],
                },
                {"label": "完整响应 P95", "values": extra_series(grafana_csv, "request_p95_s", decimals=1)},
                {"label": "完整响应 P99", "values": extra_series(grafana_csv, "request_p99_s", decimals=1)},
            ],
        },
        {
            "title": "Agent 失败",
            "items": [
                {"label": "消息失败数", "values": message_failed},
                {"label": "消息失败率 (百分比)", "kind": "derived", "values": ratio_values(message_failed, messages)},
                {"label": "流式错误数", "source": "stream_errors", "values": stream_errors},
                {
                    "label": "AI Gateway 错误",
                    "values": gateway_errors,
                    "children": gateway_error_children,
                },
                {"label": "AI Gateway 错误率 (百分比)", "kind": "derived", "values": ratio_values(gateway_errors, gateway_requests)},
            ],
        },
        {
            "title": "模型消耗",
            "items": [
                {
                    "label": "AI Gateway 请求",
                    "values": gateway_requests,
                    "children": [
                        {"label": "模型供应商请求", "values": blank_values(), "children": provider_children},
                        {"label": "模型请求", "values": blank_values(), "children": model_children},
                    ],
                },
                {"label": "credits", "values": credits},
                {"label": "人均 credits", "kind": "derived", "values": division_values(credits, dau)},
                {
                    "label": "tokens",
                    "values": blank_values(),
                    "children": [
                        {"label": "input tokens", "values": input_tokens},
                        {"label": "output tokens", "values": output_tokens},
                        {"label": "cache read tokens", "values": cache_read_tokens},
                        {"label": "cache hit ratio (百分比)", "kind": "derived", "values": ratio_values(cache_read_tokens, sum_value_series(cache_read_tokens, input_tokens))},
                    ],
                },
            ],
        },
        {
            "title": "能力使用",
            "items": [
                {
                    "label": "tool 使用请求数",
                    "values": tool_requests,
                    "children": [
                        {"label": "tool 调用次数", "values": tool_call_count},
                    ],
                },
                {"label": "tool 使用请求占比", "kind": "derived", "values": tool_request_rate},
                {
                    "label": "connector 连接完成",
                    "values": connector_connected,
                    "children": [
                        {"label": "Telegram connected", "values": telegram_connected},
                        {"label": "Google connected", "values": google_connected},
                        {"label": "Slack connected", "values": slack_connected},
                        {"label": "Discord connected", "values": discord_connected},
                        {"label": "Lark connected", "values": lark_connected},
                    ],
                },
                {"label": "Google Workspace 设置完成", "values": gw_set},
                {"label": "Google Workspace picker unavailable", "values": gw_unavailable},
                {"label": "Google Workspace 连接率 (百分比)", "kind": "derived", "values": ratio_values(gw_set, sum_value_series(gw_set, gw_unavailable))},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_engineering_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    engineering_csv = "posthog_engineering_quality_more_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    grafana_errors_csv = "grafana_ai_gateway_errors_by_type_2026-05-15_2026-05-26.csv"
    entry_csv = "grafana_entry_security_daily_2026-05-15_2026-05-26.csv"
    sandbox_csv = "grafana_sandbox_fleet_daily_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"

    api_calls = extra_series(engineering_csv, "api_calls", decimals=0)
    api_success = extra_series(engineering_csv, "api_success", decimals=0)
    api_fail = extra_series(engineering_csv, "api_fail", decimals=0)
    env_init_count = extra_series(engineering_csv, "env_init_count", decimals=0)
    env_init_failed = extra_series(engineering_csv, "env_init_failed", decimals=0)
    upload_started = extra_series(file_csv, "upload_started", decimals=0)
    upload_completed = extra_series(file_csv, "upload_completed", decimals=0)
    upload_failed = extra_series(file_csv, "upload_failed", decimals=0)
    entry_requests = extra_series(entry_csv, "entry_requests", decimals=0)
    entry_5xx = extra_series(entry_csv, "entry_5xx", decimals=0)
    target_5xx = extra_series(entry_csv, "target_5xx", decimals=0)
    elb_5xx = extra_series(entry_csv, "elb_5xx", decimals=0)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    sandbox_checked = extra_series(sandbox_csv, "sandbox_checked", decimals=0)
    sandbox_unreachable = extra_series(sandbox_csv, "sandbox_unreachable", decimals=0)
    checkout_started_users = extra_series(checkout_csv, "checkout_started_users", decimals=0)
    checkout_failed = sum_value_series(
        extra_series(checkout_csv, "checkout_start_failed", decimals=0),
        extra_series(checkout_csv, "checkout_verify_failed", decimals=0),
        extra_series(checkout_csv, "checkout_fulfillment_failed", decimals=0),
    )
    gateway_error_children = [
        {"label": "流式中断", "values": extra_group_series(grafana_errors_csv, "error_type", "stream_abort", decimals=0)},
        {"label": "超时", "values": extra_group_series(grafana_errors_csv, "error_type", "timeout", decimals=0)},
        {"label": "上游错误", "values": extra_group_series(grafana_errors_csv, "error_type", "upstream_error", decimals=0)},
        {"label": "Bedrock 错误", "values": extra_group_series(grafana_errors_csv, "error_type", "bedrock_error", decimals=0)},
    ]

    sections = [
        {
            "title": "入口",
            "items": [
                {"label": "入口请求", "values": entry_requests},
                {
                    "label": "入口 5xx",
                    "values": entry_5xx,
                    "children": [
                        {"label": "Target 5xx", "values": target_5xx},
                        {"label": "ELB 5xx", "values": elb_5xx},
                    ],
                },
                {"label": "入口服务 5xx 率 (百分比)", "kind": "derived", "values": ratio_values(entry_5xx, entry_requests)},
            ],
        },
        {
            "title": "HTTP 服务",
            "items": [
                {"label": "API 请求", "values": api_calls},
                {"label": "API 成功", "values": api_success},
                {"label": "API 失败", "values": api_fail},
                {"label": "API 成功率 (百分比)", "kind": "derived", "values": ratio_values(api_success, api_calls)},
                {
                    "label": "API 耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "API 耗时 P50", "values": extra_series(engineering_csv, "api_p50_ms", decimals=1)},
                        {"label": "API 耗时 P95", "values": extra_series(engineering_csv, "api_p95_ms", decimals=1)},
                        {"label": "API 耗时 P99", "values": extra_series(engineering_csv, "api_p99_ms", decimals=1)},
                    ],
                },
            ],
        },
        {
            "title": "实时连接",
            "items": [
                {"label": "实时连接错误", "values": extra_series(engineering_csv, "ws_error", decimals=0)},
                {"label": "消息发送失败", "values": extra_series(engineering_csv, "message_failed", decimals=0)},
                {"label": "对话流中断", "values": extra_series(engineering_csv, "chat_stream_errored", decimals=0)},
            ],
        },
        {
            "title": "AI Gateway",
            "items": [
                {"label": "AI Gateway 请求", "values": gateway_requests},
                {
                    "label": "AI Gateway 错误",
                    "values": gateway_errors,
                    "children": gateway_error_children,
                },
                {"label": "AI Gateway 错误率 (百分比)", "kind": "derived", "values": ratio_values(gateway_errors, gateway_requests)},
                {
                    "label": "LLM 首 token",
                    "values": blank_values(),
                    "children": [
                        {"label": "首 token P50", "values": extra_series(grafana_csv, "ttfb_p50_s", decimals=1)},
                        {"label": "首 token P95", "values": extra_series(grafana_csv, "ttfb_p95_s", decimals=1)},
                    ],
                },
                {
                    "label": "完整响应耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "完整响应 P95", "values": extra_series(grafana_csv, "request_p95_s", decimals=1)},
                        {"label": "完整响应 P99", "values": extra_series(grafana_csv, "request_p99_s", decimals=1)},
                    ],
                },
            ],
        },
        {
            "title": "沙盒",
            "items": [
                {"label": "沙盒检查次数", "values": sandbox_checked},
                {"label": "沙盒不可达", "values": sandbox_unreachable},
                {"label": "沙盒不可达率 (百分比)", "kind": "derived", "values": ratio_values(sandbox_unreachable, sandbox_checked)},
                {
                    "label": "沙盒启动耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "沙盒启动 P50", "values": extra_series(engineering_csv, "env_init_p50_ms", decimals=1)},
                        {"label": "沙盒启动 P95", "values": extra_series(engineering_csv, "env_init_p95_ms", decimals=1)},
                    ],
                },
                {"label": "沙盒重启失败", "values": extra_series(engineering_csv, "sandbox_restart_failed", decimals=0)},
            ],
        },
        {
            "title": "附件与支付",
            "items": [
                {"label": "附件上传开始", "values": upload_started},
                {"label": "附件上传成功", "values": upload_completed},
                {"label": "附件上传成功率 (百分比)", "kind": "derived", "values": ratio_values(upload_completed, upload_started)},
                {"label": "附件上传失败", "values": upload_failed},
                {"label": "附件上传失败率 (百分比)", "kind": "derived", "values": ratio_values(upload_failed, upload_started)},
                {"label": "发起结账用户", "values": checkout_started_users},
                {"label": "结账失败", "values": checkout_failed},
                {"label": "结账失败率 (百分比)", "kind": "derived", "values": ratio_values(checkout_failed, checkout_started_users)},
                {"label": "结账发起失败", "values": extra_series(engineering_csv, "checkout_start_failed", decimals=0)},
                {"label": "结账验证失败", "values": extra_series(engineering_csv, "checkout_verify_failed", decimals=0)},
                {"label": "结账履约失败", "values": extra_series(engineering_csv, "checkout_fulfillment_failed", decimals=0)},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_finance_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    funnel_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"
    flow_csv = "posthog_finance_checkout_flow_cohort_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    audit_rows = extra_csv_rows("calc_finance_stripe_trial_audit.csv")
    audit = audit_rows[0] if audit_rows else {}
    audit_day = str(audit.get("cash_day_utc") or METRIC_DAYS[-1])

    pricing_users = extra_series(funnel_csv, "pricing_opened_users", decimals=0)
    checkout_started_users = extra_series(funnel_csv, "checkout_started_users", decimals=0)
    session_created_flows = extra_series(flow_csv, "has_session_created", decimals=0)
    verified_flows = extra_series(flow_csv, "has_verified", decimals=0)
    fulfilled_flows = extra_series(flow_csv, "has_fulfilled", decimals=0)
    start_failed = extra_series(funnel_csv, "checkout_start_failed", decimals=0)
    verify_failed = extra_series(funnel_csv, "checkout_verify_failed", decimals=0)
    fulfillment_failed = extra_series(funnel_csv, "checkout_fulfillment_failed", decimals=0)
    checkout_failed = sum_value_series(start_failed, verify_failed, fulfillment_failed)
    active = extra_series(finance_csv, "status_active", decimals=0)
    trialing = extra_series(finance_csv, "status_trialing", decimals=0)
    past_due = extra_series(finance_csv, "status_past_due", decimals=0)
    mrr = extra_series(finance_csv, "mrr_active", kind="currency", decimals=0)
    new_trial = extra_series(funnel_csv, "fulfilled_subscription", decimals=0)

    d3_bind = one_day_series(audit_day, numeric_raw(audit.get("cohort_bind_success_unique_customers")), decimals=0)
    d3_paid = one_day_series(audit_day, numeric_raw(audit.get("cohort_succeeded_charge_customers")), decimals=0)
    d3_past_due = one_day_series(audit_day, numeric_raw(audit.get("cohort_past_due_customers")), decimals=0)
    d3_canceled = one_day_series(audit_day, numeric_raw(audit.get("cohort_canceled_customers")), decimals=0)

    sections = [
        {
            "title": "freetrial 规模",
            "items": [
                {"label": "新增freetrial", "values": new_trial},
                {"label": "查看价格用户", "values": pricing_users},
                {"label": "发起结账用户", "values": checkout_started_users},
                {"label": "价格页发起结账率 (百分比)", "kind": "derived", "values": ratio_values(checkout_started_users, pricing_users)},
            ],
        },
        {
            "title": "结账过程",
            "items": [
                {"label": "创建结账会话", "values": session_created_flows},
                {"label": "支付验证完成", "values": verified_flows},
                {"label": "结账验证率 (百分比)", "kind": "derived", "values": ratio_values(verified_flows, session_created_flows)},
                {"label": "订单完成", "values": fulfilled_flows},
                {"label": "订单完成率 (百分比)", "kind": "derived", "values": ratio_values(fulfilled_flows, session_created_flows)},
                {
                    "label": "结账失败",
                    "values": checkout_failed,
                    "children": [
                        {"label": "发起失败", "values": start_failed},
                        {"label": "验证失败", "values": verify_failed},
                        {"label": "履约失败", "values": fulfillment_failed},
                    ],
                },
                {"label": "订阅订单", "values": extra_series(flow_csv, "subscription_flows", decimals=0)},
                {"label": "点数包订单", "values": extra_series(flow_csv, "credit_pack_flows", decimals=0)},
            ],
        },
        {
            "title": "D+3 付费结果",
            "items": [
                {"label": "到期 freetrial 用户", "values": d3_bind},
                {"label": "D+3 新增付费用户", "values": d3_paid},
                {"label": "扣款成功率 (百分比)", "kind": "derived", "values": ratio_values(d3_paid, d3_bind)},
                {"label": "扣款失败用户", "values": d3_past_due},
                {"label": "扣款失败率 (百分比)", "kind": "derived", "values": ratio_values(d3_past_due, d3_bind)},
                {"label": "取消用户", "values": d3_canceled},
                {"label": "取消率 (百分比)", "kind": "derived", "values": ratio_values(d3_canceled, d3_bind)},
            ],
        },
        {
            "title": "订阅状态",
            "items": [
                {"label": "付费有效订阅", "values": active},
                {"label": "freetrial 中订阅", "values": trialing},
                {"label": "扣款失败订阅", "values": past_due},
                {"label": "新增付费用户", "values": extra_series(finance_csv, "new_paid_users", decimals=0)},
                {"label": "累计付费用户", "values": extra_series(finance_csv, "cumulative_paid_users", decimals=0)},
                {"label": "付费订阅流失", "values": extra_series(finance_csv, "paid_churn_subs", decimals=0)},
            ],
        },
        {
            "title": "收入",
            "items": [
                {"label": "净现金收入", "values": extra_series(finance_csv, "cash_net", kind="currency", decimals=0)},
                {"label": "现金收入", "values": extra_series(finance_csv, "cash_gross", kind="currency", decimals=0)},
                {"label": "退款", "values": extra_series(finance_csv, "cash_refund", kind="currency", decimals=0)},
                {"label": "月经常性收入", "values": mrr},
                {"label": "付费用户平均收入", "kind": "derived", "values": division_values(mrr, active)},
            ],
        },
        {
            "title": "成本消耗",
            "items": [
                {"label": "推理成本 credits", "values": extra_series(grafana_csv, "credits_consumed", decimals=0)},
                {"label": "云服务成本（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "总成本（待数据）", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "结果",
            "items": [
                {"label": "毛利（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "毛利率（待数据）", "kind": "derived", "values": blank_values(), "keep_empty": True},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def cell_html(row: dict, value: str, first: bool, *, show_info: bool = True) -> str:
    cls = ["l"] if first else ["num"]
    if row.get("ratio"):
        cls.append("ratio")
    attrs = tooltip_attrs(row.get("definition")) if first else ""
    if row["kind"] == "section":
        content = section_title_html(value) if first else ""
    elif first:
        toggle = '<span class="tg"></span>' if row["has_children"] else '<span class="tg empty"></span>'
        info = ""
        if show_info and row.get("definition") and row["kind"] not in {"header", "section"}:
            info = '<span class="info-dot" aria-hidden="true">i</span>'
        content = toggle + label_html(value) + info
    else:
        content = value_html(value)
    return f'<td class="{" ".join(cls)}"{attrs}>{content}</td>'


def label_html(value: str) -> str:
    escaped = html.escape(display_label_text(value))
    return re.sub(r"(\d+(?:[.,:：/-]\d+)*(?:\.\d+)?%?)", r'<span class="inline-num">\1</span>', escaped)


def value_html(value: str) -> str:
    value = display_label_text(value)
    match = re.match(r"^(.+?)\s+\((-?\d+(?:\.\d+)?)%\)$", value)
    if not match:
        return html.escape(value)
    return f'{html.escape(match.group(1))} <span class="share-pct">({html.escape(match.group(2))}%)</span>'


def section_title_html(value: str) -> str:
    value = display_label_text(value)
    match = re.match(r"^\s*(\d+)[\.\、]\s*(.+?)\s*$", value)
    if match:
        return f"「{html.escape(match.group(1))}」{html.escape(match.group(2))}"
    return f"「{html.escape(value)}」"


def section_plain_title(value: str) -> str:
    match = re.match(r"^\s*\d+[\.\、]\s*(.+?)\s*$", value)
    return match.group(1).strip() if match else value.strip("「」 ")


def make_section_row(sheet_title: str, number: int, title: str) -> dict:
    row_id = f"{sheet_title}-section-{number}".replace(" ", "_")
    return {
        "id": row_id,
        "parent": "",
        "level": 0,
        "kind": "section",
        "ratio": False,
        "value_kind": "number",
        "values": [f"{number}. {title}"] + [""] * METRIC_VALUE_COUNT,
        "comments": [],
        "definition": f"分区标题：{title}。",
        "has_children": False,
        "hidden": False,
    }


def row_has_data(row: dict) -> bool:
    if row["kind"] in {"header", "section"}:
        return True
    return any(has_present_value(value) for value in row["values"][1:])


def remove_or_promote_empty_rows(rows: list[dict], *, keep_empty_parents: bool = False) -> list[dict]:
    rows_by_id = {row["id"]: row for row in rows}
    parent_children: dict[str, list[dict]] = {}
    for row in rows:
        if row.get("parent"):
            parent_children.setdefault(row["parent"], []).append(row)

    removed: set[str] = set()
    for row in rows:
        if row["kind"] in {"header", "section"}:
            continue
        if row_has_data(row):
            continue
        if keep_empty_parents and parent_children.get(row["id"]):
            continue
        removed.add(row["id"])

    for row in rows:
        while row.get("parent") in removed:
            old_parent = rows_by_id[row["parent"]]
            row["parent"] = old_parent.get("parent", "")
            row["level"] = max(0, row.get("level", 0) - 1)
            if row["kind"] == "child" and not row["parent"]:
                row["kind"] = "derived" if row.get("ratio") else "row"
                row["hidden"] = False

    kept = [row for row in rows if row["id"] not in removed]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


def normalize_sections(sheet_title: str, rows: list[dict]) -> list[dict]:
    if sheet_title.startswith("附录"):
        return rows

    output: list[dict] = []
    inserted_leading = False
    seen_section = False
    section_number = 0

    for row in rows:
        if row["kind"] == "header":
            output.append(row)
            continue

        if row["kind"] != "section" and not seen_section and sheet_title in LEADING_SECTION_TITLES and not inserted_leading:
            title = section_plain_title(LEADING_SECTION_TITLES[sheet_title])
            section_number += 1
            output.append(make_section_row(sheet_title, section_number, title))
            seen_section = True
            inserted_leading = True

        if row["kind"] == "section":
            section_number += 1
            title = section_plain_title(row["values"][0])
            row = row.copy()
            row.update(
                {
                    "parent": "",
                    "level": 0,
                    "has_children": False,
                    "hidden": False,
                    "values": [f"{section_number}. {title}"] + [""] * (len(row["values"]) - 1),
                }
            )
            seen_section = True
        output.append(row)

    section_ids = {row["id"] for row in output if row["kind"] == "section"}
    for row in output:
        if row.get("parent") in section_ids:
            row["parent"] = ""
            row["hidden"] = False
            if row["kind"] == "child":
                row["kind"] = "derived" if row.get("ratio") else "row"
            row["level"] = 0

    child_ids = {row["parent"] for row in output if row.get("parent")}
    for row in output:
        if row["kind"] == "section":
            row["has_children"] = False
        else:
            row["has_children"] = row["id"] in child_ids
            row["hidden"] = row["kind"] == "child"
    return output


def activation_group_for_label(label: str) -> str:
    normalized = normalize_label_for_definition(label)
    if normalized in {"UV", "新增 UV", "回访 UV"} or normalized in CHANNEL_LABELS:
        return "网站访问"
    if "CTA" in normalized or "官网主页" in normalized or "引导页" in normalized:
        return "CTA 点击"
    if normalized in {"新增注册用户"} or "登录完成" in normalized:
        return "登录完成"
    if (
        "发起绑卡" in normalized
        or "新增试用" in normalized
        or "绑卡失败" in normalized
        or "扣款" in normalized
        or normalized == "失败率"
    ):
        return "freetrial"
    if "登录失败" in normalized or normalized.startswith("失败：") or "登录页连续点击" in normalized:
        return "登录失败原因"
    return "首次使用"


def engineering_group_for_label(label: str) -> str:
    normalized = normalize_label_for_definition(label)
    if normalized in {"总请求", "受影响用户", "客户端异常", "客户端连续点击"}:
        return "用户感知"
    if any(token in normalized for token in ["端到端", "页面", "输入响应", "首字节"]):
        return "响应速度"
    if any(token in normalized for token in ["断连", "连接时长", "环境启动", "冷启动"]):
        return "连接与环境"
    return "完成与中止"


def restructure_engineering(rows: list[dict]) -> list[dict]:
    header = [row for row in rows if row["kind"] == "header"]
    body = [row for row in rows if row["kind"] != "header" and row["kind"] != "section"]
    groups = ["用户感知", "响应速度", "连接与环境", "完成与中止"]
    buckets = {group: [] for group in groups}
    row_group: dict[str, str] = {}

    for row in body:
        if row.get("parent") and row["parent"] in row_group:
            group = row_group[row["parent"]]
        else:
            group = engineering_group_for_label(row["values"][0])
        row_group[row["id"]] = group
        buckets.setdefault(group, []).append(row)

    output = header[:]
    number = 0
    for group in groups:
        group_rows = buckets.get(group, [])
        if not group_rows:
            continue
        number += 1
        output.append(make_section_row("工程质量", number, group))
        output.extend(group_rows)
    return normalize_sections("工程质量", output)


def restructure_activation(rows: list[dict]) -> list[dict]:
    header = [row for row in rows if row["kind"] == "header"]
    body = [row for row in rows if row["kind"] != "header" and row["kind"] != "section"]
    groups = ["网站访问", "CTA 点击", "登录完成", "freetrial", "登录失败原因", "首次使用"]
    buckets = {group: [] for group in groups}
    row_group: dict[str, str] = {}

    for row in body:
        if row.get("parent") and row["parent"] in row_group:
            group = row_group[row["parent"]]
        else:
            group = activation_group_for_label(row["values"][0])
        row_group[row["id"]] = group
        buckets.setdefault(group, []).append(row)

    output = header[:]
    number = 0
    for group in groups:
        group_rows = buckets.get(group, [])
        if not group_rows:
            continue
        number += 1
        output.append(make_section_row("用户激活与转化", number, group))
        output.extend(group_rows)
    return normalize_sections("用户激活与转化", output)


def row_classes(row: dict, *, interactive: bool = True) -> list[str]:
    classes = ["grid-row", row["kind"]]
    if interactive and row["has_children"]:
        classes.append("has-children")
    if row["hidden"]:
        classes.append("hidden-row")
    if row["ratio"]:
        classes.append("ratio-row")
    if row["level"]:
        classes.append(f"level-{min(row['level'], 3)}")
    return classes


def row_attrs(row: dict, *, interactive: bool = True) -> str:
    attrs = [f'class="{" ".join(row_classes(row, interactive=interactive))}"', f'data-id="{html.escape(row["id"], quote=True)}"']
    if row["parent"]:
        attrs.append(f'data-parent="{html.escape(row["parent"], quote=True)}"')
    if row["hidden"]:
        attrs.append("hidden")
    return " ".join(attrs)


def left_cell_html(
    row: dict,
    *,
    sheet_title: str,
    section_row: dict | None = None,
    show_info: bool = True,
) -> str:
    value = row["values"][0]
    attrs = ""
    if row["kind"] == "section":
        attrs = anchor_attrs(
            "section",
            sheet_title=sheet_title,
            row=row,
            section_row=row,
            section_label=value,
        )
        content = section_title_html(value)
    else:
        if row["kind"] == "header":
            attrs = anchor_attrs(
                "row",
                sheet_title=sheet_title,
                row=row,
                row_key="header",
                section_key="header",
                row_label=value,
                section_label=value,
            )
        else:
            attrs = anchor_attrs(
                "row",
                sheet_title=sheet_title,
                row=row,
                section_row=section_row,
                row_label=value,
            )
        toggle = '<span class="tg"></span>' if row["has_children"] else '<span class="tg empty"></span>'
        info = ""
        if show_info and row.get("definition") and row["kind"] not in {"header", "section"}:
            info = '<span class="info-dot" aria-hidden="true">i</span>'
        content = toggle + label_html(value) + info
    return f'<div class="cell left-cell"{attrs}{tooltip_attrs(row.get("definition"))}>{content}</div>'


def right_cells_html(row: dict, *, sheet_title: str, column_labels: list[str], section_row: dict | None = None) -> str:
    cells = []
    for index, value in enumerate(row["values"][1:], start=1):
        attrs = ""
        if row["kind"] != "section":
            label = column_labels[index] if index < len(column_labels) else ""
            header_kwargs = {"row_key": "header", "section_key": "header"} if row["kind"] == "header" else {}
            attrs = anchor_attrs(
                "cell",
                sheet_title=sheet_title,
                row=row,
                column=column_key(label, index),
                section_row=section_row,
                row_label=row["values"][0],
                column_label=label,
                **header_kwargs,
            )
        cells.append(f'<div class="cell num"{attrs}>{value_html(value)}</div>')
    return "".join(cells)


def render_main_block(rows: list[dict], kind: str, *, sheet_title: str, column_labels: list[str]) -> str:
    left = []
    right = []
    current_section: dict | None = None
    for row in rows:
        if row["kind"] == "section":
            current_section = row
        left.append(f'<div {row_attrs(row, interactive=True)}>{left_cell_html(row, sheet_title=sheet_title, section_row=current_section)}</div>')
        right.append(f'<div {row_attrs(row, interactive=False)}>{right_cells_html(row, sheet_title=sheet_title, column_labels=column_labels, section_row=current_section)}</div>')
    return (
        f'<div class="table-block {kind}-block">'
        '<div class="freeze-grid">'
        f'<div class="left-pane">{"".join(left)}</div>'
        f'<div class="right-pane"><div class="right-inner">{"".join(right)}</div></div>'
        '</div></div>'
    )


APPENDIX2_SEGMENTS = [
    ("反馈对象", "1. 个人反馈", ["反馈对象", "诉求"]),
    ("高频诉求", "2. 高频诉求归属", ["主题", "涉及的人", "归属分类"]),
    ("1. 已开发 / 待开发", "3. 已开发 / 待开发", ["维度", "指标", "提出人", "状态", "备注"]),
]


def _is_blank_appendix(v: str) -> bool:
    s = (v or "").strip()
    return s in {"", "—", "-", "–", "None"}


def _segment_appendix2_rows(rows: list[dict]) -> list[dict]:
    """Split 附录2 rows into labeled segments at known marker labels."""
    starts = {label: (title, cols) for label, title, cols in APPENDIX2_SEGMENTS}
    skip_labels = {"维度"}  # source-row column labels we replace with synthetic headers
    segments: list[dict] = []
    current: dict | None = None
    for row in rows:
        first = (row["values"][0] if row["values"] else "").strip()
        if first in starts:
            title, cols = starts[first]
            current = {"title": title, "columns": cols, "rows": []}
            segments.append(current)
            continue
        if first in skip_labels:
            continue
        if current is None:
            current = {"title": None, "columns": None, "rows": []}
            segments.append(current)
        current["rows"].append(row)
    return segments


def _synthetic_header_row(sheet_title: str, segment_title: str, columns: list[str]) -> dict:
    rid = f"{sheet_title}-seg-{segment_title}-hdr".replace(" ", "_")
    return {
        "id": rid,
        "parent": "",
        "level": 0,
        "kind": "header",
        "ratio": False,
        "value_kind": "text",
        "values": list(columns),
        "comments": [],
        "definition": "段落表头。",
        "has_children": False,
        "hidden": False,
    }


def _synthetic_section_row(sheet_title: str, title: str, columns: int) -> dict:
    rid = f"{sheet_title}-seg-{title}-sec".replace(" ", "_")
    return {
        "id": rid,
        "parent": "",
        "level": 0,
        "kind": "section",
        "ratio": False,
        "value_kind": "text",
        "values": [title] + [""] * max(columns - 1, 0),
        "comments": [],
        "definition": f"分区：{title}。",
        "has_children": False,
        "hidden": False,
    }


def render_appendix_sheet(sheet: dict, idx: int) -> str:
    sheet_classes = ["sheet"]
    is_appendix = sheet["title"].startswith("附录")
    if is_appendix:
        sheet_classes.append("appendix")
    out = [f'<section class="{" ".join(sheet_classes)}" data-sheet="{idx}"><div class="wrap">']

    if sheet["title"] == "附录2 个人看板反馈":
        segments = _segment_appendix2_rows(sheet["rows"])
    else:
        segments = [{"title": None, "columns": None, "rows": sheet["rows"]}]

    for seg in segments:
        seg_rows: list[dict] = []
        if seg["title"]:
            seg_rows.append(_synthetic_section_row(sheet["title"], seg["title"], len(seg["columns"]) + 1 if seg["columns"] else 1))
        if seg["columns"]:
            seg_rows.append(_synthetic_header_row(sheet["title"], seg["title"], seg["columns"]))
        seg_rows.extend(seg["rows"])

        if seg["columns"]:
            max_used = len(seg["columns"]) - 1
        else:
            max_used = 0
            for row in seg_rows:
                if row["kind"] == "header":
                    continue
                for i in range(len(row["values"]) - 1, 0, -1):
                    if not _is_blank_appendix(row["values"][i]):
                        if i > max_used:
                            max_used = i
                        break
            if max_used == 0:
                max_used = 1
        keep_cols = max_used + 1

        for row in seg_rows:
            trimmed = row["values"][:keep_cols]
            if len(trimmed) < keep_cols:
                trimmed = trimmed + [""] * (keep_cols - len(trimmed))
            if row["kind"] != "header":
                trimmed = [trimmed[0]] + ["" if _is_blank_appendix(v) else v for v in trimmed[1:]]
            row["values"] = trimmed

        colgroup = (
            "<colgroup>"
            '<col class="c-appx-label">'
            + "".join('<col class="c-appx-text">' for _ in range(max_used))
            + "</colgroup>"
        )

        table_open = False

        def close_table() -> None:
            nonlocal table_open
            if table_open:
                out.append("</tbody></table></div>")
                table_open = False

        def open_table(kind: str) -> None:
            nonlocal table_open
            close_table()
            out.append(f'<div class="table-block {kind}-block">')
            out.append(f'<table class="t">{colgroup}<tbody>')
            table_open = True

        for row in seg_rows:
            if row["kind"] == "header" and not table_open:
                open_table("header")
            elif row["kind"] == "section":
                open_table("section")
            elif not table_open:
                open_table("plain")

            classes = [row["kind"]]
            if row["has_children"]:
                classes.append("has-children")
            if row["hidden"]:
                classes.append("hidden-row")
            if row["ratio"]:
                classes.append("ratio-row")
            if row["level"]:
                classes.append(f"level-{min(row['level'], 3)}")
            attrs = [f'class="{" ".join(classes)}"', f'data-id="{html.escape(row["id"], quote=True)}"']
            if row["parent"]:
                attrs.append(f'data-parent="{html.escape(row["parent"], quote=True)}"')
            if row["hidden"]:
                attrs.append("hidden")
            cells = "".join(cell_html(row, v, i == 0, show_info=not is_appendix) for i, v in enumerate(row["values"]))
            out.append(f"<tr {' '.join(attrs)}>{cells}</tr>")
        close_table()

    out.append("</div></section>")
    return "\n".join(out)


def render_sheet(sheet: dict, idx: int) -> str:
    if sheet["title"].startswith("附录"):
        return render_appendix_sheet(sheet, idx)

    out = [f'<section class="sheet" data-sheet="{idx}"><div class="wrap">']
    current_rows: list[dict] = []
    current_kind = "plain"
    column_labels = sheet["rows"][0]["values"] if sheet["rows"] else []

    def flush() -> None:
        nonlocal current_rows
        if current_rows:
            out.append(render_main_block(current_rows, current_kind, sheet_title=sheet["title"], column_labels=column_labels))
            current_rows = []

    for row in sheet["rows"]:
        if row["kind"] == "header":
            flush()
            current_kind = "header"
            current_rows = [row]
            flush()
            current_kind = "plain"
            continue
        if row["kind"] == "section":
            flush()
            current_kind = "section"
            current_rows = [row]
            continue
        current_rows.append(row)
    flush()
    out.append("</div></section>")
    return "\n".join(out)


CSS = r"""
:root {
  color-scheme:light;
  --navy:#252b33;
  --navy-2:#343b45;
  --derived-bg:#f2f3f5;
  --derived-ink:#2f3740;
  --ratio-bg:#f7f8f9;
  --ratio-derived-bg:#f1f2f4;
  --paper:#ffffff;
  --bg:#f4f5f6;
  --line:#d8dde3;
  --line-strong:#aeb8c3;
  --ink:#0b0f14;
  --ink-2:#26313d;
  --muted:#7c8794;
  --child:#fafbfc;
  --child-ink:#3b434d;
  --child-label:#4b5563;
  --hover:#f6f9fb;
  --today:oklch(76% .2 58);
  --today-hover:oklch(70% .2 55);
  --today-text:#211205;
  --today-line:oklch(58% .16 50);
  --topbar-bg:rgba(255,255,255,.96);
  --tab-muted:#697482;
  --header-bg:#eef1f4;
  --header-ink:#111820;
  --header-line:#8d99a6;
  --section-bg:#ffffff;
  --section-ink:#252b33;
  --section-line:#b6c0cb;
  --share:#5f6975;
  --share-on-today:#4b2705;
  --info-bg:#f7f8f9;
  --info-border:#b9c1ca;
  --info-ink:#6f7a86;
  --info-hover-bg:#e9edf1;
  --info-hover-ink:#252b33;
  --info-hover-border:#8f9aa7;
  --action-bg:#252b33;
  --action-ink:#ffffff;
  --main-table-width:__MAIN_TABLE_WIDTH__px;
  --main-table-viewport-width:__MAIN_TABLE_VIEWPORT_WIDTH__px;
}
html.dark {
  color-scheme:dark;
  --navy:#d7dee8;
  --navy-2:#aeb8c3;
  --derived-bg:#1b222a;
  --derived-ink:#d5dbe3;
  --ratio-bg:#192129;
  --ratio-derived-bg:#202832;
  --paper:#141a21;
  --bg:#0d1117;
  --line:#2e3742;
  --line-strong:#485564;
  --ink:#f2f5f8;
  --ink-2:#d7dee8;
  --muted:#97a4b4;
  --child:#171e26;
  --child-ink:#c6ced8;
  --child-label:#a9b4c2;
  --hover:#202a34;
  --today:oklch(76% .2 58);
  --today-hover:oklch(70% .2 55);
  --today-text:#211205;
  --today-line:oklch(58% .16 50);
  --topbar-bg:rgba(13,17,23,.96);
  --tab-muted:#8b98a8;
  --header-bg:#1f2832;
  --header-ink:#eef3f8;
  --header-line:#596675;
  --section-bg:#151c24;
  --section-ink:#eef3f8;
  --section-line:#536171;
  --share:#99a6b6;
  --share-on-today:#4b2705;
  --info-bg:#1d2630;
  --info-border:#536171;
  --info-ink:#b2bdca;
  --info-hover-bg:#26313d;
  --info-hover-ink:#edf3f9;
  --info-hover-border:#728092;
  --action-bg:#d7dee8;
  --action-ink:#111820;
}
* { box-sizing:border-box; }
html, body {
  margin:0;
  background:var(--bg);
  color:var(--ink-2);
  overscroll-behavior-x:none;
}
body {
  font-family:"KaiTi SC","楷体-简","Kaiti SC","STKaiti",serif;
  font-size:16px;
  line-height:1.16;
  -webkit-font-smoothing:antialiased;
  text-rendering:optimizeLegibility;
}
.num, .tab, .inline-num, .tg, .info-dot {
  font-family:"KaiTi SC","楷体-简","Kaiti SC","STKaiti",serif;
  font-variant-numeric:tabular-nums lining-nums;
  font-feature-settings:"tnum" 1, "lnum" 1;
}
.topbar {
  position:sticky;
  top:0;
  z-index:20;
  background:var(--topbar-bg);
  border-bottom:1px solid var(--line);
  backdrop-filter:saturate(140%) blur(8px);
}
.tabs {
  display:flex;
  gap:18px;
  overflow-x:auto;
  overscroll-behavior-x:contain;
  overscroll-behavior-inline:contain;
  -webkit-overflow-scrolling:touch;
  touch-action:pan-x pan-y;
  padding:0 20px;
  max-width:1840px;
  margin:0 auto;
  scrollbar-width:none;
}
.tabs::-webkit-scrollbar { display:none; }
.tab {
  border:0;
  background:transparent;
  color:var(--tab-muted);
  padding:9px 0 8px;
  border-bottom:2px solid transparent;
  font-size:16px;
  white-space:nowrap;
  cursor:pointer;
  letter-spacing:0;
}
.tab.on { color:var(--navy); border-bottom-color:var(--navy); font-weight:600; }
.tab.appendix-start { margin-left:auto; }
main { max-width:1840px; margin:0 auto; padding:10px 20px 22px; }
.history-hint {
  margin:0 0 8px;
  color:var(--muted);
  font-size:14px;
  line-height:18px;
}
.sheet { display:none; }
.sheet.on { display:block; }
.wrap { overflow:visible; background:transparent; }
.table-block {
  border:1px solid var(--line-strong);
  background:var(--paper);
  box-shadow:none;
  margin:0 0 12px;
  overflow:hidden;
}
.sheet:not(.appendix) .table-block { width:100%; max-width:calc(var(--main-table-viewport-width) + 2px); }
.sheet.appendix .table-block { width:100%; }
.header-block {
  position:sticky;
  top:40px;
  z-index:18;
  margin-bottom:12px;
}
.section-block { margin-bottom:14px; }
.freeze-grid {
  display:grid;
  grid-template-columns:__LEFT_COL_WIDTH__px minmax(0, 1fr);
  width:100%;
  background:var(--paper);
}
.left-pane {
  position:relative;
  z-index:4;
  background:var(--paper);
}
.right-pane {
  width:100%;
  overflow-x:auto;
  overflow-y:hidden;
  overscroll-behavior-x:contain;
  overscroll-behavior-inline:contain;
  -webkit-overflow-scrolling:touch;
  touch-action:pan-x pan-y;
  scrollbar-width:none;
  background:var(--paper);
}
.right-pane::-webkit-scrollbar { display:none; }
.right-inner { width:__RIGHT_INNER_WIDTH__px; min-width:__RIGHT_INNER_WIDTH__px; }
.grid-row {
  display:grid;
  height:24px;
  min-height:24px;
}
.grid-row[hidden],
tr[hidden] {
  display:none !important;
}
.left-pane .grid-row { grid-template-columns:__LEFT_COL_WIDTH__px; }
.right-pane .grid-row {
  grid-template-columns:__RIGHT_GRID_TEMPLATE__;
}
.cell {
  min-width:0;
  height:24px;
  padding:3px 8px;
  border-bottom:1px solid var(--line);
  color:var(--ink);
  font-size:16px;
  font-weight:400;
  line-height:18px;
  overflow:hidden;
  white-space:nowrap;
  text-overflow:clip;
}
.left-cell {
  text-align:left;
  background:inherit;
  border-right:1px solid var(--line-strong);
}
.num {
  text-align:right;
  color:var(--ink);
  font-size:16px;
  font-weight:400;
  font-variant-numeric:tabular-nums lining-nums;
}
.header-block .grid-row,
.header-block .cell {
  height:28px;
  min-height:28px;
  line-height:21px;
}
.header-block .cell {
  background:var(--header-bg);
  color:var(--header-ink);
  font-weight:600;
  border-bottom:1px solid var(--header-line);
}
.header-block .right-pane .grid-row > .cell:nth-child(__TODAY_GRID_NTH__),
.right-pane .grid-row:not(.section) > .cell:nth-child(__TODAY_GRID_NTH__) {
  background:var(--today);
  color:var(--today-text);
  font-weight:700;
  border-bottom-color:var(--today-line);
}
.section-block {
  border-color:var(--line-strong);
}
.section-block .grid-row.section,
.section-block .grid-row.section .cell {
  height:28px;
  min-height:28px;
  line-height:21px;
}
.grid-row.section .cell {
  background:var(--section-bg);
  color:var(--section-ink);
  font-weight:600;
  border-bottom-color:var(--section-line);
}
.grid-row.section .left-cell {
  padding-left:8px;
}
.grid-row.row .left-cell,
.grid-row.row.has-children .left-cell {
  font-weight:600;
}
.grid-row.derived .cell {
  background:var(--derived-bg);
  color:var(--derived-ink);
  font-size:16px;
  font-style:normal;
  font-weight:400;
}
.grid-row.derived .left-cell {
  color:var(--section-ink);
  font-weight:500;
}
.grid-row.child .cell {
  background:var(--child);
  color:var(--child-ink);
  font-size:16px;
  font-style:normal;
  font-weight:400;
}
.grid-row.child .left-cell {
  color:var(--child-label);
  font-weight:400;
}
.grid-row.ratio-row .cell { background:var(--ratio-bg); }
.grid-row.derived.ratio-row .cell { background:var(--ratio-derived-bg); }
.right-pane .grid-row:not(.section) > .cell:nth-child(__TODAY_GRID_NTH__) .share-pct { color:var(--share-on-today); }
.grid-row.level-1 .left-cell { padding-left:28px; }
.grid-row.level-2 .left-cell { padding-left:44px; }
.grid-row.level-3 .left-cell { padding-left:60px; }
.grid-row.has-children { cursor:pointer; }
.grid-row.has-children:hover .cell { background:var(--hover); }
.right-pane .grid-row.has-children:hover > .cell:nth-child(__TODAY_GRID_NTH__) { background:var(--today-hover); color:var(--today-text); }
.grid-row.open .left-cell .tg::before { content:"−"; }
.grid-row:last-child .cell { border-bottom:0; }
table.t { width:var(--main-table-width); min-width:var(--main-table-width); table-layout:fixed; border-collapse:separate; border-spacing:0; background:var(--paper); }
.sheet.appendix table.t { width:100%; min-width:0; table-layout:auto; }
col.c-appx-label { width:220px; }
col.c-appx-text { width:auto; }
.sheet.appendix td { vertical-align:top; line-height:1.55; }
.sheet.appendix td:first-child {
  position:static !important;
  border-right:1px solid var(--line);
  width:auto;
  min-width:0;
  max-width:none;
  background:inherit !important;
}
.sheet.appendix td,
.sheet.appendix td.num {
  text-align:left;
  white-space:normal;
}
col.c-label { width:__LEFT_COL_WIDTH__px; }
col.c-date { width:76px; }
col.c-today { width:108px; }
col.c-comp { width:128px; }
col.c-wow { width:78px; }
col.c-mom { width:78px; }
td { border-bottom:1px solid var(--line); padding:3px 8px; height:24px; vertical-align:middle; font-size:16px; font-weight:400; }
td:first-child {
  position:sticky;
  left:0;
  z-index:10;
  border-right:1px solid var(--line-strong);
  width:__LEFT_COL_WIDTH__px;
  min-width:__LEFT_COL_WIDTH__px;
  max-width:320px;
  color:var(--ink);
  text-align:left;
  background:var(--paper);
}
td.num { text-align:right; color:var(--ink); white-space:nowrap; font-size:16px; font-weight:400; letter-spacing:0; }
tr.header td {
  background:var(--header-bg);
  color:var(--header-ink);
  font-weight:600;
  border-bottom:1px solid var(--header-line);
  box-shadow:none;
}
tr.header td:first-child { color:var(--header-ink); text-align:left; padding-left:8px; z-index:13; }
tr.header td:nth-child(__TODAY_TABLE_NTH__),
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__) { background:var(--today); color:var(--today-text); border-bottom-color:var(--today-line); }
tr.header td:nth-child(__TODAY_TABLE_NTH__) { font-weight:700; }
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__).num { font-weight:700; color:var(--today-text); }
tr.section td {
  background:var(--section-bg);
  color:var(--section-ink);
  font-weight:600;
  border-top:0;
  border-bottom-color:var(--section-line);
  box-shadow:none;
  height:28px;
}
tr.section td:first-child {
  z-index:12;
  text-align:left;
  padding-left:6px;
}
tr:not(.section):last-child td { border-bottom:0; }
tr.row td { background:var(--paper); color:var(--ink); font-weight:400; }
tr.row td:first-child { font-weight:600; }
tr.row.has-children td:first-child { font-weight:600; }
tr.derived td { background:var(--derived-bg); color:var(--derived-ink); font-size:16px; font-style:normal; font-weight:400; }
tr.derived td:first-child { color:var(--section-ink); font-style:normal; font-weight:500; }
tr.derived td.num { color:var(--section-ink); font-size:16px; font-weight:400; }
tr.child td { background:var(--child); color:var(--child-ink); font-size:16px; font-style:normal; font-weight:400; }
tr.child td:first-child { color:var(--child-label); font-style:normal; font-weight:400; }
tr.child td.num { color:var(--section-ink); font-size:16px; font-style:normal; font-weight:400; }
.share-pct {
  color:var(--share);
  font-size:12px;
  font-style:italic;
  font-weight:400;
}
tr.ratio-row td { background:var(--ratio-bg); }
tr.derived.ratio-row td { background:var(--ratio-derived-bg); }
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__),
tr.derived td:nth-child(__TODAY_TABLE_NTH__).num,
tr.child td:nth-child(__TODAY_TABLE_NTH__).num,
tr.ratio-row td:nth-child(__TODAY_TABLE_NTH__),
tr.derived.ratio-row td:nth-child(__TODAY_TABLE_NTH__).num {
  background:var(--today);
  color:var(--today-text);
  font-weight:700;
}
tr.section td:nth-child(__TODAY_TABLE_NTH__),
tr.section td:nth-child(__TODAY_TABLE_NTH__).num {
  background:var(--section-bg);
  color:var(--section-ink);
  font-weight:600;
  border-bottom-color:var(--section-line);
}
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__) .share-pct { color:var(--share-on-today); }
tr.level-1 td:first-child { padding-left:28px; }
tr.level-2 td:first-child { padding-left:44px; font-style:normal; }
tr.level-3 td:first-child { padding-left:60px; font-style:normal; }
tr.has-children { cursor:pointer; }
tr.has-children:hover td { background:var(--hover); }
tr.has-children:hover td:nth-child(__TODAY_TABLE_NTH__) { background:var(--today-hover); color:var(--today-text); }
tr.open > td:first-child .tg::before { content:"−"; }
.tg {
  display:inline-flex;
  width:12px;
  height:10px;
  margin-right:6px;
  align-items:center;
  justify-content:center;
  color:#a5aeb8;
  font-size:12px;
  font-weight:400;
  line-height:1;
  vertical-align:1px;
}
.tg::before { content:"+"; }
.tg.empty { visibility:hidden; }
.inline-num { color:inherit; opacity:.72; }
.info-dot {
  display:inline-flex;
  margin-left:7px;
  width:12px;
  height:12px;
  border-radius:50%;
  align-items:center;
  justify-content:center;
  border:1px solid var(--info-border);
  background:var(--info-bg);
  color:var(--info-ink);
  font-size:8px;
  font-weight:500;
  line-height:1;
  vertical-align:2px;
  cursor:help;
  opacity:.82;
}
.info-dot:hover {
  background:var(--info-hover-bg);
  color:var(--info-hover-ink);
  border-color:var(--info-hover-border);
}
.metric-tooltip {
  position:fixed;
  z-index:90;
  max-width:min(360px, calc(100vw - 32px));
  padding:8px 10px;
  border:1px solid var(--line-strong);
  background:var(--paper);
  color:var(--ink-2);
  box-shadow:0 8px 22px rgba(11, 15, 20, .18);
  font-family:"KaiTi SC","楷体-简","Kaiti SC","STKaiti",serif;
  font-size:14px;
  line-height:1.35;
  pointer-events:none;
}
@media (max-width:900px) {
  body { font-size:16px; }
  main { padding:10px; }
  .tabs { padding-left:10px; padding-right:10px; }
  td:first-child { width:240px; min-width:240px; }
  td { padding:3px 7px; height:24px; }
}
"""


THEME_JS = r"""
(function initDashboardTheme() {
  const THEME_KEY = "maxcenter_theme";
  function normalizeTheme(value) {
    return value === "dark" ? "dark" : "light";
  }
  function applyTheme(value) {
    document.documentElement.classList.toggle("dark", normalizeTheme(value) === "dark");
  }
  function storedTheme() {
    try { return window.localStorage && window.localStorage.getItem(THEME_KEY); } catch (_error) { return null; }
  }
  function parentTheme() {
    try {
      if (window.parent && window.parent !== window && window.parent.document) {
        return window.parent.document.documentElement.classList.contains("dark") ? "dark" : "light";
      }
    } catch (_error) {}
    return null;
  }
  function systemTheme() {
    return window.matchMedia && window.matchMedia("(prefers-color-scheme: dark)").matches ? "dark" : "light";
  }
  applyTheme(storedTheme() || parentTheme() || systemTheme());
  window.addEventListener("message", (event) => {
    if (event.origin !== window.location.origin) return;
    if (!event.data || event.data.type !== "maxcenter-theme") return;
    applyTheme(event.data.theme);
  });
  window.addEventListener("storage", (event) => {
    if (event.key === THEME_KEY) applyTheme(event.newValue);
  });
})();
"""


JS = r"""
(function initDashboardBase() {
const sheets = __SHEETS__;
const DEFAULT_DATE_SCROLL_LEFT = __DEFAULT_DATE_SCROLL_LEFT__;
const baseState = window.dashboardBaseState || { currentSheet: 0, initializedScroll: new Set(), wired: false };
if (!(baseState.initializedScroll instanceof Set)) {
  baseState.initializedScroll = new Set(baseState.initializedScroll || []);
}
window.dashboardBaseState = baseState;

function refreshCommentOverlay(options = {}) {
  const overlay = window.dashboardCommentOverlay;
  if (!overlay) return;
  if (options.closeFloating) {
    if (typeof overlay.closeComposer === 'function') overlay.closeComposer();
    if (typeof overlay.closePopover === 'function') overlay.closePopover();
  }
  if (typeof overlay.scheduleRenderPins === 'function') overlay.scheduleRenderPins();
  else if (typeof overlay.renderPins === 'function') overlay.renderPins();
}

let metricTooltipEl = null;

function ensureMetricTooltip() {
  if (metricTooltipEl) return metricTooltipEl;
  metricTooltipEl = document.createElement('div');
  metricTooltipEl.className = 'metric-tooltip';
  metricTooltipEl.hidden = true;
  document.body.appendChild(metricTooltipEl);
  return metricTooltipEl;
}

function positionMetricTooltip(event, target) {
  const tooltip = ensureMetricTooltip();
  const rect = target.getBoundingClientRect();
  const sourceX = event && Number.isFinite(event.clientX) ? event.clientX : rect.left + Math.min(rect.width, 180);
  const sourceY = event && Number.isFinite(event.clientY) ? event.clientY : rect.top + rect.height / 2;
  const margin = 12;
  const tooltipRect = tooltip.getBoundingClientRect();
  let left = sourceX + margin;
  let top = sourceY + margin;
  if (left + tooltipRect.width + margin > window.innerWidth) {
    left = Math.max(margin, window.innerWidth - tooltipRect.width - margin);
  }
  if (top + tooltipRect.height + margin > window.innerHeight) {
    top = Math.max(margin, sourceY - tooltipRect.height - margin);
  }
  tooltip.style.left = `${left}px`;
  tooltip.style.top = `${top}px`;
}

function showMetricTooltip(target, event) {
  const text = target.dataset.tooltip;
  if (!text) return;
  const tooltip = ensureMetricTooltip();
  tooltip.textContent = text;
  tooltip.hidden = false;
  positionMetricTooltip(event, target);
}

function hideMetricTooltip() {
  if (metricTooltipEl) metricTooltipEl.hidden = true;
}

function applyDefaultDateScroll(sheetIndex = baseState.currentSheet) {
  if (baseState.initializedScroll.has(sheetIndex)) return;
  baseState.initializedScroll.add(sheetIndex);
  document.querySelectorAll(`.sheet[data-sheet="${sheetIndex}"] .right-pane`).forEach(pane => {
    pane.scrollLeft = DEFAULT_DATE_SCROLL_LEFT;
  });
}

function sync() {
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('on', Number(t.dataset.sheet) === baseState.currentSheet));
  document.querySelectorAll('.sheet').forEach(s => s.classList.toggle('on', Number(s.dataset.sheet) === baseState.currentSheet));
  applyDefaultDateScroll(baseState.currentSheet);
  window.scrollTo({ top: 0, behavior: 'instant' });
  hideMetricTooltip();
  refreshCommentOverlay({ closeFloating: true });
}

function closeChildren(id) {
  document.querySelectorAll(`.grid-row[data-parent="${CSS.escape(id)}"]`).forEach(child => {
    child.hidden = true;
    child.classList.remove('open');
    closeChildren(child.dataset.id);
  });
}

function horizontalWheelDelta(event) {
  const absX = Math.abs(event.deltaX);
  const absY = Math.abs(event.deltaY);
  if (absX >= absY) return event.deltaX;
  if (event.shiftKey && absY > 0) return event.deltaY;
  return 0;
}

function guardHorizontalScroll(element) {
  element.addEventListener('wheel', event => {
    const delta = horizontalWheelDelta(event);
    if (!delta || element.scrollWidth <= element.clientWidth) return;
    event.preventDefault();
    const maxLeft = element.scrollWidth - element.clientWidth;
    element.scrollLeft = Math.max(0, Math.min(maxLeft, element.scrollLeft + delta));
  }, { passive: false });
}

function wire() {
  if (baseState.wired) return;
  baseState.wired = true;
  document.querySelectorAll('.tab').forEach(btn => {
    btn.addEventListener('click', () => {
      baseState.currentSheet = Number(btn.dataset.sheet);
      sync();
    });
  });
  document.querySelectorAll('.tabs').forEach(guardHorizontalScroll);
  document.querySelectorAll('[data-tooltip]').forEach(target => {
    target.addEventListener('mouseenter', event => showMetricTooltip(target, event));
    target.addEventListener('mousemove', event => positionMetricTooltip(event, target));
    target.addEventListener('mouseleave', hideMetricTooltip);
    target.addEventListener('focus', event => showMetricTooltip(target, event));
    target.addEventListener('blur', hideMetricTooltip);
  });
  window.addEventListener('scroll', hideMetricTooltip, { passive: true });
  window.addEventListener('resize', hideMetricTooltip);
  let syncing = false;
  document.querySelectorAll('.right-pane').forEach(pane => {
    guardHorizontalScroll(pane);
    pane.addEventListener('scroll', () => {
      if (syncing) return;
      syncing = true;
      const left = pane.scrollLeft;
      document.querySelectorAll('.sheet.on .right-pane').forEach(other => {
        if (other !== pane) other.scrollLeft = left;
      });
      syncing = false;
      hideMetricTooltip();
      refreshCommentOverlay();
    });
  });
  document.querySelectorAll('.left-pane .grid-row.has-children').forEach(row => {
    row.addEventListener('click', event => {
      event.stopPropagation();
      const id = row.dataset.id;
      const children = Array.from(document.querySelectorAll(`.grid-row[data-parent="${CSS.escape(id)}"]`));
      const leftChildren = children.filter(c => c.closest('.left-pane'));
      const anyVisible = leftChildren.some(c => !c.hidden);
      document.querySelectorAll(`.left-pane .grid-row[data-id="${CSS.escape(id)}"]`).forEach(parent => parent.classList.toggle('open', !anyVisible));
      if (anyVisible) closeChildren(id);
      else children.forEach(c => { c.hidden = false; });
      refreshCommentOverlay();
    });
  });
}

wire();
sync();

window.initDashboardCommentOverlay = async function initDashboardCommentOverlay() {
  if (!window.DashboardCommentModel || !window.DashboardCommentStore || !window.DashboardCommentOverlay) return;
  const existing = window.dashboardCommentOverlay;
  if (existing && existing.initialized) return existing;
  if (existing && typeof existing.init === 'function') {
    await existing.init();
    return existing;
  }
  if (existing && typeof existing.destroy === 'function') existing.destroy();
  const model = window.DashboardCommentModel;
  const store = window.DashboardCommentStore.createDefaultStore();
  const overlay = new window.DashboardCommentOverlay.CommentOverlay({
    model,
    store,
    pageKey: "moclaw_operating_dashboard",
    pageVersion: "v2"
  });
  window.dashboardCommentOverlay = overlay;
  await overlay.init();
  return overlay;
};

function scheduleDashboardCommentOverlayInit() {
  if (document.readyState !== 'loading') {
    window.initDashboardCommentOverlay();
    return;
  }
  if (window.dashboardCommentOverlayInitScheduled) return;
  window.dashboardCommentOverlayInitScheduled = true;
  window.addEventListener('DOMContentLoaded', () => {
    window.dashboardCommentOverlayInitScheduled = false;
    window.initDashboardCommentOverlay();
  }, { once: true });
}

scheduleDashboardCommentOverlayInit();
})();
"""


def render_css() -> str:
    return (
        CSS.replace("__MAIN_TABLE_WIDTH__", str(MAIN_TABLE_WIDTH))
        .replace("__MAIN_TABLE_VIEWPORT_WIDTH__", str(MAIN_TABLE_VIEWPORT_WIDTH))
        .replace("__LEFT_COL_WIDTH__", str(LEFT_COL_WIDTH))
        .replace("__RIGHT_INNER_WIDTH__", str(RIGHT_INNER_WIDTH))
        .replace("__RIGHT_GRID_TEMPLATE__", RIGHT_GRID_TEMPLATE)
        .replace("__TODAY_GRID_NTH__", str(TODAY_GRID_NTH))
        .replace("__TODAY_TABLE_NTH__", str(TODAY_TABLE_NTH))
    )


def render_js(sheets: list[dict]) -> str:
    payload = json.dumps([{"title": display_sheet_title(s["title"])} for s in sheets], ensure_ascii=False)
    return JS.replace("__SHEETS__", payload).replace("__DEFAULT_DATE_SCROLL_LEFT__", str(DEFAULT_DATE_SCROLL_LEFT))


def main() -> None:
    MISSING_DEFINITIONS.clear()
    sheets = workbook_model()
    comments_css = COMMENTS_CSS_PATH.read_text(encoding="utf-8")
    comment_scripts = [path.read_text(encoding="utf-8") for path in COMMENT_SCRIPT_PATHS]
    parts = [
        "<!doctype html>",
        '<html lang="zh-CN"><head><meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>MoClaw Operating Dashboard · Final</title>",
        f"<script>{THEME_JS}</script>",
        f"<style>{render_css()}</style>",
        f"<style>{comments_css}</style></head><body>",
        '<nav class="topbar"><div class="tabs">',
    ]
    for i, sheet in enumerate(sheets):
        appendix_class = " appendix-start" if sheet["title"].startswith("附录1") else ""
        parts.append(f'<button class="tab{appendix_class}" data-sheet="{i}">{html.escape(display_sheet_title(sheet["title"]))}</button>')
    parts.extend(['</div></nav>', '<main><p class="history-hint">左滑查看更多历史数据</p>'])
    for i, sheet in enumerate(sheets):
        parts.append(render_sheet(sheet, i))
    parts.extend(
        [
            "</main>",
            *[f"<script>{script}</script>" for script in comment_scripts],
            f"<script>{render_js(sheets)}</script>",
            "</body></html>",
        ]
    )
    OUT.write_text("\n".join(parts), encoding="utf-8")
    write_missing_definitions()
    print(OUT)
    print("sheets", len(sheets), "rows", sum(len(s["rows"]) for s in sheets))
    print("missing_definitions", len(set(MISSING_DEFINITIONS)), MISSING_DEFS_OUT)


def write_missing_definitions() -> None:
    by_sheet: dict[str, list[str]] = {}
    for sheet, label in sorted(set(MISSING_DEFINITIONS)):
        by_sheet.setdefault(sheet, []).append(label)
    lines = [
        "MoClaw Operating Dashboard missing field definitions",
        f"Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Count: {sum(len(v) for v in by_sheet.values())}",
        "",
    ]
    for sheet in ORDER:
        labels = by_sheet.get(sheet)
        if not labels:
            continue
        lines.append(f"[{sheet}]")
        lines.extend(f"- {label}" for label in labels)
        lines.append("")
    MISSING_DEFS_OUT.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
