#!/usr/bin/env python3
"""Build the final table-first MoClaw operating dashboard web page."""
from __future__ import annotations

import csv
import datetime as dt
import html
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
SRC = ROOT / "web_restore_sources" / "JJHV_MoClaw_Dashboard_v31.xlsx"
OUT = ROOT / "MoClaw_Operating_Dashboard_Web.html"
MISSING_DEFS_OUT = ROOT / "MoClaw_Operating_Dashboard_missing_definitions.txt"
GENERATED_METRICS = ROOT / "generated_posthog_metrics"
DASHBOARD_EXPORTS = ROOT.parent / "dashboard"
COMMENTS_CSS_PATH = ROOT / "dashboard_comments" / "comment_overlay.css"
COMMENT_SCRIPT_PATHS = [
    ROOT / "dashboard_comments" / "comment_model.js",
    ROOT / "dashboard_comments" / "comment_store.js",
    ROOT / "dashboard_comments" / "comment_overlay.js",
]
PAGE_KEY = "moclaw_operating_dashboard"
PAGE_VERSION = "v1"

ORDER = [
    "Dashboard",
    "用户获取",
    "用户激活与转化",
    "用户活跃与使用分布",
    "Agent 质量",
    "财务",
    "附录1_口径裁决",
    "附录2 个人看板反馈",
]

SHEET_DISPLAY_NAMES = {
    "用户留存与流失": "留存与流失",
    "用户活跃与使用分布": "用户活跃与使用",
    "Agent 质量": "Agent & 工程质量",
    "财务": "业务经营",
    "附录1_口径裁决": "附录 1 口径字典",
    "附录2 个人看板反馈": "附录 2 个人看板反馈",
}

LEADING_SECTION_TITLES = {
    "Agent 质量": "1. Agent 响应体验",
}

TERMINOLOGY_DICTIONARY = {
    # freetrial lifecycle
    "新增 freetrial 用户": {
        "definition": "当日创建 freetrial subscription 的去重用户数。",
        "aliases": [
            "绑卡成功",
            "新增试用用户",
            "新增试订阅用户",
            "新增freetrial",
            "新增 freetrial 用户数",
        ],
    },
    "freetrial 用户": {
        "definition": "当前处于 freetrial 状态的去重用户数。",
        "aliases": ["试用中用户", "freetrial 中用户"],
    },
    "当前 freetrial 订阅数": {
        "definition": "当前 Stripe status=trialing 的 subscription 数。",
        "aliases": ["freetrial 中订阅", "trialing 订阅", "试用中订阅"],
    },
    "到期 freetrial 用户": {
        "definition": "3 天前进入 freetrial、在当日进入 D+3 扣款观察窗口的去重 Stripe customer。",
        "aliases": ["D-3 绑卡成功客户", "到期试用用户"],
    },
    # paid subscription lifecycle
    "新增付费用户": {
        "definition": "当日首次成功扣款的去重 Stripe customer。",
        "aliases": ["当日首次付费用户"],
    },
    "付费用户": {
        "definition": "当前有效且成功扣款过的去重付费 customer。",
        "aliases": ["付费用户数", "付费中用户"],
    },
    "当前有效付费订阅数": {
        "definition": "当前 Stripe status=active，且 customer 历史上至少有一次 successful paid charge 的订阅数；不含 trialing、past_due、canceled。",
        "aliases": ["付费有效订阅", "付费有效订阅数", "付费中订阅", "active 订阅"],
    },
    "累计付费用户": {
        "definition": "历史首次成功扣款过的去重 customer 累计数。",
        "aliases": ["历史付费用户", "曾付费用户"],
    },
    "当前扣款失败订阅数": {
        "definition": "当前扣款失败的 subscription 数。",
        "aliases": ["扣款失败订阅", "past_due 订阅", "逾期订阅 past_due"],
    },
    "扣款失败用户": {
        "definition": "当前或指定观察窗口内扣款失败的去重用户数。",
        "aliases": ["past_due 用户"],
    },
    "取消订阅用户": {
        "definition": "曾有 freetrial 或 paid subscription，但当前已取消或结束的去重用户数。",
        "aliases": ["取消用户", "已结束用户"],
    },
    "截至当日已取消/结束订阅数": {
        "definition": "截至当日结束前，已 ended_at 或 canceled_at 的 subscription 数；当前实现使用 paid_churn_subs/all_churn_subs，是累计存量，不是当日取消量。",
        "aliases": ["取消订阅数", "付费订阅流失", "订阅流失", "paid_churn_subs", "all_churn_subs"],
    },
    # unpaid / non-subscription state
    "未订阅用户": {
        "definition": "历史活跃过，但从未创建过 subscription 的去重用户数。",
        "aliases": ["未付费用户", "未订阅"],
    },
    "未订阅 DAU": {
        "definition": "当日发过至少 1 条消息且不属于 freetrial、付费、扣款失败或取消订阅状态的去重用户数。",
        "aliases": ["未付费 DAU"],
    },
    "未订阅 D1 留存率": {
        "definition": "未订阅新增活跃 cohort 在次日再次活跃的比例。",
        "aliases": ["未付费 D1 留存率"],
    },
    # acquisition / activation denominators
    "新增访客用户数": {
        "definition": "当日首次访问官网或获客入口的去重用户数。",
        "aliases": ["新增 UV", "新增访客 UV"],
    },
    "注册率": {
        "definition": "注册完成用户数 / 对应访问或点击分母。",
        "aliases": ["访问→注册"],
    },
    "开口率": {
        "definition": "新访 DAU / 新增 freetrial 用户。",
        "aliases": ["访问→开口", "同批 freetrial 开口率", "同批新增 freetrial 开口率", "同批 freetrial 开口率 (百分比)"],
    },
    "发消息 DAU": {
        "definition": "当日发过至少 1 条消息的去重用户数。",
        "aliases": ["DAU"],
    },
    "回访发消息 DAU": {
        "definition": "当日发消息但不是首次发消息的去重用户数。",
        "aliases": ["回访 DAU"],
    },
    "D1 回访用户": {
        "definition": "D0 cohort 用户在次日发生页面访问、对话、任务发起或响应任一行为的去重用户数。",
        "aliases": ["次日回访用户数", "次日回访用户"],
    },
    "D1 留存率": {
        "definition": "首次发消息 cohort 在次日再次发消息的比例。",
        "aliases": ["次日留存", "D1 留存"],
    },
    "DAU D1 留存率": {
        "definition": "新增 DAU 次日再次活跃的比例。",
        "aliases": ["D1 留存"],
    },
    "DAU D7 留存率": {
        "definition": "新增 DAU 第 7 天再次活跃的比例。",
        "aliases": ["D7 留存"],
    },
    # finance / revenue labels
    "现金收入": {
        "definition": "当日现金收入。",
        "aliases": [],
    },
    "Stripe 现金收入": {
        "definition": "Stripe 当日成功扣款金额，未扣除退款、手续费和税。",
        "aliases": ["cash_gross"],
    },
    "净现金收入": {
        "definition": "当日现金收入扣除退款后的净额。",
        "aliases": [],
    },
    "付费用户平均收入": {
        "definition": "月经常性收入 / 当前有效付费订阅数。",
        "aliases": ["ARPU"],
    },
    "AI Gateway credits 消耗": {
        "definition": "AI Gateway 请求消耗的 credits 数。",
        "aliases": ["推理成本 credits", "credits_consumed"],
    },
    "含工具调用的 LLM 请求数": {
        "definition": "包含 tool use 的 LLM 请求数。",
        "aliases": ["tool use 请求数", "requests_with_tool_use"],
    },
    "LLM 工具调用次数": {
        "definition": "LLM 响应中实际触发 tool use 的次数。",
        "aliases": ["tool use 次数", "tool_use_count"],
    },
    "tool调用数": {
        "definition": "LLM 响应中实际触发 tool use 的次数。",
        "aliases": ["Tool 调用数", "tool调用数", "tool_use_count"],
    },
    "每条消息平均 tool 调用数": {
        "definition": "tool调用数 / 消息数。",
        "aliases": [],
    },
    "消息 tool 渗透率": {
        "definition": "含 tool 调用的 LLM 请求数 / LLM 请求数。",
        "aliases": ["消息 tool 使用率"],
    },
    "含 tool 消息平均 tool 调用数": {
        "definition": "tool调用数 / 含 tool 调用的 LLM 请求数。",
        "aliases": [],
    },
    "附件上传开始次数": {
        "definition": "附件上传开始事件次数。",
        "aliases": ["附件上传开始", "upload_started"],
    },
    "附件上传用户数": {
        "definition": "发起附件上传的去重用户数。",
        "aliases": ["附件上传用户数", "upload_started_users"],
    },
    # traffic / acquisition
    "UV": {
        "definition": "官网或获客入口的去重访问用户数。",
        "aliases": ["uv"],
    },
    "网站访问会话数": {
        "definition": "官网或获客入口的 pageview session 去重数。",
        "aliases": ["访问会话数", "web sessions"],
    },
    "付费渠道新增 UV": {
        "definition": "新增访客中来自付费投放渠道的去重用户数。",
        "aliases": ["付费渠道新访客 UV", "paid_uv"],
    },
    "付费新增 UV 成本": {
        "definition": "广告花费 / 付费渠道新增 UV。",
        "aliases": [],
    },
    "自然渠道新增 UV": {
        "definition": "新增访客中不属于付费投放渠道的去重用户数。",
        "aliases": ["自然渠道新访客 UV", "organic_uv"],
    },
    "Direct": {
        "definition": "没有可识别来源参数或推荐来源的访问。",
        "aliases": ["直接访问"],
    },
    "SEO / Organic Search": {
        "definition": "来自自然搜索的访问。",
        "aliases": ["SEO / 自然搜索"],
    },
    "Organic Social": {
        "definition": "来自非付费社交渠道的访问。",
        "aliases": ["自然社交"],
    },
    "Referral": {
        "definition": "来自第三方链接推荐的访问。",
        "aliases": ["外链推荐"],
    },
    "Google Ads": {
        "definition": "来自 Google Ads 的广告流量或广告平台数据。",
        "aliases": [],
    },
    "Meta Ads": {
        "definition": "来自 Meta Ads 的广告流量或广告平台数据。",
        "aliases": ["Meta Ads（待数据）"],
    },
    "KOC / KOL（待映射）": {
        "definition": "KOC / KOL 渠道；成本和归因映射接入前显示为空。",
        "aliases": ["KOC / KOL", "KOC / KOL（待成本与映射）"],
    },
    "campaign（待数据）": {
        "definition": "广告 campaign 维度；接入 campaign 映射前显示为空。",
        "aliases": [],
    },
    "CTR": {
        "definition": "广告点击数 / 广告曝光数。",
        "aliases": ["CTR (百分比)"],
    },
    "新增注册": {
        "definition": "新增 UV 中当日完成注册的去重用户数。",
        "aliases": ["新增注册用户", "registered_d0_users", "registered_users"],
    },
    "登录入口点击用户数": {
        "definition": "新增访客中点击登录或注册入口的去重用户数。",
        "aliases": ["new_uv_landing_cta_users", "登录入口点击用户"],
    },
    "登录入口点击率": {
        "definition": "登录入口点击用户数 / 新增访客用户数。",
        "aliases": ["新增 UV 登录 CTA 点击率"],
    },
    "注册完成用户数": {
        "definition": "完成注册的去重用户数。",
        "aliases": ["注册完成", "reg_done"],
    },
    "freetrial 率": {
        "definition": "新增 freetrial 用户 / 新增注册用户。",
        "aliases": ["试订阅率", "注册到 freetrial 率"],
    },
    "付费率": {
        "definition": "新增付费用户 / 新增 UV cohort；缺少同 cohort 付费映射时显示为空。",
        "aliases": ["新增付费率"],
    },
    "注册 CPA": {
        "definition": "广告花费 / 新增注册用户。",
        "aliases": [],
    },
    "广告花费": {
        "definition": "获客投放花费；当前主要来自 Google Ads。",
        "aliases": ["广告花费 (美元)"],
    },
    "付费 CAC": {
        "definition": "广告花费 / 新增付费用户；当前只在成本和归因完整时展示。",
        "aliases": [],
    },
    # activation and retention cohorts
    "新访 DAU": {
        "definition": "D0 新增 freetrial 用户中，D0 发过至少一条 chat:message_sent 的去重用户数。",
        "aliases": [
            "新增 DAU（新访 freetrial 开口）",
            "新增 DAU (新访 freetrial 开口)",
            "同批新增 freetrial 开口用户",
            "同批 freetrial 开口用户",
            "freetrial_message_d0_users",
        ],
    },
    "新增 DAU": {
        "definition": "当日首次发消息的全量去重用户，不限当天是否新访。",
        "aliases": ["new_dau"],
    },
    "回访 DAU": {
        "definition": "当日发消息但不是首次发消息的去重用户数。",
        "aliases": ["returning_dau"],
    },
    "新增 WAU": {
        "definition": "近 7 日活跃用户中，首次发消息日在近 7 日窗口内的去重用户数。",
        "aliases": ["new_wau"],
    },
    "回访 WAU": {
        "definition": "近 7 日活跃用户中，首次发消息日早于近 7 日窗口的去重用户数。",
        "aliases": ["returning_wau"],
    },
    "新增 MAU": {
        "definition": "近 30 日活跃用户中，首次发消息日在近 30 日窗口内的去重用户数。",
        "aliases": ["new_mau"],
    },
    "回访 MAU": {
        "definition": "近 30 日活跃用户中，首次发消息日早于近 30 日窗口的去重用户数。",
        "aliases": ["returning_mau"],
    },
    "WAU": {
        "definition": "近 7 日发过至少 1 条消息的去重用户数。",
        "aliases": [],
    },
    "MAU": {
        "definition": "近 30 日发过至少 1 条消息的去重用户数。",
        "aliases": [],
    },
    "DAU / WAU": {
        "definition": "DAU / WAU，用于观察短周期使用粘性。",
        "aliases": [],
    },
    "DAU / MAU": {
        "definition": "DAU / MAU，用于观察月度使用粘性。",
        "aliases": [],
    },
    "首日人均消息数": {
        "definition": "D0 消息数 / D0 首次发消息用户数。",
        "aliases": ["d0_messages"],
    },
    "前三日人均消息数": {
        "definition": "D0-D2 消息数 / D0 首次发消息用户数。",
        "aliases": ["d0_d2_messages"],
    },
    "首次任务开始用户数": {
        "definition": "首日首次发起任务的去重用户数。",
        "aliases": ["first_task_start_d0_users", "首次发起任务用户数"],
    },
    "首次任务完成用户数": {
        "definition": "首日首次完成任务的去重用户数。",
        "aliases": ["first_task_done_d0_users"],
    },
    "首次发起任务率": {
        "definition": "首次任务开始用户数 / 新增 DAU。",
        "aliases": ["首次发起任务率 (百分比)"],
    },
    "首次任务完成率": {
        "definition": "首次任务完成用户数 / 首次任务开始用户数。",
        "aliases": ["首次任务完成率 (百分比)"],
    },
    "次日发起会话用户数": {
        "definition": "D1 回访用户中再次发起会话的去重用户数。",
        "aliases": ["next_chat_d1_users", "next_chat_users"],
    },
    "次日发起任务用户数": {
        "definition": "D1 回访用户中发起任务的去重用户数。",
        "aliases": ["next_task_start_d1_users", "next_task_start_users"],
    },
    "次日回访发起任务率": {
        "definition": "次日发起任务用户数 / 次日回访用户数。",
        "aliases": ["次日回访发起任务率 (百分比)"],
    },
    "次日完成任务用户数": {
        "definition": "D1 回访用户中完成任务的去重用户数。",
        "aliases": ["next_task_done_d1_users", "next_task_done_users"],
    },
    "次日任务完成率": {
        "definition": "次日完成任务用户数 / 次日发起任务用户数。",
        "aliases": ["次日任务完成率 (百分比)"],
    },
    "次日回访率": {
        "definition": "次日回访用户数 / 新增 DAU。",
        "aliases": ["次日回访率 (百分比)"],
    },
    "次日回访开口率": {
        "definition": "次日发起会话用户数 / 次日回访用户数。",
        "aliases": ["次日回访开口率 (百分比)"],
    },
    "D1 留存用户": {
        "definition": "新增 DAU cohort 在 D1 再次成为 DAU 的去重用户数。",
        "aliases": ["next_return_users"],
    },
    "新增 DAU cohort": {
        "definition": "当日首次发消息用户，作为 DAU 口径留存 cohort 分母。",
        "aliases": ["new_dau_cohort"],
    },
    "D7 留存用户": {
        "definition": "新增 DAU cohort 在 D7 再次成为 DAU 的去重用户数；尚未完整到达 D7 观察窗口的日期显示为空。",
        "aliases": ["d7_chat_users"],
    },
    "D30 留存用户": {
        "definition": "首次发消息 cohort 在 D30 再次发消息的去重用户数；当前数据未接入时显示为空。",
        "aliases": [],
    },
    "D7 留存率": {
        "definition": "新增 DAU cohort 在 D7 再次成为 DAU 的比例。",
        "aliases": ["7 日留存", "D7 留存"],
    },
    "D30 留存率": {
        "definition": "新增 DAU cohort 在 D30 再次成为 DAU 的比例。",
        "aliases": ["D30 留存"],
    },
    # activity frequency
    "活跃频次分层": {
        "definition": "近 30 日发过至少 1 条消息的 MAU 作为分母，按近 7 日发消息频次划分。",
        "aliases": [],
    },
    "重度日用": {
        "definition": "MAU 中近 7 日发消息 >=10 的用户数。",
        "aliases": ["mau_freq_heavy_daily"],
    },
    "中度": {
        "definition": "MAU 中近 7 日发消息 5-9 的用户数。",
        "aliases": ["mau_freq_medium"],
    },
    "轻度": {
        "definition": "MAU 中近 7 日发消息 2-4 的用户数。",
        "aliases": ["mau_freq_light"],
    },
    "低频": {
        "definition": "MAU 中近 7 日发消息 1 的用户数。",
        "aliases": ["mau_freq_low"],
    },
    "冷却": {
        "definition": "MAU 中近 7 日没有发消息，但近 8-30 日发过消息的用户数。",
        "aliases": ["mau_freq_cooling"],
    },
    # capability adoption
    "核心能力采纳": {
        "definition": "核心能力的使用用户数和使用次数分布。",
        "aliases": [],
    },
    "DAU地域分布": {
        "definition": "DAU 按地域维度的分布；国家和地区不拆成两个独立概念。",
        "aliases": [],
    },
    "DAU Interface 分布": {
        "definition": "DAU 按用户入口或交互界面维度的分布。",
        "aliases": [],
    },
    "Web App": {
        "definition": "通过 Web App 与产品交互的 DAU。",
        "aliases": [],
    },
    "IM": {
        "definition": "通过 IM 发消息的 DAU。",
        "aliases": [],
    },
    "Telegram": {
        "definition": "Telegram IM 连接使用用户数及其在 IM 连接用户中的占比。",
        "aliases": [],
    },
    "Slack": {
        "definition": "Slack IM 连接使用用户数及其在 IM 连接用户中的占比。",
        "aliases": [],
    },
    "Discord": {
        "definition": "Discord IM 连接使用用户数及其在 IM 连接用户中的占比。",
        "aliases": [],
    },
    "Google Workspace": {
        "definition": "Google Workspace connector 使用用户数及其在 connector 使用用户中的占比。",
        "aliases": [],
    },
    "Lark": {
        "definition": "Lark connector 使用用户数及其在 connector 使用用户中的占比。",
        "aliases": [],
    },
    "新增使用用户": {
        "definition": "当日首次使用该能力的去重用户数。",
        "aliases": ["new_users"],
    },
    "存量使用用户": {
        "definition": "当日使用该能力、但不是首次使用该能力的去重用户数。",
        "aliases": ["existing_users"],
    },
    "使用次数": {
        "definition": "该能力当天发生的使用事件次数。",
        "aliases": ["events"],
    },
    "IM 连接": {
        "definition": "IM 连接能力的采纳与使用。",
        "aliases": ["IM连接能力"],
    },
    "Connector 连接": {
        "definition": "Workspace connector 等连接能力的采纳与使用。",
        "aliases": ["Workspace connector 连接完成", "connector 连接完成"],
    },
    "定时任务": {
        "definition": "使用定时任务能力的去重用户数。",
        "aliases": ["自动化"],
    },
    "附件上传": {
        "definition": "附件上传能力的采纳与使用。",
        "aliases": [],
    },
    "附件处理次数": {
        "definition": "附件被处理完成的次数。",
        "aliases": ["attach_processed"],
    },
    "附件查看次数": {
        "definition": "附件或文件被查看的次数。",
        "aliases": ["file_viewed"],
    },
    # checkout / finance flow
    "查看价格用户": {
        "definition": "打开价格页或价格入口的去重用户数。",
        "aliases": ["pricing_opened_users", "打开价格页用户"],
    },
    "发起结账用户": {
        "definition": "发起结账流程的去重用户数。",
        "aliases": ["checkout_started_users"],
    },
    "创建结账会话": {
        "definition": "已创建 Stripe 结账会话的去重 checkout flow 数。",
        "aliases": ["has_session_created"],
    },
    "支付验证完成": {
        "definition": "完成支付验证的去重 checkout flow 数。",
        "aliases": ["has_verified"],
    },
    "订单完成": {
        "definition": "完成订单履约的去重 checkout flow 数。",
        "aliases": ["has_fulfilled"],
    },
    "订阅订单": {
        "definition": "订阅类结账 flow 数。",
        "aliases": ["subscription_flows", "fulfilled_subscription"],
    },
    "点数包订单": {
        "definition": "点数包类结账 flow 数。",
        "aliases": ["credit_pack_flows"],
    },
    "结账失败": {
        "definition": "结账发起、支付验证、订单履约失败的合计。",
        "aliases": ["checkout_failed"],
    },
    "结账发起失败": {
        "definition": "结账发起阶段失败数。",
        "aliases": ["checkout_start_failed", "发起失败"],
    },
    "结账验证失败": {
        "definition": "支付验证阶段失败数。",
        "aliases": ["checkout_verify_failed", "验证失败"],
    },
    "结账履约失败": {
        "definition": "订单履约阶段失败数。",
        "aliases": ["checkout_fulfillment_failed", "履约失败"],
    },
    "D+3 新增付费用户": {
        "definition": "D+3 观察日首次成功扣款的 Stripe customer 数。",
        "aliases": ["D+3 成功扣款客户", "cohort_succeeded_charge_customers"],
    },
    "D+3 扣款失败用户": {
        "definition": "D+3 观察日 freetrial 到期扣款失败的 Stripe customer 数。",
        "aliases": ["D+3 扣款失败客户", "cohort_past_due_customers"],
    },
    "D+3 取消订阅用户": {
        "definition": "D+3 观察日 freetrial 到期前后取消订阅的 Stripe customer 数。",
        "aliases": ["D+3 取消客户", "cohort_canceled_customers"],
    },
    "月经常性收入": {
        "definition": "当前有效付费订阅折算的月化经常性收入。",
        "aliases": ["MRR", "mrr_active"],
    },
    "退款": {
        "definition": "当日退款金额。",
        "aliases": ["cash_refund"],
    },
    "净收入": {
        "definition": "当日 Stripe 现金收入 - 当日退款；未扣除支付手续费和税。",
        "aliases": ["cash_net"],
    },
    "月经常性收入 MRR": {
        "definition": "当前有效付费订阅折算的月化经常性收入。",
        "aliases": ["MRR", "mrr_active", "月经常性收入"],
    },
    "推理成本": {
        "definition": "模型推理消耗的 AI Gateway credits。",
        "aliases": ["LLM 成本", "模型推理成本"],
    },
    "推理成本 credits": {
        "definition": "AI Gateway 当日消耗的 credits 汇总。",
        "aliases": [],
    },
    "LLM Cost Dashboard credits": {
        "definition": "PostHog LLM Cost Dashboard 导出的 llm:request_completed credits_consumed。",
        "aliases": [],
    },
    "anthropic credits": {
        "definition": "LLM Cost Dashboard 中 anthropic provider 消耗的 credits。",
        "aliases": [],
    },
    "fal-run credits": {
        "definition": "LLM Cost Dashboard 中 fal-run provider 消耗的 credits。",
        "aliases": [],
    },
    "deepseek credits": {
        "definition": "LLM Cost Dashboard 中 deepseek provider 消耗的 credits。",
        "aliases": [],
    },
    "fal-queue credits": {
        "definition": "LLM Cost Dashboard 中 fal-queue provider 消耗的 credits。",
        "aliases": [],
    },
    "credits": {
        "definition": "AI Gateway 请求消耗的 credits 汇总。",
        "aliases": ["credits_consumed"],
    },
    "credits / 消息": {
        "definition": "AI Gateway credits / 消息数。",
        "aliases": [],
    },
    "credits / 任务": {
        "definition": "AI Gateway credits / 活跃任务数。",
        "aliases": [],
    },
    "云服务成本（待数据）": {
        "definition": "云服务与基础设施成本；当前数据未接入。",
        "aliases": ["云服务成本"],
    },
    "支付手续费（待数据）": {
        "definition": "支付通道手续费；当前数据未接入。",
        "aliases": ["支付手续费"],
    },
    "总服务成本（待数据）": {
        "definition": "推理、云服务、支付手续费等直接服务成本合计。",
        "aliases": ["总服务成本", "总成本（待数据）", "总成本"],
    },
    "沙盒成本（待数据）": {
        "definition": "沙盒或执行环境相关成本；当前数据未接入。",
        "aliases": ["沙盒成本"],
    },
    "服务器成本（待数据）": {
        "definition": "服务器与基础设施运行成本；当前数据未接入。",
        "aliases": ["服务器成本"],
    },
    "毛利（待数据）": {
        "definition": "净收入 - 总服务成本；当前总服务成本未完整接入。",
        "aliases": ["毛利"],
    },
    "毛利率（待数据）": {
        "definition": "毛利 / 净收入；当前总服务成本未完整接入。",
        "aliases": ["毛利率"],
    },
    "广告支出": {
        "definition": "当日获客投放成本；当前已接 Google Ads，其他渠道待接入。",
        "aliases": ["ad_spend"],
    },
    "广告成本": {
        "definition": "当日获客投放成本；当前已接 Google Ads，其他渠道待接入。",
        "aliases": [],
    },
    "Google Ads 花费": {
        "definition": "Google Ads 当日广告花费。",
        "aliases": [],
    },
    "KOC / KOL 花费（待数据）": {
        "definition": "KOC / KOL 渠道花费；当前数据未接入。",
        "aliases": ["KOC / KOL 花费"],
    },
    "其他渠道花费（待数据）": {
        "definition": "其他获客渠道花费，当前未接入。",
        "aliases": ["其他渠道花费"],
    },
    "广告支出 / 新增注册": {
        "definition": "广告支出 / 新增注册用户，用于解释注册 CPA。",
        "aliases": [],
    },
    "广告支出 / 新增付费用户": {
        "definition": "广告支出 / 新增付费用户，用于解释 CAC。",
        "aliases": [],
    },
    "CAC": {
        "definition": "广告支出 / 新增付费用户。",
        "aliases": ["付费 CAC"],
    },
    "贡献利润（待数据）": {
        "definition": "毛利 - 广告支出；当前毛利所需成本数据未完整接入。",
        "aliases": ["贡献利润"],
    },
    # quality / engineering
    "页面首次加载次数": {
        "definition": "页面首次加载事件数量，用于观察前端性能样本量。",
        "aliases": ["page_load_count"],
    },
    "页面首次加载 P50": {
        "definition": "页面首次加载耗时 P50。",
        "aliases": ["page_load_p50_ms", "Chat 页面首次加载 P50"],
    },
    "页面首次加载 P95": {
        "definition": "页面首次加载耗时 P95。",
        "aliases": ["page_load_p95_ms", "Chat 页面首次加载 P95"],
    },
    "消息失败数": {
        "definition": "消息失败事件数。",
        "aliases": ["message_failed", "消息发送失败"],
    },
    "消息数": {
        "definition": "用户发送消息事件数。",
        "aliases": ["messages"],
    },
    "消息失败率": {
        "definition": "消息失败数 / 消息数。",
        "aliases": [],
    },
    "流式错误数": {
        "definition": "AI Gateway 或流式响应相关错误事件数。",
        "aliases": ["stream_errors"],
    },
    "流式错误率": {
        "definition": "流式错误数 / 消息数。",
        "aliases": [],
    },
    "对话流中断": {
        "definition": "对话流式响应中断事件数。",
        "aliases": ["chat_stream_errored"],
    },
    "对话流中断率": {
        "definition": "对话流中断 / 消息数。",
        "aliases": [],
    },
    "实时连接错误": {
        "definition": "实时连接或 WebSocket 相关错误事件数。",
        "aliases": ["ws_error"],
    },
    "实时连接错误率": {
        "definition": "发生实时连接错误的去重用户数 / DAU。",
        "aliases": ["ws_error_users"],
    },
    "API 请求": {
        "definition": "后端 API 请求总数。",
        "aliases": ["api_calls"],
    },
    "API 成功": {
        "definition": "后端 API 成功请求数。",
        "aliases": ["api_success"],
    },
    "API 失败": {
        "definition": "后端 API 失败请求数。",
        "aliases": ["api_fail"],
    },
    "API 失败率": {
        "definition": "后端 API 失败请求数 / 后端 API 请求总数。",
        "aliases": ["API 失败率 (百分比)"],
    },
    "AI Gateway 请求": {
        "definition": "AI Gateway 当日模型请求数。",
        "aliases": ["gateway_requests"],
    },
    "AI Gateway 错误": {
        "definition": "AI Gateway 当日模型请求错误总数。",
        "aliases": ["gateway_errors"],
    },
    "AI Gateway 错误率": {
        "definition": "AI Gateway 错误 / AI Gateway 请求。",
        "aliases": [],
    },
    "流式中断": {
        "definition": "AI Gateway 返回流式响应过程中断的错误次数。",
        "aliases": [],
    },
    "超时": {
        "definition": "AI Gateway 调用模型上游或返回响应超时的错误次数。",
        "aliases": [],
    },
    "上游错误": {
        "definition": "模型 provider 返回错误或不可用导致的 AI Gateway 错误次数。",
        "aliases": [],
    },
    "Bedrock 错误": {
        "definition": "AWS Bedrock 代理或上游返回错误的次数。",
        "aliases": [],
    },
    "首 token P50": {
        "definition": "LLM 首 token 返回耗时 P50。",
        "aliases": ["ttfb_p50_s", "TTFB P50", "LLM 首 token P50", "LLM 首 token P50（秒）"],
    },
    "首 token P95": {
        "definition": "LLM 首 token 返回耗时 P95。",
        "aliases": ["ttfb_p95_s", "TTFB P95", "LLM 首 token P95", "LLM 首 token P95（秒）"],
    },
    "完整响应 P95": {
        "definition": "LLM 完整响应耗时 P95。",
        "aliases": ["request_p95_s"],
    },
    "完整响应 P50": {
        "definition": "LLM 完整响应耗时 P50。",
        "aliases": ["request_p50_s"],
    },
    "完整响应 P99": {
        "definition": "LLM 完整响应耗时 P99。",
        "aliases": ["request_p99_s"],
    },
    "用户首条消息 首 token P50": {
        "definition": "用户首条消息触发的 LLM 首 token 返回耗时 P50。",
        "aliases": ["first_message_ttfb_p50_ms", "用户首条消息 首 token P50（秒）"],
    },
    "用户首条消息 首 token P95": {
        "definition": "用户首条消息触发的 LLM 首 token 返回耗时 P95。",
        "aliases": ["first_message_ttfb_p95_ms", "用户首条消息 首 token P95（秒）"],
    },
    "用户首条消息 完整响应 P50": {
        "definition": "用户首条消息触发的 LLM 完整响应耗时 P50。",
        "aliases": ["first_message_llm_p50_ms", "用户首条消息 完整响应 P50（秒）"],
    },
    "用户首条消息 完整响应 P95": {
        "definition": "用户首条消息触发的 LLM 完整响应耗时 P95。",
        "aliases": ["first_message_llm_p95_ms", "用户首条消息 完整响应 P95（秒）"],
    },
    "沙盒检查次数": {
        "definition": "沙盒或执行环境可用性检查次数。",
        "aliases": ["sandbox_checked"],
    },
    "沙盒可达": {
        "definition": "沙盒或执行环境检查可达的次数。",
        "aliases": ["sandbox_reachable"],
    },
    "沙盒不可达": {
        "definition": "沙盒或执行环境检查不可达的次数。",
        "aliases": ["sandbox_unreachable"],
    },
    "沙盒启动耗时": {
        "definition": "沙盒或执行环境启动耗时分位数。",
        "aliases": [],
    },
    "沙盒启动 P50": {
        "definition": "沙盒或执行环境启动耗时 P50。",
        "aliases": ["env_init_p50_ms"],
    },
    "沙盒启动 P95": {
        "definition": "沙盒或执行环境启动耗时 P95。",
        "aliases": ["env_init_p95_ms"],
    },
    "用户每日首次沙盒启动 P50": {
        "definition": "每个用户每天第一次触发沙盒或执行环境启动的耗时 P50。",
        "aliases": ["env_init_first_daily_p50_ms"],
    },
    "用户每日首次沙盒启动 P95": {
        "definition": "每个用户每天第一次触发沙盒或执行环境启动的耗时 P95。",
        "aliases": ["env_init_first_daily_p95_ms"],
    },
    "沙盒启动次数": {
        "definition": "沙盒或执行环境启动尝试次数。",
        "aliases": ["env_init_count"],
    },
    "沙盒启动成功": {
        "definition": "沙盒或执行环境启动成功次数。",
        "aliases": ["env_init_success"],
    },
    "沙盒启动失败": {
        "definition": "沙盒或执行环境启动失败次数。",
        "aliases": ["env_init_failed"],
    },
    "沙盒启动失败率": {
        "definition": "沙盒或执行环境启动失败次数 / 沙盒启动次数。",
        "aliases": ["沙盒启动失败率 (百分比)"],
    },
    "沙盒重启失败": {
        "definition": "沙盒或执行环境重启失败次数。",
        "aliases": ["sandbox_restart_failed"],
    },
    "沙盒重启失败率（待重启尝试数）": {
        "definition": "沙盒重启失败 / 沙盒重启尝试；当前缺少重启尝试分母，暂不计算。",
        "aliases": ["沙盒重启失败率"],
    },
    "缓存命中率": {
        "definition": "cache read tokens / 输入相关 tokens，按当前数据源口径计算。",
        "aliases": ["cache hit ratio"],
    },
    # retention / churn
    "历史活跃用户": {
        "definition": "历史上至少发过 1 条消息的去重用户数，按当日订阅状态分层。",
        "aliases": [],
    },
    "7+ 天未活跃": {
        "definition": "所在订阅状态用户中，近 7 天没有发消息的用户数。",
        "aliases": ["inactive_7d"],
    },
    "首次发消息用户": {
        "definition": "当日首次发消息的去重用户数，作为新增活跃留存 cohort 分母。",
        "aliases": ["d1_cohort"],
    },
    "未订阅 D1 留存用户": {
        "definition": "未订阅首次发消息用户在次日再次发消息的人数。",
        "aliases": [],
    },
    "freetrial D1 留存用户": {
        "definition": "freetrial 首次发消息用户在次日再次发消息的人数。",
        "aliases": [],
    },
    "付费 D1 留存用户": {
        "definition": "付费首次发消息用户在次日再次发消息的人数。",
        "aliases": [],
    },
    "扣款失败 D1 留存用户": {
        "definition": "扣款失败首次发消息用户在次日再次发消息的人数。",
        "aliases": [],
    },
    "取消订阅 D1 留存用户": {
        "definition": "取消订阅首次发消息用户在次日再次发消息的人数。",
        "aliases": [],
    },
    "D7 留存用户（待数据）": {
        "definition": "首次发消息用户在 D7 再次发消息的人数；当前数据未接入。",
        "aliases": ["D7 留存用户"],
    },
    "D30 留存用户（待数据）": {
        "definition": "首次发消息用户在 D30 再次发消息的人数；当前数据未接入。",
        "aliases": ["D30 留存用户"],
    },
    "回流用户数": {
        "definition": "长期未活跃后再次发消息的用户数。",
        "aliases": ["当周回流"],
    },
    "回流后活跃用户数": {
        "definition": "回流后 7 天内再次达到活跃标准的用户数。",
        "aliases": ["回流后 7 天再激活"],
    },
    "回流后活跃率": {
        "definition": "回流后活跃用户数 / 回流用户数。",
        "aliases": [],
    },
    "回流后 7 天未活跃": {
        "definition": "回流后 7 天内未再次达到活跃标准的用户数。",
        "aliases": ["回流后又流失"],
    },
    # usage depth / tool adoption
    "活跃任务数": {
        "definition": "当日发生任务开始或任务执行事件的任务数。",
        "aliases": [],
    },
    "任务数": {
        "definition": "当日发生任务开始或任务执行事件的任务数。",
        "aliases": [],
    },
    "人均日任务数": {
        "definition": "活跃任务数 / DAU。",
        "aliases": ["人均任务数"],
    },
    "人均任务完成率": {
        "definition": "任务完成数 / 活跃任务数；当前缺少稳定任务完成事件时显示为空。",
        "aliases": [],
    },
    "人均消息数": {
        "definition": "消息数 / DAU。",
        "aliases": [],
    },
    "人均日消息数": {
        "definition": "消息数 / DAU。",
        "aliases": [],
    },
    "Tool / Skill 使用": {
        "definition": "Agent 执行过程中触发 tool 或 skill 的次数。",
        "aliases": [],
    },
    "tool use 请求占比": {
        "definition": "包含 tool use 的 LLM 请求数 / 可使用 tool 的 LLM 请求数。",
        "aliases": [],
    },
    "tool call 成功率": {
        "definition": "tool call 成功次数 / tool call 总次数；待 tool result 明细接入。",
        "aliases": [],
    },
    "tool call 失败率": {
        "definition": "tool call 失败次数 / tool call 总次数；待 tool result 明细接入。",
        "aliases": [],
    },
    "附件上传成功率": {
        "definition": "附件上传完成次数 / 附件上传开始次数。",
        "aliases": [],
    },
    "附件上传失败率": {
        "definition": "附件上传失败次数 / 附件上传开始次数。",
        "aliases": [],
    },
    "IM 连接比例": {
        "definition": "当日通过 IM 连接使用产品的去重用户数 / DAU。",
        "aliases": [],
    },
    "Connector 连接比例": {
        "definition": "当日使用 Connector 连接的去重用户数 / DAU。",
        "aliases": [],
    },
    "定时任务比例": {
        "definition": "当日使用定时任务能力的去重用户数 / DAU。",
        "aliases": [],
    },
    "附件上传比例": {
        "definition": "当日上传附件的去重用户数 / DAU。",
        "aliases": [],
    },
    "会话时长 P50": {
        "definition": "chat session 时长 P50，按 30 分钟无对话活动切分，单位分钟。",
        "aliases": ["对话时长 P50", "活跃时长 P50"],
    },
    "会话时长 P90": {
        "definition": "chat session 时长 P90，按 30 分钟无对话活动切分，单位分钟。",
        "aliases": ["对话时长 P90", "活跃时长 P90"],
    },
}


CONTEXT_SENSITIVE_DISPLAY_ALIASES = {
    "注册率",
    "开口率",
    "新增 UV",
    "DAU",
    "UV",
    "新增 DAU",
    "回访 DAU",
    "新增注册用户",
    "活跃时长 P50",
    "活跃时长 P90",
    "注册完成",
    "注册完成用户数",
    "新增 freetrial 用户",
    "当日首次付费用户",
    "freetrial 率",
    "付费率",
    "次日留存",
    "D1 留存",
    "D1 留存率",
    "ARPU",
    "现金收入",
    "净现金收入",
    "推理成本 credits",
    "订阅订单",
    "使用次数",
    "页面首次加载 P50",
    "页面首次加载 P95",
    "AI Gateway 请求",
}


def terminology_aliases(*, include_context_sensitive: bool = False) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for canonical, spec in TERMINOLOGY_DICTIONARY.items():
        aliases[canonical] = canonical
        for alias in spec["aliases"]:
            if not include_context_sensitive and alias in CONTEXT_SENSITIVE_DISPLAY_ALIASES:
                continue
            aliases[alias] = canonical
    return aliases


LABEL_RENAMES = {
    "绑卡成功": "新增 freetrial 用户",
    "试用订阅": "freetrial",
    "新增试用用户": "新增 freetrial 用户",
    "累计付费用户": "累计付费用户",
    "付费用户数": "付费用户",
    "净 MRR": "MRR",
    "现金收入": "现金收入",
    "净现金收入": "净现金收入",
    "退款": "退款",
    "注册 CPA": "注册 CPA",
    "付费 CAC": "付费 CAC",
    "新增 UV 占比": "新增 UV 占比",
    "CTA CTR": "CTA CTR",
    "对话时长 P50 (分钟)": "对话时长 P50 (分钟)",
    "对话时长 P90 (分钟)": "对话时长 P90 (分钟)",
}

TEXT_RENAMES = {
    "绑卡成功用户总数": "freetrial 用户总数",
    "绑卡漏斗": "freetrial 漏斗",
    "发起绑卡→绑卡成功": "发起绑卡→新增 freetrial",
    "访问→绑卡成功": "访问→新增 freetrial",
    "注册→绑卡成功": "注册→新增 freetrial",
    "绑卡成功": "新增 freetrial",
    "逾期订阅 past_due": "当前扣款失败订阅数",
    "past_due": "扣款失败",
    "Trialing": "freetrial",
    "trialing": "freetrial",
    "LCP P75": "页面加载体验 P75",
    "INP P75": "输入响应 P75",
    "CLS P75": "页面稳定 P75",
    "LLM — 流式完成率": "LLM 流式完成率",
    "TTFB P50": "首字节时间 P50",
    "TTFB P95": "首字节时间 P95",
    "环境初始化失败率": "环境启动失败率",
}

DISPLAY_LABEL_RENAMES = {
    **terminology_aliases(),
    "试订阅率": "freetrial 率",
    "试用订阅": "freetrial",
    "试用用户总数": "freetrial 用户总数",
}

FIELD_DEFINITIONS = {
    "指标": "表格第一列，展示业务指标、分区或下钻维度。",
    "UV": "官网或获客入口的去重访问用户数；不是 DAU，也不是全域活跃。",
    "新增 UV": "当日首次访问官网或获客入口的去重用户数。",
    "新增 UV 占比": "新增 UV / UV。",
    "付费渠道新访客 UV": "来自付费投放渠道的新访问去重用户数。",
    "新增注册用户": "新增 UV cohort 中当日完成注册的去重用户数。",
    "新增注册": "新增 UV cohort 中当日完成注册的去重用户数。",
    "注册开始": "开始注册流程的去重用户数。",
    "注册完成": "完成注册流程的去重用户数。",
    "注册完成率": "注册完成 / 注册开始。",
    "发起绑卡": "开始绑定支付方式的去重用户数。",
    "绑卡开始": "开始绑定支付方式的去重用户数。",
    "新增试用用户": "Stripe 创建 freetrial subscription 的去重用户数。",
    "新增试订阅用户": "Stripe 创建 trial subscription 的去重用户数。",
    "新增 freetrial": "同 cohort 当日创建 freetrial subscription 的去重用户数。",
    "新增freetrial": "同 cohort 当日创建 freetrial subscription 的去重用户数。",
    "新增 freetrial 用户": "同 cohort 当日创建 freetrial subscription 的去重用户数。",
    "新增 freetrial 用户数": "同 cohort 当日创建 freetrial subscription 的去重用户数。",
    "注册完成用户数": "完成注册流程的去重用户数。",
    "首次任务开始用户数": "首次发起任务的去重用户数。",
    "首次任务完成用户数": "首次完成任务的去重用户数。",
    "freetrial 率": "同一批注册用户中当日进入 freetrial 的比例。",
    "新增付费用户": "当日首次成功扣款的去重 Stripe customer。",
    "当日首次付费用户": "Stripe 当日首次成功扣款的去重 customer；不等同新增 UV cohort 付费。",
    "累计付费用户": "历史首次成功扣款过的去重 customer 累计数。",
    "付费用户": "当前有效且成功扣款过的去重付费 customer。",
    "付费用户数": "当前有效且成功扣款过的去重付费 customer。",
    "试用用户总数": "历史或当前口径下创建 freetrial subscription 的去重用户总数，按所在表定义窗口读取。",
    "freetrial 用户总数": "历史或当前口径下创建 freetrial subscription 的去重用户总数，按所在表定义窗口读取。",
    "广告花费": "广告平台当日花费；当前主要来自 Google Ads。",
    "广告花费 (美元)": "广告平台当日花费，单位美元。",
    "付费曝光": "广告平台展示次数。",
    "付费点击": "广告平台点击次数。",
    "KOC / KOL": "KOC / KOL 渠道带来的投放或流量，成本和映射接入前显示 —。",
    "campaign": "广告平台 campaign 维度下钻。",
    "CPM": "广告花费 / 曝光数 * 1000。",
    "CPC": "广告花费 / 点击数。",
    "CTR": "广告点击数 / 广告曝光数。",
    "注册 CPA": "广告花费 / 新增注册用户。",
    "付费 CAC": "广告花费 / 新增付费用户；当前只在成本和归因完整时展示。",
    "DAU": "当日发过至少 1 条消息的去重用户数。",
    "WAU": "近 7 日发过至少 1 条消息的去重用户数。",
    "MAU": "近 30 日发过至少 1 条消息的去重用户数。",
    "新访 DAU": "D0 新增 freetrial 用户中，D0 发过至少一条 chat:message_sent 的去重用户数。",
    "新增 DAU": "当日首次发消息的全量去重用户，不限当天是否新访。",
    "新增 DAU 人均消息数": "新增 DAU 当日平均发送的消息数。",
    "新增 DAU 前三日人均消息数": "新增 DAU 队列在前三日内平均发送的消息数。",
    "首日人均消息数": "新增 DAU 当日平均发送的消息数。",
    "前三日人均消息数": "新增 DAU 队列在前三日内平均发送的消息数。",
    "人均消息数": "消息数 / DAU。",
    "活跃任务数": "当日发生任务开始或任务执行事件的任务数。",
    "人均任务数": "活跃任务数 / DAU。",
    "AI Gateway 请求": "AI Gateway 当日处理的模型请求总数；一条用户消息可能触发多次模型请求。",
    "模型供应商请求": "AI Gateway 按模型供应商拆分的请求数。",
    "模型请求": "AI Gateway 按模型名称拆分的请求数。",
    "AI Gateway 错误": "AI Gateway 当日模型请求错误总数。",
    "入口服务 5xx 率": "入口 ALB 返回 5xx 的比例，(Target 5xx + ELB 5xx) / ALB RequestCount。",
    "入口请求": "入口 ALB 当日收到的请求数。",
    "入口 5xx": "入口 ALB 当日 5xx 总数，包含 Target 5xx 和 ELB 5xx。",
    "Target 5xx": "应用目标服务返回给 ALB 的 5xx 次数。",
    "ELB 5xx": "ALB 自身返回的 5xx 次数。",
    "流式中断": "AI Gateway 返回流式响应过程中断的错误次数。",
    "超时": "AI Gateway 调用模型上游或返回响应超时的错误次数。",
    "上游错误": "模型 provider 返回错误或不可用导致的 AI Gateway 错误次数。",
    "Bedrock 错误": "AWS Bedrock 代理或上游返回错误的次数。",
    "LLM 首 token": "AI Gateway 从收到请求到返回首个 token 的耗时分位数。",
    "首 token P50": "AI Gateway 首 token 耗时 P50。",
    "首 token P95": "AI Gateway 首 token 耗时 P95。",
    "LLM 首 token P95": "AI Gateway 从收到请求到返回首个 token 的 P95 秒数。",
    "完整响应耗时": "AI Gateway 从收到请求到完整响应结束的耗时分位数。",
    "完整响应 P95": "AI Gateway 完整响应耗时 P95。",
    "完整响应 P99": "AI Gateway 完整响应耗时 P99。",
    "页面首次加载次数": "页面首次加载事件数量，用于观察前端性能样本量。",
    "沙盒不可达率": "Sandbox Fleet 监控中 unreachable / checked 的比例。",
    "沙盒检查次数": "Sandbox Fleet 监控检查次数。",
    "沙盒可达": "Sandbox Fleet 监控检查中未发现不可达的次数。",
    "沙盒不可达": "Sandbox Fleet 监控发现 sandbox 不可达的次数。",
    "结账失败率": "结账发起、验证或履约失败 / 发起结账用户。",
    "发起结账用户": "进入 Stripe Checkout 或发起结账流程的去重用户数。",
    "结账失败": "结账发起、验证或履约失败的总次数。",
    "推理成本 credits": "AI Gateway 当日消耗的 credits 汇总。",
    "新增任务完成用户": "当日首次完成任务的新增活跃用户数。",
    "任务完成率": "新增任务完成用户 / 新增 DAU。",
    "首次任务开始": "当日首次发起任务的去重用户数。",
    "首次任务完成": "当日首次完成任务的去重用户数。",
    "首次发起任务率": "首次任务开始用户数 / 新增 DAU。",
    "首次任务完成率": "首次任务完成用户数 / 首次任务开始用户数。",
    "存量国家 / 地区分布": "回访 DAU 按国家或地区拆分的分布。",
    "新增国家 / 地区分布": "新增 DAU 按国家或地区拆分的分布。",
    "存量端类型分布": "回访 DAU 按 Web、移动、IM 等入口拆分的分布。",
    "新增端类型分布": "新增 DAU 按 Web、移动、IM 等入口拆分的分布。",
    "DAU地域分布": "DAU 按地域维度的分布；国家和地区不拆成两个独立概念。",
    "DAU Interface 分布": "DAU 按用户入口或交互界面维度的分布。",
    "新增": "新增 DAU 的分布数量，括号为占比。",
    "回访": "回访 DAU 的分布数量，括号为占比。",
    "历史活跃用户": "历史上至少发过 1 条消息的去重用户数，按当日订阅状态分层。",
    "未订阅用户": "历史活跃过、但没有匹配到 Stripe subscription 的用户数。",
    "freetrial 用户": "历史活跃过、且当日处于 Stripe trialing subscription 的用户数。",
    "付费用户": "历史活跃过、且当日处于 Stripe active subscription 的用户数。",
    "扣款失败用户": "历史活跃过、且当日处于 Stripe past_due 或 unpaid subscription 的用户数。",
    "取消订阅用户": "历史活跃过、曾有 subscription、且当日已取消或结束订阅的用户数。",
    "7+ 天未活跃": "所在订阅状态用户中，近 7 天没有发消息的人数。",
    "首次发消息用户": "当日首次发消息的去重用户数。",
    "D1 留存用户": "首次发消息用户在次日再次发消息的人数。",
    "D7 留存用户（待数据）": "首次发消息用户在 D7 再次发消息的人数；当前数据未接入。",
    "D30 留存用户（待数据）": "首次发消息用户在 D30 再次发消息的人数；当前数据未接入。",
    "未订阅 D1 留存用户": "未订阅首次发消息用户在次日再次发消息的人数。",
    "freetrial D1 留存用户": "freetrial 首次发消息用户在次日再次发消息的人数。",
    "付费 D1 留存用户": "付费首次发消息用户在次日再次发消息的人数。",
    "扣款失败 D1 留存用户": "扣款失败首次发消息用户在次日再次发消息的人数。",
    "取消订阅 D1 留存用户": "取消订阅首次发消息用户在次日再次发消息的人数。",
    "Top 地域": "地域维度下的 Top 分布。",
    "Top Interface": "Interface 维度下的 Top 分布。",
    "渠道": "渠道维度下探。",
    "freetrial DAU": "当日发过至少 1 条消息且处于 freetrial 状态的去重用户数。",
    "付费 DAU": "当日发过至少 1 条消息且已成功扣款的去重用户数。",
    "未订阅 DAU": "当日发过至少 1 条消息且不属于 freetrial、付费、扣款失败或取消订阅状态的去重用户数。",
    "未付费 DAU": "当日发过至少 1 条消息且不属于 freetrial 或付费状态的去重用户数。",
    "freetrial D1 留存率": "freetrial 新增活跃 cohort 在次日再次活跃的比例。",
    "付费 D1 留存率": "付费新增活跃 cohort 在次日再次活跃的比例。",
    "未订阅 D1 留存率": "未订阅新增活跃 cohort 在次日再次活跃的比例。",
    "未付费 D1 留存率": "未付费新增活跃 cohort 在次日再次活跃的比例。",
    "付费 D7 留存率": "付费新增活跃 cohort 在第 7 日或 7 日窗口内再次活跃的比例。",
    "付费 D30 留存率": "付费新增活跃 cohort 在第 30 日或 30 日窗口内再次活跃的比例。",
    "页面首次加载": "页面首次加载事件数量，用于观察前端性能样本量。",
    "页面首次加载 P50": "页面首次加载耗时 P50。",
    "页面首次加载 P95": "页面首次加载耗时 P95。",
        "次日回访用户数": "新增 DAU cohort 次日发生页面访问、对话、任务发起或响应任一行为的去重用户数。",
        "次日发起会话用户数": "新增 DAU cohort 次日发生 chat:session_start 或 chat:message_sent 的去重用户数。",
        "次日发起任务用户数": "新增 DAU cohort 次日发起任务的去重用户数。",
        "次日完成任务用户数": "新增 DAU cohort 次日完成任务的去重用户数。",
    "回访 DAU": "当日发消息但不是首次发消息的去重用户数。",
    "回访 UV": "当日访问官网或获客入口、但不是首次访问的去重用户数。",
    "持续活跃 DAU": "当日发消息，且上一个周期也发过消息的去重用户数。",
    "复活 DAU": "当日发消息、上一个周期未发消息、但历史曾发过消息的去重用户数。",
    "新增流失用户": "本周期未活跃、上一周期仍活跃，因此新进入流失状态的用户数。",
    "DAU / WAU": "DAU / WAU，用于观察短周期使用粘性。",
    "DAU / MAU": "DAU / MAU，用于观察月度使用粘性。",
    "WAU / MAU": "WAU / MAU，用于观察周活跃用户在月活中的覆盖度。",
    "新增活跃用户": "近 7 日活跃用户中，本周首次发消息的去重用户数。",
    "持续活跃用户": "本周活跃、且上周也活跃的去重用户数。",
    "复活用户": "本周活跃、上周未活跃、但历史曾活跃的去重用户数。",
    "新增 WAU": "近 7 日活跃用户中，首次发消息日在近 7 日窗口内的去重用户数。",
    "回访 WAU": "近 7 日活跃用户中，首次发消息日早于近 7 日窗口的去重用户数。",
    "新增 MAU": "近 30 日活跃用户中，首次发消息日在近 30 日窗口内的去重用户数。",
    "回访 MAU": "近 30 日活跃用户中，首次发消息日早于近 30 日窗口的去重用户数。",
    "新增活跃用户占 WAU": "新增活跃用户 / WAU。",
    "持续活跃用户占 WAU": "持续活跃用户 / WAU。",
    "复活用户占 WAU": "复活用户 / WAU。",
    "对话数": "chat session 数，按 30 分钟无对话活动切分。",
    "会话数": "chat session 数，按 30 分钟无对话活动切分。",
    "人均会话数": "会话数 / DAU。",
    "单对话消息数": "chat session 内平均消息数。",
    "单消息对话数": "只有 1 条消息的 chat session 数。",
    "对话时长": "chat session 内首尾活动时间差，按 30 分钟无对话活动切分。",
    "对话时长 P50": "chat session 时长 P50，按 30 分钟无对话活动切分，单位分钟。",
    "对话时长 P90": "chat session 时长 P90，按 30 分钟无对话活动切分，单位分钟。",
    "总对话段": "chat session 总数，按 30 分钟无对话活动切分。",
    "总对话段 (托付量)": "chat session 总数，按 30 分钟无对话活动切分。",
    "次日留存": "首次发消息用户在次日再次发消息的比例。",
    "次日回访用户": "新增 DAU 在次日再次访问或再次使用的去重用户数。",
    "次日回访用户数": "新增 DAU 在次日再次访问或再次使用的去重用户数。",
    "次日回访率": "次日回访用户数 / 新增 DAU。",
    "次日发起对话用户": "新增 DAU 在次日再次发起对话的去重用户数。",
    "次日发起会话用户数": "次日回访用户中发起会话的去重用户数。",
    "次日回访开口率": "次日发起会话用户数 / 次日回访用户数。",
    "次日发起任务用户": "新增 DAU 在次日发起任务的去重用户数。",
    "次日发起任务用户数": "次日回访用户中发起任务的去重用户数。",
    "次日回访发起任务率": "次日发起任务用户数 / 次日回访用户数。",
    "次日完成任务用户": "新增 DAU 在次日完成任务的去重用户数。",
    "次日完成任务用户数": "次日回访用户中完成任务的去重用户数。",
    "次日任务完成率": "次日完成任务用户数 / 次日发起任务用户数。",
    "7 日留存": "新增 DAU cohort 在 D7 再次成为 DAU 的比例。",
    "D1 留存": "首次发消息队列在 D1 再次发消息的比例。",
    "D1 留存率": "首次发消息队列在 D1 再次发消息的比例。",
    "首次发消息队列人数": "当日首次发消息的去重用户数，作为 cohort 留存分母。",
    "D1 留存用户": "首次发消息队列在 D1 留存的用户数。",
    "D7 留存用户": "首次发消息队列在 D7 留存的用户数。",
    "D30 留存用户": "首次发消息队列在 D30 留存的用户数。",
    "D3 留存": "首次发消息队列在 D3 再次发消息的比例。",
    "D3 留存率": "首次发消息队列在 D3 再次发消息的比例。",
    "D7 留存": "首次发消息队列在 D7 再次发消息的比例。",
    "D7 留存率": "首次发消息队列在 D7 再次发消息的比例。",
    "D14 留存": "首次发消息队列在 D14 再次发消息的比例。",
    "D14 留存率": "首次发消息队列在 D14 再次发消息的比例。",
    "D30 留存": "首次发消息队列在 D30 再次发消息的比例。",
    "D30 留存率": "首次发消息队列在 D30 再次发消息的比例。",
    "当日次日留存": "当日观测到的次日留存率；不是 cohort D1 用户数。",
    "当日7日留存": "当日观测到的 7 日窗口留存率；不是 cohort D7 用户数。",
    "CTA 点击": "获客或激活路径中 CTA 被点击的次数或人数，按所在行口径读取。",
    "新增 UV CTA 点击": "新增 UV 用户触发 CTA 点击的总次数；这是事件次数，不是去重点击用户数。",
    "新增 UV CTA 点击次数": "新增 UV 用户触发 CTA 点击的总次数；这是事件次数，不是去重点击用户数。",
    "新增 UV 人均 CTA 点击": "新增 UV CTA 点击次数 / 新增 UV。",
    "新增 UV CTA 点击率": "新增 UV 中至少点击一次 CTA 的用户数 / 新增 UV；缺少去重点用户数时不展示。",
    "新增 UV 登录 CTA 点击次数": "日新增 UV 用户触发登录/注册入口 CTA 的总次数；这是事件次数，不是去重点击用户数。",
    "新增 UV 登录 CTA 点击率": "日新增 UV 中至少点击一次登录/注册入口 CTA 的用户数 / 日新增 UV。",
    "登录入口点击用户数": "日新增 UV 中至少点击一次登录/注册入口 CTA 的去重用户数。",
    "登录入口点击率": "日新增 UV 中至少点击一次登录/注册入口 CTA 的用户数 / 日新增 UV。",
    "注册失败率": "注册失败用户数 / 日新增 UV；用于解释注册率波动。",
    "onboarding 到达用户数": "到达 onboarding 入口页的去重用户数。",
    "onboarding auth gate 用户数": "在 onboarding 路径触发 auth gate 的去重用户数。",
    "usage gate CTA 点击次数": "在 usage gate 触发 CTA 的次数。",
    "referral sidebar CTA 点击次数": "在 referral sidebar 触发 CTA 的次数。",
    "官网主页 CTA 点击": "官网主页 CTA 被点击的次数。",
    "引导页内 CTA 点击": "引导页内 CTA 被点击的次数。",
    "CTA 点击次数 - 回访": "回访用户触发 CTA 点击的次数。",
    "CTA 点击次数 - 新访": "新访用户触发 CTA 点击的次数。",
    "CTA CTR": "CTA 点击数 / 对应访问或曝光分母。",
    "CTA CTR - 回访": "回访用户 CTA 点击数 / 回访 UV。",
    "CTA CTR - 新访": "新增访问用户 CTA 点击数 / 新增 UV。",
    "CTA 点击用户 → 登录完成": "CTA 点击用户中完成登录/注册流程的比例。",
    "访问→注册": "新增注册用户 / UV。",
    "注册率": "登录入口点击用户中完成注册的比例，注册完成用户数 / 登录入口点击用户数。",
    "试订阅率": "完成注册后进入 freetrial 的比例，新增 freetrial / 注册完成。",
    "付费率": "新增 UV cohort 到付费的比例；缺少同 cohort 付费映射时不展示。",
    "注册→发起绑卡": "发起绑卡用户 / 新增注册用户。",
    "访问→发起绑卡": "发起绑卡用户 / UV。",
    "注册→发起绑卡 P50 时长": "从注册完成到发起绑卡的 P50 用时。",
    "注册→发起绑卡 P95 时长": "从注册完成到发起绑卡的 P95 用时。",
    "发起绑卡→新增试用": "新增 freetrial / 发起绑卡用户。",
    "发起绑卡→新增 freetrial": "新增 freetrial / 发起绑卡用户。",
    "绑卡完成率": "同一批注册用户中，进入 freetrial 的用户 / 开始绑卡的用户。",
    "开口率": "新访 DAU / 新增 freetrial 用户。",
    "首次任务开始": "当日首次发起任务的去重用户数。",
    "首次发起任务率": "首次任务开始用户数 / 新增 DAU。",
    "首次任务完成": "当日首次完成任务的去重用户数。",
    "首次任务完成率": "首次任务完成用户数 / 首次任务开始用户数。",
    "新增 DAU 次日回访率": "次日回访用户 / 新增 DAU。",
    "访问→新增试用": "新增 freetrial / UV。",
    "注册→新增试用": "新增 freetrial / 新增注册。",
    "访问→新增 freetrial": "新增 freetrial / UV。",
    "注册→新增 freetrial": "新增 freetrial / 新增注册。",
    "Stripe 现金收入": "Stripe 当日成功扣款金额，未扣除手续费和税。",
    "Stripe 净现金收入": "Stripe 当日成功扣款 - 退款；未扣除手续费和税。",
    "退款": "Stripe 当日退款金额。",
    "D3 应扣款用户": "3 天前新增 freetrial、理论上当日应进入扣款窗口的用户数。",
    "D3 付费用户": "D3 扣款窗口内首次成功扣款用户数。",
    "D3 付费率": "D3 付费用户 / D3 应扣款用户。",
    "D+3 应扣款用户": "3 天前新增 freetrial、理论上当日应进入扣款窗口的用户数。",
    "D+3 付费用户": "D+3 扣款窗口内首次成功扣款用户数。",
    "D+3 付费率": "D+3 付费用户 / D+3 应扣款用户。",
    "D3 扣款成功率": "D3 成功扣款用户 / D3 应扣款用户。",
    "到期扣款失败用户": "freetrial 到期后扣款失败的用户数。",
    "试用期取消用户": "freetrial 期内取消或预约取消的用户数。",
    "MRR": "当前有效付费订阅对应的月化经常性收入。",
    "ARPU": "月经常性收入 / 付费用户。",
    "总成本": "经营成本汇总；成本源未接通时显示 —。",
    "LLM 成本": "模型推理成本。",
    "云服务成本": "云服务与基础设施成本。",
    "毛利": "收入 - 成本。",
    "毛利率": "毛利 / 收入。",
    "任务成功率": "任务完成数 / 任务发起数；不得用 API 成功率冒充。",
    "活跃用户国家 / 地区分布": "活跃用户按国家和地区的分布。",
    "活跃用户端类型分布": "活跃用户按客户端类型的分布。",
    "文件上传用户数": "当日上传文件的去重用户数。",
    "文件上传次数": "当日文件上传次数。",
    "文件上传成功率": "文件上传完成次数 / 文件上传开始次数。",
    "文件下载用户数": "当日下载文件的去重用户数。",
    "文件下载次数": "当日文件下载次数。",
    "定时任务创建用户数": "当日创建定时任务的去重用户数。",
    "定时任务创建数": "当日创建定时任务次数。",
    "定时任务执行用户数": "当日触发定时任务执行的去重用户数。",
    "定时任务执行数": "当日定时任务执行次数。",
    "skill 使用用户数": "当日使用 skill 的去重用户数。",
    "skill 使用比例": "skill 使用用户数 / DAU。",
    "connector 使用用户数": "当日使用 connector 的去重用户数。",
    "connector 使用比例": "connector 使用用户数 / DAU。",
    "流失用户数": "到达应回访窗口但未回访的用户数。",
    "流失率": "流失用户数 / 应回访用户数。",
    "新增流失用户 / WAU": "新增流失用户 / WAU。",
    "付费流失用户数": "真付费有效用户中进入流失状态的人数。",
    "付费流失率": "付费流失用户数 / 真付费有效用户。",
    "回流用户数": "流失后再次活跃的用户数。",
    "回流后活跃用户数": "回流后达到活跃标准的用户数。",
    "回流后活跃率": "回流后活跃用户数 / 回流用户数。",
    "回流后 7 天未活跃": "回流后 7 天内未再次达到活跃标准的人数。",
    "首个任务发起用户": "首次进入有效 Agent 任务流的用户数。",
    "首个任务完成用户": "首次完成有效 Agent 任务的用户数。",
    "首条消息→首个任务完成": "首次发消息用户中完成首个有效 Agent 任务的比例。",
    "端到端成功率": "用户请求从入口到最终完成的成功比例。",
    "消息错误率": "消息链路中出现错误的比例。",
    "WS 连接错误率": "WebSocket 连接错误数 / WebSocket 连接尝试数。",
    "API 调用量": "对应 API 的请求数。",
    "API P50 延迟": "API 响应时间 P50。",
    "API P95 延迟": "API 响应时间 P95。",
    "API — 整体成功率": "API 2xx 成功请求 / 全部请求。",
    "LLM — 流式完成率": "LLM 流式响应正常完成数 / LLM 流式请求数。",
    "LLM 流式完成率": "LLM 流式响应正常完成数 / LLM 流式请求数。",
    "LLM 全失败": "LLM 请求完全失败的次数。",
    "沙盒失败": "沙盒创建、执行或健康检查失败次数。",
    "沙盒失败率": "沙盒失败数 / 沙盒请求数。",
    "沙盒池饱和度": "沙盒资源占用 / 可用容量。",
    "用户主动中止率": "用户主动中断响应或任务的比例。",
    "流式错误": "流式响应过程中发生错误的次数。",
    "中位中止时点": "用户或系统中止发生时，在响应生命周期中的中位位置。",
    "中位循环深度": "Agent 任务执行过程中的中位循环轮次。",
    "中位消息长度": "用户消息长度的中位数。",
    "中位时长": "对应对象时长的中位数。",
    "有附件对话段占比": "包含附件的 chat session / 全部 chat session。",
    "24 小时续聊率": "首次对话后 24 小时内再次发消息的用户比例。",
    "授权通过率": "发起授权后完成授权的比例。",
    "缓存命中率": "命中缓存的请求 / 可缓存请求。",
    "页面加载体验 P75": "页面最大内容完成渲染的 P75 用时。",
    "输入响应 P75": "页面输入交互响应的 P75 用时。",
    "页面稳定 P75": "页面布局稳定性的 P75 指标。",
    "首字节时间 P50": "请求发出到收到首个响应字节的 P50 用时。",
    "首字节时间 P95": "请求发出到收到首个响应字节的 P95 用时。",
    "环境启动失败率": "运行环境初始化失败数 / 初始化请求数。",
    "不可回滚副作用": "执行后不能自动撤销的高风险动作数量。",
    "数量": "对应分组下的计数。",
    "占比": "对应分组数量 / 所在分组总量。",
    "打开账单门户": "打开 Stripe billing portal 的去重用户数。",
    "撞额度墙": "触发额度不足或用量限制提示的去重用户数。",
    "当周回流": "长期未活跃后本周再次发消息的用户数。",
    "回流后 7 天再激活": "回流用户在 7 天内再次达到活跃标准的人数。",
    "回流后又流失": "回流后短期内再次停止活跃的人数。",
    "重度日用": "MAU 用户频次分桶：近 7 天发消息次数 >= 10。",
    "中度": "MAU 用户频次分桶：近 7 天发消息次数 5-9 次。",
    "轻度": "MAU 用户频次分桶：近 7 天发消息次数 2-4 次。",
    "低频": "MAU 用户频次分桶：近 7 天发消息次数 1 次。",
    "冷却": "MAU 用户频次分桶：近 7 天 0 次发消息，但 8-30 天内发过消息。",
    "流失中": "freetrial 用户频次分桶：14-30 天没有发消息。",
    "深度流失": "freetrial 用户频次分桶：>=30 天没有发消息，或从未发过消息。",
    "中断异常": "个人 chat:response_aborted 次数高于自身历史均值显著区间的用户或事件。",
    "之后 30 天真取消": "触发前兆事件后 30 天内真实取消订阅的人数。",
    "7-13 天前撞墙队列后 7 天升级": "7-13 天前撞额度墙的队列，在随后 7 天内升级的人数。",
    "7-13 天前撞墙队列后 7 天未活跃": "7-13 天前撞额度墙的队列，在随后 7 天内未再发消息的人数。",
    "直接访问": "没有可识别来源参数或推荐来源的访问。",
    "Direct": "没有可识别来源参数或推荐来源的访问。",
    "SEO / 自然搜索": "来自自然搜索的访问。",
    "SEO / Organic Search": "来自自然搜索的访问。",
    "自然社交": "来自非付费社交渠道的访问。",
    "Organic Social": "来自非付费社交渠道的访问。",
    "外链推荐": "来自第三方链接推荐的访问。",
    "Referral": "来自第三方链接推荐的访问。",
    "Google Ads": "来自 Google Ads 的广告流量或广告平台数据。",
    "Meta Ads": "来自 Meta Ads 的广告流量或广告平台数据。",
    "Twitter / X": "来自 Twitter / X 的访问。",
    "Reddit": "来自 Reddit 的访问。",
    "Product Hunt": "来自 Product Hunt 的访问。",
    "TikTok": "来自 TikTok 的访问。",
    "YouTube": "来自 YouTube 的访问。",
    "未识别 UTM": "有 UTM 但无法归入已知渠道的访问。",
    "KOC / KOL": "KOC / KOL 渠道；成本和归因映射接入前显示 —。",
    "KOC / KOL（待映射）": "KOC / KOL 渠道；成本和归因映射接入前显示 —。",
    "KOC / KOL（待成本与映射）": "KOC / KOL 渠道；成本和归因映射接入前显示 —。",
    "onboarding 页面": "经 onboarding 路径进入注册或激活流程的用户。",
    "/start": "经 /start 页面进入注册或激活流程的用户。",
    "/usecase": "经 /usecase 页面进入注册或激活流程的用户。",
    "首页": "经首页进入注册或激活流程的用户。",
    "campaign（待数据）": "广告 campaign 维度；接入 campaign 映射前显示 —。",
        "绑卡开始率": "绑卡开始用户数 / 注册完成用户数。",
        "绑卡失败率": "绑卡失败用户数 / 注册完成用户数；用于解释 freetrial 率波动。",
        "绑卡失败": "绑定支付方式过程中发生发起、验证或履约失败的流程数。",
    "结账页连续点击": "结账页同一用户短时间内连续点击同一操作的次数。",
    "登录失败": "登录或注册流程失败的次数或人数，按所在行口径读取。",
    "登录页连续点击": "登录页同一用户短时间内连续点击同一操作的次数。",
    "绑卡成本": "广告花费 / 新增 freetrial 用户。",
    "ROAS": "广告归因收入 / 广告花费。",
    "WebSocket — 连接健康": "WebSocket 连接稳定性分区。",
    "沙盒 — 健康": "沙盒可用性与稳定性分区。",
    "浏览器/前端 — Web 核心指标健康": "浏览器与前端体验指标分区。",
    "端到端 P50 延迟": "用户请求从入口到最终响应完成的 P50 用时。",
    "端到端 P95 延迟": "用户请求从入口到最终响应完成的 P95 用时。",
    "完成率": "完成数 / 发起数。",
    "总请求": "对应服务或路径的请求总数。",
    "回调调用量": "回调接口被调用的次数。",
    "受影响用户": "被错误、故障或性能问题影响的去重用户数。",
    "今日发布": "当日生产环境发布次数。",
    "回滚": "当日生产环境回滚次数。",
    "发版前后 1 小时错误率 Δ": "发布后 1 小时错误率 - 发布前 1 小时错误率。",
    "重启恢复率": "重启后恢复正常的实例或任务 / 需要恢复的总数。",
    "卡顿率": "前端交互或页面渲染发生明显卡顿的比例。",
    "客户端异常": "客户端捕获到的异常次数。",
    "客户端连续点击": "同一控件短时间连续点击导致的异常交互次数。",
    "前端版本切换": "用户会话中发生前端版本切换的次数。",
    "断连率 / 小时": "每小时 WebSocket 或连接断开的比例。",
    "AI / 网络 / 工具三分": "把失败或问题按 AI、网络、工具三类归因的分布。",
    "跨桶 — 副作用治理": "跨任务桶统计不可回滚副作用与治理情况的分区。",
}

MISSING_DEFINITIONS: list[tuple[str, str]] = []

FIELD_DEFINITIONS.update(
    {
        "对话数": "chat session 数，按 30 分钟无对话活动切分。",
        "人均对话数": "对话数 / DAU。",
        "Agent 响应数": "chat:response_received 事件数。",
        "响应率": "Agent 响应数 / 消息数。",
        "消息失败数": "error:message_failed 事件数。",
        "流式错误数": "chat.stream.errored 事件数。",
        "LLM 请求": "llm:request_completed 事件数。",
        "LLM 2xx": "status_code 为 2xx 的 LLM 请求数。",
        "LLM 4xx / 5xx": "status_code 为 4xx 或 5xx 的 LLM 请求数。",
        "LLM 2xx 率": "LLM 2xx / LLM 请求。",
        "LLM 耗时": "llm:request_completed 的 duration_ms 分位数。",
        "credits": "llm:request_completed 的 credits_consumed 汇总。",
        "人均 credits": "credits / DAU。",
        "credits / 消息": "推理 credits / 消息数。",
        "credits / 任务": "推理 credits / 活跃任务数。",
        "活跃频次分层": "近 30 日发过至少 1 条消息的 MAU 作为分母，按近 7 日发消息频次划分为 5 个互斥分桶。",
        "新增 DAU 占比": "新增 DAU / DAU。",
        "回访 DAU 占比": "回访 DAU / DAU。",
        "消息 / 对话": "消息数 / 对话数。",
        "单消息对话占比": "单消息对话数 / 对话数。",
        "Tool / Skill 使用": "Agent 执行过程中触发 tool 或 skill 的次数；用户可感知能力按核心能力采纳统计。",
        "tool use 请求数": "LLM 请求中实际发生 tool use 的请求数。",
        "tool use 请求占比": "包含 tool use 的请求数 / 可使用 tool 的请求数。",
        "tool call 成功率": "tool call 成功次数 / tool call 总次数；待 tool result 明细接入。",
        "tool call 失败率": "tool call 失败次数 / tool call 总次数；待 tool result 明细接入。",
        "Chat 页面首次加载 P95": "Chat 页 perf:page_loaded 的 load_time_ms P95，单位秒。",
        "Chat 页面首次加载 P50": "Chat 页 perf:page_loaded 的 load_time_ms P50，单位秒。",
        "LLM 首 token P50": "llm:request_completed 的 ttfb_ms P50，单位秒。",
        "LLM 首 token P95": "llm:request_completed 的 ttfb_ms P95，单位秒。",
        "用户首条消息 首 token P50": "llm:request_completed 且 message_count=1 的 ttfb_ms P50，单位秒。",
        "用户首条消息 首 token P95": "llm:request_completed 且 message_count=1 的 ttfb_ms P95，单位秒。",
        "完整响应 P50": "llm:request_completed 的 duration_ms P50，单位秒。",
        "完整响应 P95": "llm:request_completed 的 duration_ms P95，单位秒。",
        "用户首条消息 完整响应 P50": "llm:request_completed 且 message_count=1 的 duration_ms P50，单位秒。",
        "用户首条消息 完整响应 P95": "llm:request_completed 且 message_count=1 的 duration_ms P95，单位秒。",
        "流式错误率": "流式错误数 / 消息数。",
        "对话流中断率": "对话流中断数 / 消息数。",
        "沙盒重启失败率（待重启尝试数）": "沙盒重启失败 / 沙盒重启尝试；当前缺少重启尝试分母，暂不计算。",
        "tokens": "LLM 请求的 token 消耗汇总。",
        "input tokens": "llm:request_completed 的 input_tokens 汇总。",
        "output tokens": "llm:request_completed 的 output_tokens 汇总。",
        "cache read tokens": "llm:request_completed 的 cache_read_tokens 汇总。",
    "IM 连接": "当日使用 IM 连接能力的去重用户数。",
    "Connector 连接": "当日使用非 IM 外部 connector 能力的去重用户数。",
    "定时任务": "当日调用 /api/schedules 的去重用户数；括号为 DAU 占比，DAU 为当日有过对话的用户。",
    "附件上传": "当日上传附件的去重用户数。",
    "DAU 占比": "使用该能力的去重用户数 / DAU。",
    "新增使用用户": "当日首次使用该能力的去重用户数。",
    "存量使用用户": "当日使用该能力但不是首次使用该能力的去重用户数。",
    "新增使用 DAU 占比": "当日首次使用该能力的去重用户数 / DAU。",
    "使用次数": "该能力对应使用事件的次数。",
    "下探项": "该能力下的具体类型或渠道下探。",
    "Web App": "通过 Web App 与产品交互的 DAU。",
    "IM": "通过 IM 与产品交互的 DAU；只统计真实对话入口，不用连接成功事件代替。",
    "Telegram": "Telegram IM 连接使用用户数及其在 IM 连接用户中的占比。",
    "Slack": "Slack IM 连接使用用户数及其在 IM 连接用户中的占比。",
    "Discord": "Discord IM 连接使用用户数及其在 IM 连接用户中的占比。",
    "Lark": "Lark IM 连接使用用户数及其在 IM 连接用户中的占比。",
    "Google Workspace": "Google Workspace connector 使用用户数及其在 Connector 连接用户中的占比。",
    "附件处理次数": "附件被后端处理的次数。",
    "附件查看次数": "附件被打开查看的次数。",
    "授权失败率": "授权失败数 / 授权尝试数。",
    "调用成功率": "调用成功数 / 调用尝试数。",
    "附件上传成功率": "附件上传成功次数 / 附件上传次数。",
    "附件上传失败率": "附件上传失败次数 / 附件上传次数。",
    "connector 连接完成": "connector:*_connected 事件数汇总。",
    "connector 连接完成用户数": "当日完成 connector 连接的去重用户数。",
    "connector 连接完成 DAU 占比": "connector 连接完成用户数 / DAU。",
    "IM 连接完成": "Telegram、Slack、Discord、Lark 等 IM connector 连接完成事件数汇总。",
    "IM 连接完成用户数": "当日完成 IM connector 连接的去重用户数。",
    "IM 连接完成 DAU 占比": "IM 连接完成用户数 / DAU。",
    "IM 存量连接用户数": "截至当日仍有 IM connector 连接关系的去重用户数；待存量快照接入。",
    "IM 存量连接 DAU 占比": "IM 存量连接用户数 / DAU；待存量快照接入。",
    "Workspace connector 连接完成": "Google Workspace 等工作区类 connector 连接完成事件数汇总。",
    "Workspace connector 连接完成用户数": "当日完成工作区类 connector 连接的去重用户数。",
    "Workspace connector 连接完成 DAU 占比": "Workspace connector 连接完成用户数 / DAU。",
    "Workspace connector 存量连接用户数": "截至当日仍有工作区类 connector 连接关系的去重用户数；待存量快照接入。",
    "Workspace connector 存量连接 DAU 占比": "Workspace connector 存量连接用户数 / DAU；待存量快照接入。",
    "附件上传用户 DAU 占比": "附件上传用户数 / DAU。",
    "附件处理次数": "附件被后端处理的次数。",
    "人均附件处理次数": "附件处理次数 / DAU。",
    "自动化任务新增数": "当日新增自动化任务数；待自动化事件接入。",
    "自动化任务新增 DAU 占比": "自动化任务新增用户数 / DAU；待自动化事件接入。",
    "自动化任务运行数": "当日自动化任务运行次数；待自动化事件接入。",
    "自动化任务运行 DAU 占比": "自动化任务运行用户数 / DAU；待自动化事件接入。",
    "skill 使用次数": "当日 skill 使用次数；待 skill 事件接入。",
    "skill 使用 DAU 占比": "skill 使用用户数 / DAU；待 skill 事件接入。",
        "Google Workspace 设置完成": "connector:google_workspace_folder_set 事件数。",
        "Google Workspace picker unavailable": "connector:google_workspace_picker_unavailable 事件数。",
        "Google Workspace 连接率": "Google Workspace 设置完成 / (Google Workspace 设置完成 + Google Workspace picker unavailable)。",
        "API 请求": "server:api_called 事件数。",
        "API 成功": "server:api_called 中 status_code 为 2xx/3xx 的请求数。",
        "API 失败": "server:api_called 中 status_code 为 4xx/5xx 的请求数。",
        "API 成功率": "API 成功 / API 请求。",
        "API 失败率": "API 失败 / API 请求。",
        "API 耗时": "server:api_called 的 latency_ms 分位数。",
        "API 耗时 P50": "server:api_called 的 latency_ms P50；展示时换算为秒。",
        "API 耗时 P95": "server:api_called 的 latency_ms P95；展示时换算为秒。",
        "API 耗时 P99": "server:api_called 的 latency_ms P99；展示时换算为秒。",
        "实时连接错误": "WebSocket 连接错误事件数。",
        "实时连接错误率": "发生 WebSocket 连接错误的去重用户数 / DAU。",
        "消息发送失败": "消息发送失败事件数。",
        "对话流中断": "对话流返回过程中断或报错的事件数。",
        "沙盒启动次数": "需要执行环境的任务触发沙盒环境初始化的次数。",
        "沙盒启动成功": "沙盒环境初始化成功次数。",
        "沙盒启动失败": "沙盒环境初始化失败次数。",
        "沙盒启动失败率": "沙盒启动失败 / 沙盒启动次数。",
        "沙盒启动耗时": "沙盒环境初始化耗时分位数。",
        "沙盒启动 P50": "沙盒环境初始化耗时 P50。",
        "沙盒启动 P95": "沙盒环境初始化耗时 P95。",
        "用户每日首次沙盒启动 P50": "每个用户每天第一次触发沙盒环境初始化的耗时 P50。",
        "用户每日首次沙盒启动 P95": "每个用户每天第一次触发沙盒环境初始化的耗时 P95。",
        "沙盒重启失败": "已有沙盒热恢复或重启失败次数。",
        "附件上传开始": "用户开始上传附件的次数。",
        "附件上传成功": "附件上传成功完成的次数。",
        "附件上传成功率": "附件上传成功 / 附件上传开始。",
        "附件上传失败": "附件上传失败次数。",
        "附件上传失败率": "附件上传失败 / 附件上传开始。",
        "结账发起失败": "支付或订阅结账发起失败次数。",
        "结账验证失败": "支付或订阅结账验证失败次数。",
        "结账履约失败": "支付或订阅订单履约失败次数。",
        "结账失败率": "结账失败数 / 创建结账会话。",
        "pricing opened 用户": "触发 pricing:dialog_opened 的去重用户数。",
        "checkout started 用户": "触发 payment:checkout_started 的去重用户数。",
        "pricing → checkout started": "checkout started 用户 / pricing opened 用户。",
        "查看价格用户": "打开价格或套餐入口的去重用户数。",
        "发起结账用户": "开始结账流程的去重用户数。",
        "价格页发起结账率": "发起结账用户 / 查看价格用户。",
        "checkout session created flows": "按 checkout_flow_id 去重的 payment:checkout_session_created flow 数。",
        "checkout verified flows": "按 checkout_flow_id 去重的 payment:checkout_verified flow 数。",
        "session created → verified": "checkout verified flows / checkout session created flows。",
        "checkout fulfilled flows": "按 checkout_flow_id 去重的 payment:checkout_fulfilled flow 数。",
        "session created → fulfilled": "checkout fulfilled flows / checkout session created flows。",
        "checkout failed": "checkout start / verify / fulfillment failed 汇总。",
        "subscription flows": "订阅类结账流程数。",
        "credit pack flows": "点数包类结账流程数。",
        "创建结账会话": "已创建 Stripe 结账会话的去重结账流程数。",
        "支付验证完成": "完成支付验证的去重结账流程数。",
        "结账验证率": "支付验证完成 / 创建结账会话。",
        "订单完成": "已完成履约的去重结账流程数。",
        "订单完成率": "订单完成 / 创建结账会话。",
        "结账失败": "结账发起、支付验证、订单履约失败的合计。",
        "发起失败": "结账发起阶段失败数。",
        "验证失败": "支付验证阶段失败数。",
        "履约失败": "订单履约阶段失败数。",
        "订阅订单": "订阅类结账流程数。",
        "点数包订单": "点数包类结账流程数。",
        "到期 freetrial 用户": "3 天前进入 freetrial、在当日进入 D+3 扣款观察窗口的去重 Stripe customer。",
        "D-3 绑卡成功客户": "3 天前进入 freetrial、在当日进入 D+3 扣款观察窗口的去重 Stripe customer。",
        "D+3 成功扣款客户": "D+3 观察日首次成功扣款的 Stripe customer。",
        "D+3 新增付费用户": "D+3 观察日首次成功扣款的 Stripe customer。",
        "扣款成功率": "D+3 新增付费用户 / 到期 freetrial 用户。",
        "D+3 past_due 客户": "D+3 观察日扣款失败并进入 past_due 的 Stripe customer。",
        "past_due 比例": "D+3 past_due 客户 / 到期 freetrial 用户。",
        "D+3 canceled 客户": "D+3 观察日已取消的 Stripe customer。",
        "canceled 比例": "D+3 canceled 客户 / 到期 freetrial 用户。",
        "D+3 扣款失败用户": "D+3 观察日 freetrial 到期扣款失败的客户数。",
        "D+3 扣款失败客户": "freetrial 到期扣款失败的客户数。",
        "扣款失败用户": "freetrial 到期扣款失败的客户数。",
        "扣款失败率": "扣款失败用户 / 到期 freetrial 用户。",
        "D+3 取消客户": "freetrial 到期前后取消的客户数。",
        "D+3 取消订阅用户": "D+3 观察日 freetrial 到期前后取消订阅的客户数。",
        "取消订阅用户": "freetrial 到期前后取消订阅的客户数。",
        "取消用户": "freetrial 到期前后取消的客户数。",
        "取消率": "D+3 取消订阅用户 / 到期 freetrial 用户。",
        "active 订阅": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "当前有效付费订阅数": "当前 Stripe status=active，且 customer 历史上至少有一次 successful paid charge 的订阅数；不含 trialing、past_due、canceled。",
        "付费有效订阅": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "付费有效订阅数": "当前 Stripe status=active 的订阅数，不含 trialing freetrial 订阅。",
        "当前 freetrial 订阅数": "当前 Stripe status=trialing 的订阅数。",
        "freetrial 中订阅": "当前 Stripe status=trialing 的订阅数。",
        "trialing 订阅": "当前 Stripe status=trialing 的订阅数。",
        "当前扣款失败订阅数": "当前扣款失败的订阅数。",
        "past_due 订阅": "当前扣款失败的订阅数。",
        "扣款失败订阅": "当前扣款失败的订阅数。",
        "月经常性收入": "当前有效付费订阅折算的月经常性收入。",
        "付费用户平均收入": "月经常性收入 / 当前有效付费订阅数。",
        "截至当日已取消/结束订阅数": "截至当日结束前，已 ended_at 或 canceled_at 的 subscription 数；当前实现使用 paid_churn_subs/all_churn_subs，是累计存量，不是当日取消量。",
        "取消订阅数": "截至当日结束前，已 ended_at 或 canceled_at 的 subscription 数；当前实现使用 paid_churn_subs/all_churn_subs，是累计存量，不是当日取消量。",
        "付费订阅流失": "截至当日结束前，已 ended_at 或 canceled_at 的 subscription 数；当前实现使用 paid_churn_subs/all_churn_subs，是累计存量，不是当日取消量。",
    }
)

FIELD_DEFINITIONS.update(
    {canonical: spec["definition"] for canonical, spec in TERMINOLOGY_DICTIONARY.items()}
)


MISSING = {"", "—", "-", "–", "None", "none", "null"}
DERIVED_METRIC_TOKENS = [
    "率",
    "占比",
    "CTR",
    "CPC",
    "CPM",
    "CPA",
    "CAC",
    "ARPU",
    "MRR",
    "P50",
    "P90",
    "P95",
    "P99",
]
PAID_EFFICIENCY_NO_DRILLDOWN = {"CPM", "CTR", "CPC"}
PAID_PLATFORM_LABELS = {"Google Ads", "Meta Ads"}
CHANNEL_LABELS = {
    "直接访问",
    "SEO / 自然搜索",
    "自然社交",
    "外链推荐",
    "Google Ads",
    "Meta Ads",
    "Twitter / X",
    "Reddit",
    "Product Hunt",
    "TikTok",
    "YouTube",
    "未识别 UTM",
}


def is_missing_text(value: str) -> bool:
    return value.strip() in MISSING


def metric_kind(label: str) -> str:
    if is_percent_label(label):
        return "percent"
    if any(k in label for k in ["美元", "现金", "收入", "退款", "MRR", "ARPU", "CAC", "CPA", "CPC", "CPM", "成本", "毛利", "广告花费"]):
        return "currency"
    if any(k in label for k in ["时长", "时点", "延迟", "P50", "P90", "P95", "P99"]):
        return "decimal"
    return "number"


def is_derived_metric_label(label: str) -> bool:
    return any(token in label for token in DERIVED_METRIC_TOKENS)


def is_paid_efficiency_label(label: str) -> bool:
    normalized = normalize_label_for_definition(label)
    return normalized in PAID_EFFICIENCY_NO_DRILLDOWN


def derived_child_label(label: str, parent_label: str) -> str:
    normalized_label = normalize_label_for_definition(label)
    if normalized_label not in CHANNEL_LABELS:
        return label
    normalized_parent = normalize_label_for_definition(parent_label)
    suffix = ""
    if "占比" in normalized_parent:
        suffix = "占比"
    elif "CTR" in normalized_parent:
        suffix = "CTR"
    elif "CPM" in normalized_parent:
        suffix = "CPM"
    elif "CPC" in normalized_parent:
        suffix = "CPC"
    elif "CPA" in normalized_parent:
        suffix = "CPA"
    elif "CAC" in normalized_parent:
        suffix = "CAC"
    elif "率" in normalized_parent:
        suffix = "率"
    if not suffix or suffix in label:
        return label
    return f"{label} {suffix}"


def fmt_number(value: float, *, decimals: int = 1, currency: bool = False) -> str:
    sign = "-" if value < 0 else ""
    abs_value = abs(value)
    if decimals == 0 or abs_value == int(abs_value):
        body = f"{int(abs_value):,}"
    else:
        body = f"{abs_value:,.{decimals}f}".rstrip("0").rstrip(".")
    return f"{sign}${body}" if currency else f"{sign}{body}"


def fmt_value(value, number_format: str = "", *, kind: str = "number") -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        nf = (number_format or "").lower()
        if "%" in nf or kind == "percent":
            return f"{value * 100:.1f}%"
        if kind == "currency":
            return fmt_number(float(value), currency=True)
        if kind == "decimal":
            return fmt_number(float(value), decimals=1)
        return fmt_number(float(value), decimals=1)
    return str(value)


def parse_display_number(value: str) -> float | None:
    cleaned = value.strip()
    if not cleaned or is_missing_text(cleaned) or "%" in cleaned:
        if "(" not in cleaned:
            return None
        cleaned = cleaned.split("(", 1)[0].strip()
        if not cleaned or is_missing_text(cleaned) or "%" in cleaned:
            return None
    cleaned = cleaned.replace("$", "").replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def append_child_share(value: str, child: float | None, parent: float | None) -> str:
    if child is None or parent in (None, 0):
        return value
    return f"{value} ({child / parent * 100:.0f}%)"


SHARE_ANNOTATION_RE = re.compile(r"^(.+?)\s+\(-?\d+(?:\.\d+)?%\)$")


def strip_share_annotation(value: str) -> str:
    match = SHARE_ANNOTATION_RE.match(value.strip())
    return match.group(1) if match else value


def has_present_value(value: str) -> bool:
    cleaned = value.strip()
    return bool(cleaned) and not is_missing_text(cleaned)


def has_quantitative_value(value: str) -> bool:
    cleaned = value.strip()
    if not cleaned or is_missing_text(cleaned):
        return False
    if "%" in cleaned:
        numeric = cleaned.split("%", 1)[0].split("(")[-1].strip()
        try:
            float(numeric.replace(",", "").replace("$", ""))
            return True
        except ValueError:
            return False
    return parse_display_number(cleaned) is not None


def has_nonzero_value(value: str) -> bool:
    if not has_present_value(value):
        return False
    number = parse_display_number(value)
    if number is None:
        return True
    return number != 0


def filter_empty_rows(rows: list[dict]) -> list[dict]:
    keep_ids: set[str] = set()
    row_by_id = {row["id"]: row for row in rows}

    for row in rows:
        if row["kind"] in {"header", "section"}:
            continue
        values = row["values"][1:]
        if row.get("keep_empty"):
            keep = True
        elif row["kind"] == "child":
            keep = any(has_nonzero_value(value) for value in values)
        else:
            keep = any(has_present_value(value) for value in values)
        if keep:
            keep_ids.add(row["id"])

    changed = True
    while changed:
        changed = False
        for row_id in list(keep_ids):
            parent_id = row_by_id.get(row_id, {}).get("parent")
            if parent_id and parent_id not in keep_ids:
                keep_ids.add(parent_id)
                changed = True

    section_has_rows: dict[str, bool] = {}
    current_section = ""
    for row in rows:
        if row["kind"] == "section":
            current_section = row["id"]
            section_has_rows[current_section] = False
            continue
        if current_section and row["id"] in keep_ids:
            section_has_rows[current_section] = True

    filtered = []
    for row in rows:
        if row["kind"] == "header":
            filtered.append(row)
        elif row["kind"] == "section":
            if section_has_rows.get(row["id"]):
                filtered.append(row)
        elif row["id"] in keep_ids:
            filtered.append(row)

    child_ids = {row["parent"] for row in filtered if row.get("parent")}
    default_open_ids = {row["id"] for row in filtered if row.get("default_open")}
    for row in filtered:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child" and row.get("parent") not in default_open_ids
    return filtered


def filter_non_quantitative_metric_rows(rows: list[dict]) -> list[dict]:
    removed = {
        row["id"]
        for row in rows
        if row["kind"] not in {"header", "section"}
        and any(has_present_value(value) for value in row["values"][1:])
        and not any(has_quantitative_value(value) for value in row["values"][1:])
    }
    if not removed:
        return rows
    for row in rows:
        while row.get("parent") in removed:
            parent = next((candidate for candidate in rows if candidate["id"] == row["parent"]), None)
            if not parent:
                row["parent"] = ""
                break
            row["parent"] = parent.get("parent", "")
            row["level"] = max(0, row.get("level", 0) - 1)
    kept = [row for row in rows if row["id"] not in removed]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


SHARE_BREAKDOWN_EXACT_LABELS = {
    "付费渠道新访客 UV",
    "新增 UV",
    "回访 UV",
    "新增 DAU",
    "回访 DAU",
    "持续活跃 DAU",
    "复活 DAU",
    "新增流失用户",
    "单消息对话数",
    "新增",
    "持续活跃",
    "复活",
    "流失转入",
    "重度日用",
    "中度",
    "轻度",
    "低频",
    "冷却",
    "流失中",
    "深度流失",
}

SHARE_BREAKDOWN_DENY_TOKENS = [
    "平均",
    "人均",
    "单对话",
    "P50",
    "P75",
    "P90",
    "P95",
    "P99",
    "时长",
    "延迟",
    "CPA",
    "CAC",
    "CPM",
    "CPC",
    "CTR",
    "ARPU",
    "MRR",
    "率",
    "占比",
]


def is_share_breakdown_child(child_label: str, parent_label: str) -> bool:
    child = normalize_label_for_definition(child_label)
    parent = normalize_label_for_definition(parent_label)
    if any(token in child for token in SHARE_BREAKDOWN_DENY_TOKENS):
        return False
    if child.startswith("失败："):
        return True
    if child in CHANNEL_LABELS:
        return True
    if child in SHARE_BREAKDOWN_EXACT_LABELS:
        return True
    if child in {"新增试用用户", "新增付费用户"} and "注册" in parent:
        return False
    return False


FLOW_OR_STAT_TOKENS = [
    "→",
    "P50",
    "P75",
    "P90",
    "P95",
    "P99",
    "平均",
    "人均",
    "单对话",
    "单消息",
    "中位",
    "时长",
    "延迟",
    "率",
    "占比",
    "CTR",
    "CPC",
    "CPM",
    "CPA",
    "CAC",
    "ARPU",
    "MRR",
]


def is_valid_drilldown(parent_label: str, child_label: str) -> bool:
    parent = normalize_label_for_definition(parent_label)
    child = normalize_label_for_definition(child_label)
    if not parent or not child:
        return False
    channel_base = next((channel for channel in CHANNEL_LABELS if child == channel or child.startswith(f"{channel} ")), "")
    if channel_base:
        return parent in {
            "UV",
            "新增 UV",
            "CTA 点击",
            "新增注册用户",
            "新增试用用户",
            "广告花费",
            "付费曝光",
            "付费点击",
            "新增 UV 占比",
            "CTR",
            "CPC",
            "CPM",
            "注册 CPA",
            "付费 CAC",
        }
    if any(token in child for token in FLOW_OR_STAT_TOKENS):
        return False
    if child in PAID_PLATFORM_LABELS:
        return parent in {"广告花费", "付费曝光", "付费点击"}
    if parent == "UV" and child in {"新增 UV", "回访 UV"}:
        return True
    if parent == "DAU" and child in {"新增 DAU", "回访 DAU"}:
        return True
    if parent in {"登录失败", "绑卡失败"} and child.startswith("失败："):
        return True
    if parent in {"试用用户总数", "活跃频次分层"} and child in {"重度日用", "中度", "轻度", "低频", "冷却"}:
        return True
    if parent == "净现金收入" and child in {"新增付费用户", "退款"}:
        return False
    return False


def normalize_drilldown_relationships(rows: list[dict]) -> list[dict]:
    by_id = {row["id"]: row for row in rows}
    for row in rows:
        parent_id = row.get("parent")
        if not parent_id:
            continue
        parent = by_id.get(parent_id)
        if not parent or is_valid_drilldown(parent["values"][0], row["values"][0]):
            continue
        row["parent"] = ""
        row["hidden"] = False
        if row["kind"] == "child":
            row["kind"] = "derived" if row.get("ratio") or is_ratio_label(row["values"][0]) else "row"
        row["level"] = 0

    child_ids = {row["parent"] for row in rows if row.get("parent")}
    for row in rows:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return rows


DASHBOARD_EXCLUDE_LABELS = {
    "付费渠道新访客 UV",
    "单对话消息数",
    "单消息对话数",
}


def prune_dashboard_summary(rows: list[dict]) -> list[dict]:
    kept = [
        row
        for row in rows
        if row["kind"] in {"header", "section"}
        or normalize_label_for_definition(row["values"][0]) not in DASHBOARD_EXCLUDE_LABELS
    ]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


def remove_single_child_share_annotations(rows: list[dict]) -> list[dict]:
    children_by_parent: dict[str, list[dict]] = {}
    for row in rows:
        if row.get("parent") and row["kind"] == "child":
            children_by_parent.setdefault(row["parent"], []).append(row)

    for children in children_by_parent.values():
        if len(children) != 1:
            continue
        child = children[0]
        child["values"] = [child["values"][0]] + [strip_share_annotation(value) for value in child["values"][1:]]
    return rows


def strip_marker(text: str) -> tuple[str, bool]:
    value = text.strip()
    had_marker = bool(re.match(r"^(└|┄|·)\s*", value))
    value = re.sub(r"^(└|┄|·)\s*", "", value)
    for src, dst in TEXT_RENAMES.items():
        value = value.replace(src, dst)
    return LABEL_RENAMES.get(value, value), had_marker


def clean_text(text: str) -> str:
    value = text
    for src, dst in TEXT_RENAMES.items():
        value = value.replace(src, dst)
    return value


def raw_cell_text(cell) -> str:
    return clean_text(fmt_value(cell.value, cell.number_format))


def used_bounds(ws) -> tuple[int, int]:
    max_row = max_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def is_section(values: list[str]) -> bool:
    first = values[0].strip() if values else ""
    return bool(re.match(r"^\d+[\.\、]\s*\S+", first)) and all(not v.strip() for v in values[1:])


def is_header_row(row_index: int, values: list[str]) -> bool:
    return row_index == 1 and values and values[0] in {"指标", "反馈对象", "口径项", "指标名称", "业务指标名"}


def is_ratio_label(label: str) -> bool:
    return any(k in label for k in ["率", "占比", "CTR", "CPC", "CPM", "CPA", "CAC", "ARPU", "MRR"])


def is_percent_label(label: str) -> bool:
    return any(k in label for k in ["率", "占比", "CTR", "留存", "转化", "粘性", "DAU / WAU", "DAU / MAU"])


UNIT_SUFFIX_RE = re.compile(r"\s*[（(](个|美元|百分比|毫秒|秒|分钟|人|次|天|月|%|\$)[）)]\s*$")
DISPLAY_UNIT_RE = re.compile(r"\s*[（(](个|美元|百分比|人|次|天|月|%|\$)[）)]")
SECTION_PREFIX_RE = re.compile(r"^\d+[\.\、]\s*")


def slug_key(value) -> str:
    text = html.unescape(str(value or "")).strip().lower()
    text = SECTION_PREFIX_RE.sub("", text)
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "-", text)
    text = text.strip("-")
    return text or "blank"


def column_key(label, index: int) -> str:
    label_key = slug_key(label)
    return f"col-{index}-{label_key}" if label_key != "blank" else f"col-{index}"


def row_anchor_key(row: dict) -> str:
    return slug_key(row.get("id") or row.get("values", [""])[0])


def section_anchor_key(section_row: dict | None) -> str:
    if not section_row:
        return ""
    return row_anchor_key(section_row)


def anchor_id_from_parts(*, sheet: str, section: str, row: str, column: str, anchor_type: str) -> str:
    return ":".join(
        [
            PAGE_KEY,
            PAGE_VERSION,
            sheet,
            section or "_",
            row or "_",
            column or "_",
            anchor_type,
        ]
    )


def anchor_label(value) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(value or "")).strip())


def anchor_attrs(
    anchor_type: str,
    *,
    sheet_title: str,
    row: dict,
    column: str = "",
    section_row: dict | None = None,
    row_key: str | None = None,
    section_key: str | None = None,
    row_label: str = "",
    section_label: str = "",
    column_label: str = "",
) -> str:
    sheet = slug_key(sheet_title)
    section = section_key if section_key is not None else section_anchor_key(section_row)
    anchor_row = row_key if row_key is not None else row_anchor_key(row)
    if anchor_type == "section":
        section = section or row_anchor_key(row)
        anchor_row = ""
        column = ""
    elif anchor_type == "row":
        column = ""
    anchor_id = anchor_id_from_parts(
        sheet=sheet,
        section=section,
        row=anchor_row,
        column=column,
        anchor_type=anchor_type,
    )
    readable_row = anchor_label(row_label or (row.get("values") or [""])[0])
    readable_section = anchor_label(section_label or ((section_row or row).get("values") or [""])[0])
    attrs = {
        "data-anchor-id": anchor_id,
        "data-anchor-type": anchor_type,
        "data-anchor-page": PAGE_KEY,
        "data-anchor-version": PAGE_VERSION,
        "data-anchor-sheet": sheet,
        "data-anchor-sheet-title": display_sheet_title(sheet_title),
        "data-anchor-section": section,
        "data-anchor-section-title": readable_section,
        "data-anchor-row": anchor_row,
        "data-anchor-row-label": readable_row,
        "data-anchor-column": column,
        "data-anchor-column-label": anchor_label(column_label),
    }
    return "".join(f' {name}="{html.escape(value, quote=True)}"' for name, value in attrs.items())


def normalize_label_for_definition(label: str) -> str:
    value = html.unescape(label).strip()
    value = SECTION_PREFIX_RE.sub("", value)
    value = re.sub(r"\s+", " ", value)
    value = UNIT_SUFFIX_RE.sub("", value).strip()
    value = value.replace("SEO / Organic Search", "SEO / 自然搜索")
    value = value.replace("Direct", "直接访问")
    value = value.replace("Organic Social", "自然社交")
    value = value.replace("Referral", "外链推荐")
    value = value.replace("Unrecognized UTM", "未识别 UTM")
    value = terminology_aliases(include_context_sensitive=True).get(value, DISPLAY_LABEL_RENAMES.get(value, value))
    return value


def display_label_text(label: str) -> str:
    value = html.unescape(label).strip()
    value = DISPLAY_UNIT_RE.sub("", value)
    value = re.sub(r"\s+", " ", value)
    value = DISPLAY_LABEL_RENAMES.get(value, value)
    value = (
        value.replace("付费有效订阅数", "当前付费订阅数")
        .replace("付费有效订阅", "当前付费订阅数")
        .replace("freetrial 中订阅", "当前 freetrial 订阅数")
        .replace("trialing 订阅", "当前 freetrial 订阅数")
        .replace("扣款失败订阅", "当前扣款失败订阅数")
        .replace("past_due 订阅", "当前扣款失败订阅数")
        .replace("付费订阅流失", "截至当日已取消/结束订阅数")
        .replace("取消订阅数", "截至当日已取消/结束订阅数")
        .replace("取消用户", "取消订阅用户")
        .replace("未付费 DAU", "未订阅 DAU")
        .replace("未付费 D1 留存率", "未订阅 D1 留存率")
        .replace("新增freetrial", "新增 freetrial 用户")
        .replace("新增 freetrial 用户数", "新增 freetrial 用户")
        .replace("试订阅", "freetrial")
        .replace("试用订阅", "freetrial")
        .replace("新增试用用户", "新增 freetrial 用户")
        .replace("新增试用", "新增 freetrial")
        .replace("试用用户", "freetrial 用户")
        .replace("试用中", "freetrial")
        .replace("试用期", "freetrial 期")
        .replace("试用", "freetrial")
        .replace("付费用户数", "付费用户")
    )
    value = (
        value.replace("当前当前扣款失败订阅数数", "当前扣款失败订阅数")
        .replace("当前当前付费订阅数数", "当前付费订阅数")
        .replace("当前当前 freetrial 订阅数数", "当前 freetrial 订阅数")
    )
    value = DISPLAY_LABEL_RENAMES.get(value, value)
    return value.strip()


def definition_for(label: str, sheet_title: str, row_kind: str) -> str | None:
    normalized = normalize_label_for_definition(label)
    if not normalized:
        return None
    if row_kind == "header":
        return FIELD_DEFINITIONS.get(normalized, "表头字段，用于说明后续数据列。")
    if row_kind == "section":
        return f"分区标题：{normalized}。"

    if normalized == "注册率":
        if sheet_title == "用户激活与转化":
            return "注册完成用户数 / 登录入口点击用户数。"
        return "新增注册用户 / 新增 UV。"
    if normalized == "开口率":
        if sheet_title in {"Dashboard", "用户获取", "用户激活与转化"}:
            return "新访 DAU / 新增 freetrial 用户。"
        return "首次发消息用户数 / 对应转化分母。"
    if normalized == "freetrial 率":
        return "新增 freetrial 用户 / 新增注册用户。"
    if normalized == "付费率":
        return "新增付费用户 / 新增 UV；缺少同 cohort 付费映射时显示为空。"
    if normalized == "新访 DAU":
        return "D0 新增 freetrial 用户中，D0 发过至少一条 chat:message_sent 的去重用户数。"
    if normalized == "截至当日已取消/结束订阅数":
        return "截至当日结束前，已 ended_at 或 canceled_at 的 subscription 数；当前实现使用 paid_churn_subs/all_churn_subs，是累计存量，不是当日取消量。"
    if normalized == "新增 DAU":
        return "当日首次发消息的全量去重用户，不限当天是否新访。"
    if normalized == "使用次数":
        return "该能力当天发生的使用事件次数。"

    direct = FIELD_DEFINITIONS.get(normalized)
    if direct:
        return direct

    if sheet_title == "附录2 个人看板反馈":
        return "反馈对象、反馈主题或反馈明细。"
    if sheet_title == "附录1_口径裁决":
        return "口径裁决表字段，用于记录指标定义和最终取舍。"

    if sheet_title == "用户活跃与使用分布" and re.match(r"^[A-Z][A-Za-z ]+$", normalized):
        return "DAU 地域分布中的国家或地区；数值为用户数，括号为占比。"
    if normalized.endswith("（待数据）"):
        return f"{normalized.removesuffix('（待数据）')}，当前数据源未接入，暂时显示为空。"

    if normalized.startswith("总对话段"):
        return "chat session 总数，按 30 分钟无对话活动切分。"

    if normalized.startswith("/api/") or normalized.startswith("/auth") or normalized.startswith("/internal"):
        if "P95" in normalized:
            return f"{normalized.replace(' P95', '')} 的 P95 响应时间。"
        if "P99" in normalized:
            return f"{normalized.replace(' P99', '')} 的 P99 响应时间。"
        if "成功率" in normalized:
            return f"{normalized.replace(' 成功率', '')} 的成功请求 / 全部请求。"
        if "调用量" in normalized:
            return f"{normalized.replace(' 调用量', '')} 的请求量。"

    if normalized.startswith("失败："):
        return f"失败原因「{normalized.removeprefix('失败：')}」对应的次数。"
    if re.match(r"^(断连原因|履约失败原因|慢请求路径)第\s*\d+", normalized):
        return "按发生次数或影响排序的原因/路径占位，具体名称来自对应数据源。"
    if re.search(r"P(50|75|90|95|99)$", normalized):
        return f"{normalized} 对应分位数。"
    if re.search(r"P(50|75|90|95|99)\s*时长$", normalized):
        return f"{normalized}，按对应起止事件计算的分位用时。"
    if "CTA CTR" in normalized:
        return f"{normalized}，CTA 点击数 / 对应访问分母。"
    if "CTA 点击" in normalized:
        return f"{normalized}，按对应页面或人群统计 CTA 点击。"
    if "扣款" in normalized and normalized.endswith("用户"):
        return f"{normalized}，按 Stripe 扣款状态统计的去重用户数。"
    if normalized.endswith("占比"):
        return f"{normalized}，对应分组数量 / 所在分组总量。"
    if normalized.endswith("成功率"):
        return f"{normalized.removesuffix('成功率')}成功数 / 对应总数。"
    if normalized.endswith("失败率"):
        return f"{normalized.removesuffix('失败率')}失败数 / 对应总数。"
    if normalized.endswith("数"):
        return f"{normalized}，按本表口径统计的数量。"

    return None


def sheet_model(ws) -> dict:
    max_row, max_col = used_bounds(ws)
    max_col = min(max_col, 13)
    rows = []
    stack: dict[int, str] = {}
    stack_metric_kind: dict[int, str] = {}
    active_child_parent: dict[int, str] = {}
    active_child_kind: dict[int, str] = {}
    child_counts: dict[str, int] = {}
    row_labels: dict[str, str] = {}

    for r in range(1, max_row + 1):
        cells = [ws.cell(r, c) for c in range(1, max_col + 1)]
        raw_label = raw_cell_text(cells[0])
        raw_rest_probe = [raw_cell_text(c) for c in cells[1:]]
        if not raw_label:
            continue
        if not any(not is_missing_text(v) for v in [raw_label] + raw_rest_probe):
            continue

        label, had_marker = strip_marker(raw_label)
        outline = ws.row_dimensions[r].outlineLevel or 0
        level = outline if outline else (1 if had_marker else 0)
        probe_values = [label] + raw_rest_probe
        kind = "header" if is_header_row(r, probe_values) else ("section" if is_section(probe_values) else ("child" if level else "row"))
        if kind == "child" and is_derived_metric_label(label):
            kind = "derived"
        row_id = f"{ws.title}-{r}".replace(" ", "_")
        parent = ""
        own_kind = metric_kind(label)
        display_level = level

        if kind == "derived":
            parent = ""
            inherited_kind = own_kind
            display_level = max(level, 1)
        elif kind == "child":
            if own_kind == "number" and active_child_parent.get(level):
                parent = active_child_parent[level]
                inherited_kind = active_child_kind.get(level, "number")
                display_level = level + 1
            else:
                parent = stack.get(max(level - 1, 0), "")
                inherited_kind = stack_metric_kind.get(max(level - 1, 0), "number")
        else:
            inherited_kind = "number"

        parent_label = row_labels.get(parent, "")
        display_label = derived_child_label(label, parent_label) if kind == "child" and parent_label and is_derived_metric_label(parent_label) else label
        kind_for_row = inherited_kind if kind == "child" and own_kind == "number" else own_kind
        rest_values = [clean_text(fmt_value(c.value, c.number_format, kind=kind_for_row)) for c in cells[1:]]
        values = [display_label] + rest_values

        if parent:
            child_counts[parent] = child_counts.get(parent, 0) + 1

        if kind in {"child", "derived"}:
            stack[level] = row_id
            stack_metric_kind[level] = kind_for_row
            if own_kind != "number":
                active_child_parent[level] = row_id
                active_child_kind[level] = kind_for_row
        else:
            stack[0] = row_id
            stack_metric_kind[0] = kind_for_row
            for k in list(stack):
                if k > 0:
                    del stack[k]
                    stack_metric_kind.pop(k, None)
            active_child_parent.clear()
            active_child_kind.clear()

        comments = [c.comment.text.strip() for c in cells if c.comment and c.comment.text]
        definition = definition_for(label, ws.title, kind)
        if not definition:
            MISSING_DEFINITIONS.append((ws.title, label))
            definition = "待定义：请补充这个字段的业务定义和分母。"
        rows.append(
            {
                "id": row_id,
                "parent": parent,
                "level": display_level,
                "kind": kind,
                "ratio": kind in {"child", "derived"} and is_ratio_label(label),
                "value_kind": kind_for_row,
                "values": values,
                "comments": comments,
                "definition": definition,
            }
        )
        row_labels[row_id] = label

    for row in rows:
        row["has_children"] = child_counts.get(row["id"], 0) > 0
        row["hidden"] = row["kind"] == "child"
    row_by_id = {row["id"]: row for row in rows}
    for row in rows:
        parent_row = row_by_id.get(row.get("parent", ""))
        if not parent_row:
            continue
        if row["kind"] != "child":
            continue
        if parent_row["kind"] == "derived":
            continue
        if row.get("ratio") or row.get("value_kind") not in {"number", "currency"}:
            continue
        if parent_row.get("value_kind") not in {"number", "currency"}:
            continue
        if not is_share_breakdown_child(row["values"][0], parent_row["values"][0]):
            continue
        values = row["values"][:]
        for idx in range(1, len(values)):
            child_number = parse_display_number(values[idx])
            parent_number = parse_display_number(parent_row["values"][idx]) if idx < len(parent_row["values"]) else None
            values[idx] = append_child_share(values[idx], child_number, parent_number)
        row["values"] = values
    if not ws.title.startswith("附录"):
        rows = filter_non_quantitative_metric_rows(rows)
        rows = filter_empty_rows(rows)
        rows = remove_or_promote_empty_rows(rows)
    else:
        # Keep single-column marker rows (e.g., "高频诉求") so appendix segmentation works.
        pass
    rows = normalize_sections(ws.title, rows)
    if not ws.title.startswith("附录"):
        rows = normalize_drilldown_relationships(rows)
    if ws.title == "Dashboard":
        rows = prune_dashboard_summary(rows)
    if ws.title == "用户激活与转化":
        rows = restructure_activation(rows)
    if ws.title == "工程质量":
        rows = restructure_engineering(rows)
    if not ws.title.startswith("附录"):
        rows = remove_single_child_share_annotations(rows)
    return {"title": ws.title, "rows": rows, "max_col": max_col}


def workbook_model() -> list[dict]:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    source_order = ORDER + ["用户留存与流失"]
    sheets = [sheet_model(wb[name]) for name in source_order if name in wb.sheetnames]
    return apply_operating_schema(sheets)


def display_sheet_title(title: str) -> str:
    return SHEET_DISPLAY_NAMES.get(title, title)


def blank_values() -> list[str]:
    return ["—"] * METRIC_VALUE_COUNT


METRIC_DAYS = [
    "2026-05-15",
    "2026-05-16",
    "2026-05-17",
    "2026-05-18",
    "2026-05-19",
    "2026-05-20",
    "2026-05-21",
    "2026-05-22",
    "2026-05-23",
    "2026-05-24",
    "2026-05-25",
]
SOURCE_METRIC_DAYS = [
    "2026-05-15",
    "2026-05-16",
    "2026-05-17",
    "2026-05-18",
    "2026-05-19",
    "2026-05-20",
    "2026-05-21",
    "2026-05-22",
]
METRIC_WOW_REF = "2026-05-18"
METRIC_MOM_REF = "2026-04-25"
METRIC_VALUE_COUNT = len(METRIC_DAYS) + 4
DATE_COL_WIDTH = 76
TODAY_COL_WIDTH = 108
COMP_COL_WIDTH = 128
WOW_COL_WIDTH = 78
MOM_COL_WIDTH = 78
LEFT_COL_WIDTH = 240
TODAY_GRID_NTH = len(METRIC_DAYS)
TODAY_TABLE_NTH = len(METRIC_DAYS) + 1
DEFAULT_FIRST_VISIBLE_DATE = "2026-05-18"
DEFAULT_DATE_SCROLL_LEFT = max(0, METRIC_DAYS.index(DEFAULT_FIRST_VISIBLE_DATE)) * DATE_COL_WIDTH
RIGHT_GRID_TEMPLATE = (
    f"repeat({len(METRIC_DAYS) - 1}, {DATE_COL_WIDTH}px) "
    f"{TODAY_COL_WIDTH}px {COMP_COL_WIDTH}px {WOW_COL_WIDTH}px {COMP_COL_WIDTH}px {MOM_COL_WIDTH}px"
)
RIGHT_INNER_WIDTH = (
    (len(METRIC_DAYS) - 1) * DATE_COL_WIDTH
    + TODAY_COL_WIDTH
    + COMP_COL_WIDTH
    + WOW_COL_WIDTH
    + COMP_COL_WIDTH
    + MOM_COL_WIDTH
)
MAIN_TABLE_WIDTH = LEFT_COL_WIDTH + RIGHT_INNER_WIDTH
_EXTRA_CSV_CACHE: dict[str, list[dict[str, str]]] = {}
_DASHBOARD_JSON_CACHE: dict[str, dict] = {}


def normalize_metric_values(values: list[str] | None) -> list[str]:
    raw = list(values or [])
    if len(raw) == METRIC_VALUE_COUNT:
        return raw
    if len(raw) == len(SOURCE_METRIC_DAYS) + 4 and len(METRIC_DAYS) >= len(SOURCE_METRIC_DAYS):
        date_values = raw[: len(SOURCE_METRIC_DAYS)]
        return date_values + ["—"] * (len(METRIC_DAYS) - len(SOURCE_METRIC_DAYS)) + ["—"] * 4
    if len(raw) < METRIC_VALUE_COUNT:
        return raw + ["—"] * (METRIC_VALUE_COUNT - len(raw))
    return raw[:METRIC_VALUE_COUNT]


def schema_header_row(sheet_title: str) -> dict:
    labels = [
        "指标",
        *[day[5:] for day in METRIC_DAYS[:-1]],
        f"当日 {METRIC_DAYS[-1][5:]}",
        f"上周同日 {METRIC_WOW_REF[5:]}",
        "WoW %",
        f"上月同日 {METRIC_MOM_REF[5:]}",
        "MoM %",
    ]
    return {
        "id": f"{sheet_title}-schema-header".replace(" ", "_"),
        "parent": "",
        "level": 0,
        "kind": "header",
        "ratio": False,
        "value_kind": "number",
        "values": labels,
        "comments": [],
        "definition": "",
        "has_children": False,
        "hidden": False,
    }


def extra_csv_rows(filename: str) -> list[dict[str, str]]:
    if filename not in _EXTRA_CSV_CACHE:
        path = GENERATED_METRICS / filename
        if not path.exists():
            match = re.match(r"^(.+)_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.csv$", filename)
            candidates = sorted(GENERATED_METRICS.glob(f"{match.group(1)}_*.csv")) if match else []
            path = candidates[-1] if candidates else path
        if not path.exists():
            _EXTRA_CSV_CACHE[filename] = []
        else:
            with path.open(encoding="utf-8-sig", newline="") as handle:
                _EXTRA_CSV_CACHE[filename] = list(csv.DictReader(handle))
    return _EXTRA_CSV_CACHE[filename]


def dashboard_json(filename: str) -> dict:
    if filename not in _DASHBOARD_JSON_CACHE:
        path = DASHBOARD_EXPORTS / filename
        if not path.exists():
            _DASHBOARD_JSON_CACHE[filename] = {}
        else:
            _DASHBOARD_JSON_CACHE[filename] = json.loads(path.read_text(encoding="utf-8"))
    return _DASHBOARD_JSON_CACHE[filename]


def numeric_raw(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or is_missing_text(text):
        return None
    try:
        return float(text.replace(",", "").replace("$", "").replace("%", ""))
    except ValueError:
        return None


def fmt_extra_value(value: float | None, *, kind: str = "number", decimals: int = 1) -> str:
    if value is None:
        return "—"
    if kind == "percent":
        return f"{value * 100:.1f}%"
    if kind == "currency":
        return fmt_number(value, decimals=decimals, currency=True)
    return fmt_number(value, decimals=decimals)


def extra_series_from_daily(daily: dict[str, float | None], *, kind: str = "number", decimals: int = 1) -> list[str]:
    values = [fmt_extra_value(daily.get(day), kind=kind, decimals=decimals) for day in METRIC_DAYS]
    current = daily.get(METRIC_DAYS[-1])
    wow_ref = daily.get(METRIC_WOW_REF)
    mom_ref = daily.get(METRIC_MOM_REF)
    values.append(fmt_extra_value(wow_ref, kind=kind, decimals=decimals))
    if current is not None and wow_ref not in (None, 0):
        values.append(f"{(current / wow_ref - 1) * 100:.1f}%")
    else:
        values.append("—")
    values.append(fmt_extra_value(mom_ref, kind=kind, decimals=decimals))
    if current is not None and mom_ref not in (None, 0):
        values.append(f"{(current / mom_ref - 1) * 100:.1f}%")
    else:
        values.append("—")
    return values


def extra_series(filename: str, column: str, *, kind: str = "number", decimals: int = 1, date_column: str = "date") -> list[str]:
    daily = {
        str(row.get(date_column, ""))[:10]: numeric_raw(row.get(column))
        for row in extra_csv_rows(filename)
        if row.get(date_column)
    }
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def dashboard_trend_series(
    filename: str,
    insight_name_contains: str,
    *,
    series_label: str | None = None,
    breakdown_value: str | None = None,
    kind: str = "number",
    decimals: int = 0,
) -> list[str]:
    dashboard = dashboard_json(filename)
    needle = insight_name_contains.lower()
    for tile in dashboard.get("tiles", []):
        insight = tile.get("insight") or {}
        if needle not in str(insight.get("name") or "").lower():
            continue
        for series in insight.get("result") or []:
            if not isinstance(series, dict):
                continue
            if series_label is not None and str(series.get("label") or "") != series_label:
                continue
            if breakdown_value is not None:
                breakdown = series.get("breakdown_value")
                breakdown_values = breakdown if isinstance(breakdown, list) else [breakdown]
                if breakdown_value not in {str(value) for value in breakdown_values}:
                    continue
            labels = series.get("days") or series.get("labels") or []
            data = series.get("data") or []
            daily: dict[str, float] = {}
            for label, value in zip(labels, data):
                day = str(label)[:10]
                number = numeric_raw(value)
                if re.match(r"^\d{4}-\d{2}-\d{2}$", day) and number is not None:
                    daily[day] = number
            return extra_series_from_daily(daily, kind=kind, decimals=decimals)
    return blank_values()


def extra_text_series(filename: str, column: str, *, date_column: str = "date") -> list[str]:
    daily = {
        str(row.get(date_column, ""))[:10]: (str(row.get(column) or "—").strip() or "—")
        for row in extra_csv_rows(filename)
        if row.get(date_column)
    }
    values = [daily.get(day, "—") for day in METRIC_DAYS]
    values.append(daily.get(METRIC_WOW_REF, "—"))
    values.append("—")
    values.append(daily.get(METRIC_MOM_REF, "—"))
    values.append("—")
    return values


def fmt_distribution_value(users: float | None, share: float | None) -> str:
    if users is None:
        return "—"
    if share is None:
        return fmt_number(users, decimals=0)
    return f"{fmt_number(users, decimals=0)} ({share * 100:.0f}%)"


def distribution_children(
    filename: str,
    cohort: str,
    dimension: str,
    *,
    limit: int = 3,
    required_labels: list[str] | None = None,
) -> list[dict]:
    by_value: dict[str, dict[str, tuple[float, float | None]]] = defaultdict(dict)
    totals: dict[str, float] = defaultdict(float)
    current_counts: dict[str, float] = {}
    for row in extra_csv_rows(filename):
        if row.get("cohort") != cohort or row.get("dimension") != dimension:
            continue
        value = str(row.get("value") or "").strip()
        day = str(row.get("date", ""))[:10]
        users = numeric_raw(row.get("users"))
        share = numeric_raw(row.get("share"))
        if not value or not day or users is None:
            continue
        by_value[value][day] = (users, share)
        totals[value] += users
        if day == METRIC_DAYS[-1]:
            current_counts[value] = users

    if current_counts:
        labels = [label for label, _ in sorted(current_counts.items(), key=lambda item: (-item[1], item[0]))[:limit]]
    else:
        labels = [label for label, _ in sorted(totals.items(), key=lambda item: (-item[1], item[0]))[:limit]]
    for required_label in required_labels or []:
        if required_label not in labels:
            labels.append(required_label)

    children: list[dict] = []
    for label in labels:
        values: list[str] = []
        for day in METRIC_DAYS:
            users, share = by_value[label].get(day, (None, None))
            values.append(fmt_distribution_value(users, share))
        wow_users, wow_share = by_value[label].get(METRIC_WOW_REF, (None, None))
        mom_users, mom_share = by_value[label].get(METRIC_MOM_REF, (None, None))
        values.extend([fmt_distribution_value(wow_users, wow_share), "—", fmt_distribution_value(mom_users, mom_share), "—"])
        children.append({"label": label, "values": values, "level": 2, "keep_empty": label in (required_labels or [])})
    return children


def distribution_value_series(
    filename: str,
    cohort: str,
    dimension: str,
    value: str,
    *,
    decimals: int = 0,
) -> list[str]:
    daily: dict[str, float | None] = {}
    for row in extra_csv_rows(filename):
        if row.get("cohort") != cohort or row.get("dimension") != dimension:
            continue
        if str(row.get("value") or "").strip() != value:
            continue
        day = str(row.get("date", ""))[:10]
        if not day:
            continue
        daily[day] = numeric_raw(row.get("users"))
    return extra_series_from_daily(daily, decimals=decimals)


def category_series(filename: str, category: str, column: str, *, decimals: int = 0) -> list[str]:
    daily = {
        str(row.get("date", ""))[:10]: numeric_raw(row.get(column))
        for row in extra_csv_rows(filename)
        if row.get("category") == category and row.get("date")
    }
    return extra_series_from_daily(daily, decimals=decimals)


def capability_detail_children(filename: str, category: str, *, cohort: str = "all", level: int = 1) -> list[dict]:
    by_detail: dict[str, dict[str, tuple[float, float | None]]] = defaultdict(dict)
    totals: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        if row.get("category") != category:
            continue
        if row.get("cohort", "all") != cohort:
            continue
        detail = str(row.get("detail") or "").strip()
        day = str(row.get("date", ""))[:10]
        users = numeric_raw(row.get("users"))
        share = numeric_raw(row.get("share"))
        if not detail or not day or users is None:
            continue
        by_detail[detail][day] = (users, share)
        totals[detail] += users

    children: list[dict] = []
    for label, _ in sorted(totals.items(), key=lambda item: (-item[1], item[0])):
        values: list[str] = []
        for day in METRIC_DAYS:
            users, share = by_detail[label].get(day, (None, None))
            values.append(fmt_distribution_value(users, share))
        wow_users, wow_share = by_detail[label].get(METRIC_WOW_REF, (None, None))
        mom_users, mom_share = by_detail[label].get(METRIC_MOM_REF, (None, None))
        values.extend([fmt_distribution_value(wow_users, wow_share), "—", fmt_distribution_value(mom_users, mom_share), "—"])
        children.append({"label": label, "values": values, "level": level})
    return children


def scaled_extra_series(
    filename: str,
    column: str,
    *,
    scale: float,
    kind: str = "number",
    decimals: int = 1,
    date_column: str = "date",
) -> list[str]:
    daily = {
        str(row.get(date_column, ""))[:10]: (
            None if numeric_raw(row.get(column)) is None else numeric_raw(row.get(column)) * scale
        )
        for row in extra_csv_rows(filename)
        if row.get(date_column)
    }
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def seconds_from_ms_series(filename: str, column: str, *, decimals: int = 1) -> list[str]:
    return scaled_extra_series(filename, column, scale=0.001, decimals=decimals)


def extra_event_series(filename: str, event: str, column: str = "events", *, kind: str = "number", decimals: int = 0) -> list[str]:
    daily: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        if row.get("event") != event:
            continue
        day = str(row.get("day", ""))[:10]
        value = numeric_raw(row.get(column))
        if day and value is not None:
            daily[day] += value
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def extra_group_series(
    filename: str,
    group_column: str,
    group_value: str,
    value_column: str = "value",
    *,
    kind: str = "number",
    decimals: int = 0,
) -> list[str]:
    daily: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        if row.get(group_column) != group_value:
            continue
        day = str(row.get("date", ""))[:10]
        value = numeric_raw(row.get(value_column))
        if day and value is not None:
            daily[day] += value
    return extra_series_from_daily(daily, kind=kind, decimals=decimals)


def nonzero_groups(filename: str, group_column: str, *, value_column: str = "value") -> list[str]:
    totals: dict[str, float] = defaultdict(float)
    for row in extra_csv_rows(filename):
        key = row.get(group_column)
        value = numeric_raw(row.get(value_column))
        if key and value is not None:
            totals[key] += value
    return [key for key, value in sorted(totals.items()) if value > 0]


def sum_value_series(*series: list[str], kind: str = "number", decimals: int = 0) -> list[str]:
    values: list[str] = []
    for columns in zip(*series):
        numbers = [parse_display_number(value) for value in columns]
        if any(number is not None for number in numbers):
            values.append(fmt_extra_value(sum(number or 0 for number in numbers), kind=kind, decimals=decimals))
        else:
            values.append("—")
    return values


def append_share_series(child: list[str], parent: list[str]) -> list[str]:
    values: list[str] = []
    for child_value, parent_value in zip(child, parent):
        child_number = parse_display_number(child_value)
        parent_number = parse_display_number(parent_value)
        values.append(append_child_share(child_value, child_number, parent_number))
    return values


def has_any_number(series: list[str]) -> bool:
    return any(parse_display_number(value) is not None for value in series)


def first_present_series(primary: list[str], fallback: list[str]) -> list[str]:
    return primary if has_any_number(primary) else fallback


def one_day_series(day: str, value: float | int | None, *, kind: str = "number", decimals: int = 0) -> list[str]:
    return extra_series_from_daily({day: None if value is None else float(value)}, kind=kind, decimals=decimals)


def row_lookup(sheets: list[dict]) -> list[tuple[str, dict, str]]:
    output: list[tuple[str, dict, str]] = []
    for sheet in sheets:
        by_id = {row["id"]: row for row in sheet["rows"]}
        for row in sheet["rows"]:
            if row["kind"] in {"header", "section"}:
                continue
            parent = by_id.get(row.get("parent", ""))
            parent_label = normalize_label_for_definition(parent["values"][0]) if parent else ""
            output.append((sheet["title"], row, parent_label))
    return output


def find_source_values(
    lookup: list[tuple[str, dict, str]],
    labels: list[str] | str,
    *,
    titles: list[str] | None = None,
    parent: str | None = None,
) -> list[str]:
    wanted = [labels] if isinstance(labels, str) else labels
    wanted_norm = {normalize_label_for_definition(label) for label in wanted}
    title_set = set(titles or [])
    parent_norm = normalize_label_for_definition(parent) if parent else None
    for title, row, parent_label in lookup:
        if title_set and title not in title_set:
            continue
        if parent_norm is not None and parent_label != parent_norm:
            continue
        if normalize_label_for_definition(row["values"][0]) in wanted_norm:
            return normalize_metric_values(row["values"][1:])
    for title, row, parent_label in lookup:
        if parent_norm is not None and parent_label != parent_norm:
            continue
        if normalize_label_for_definition(row["values"][0]) in wanted_norm:
            return normalize_metric_values(row["values"][1:])
    return blank_values()


def ratio_values(numerator: list[str], denominator: list[str]) -> list[str]:
    values: list[str] = []
    for n_value, d_value in zip(numerator, denominator):
        n = parse_display_number(n_value)
        d = parse_display_number(d_value)
        values.append(f"{n / d * 100:.1f}%" if n is not None and d not in (None, 0) else "—")
    return values


def zero_missing_by_denominator(values: list[str], denominator: list[str]) -> list[str]:
    output: list[str] = []
    for value, denominator_value in zip(values, denominator):
        d = parse_display_number(denominator_value)
        output.append("0" if parse_display_number(value) is None and d not in (None, 0) else value)
    return output


def division_values(numerator: list[str], denominator: list[str]) -> list[str]:
    values: list[str] = []
    for n_value, d_value in zip(numerator, denominator):
        n = parse_display_number(n_value)
        d = parse_display_number(d_value)
        values.append(fmt_number(n / d, decimals=1) if n is not None and d not in (None, 0) else "—")
    return values


def division_values_zero_when_no_numerator(numerator: list[str], denominator: list[str]) -> list[str]:
    values: list[str] = []
    for n_value, d_value in zip(numerator, denominator):
        n = parse_display_number(n_value)
        d = parse_display_number(d_value)
        if n == 0:
            values.append("0")
        elif n is not None and d not in (None, 0):
            values.append(fmt_number(n / d, decimals=1))
        else:
            values.append("—")
    return values


def strip_share_values(values: list[str]) -> list[str]:
    return [strip_share_annotation(value) for value in values]


def metric_row(
    sheet_title: str,
    seq: int,
    label: str,
    values: list[str] | None,
    *,
    kind: str = "row",
    parent: str = "",
    level: int = 0,
    keep_empty: bool = False,
    ratio: bool | None = None,
) -> dict:
    if kind == "derived" and not parent and level == 0:
        level = 1
    definition = definition_for(label, sheet_title, kind) or f"{normalize_label_for_definition(label)}。"
    if is_task_metric_label(label):
        values = blank_values()
        keep_empty = True
    return {
        "id": f"{sheet_title}-schema-{seq}".replace(" ", "_"),
        "parent": parent,
        "level": level,
        "kind": kind,
        "ratio": (kind == "derived" or is_ratio_label(label)) if ratio is None else ratio,
        "value_kind": metric_kind(label),
        "values": [label] + normalize_metric_values(values),
        "comments": [],
        "definition": definition,
        "has_children": False,
        "hidden": kind == "child",
        "keep_empty": keep_empty,
        "default_open": False,
    }


def is_task_metric_label(label: str) -> bool:
    normalized = normalize_label_for_definition(label)
    return "任务" in normalized or normalized == "credits / 任务"


def finalize_schema_rows(rows: list[dict]) -> list[dict]:
    child_ids = {row["parent"] for row in rows if row.get("parent")}
    default_open_ids = {row["id"] for row in rows if row.get("default_open")}
    for row in rows:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child" and row.get("parent") not in default_open_ids
    return remove_single_child_share_annotations(rows)


def build_schema_sheet(source_sheet: dict, lookup: list[tuple[str, dict, str]], sections: list[dict]) -> dict:
    rows = [schema_header_row(source_sheet["title"])]
    seq = 0
    for section_idx, section in enumerate(sections, start=1):
        rows.append(make_section_row(source_sheet["title"], section_idx, section["title"]))
        for item in section["items"]:
            seq += 1
            label = item["label"]
            values = item.get("values")
            if values is None:
                values = find_source_values(
                    lookup,
                    item.get("source", label),
                    titles=item.get("titles", [source_sheet["title"]]),
                    parent=item.get("source_parent"),
                )
            row = metric_row(
                source_sheet["title"],
                seq,
                label,
                values,
                kind=item.get("kind", "row"),
                parent="",
                level=item.get("level", 0),
                keep_empty=item.get("keep_empty", False),
                ratio=item.get("ratio"),
            )
            if item.get("default_open"):
                row["default_open"] = True
            rows.append(row)
            parent_id = row["id"]
            for child in item.get("children", []):
                seq += 1
                child_label = child["label"]
                child_values = child.get("values")
                if child_values is None:
                    child_values = find_source_values(
                        lookup,
                        child.get("source", child_label),
                        titles=child.get("titles", [source_sheet["title"]]),
                        parent=child.get("source_parent", label),
                    )
                child_level = child.get("level", row.get("level", 0) + 1)
                child_row = metric_row(
                    source_sheet["title"],
                    seq,
                    child_label,
                    child_values,
                    kind="child",
                    parent=parent_id,
                    level=child_level,
                    keep_empty=child.get("keep_empty", item.get("keep_empty_children", False)),
                    ratio=child.get("ratio"),
                )
                if child.get("default_open"):
                    child_row["default_open"] = True
                rows.append(child_row)
                child_parent_id = child_row["id"]
                for grandchild in child.get("children", []):
                    seq += 1
                    grand_values = grandchild.get("values")
                    if grand_values is None:
                        grand_values = find_source_values(
                            lookup,
                            grandchild.get("source", grandchild["label"]),
                            titles=grandchild.get("titles", [source_sheet["title"]]),
                            parent=grandchild.get("source_parent", child_label),
                        )
                    grand_row = metric_row(
                        source_sheet["title"],
                        seq,
                        grandchild["label"],
                        grand_values,
                        kind="child",
                        parent=child_parent_id,
                        level=grandchild.get("level", child_row.get("level", 0) + 1),
                        keep_empty=grandchild.get("keep_empty", child.get("keep_empty_children", item.get("keep_empty_children", False))),
                        ratio=grandchild.get("ratio"),
                    )
                    if grandchild.get("default_open"):
                        grand_row["default_open"] = True
                    rows.append(grand_row)
                    grand_parent_id = grand_row["id"]
                    for great_grandchild in grandchild.get("children", []):
                        seq += 1
                        great_values = great_grandchild.get("values")
                        if great_values is None:
                            great_values = find_source_values(
                                lookup,
                                great_grandchild.get("source", great_grandchild["label"]),
                                titles=great_grandchild.get("titles", [source_sheet["title"]]),
                                parent=great_grandchild.get("source_parent", grandchild["label"]),
                            )
                        rows.append(
                            metric_row(
                                source_sheet["title"],
                                seq,
                                great_grandchild["label"],
                                great_values,
                                kind="child",
                                parent=grand_parent_id,
                                level=great_grandchild.get("level", grand_row.get("level", 0) + 1),
                                keep_empty=great_grandchild.get(
                                    "keep_empty",
                                    grandchild.get(
                                        "keep_empty_children",
                                        child.get("keep_empty_children", item.get("keep_empty_children", False)),
                                    ),
                                ),
                                ratio=great_grandchild.get("ratio"),
                            )
                        )
    rows = finalize_schema_rows(rows)
    rows = filter_empty_rows(rows)
    return {**source_sheet, "rows": rows}


EMPLOYEE_FEEDBACK_ROWS = [
    (
        "Zhi Tian",
        "付费用户必须拆清 freetrial、真付费、扣款失败和取消, 不能把绑卡直接当作付费; 相关看板项需要进入研发和运营的跟进流程。",
    ),
    (
        "汪健",
        "看板需要建立埋点体系和 schema 基础, 并跟踪端到端成功率、工具失败率、SLA、故障恢复、功能展示 / 点击、skill 与 connector 漏斗。",
    ),
    (
        "黄彬",
        "核心诉求是用户任务错误率和错误类型拆分; WS 错误要能归因到后端、网络、工具等根因, 不能都停在连接层。",
    ),
    (
        "张立鹏",
        "需要看用户对话轮次、任务完成数、UE 单位经济、负反馈指标和用户国家分布。",
    ),
    (
        "王泉澴",
        "需要讲清 UV vs DAU、任务是否等于 query、留存窗口等口径; 想看激活用户数, 最关心数据链路。",
    ),
    (
        "陈作君",
        "希望增加用户停留时长, 并细分付费用户的 DAU 和留存。",
    ),
    (
        "江伟忠",
        "付费转化希望细化到各渠道 CPA、CPM 等指标; 同时需要团队对 PMF、PTF 和商业经营三件事形成统一理解。",
    ),
    (
        "吕枫",
        "关注对话稳定性, 包括人均成功轮次、人均失败次数、首次失败轮次、每 100 轮失败比例; 还要工具调用次数 / 成功率、skill 引用、IM / 移动 / Web 占比和 API 成功率。",
    ),
    (
        "程志达",
        "关注移动端 vs PC、上传 / 下载文件数、定时任务创建 / 执行、任务成功率、用户情绪曲线和用户关注对话间隔。",
    ),
    (
        "马为峰",
        "当前大指标够用, 核心是 trace, 即用户路径和用户旅程; 满意度应从行为 trace 与对话两条线看。",
    ),
    (
        "idoubi",
        "需要看 UV、来源、对话次数、留存、付费率、流失率和流失归因。",
    ),
    (
        "王思琦",
        "campaign 太多需要搜索, 用户明细需要下载, 并希望有自助分析能力; 同时要增加大盘均值、发起支付 / 点击登录等第一步、曝光 / 点击等原始数据, 并支持国家和 campaign 筛选。",
    ),
    (
        "徐瑞",
        "先要保证数据准确; 关注页面首次加载、API 响应、沙盒挂率、bug 触发比例、bug 反馈占比、bug 后流失率、渠道到注册 / 付费转化和来源地区分布。",
    ),
    (
        "李林鑫",
        "不建议一人一定制, 更适合按角色提供 dashboard, 复杂查询交给 Metabase 自助; pipeline / 数仓基础设施和业务指标需求收集需要分工清楚。",
    ),
    (
        "许睿",
        "CAC 口径应为广告花费 / 付费用户, 原广告花费 / 注册应叫 CPA, 并需要补 KOC 投放成本; 渠道侧要支持 YouTube 单渠道、KOC / KOL 渠道组、付费 / 免费渠道组筛选。",
    ),
    (
        "糯米鸡",
        "需要常规产品指标 DAU / WAU、留存、付费、毛利, 使用情况里的任务类型 / 轮数 / 时长 / AI agent server 指标, 以及获客转化漏斗。",
    ),
    (
        "宋伟",
        "DAU 与任务定义必须简单可解释且数据准确, 可以先用 tool_use 趋势类规则; 工程质量应面向业务核心路径, 实时工程细节回 Grafana; 泛工程块应换成 agent view 的单用户 / 单 agent token 消耗与工具调用率。",
    ),
]

FEEDBACK_THEME_ROWS = [
    (
        "口径可信与数据链路",
        "王泉澴, 宋伟, 李林鑫",
        "数据基础 / 口径裁决",
    ),
    (
        "付费转化与 D+3 cohort",
        "Zhi Tian, 江伟忠, 许睿, 王思琦, idoubi, 糯米鸡",
        "用户激活与转化 / 业务经营",
    ),
    (
        "首消息 / 首任务 / 激活",
        "王泉澴, 张立鹏, 宋伟, 糯米鸡",
        "用户激活与转化 / 用户活跃与使用分布",
    ),
    (
        "任务成功率 / 错误率",
        "黄彬, 张立鹏, 吕枫, 程志达, 汪健, 宋伟",
        "Agent & 工程质量",
    ),
    (
        "trace / 用户旅程 / 满意度",
        "马为峰, 程志达",
        "Agent & 工程质量 / 未来用户旅程分析",
    ),
    (
        "渠道 / 国家 / 端类型 / campaign",
        "许睿, 王思琦, 徐瑞, 张立鹏, 吕枫, 程志达, idoubi, 江伟忠",
        "用户获取 / 用户活跃与使用分布",
    ),
    (
        "工程质量业务化",
        "汪健, 徐瑞, 黄彬, 宋伟, 吕枫",
        "Agent & 工程质量",
    ),
    (
        "功能使用与 Agent 资源消耗",
        "汪健, 程志达, 吕枫, 宋伟, 徐瑞, 糯米鸡",
        "用户活跃与使用分布 / 业务经营 / Agent & 工程质量",
    ),
    (
        "使用深度 / 停留时长 / 付费活跃",
        "陈作君, 糯米鸡, idoubi, 宋伟, 吕枫",
        "用户活跃与使用分布 / 留存与流失",
    ),
    (
        "留存 / 流失 / 流失归因",
        "idoubi, 徐瑞, 糯米鸡, 王泉澴, Zhi Tian, 陈作君",
        "留存与流失 / 业务经营",
    ),
    (
        "自助分析与下载筛选",
        "王思琦, 李林鑫",
        "Dashboard 交互 / 外部 BI",
    ),
]

FEEDBACK_COVERAGE_ROWS = [
    (
        "数据基础",
        "tooltip / 口径解释 / 口径裁决",
        "王泉澴, 宋伟, 李林鑫",
        "已开发",
        "当前 tooltip 缺失为 0, 附录1保留口径裁决; 数据链路和自助数仓能力仍需要在看板外产品化。",
    ),
    (
        "Dashboard",
        "五区总览主表",
        "综合需求",
        "已开发",
        "主表已按新用户链路、活跃与使用、Agent & 工程质量、业务经营组织核心指标。",
    ),
    (
        "用户获取",
        "广告花费、曝光、点击、CTR、CPC、CPM",
        "王思琦, 许睿, 江伟忠, idoubi",
        "已开发",
        "投放区已展示广告花费、付费曝光、付费点击、CPM、CTR、CPC; Google Ads 可下钻。",
    ),
    (
        "用户获取",
        "新访质量漏斗",
        "王思琦, 徐瑞, idoubi, 糯米鸡",
        "已开发",
        "新增 UV、注册、freetrial、新访 DAU、新增付费和对应转化率已进入用户获取页。",
    ),
    (
        "用户激活与转化",
        "访问、登录 CTA、注册、freetrial、次日回访",
        "Zhi Tian, 王思琦, idoubi, 糯米鸡",
        "已开发",
        "激活页已拆访问到注册、注册到 freetrial、freetrial 到首次任务、次日回访, 并支持 onboarding、/start、/usecase、首页路径下钻。",
    ),
    (
        "用户激活与转化",
        "首消息、首任务开始、首任务完成",
        "王泉澴, 张立鹏, 宋伟, 糯米鸡",
        "已开发",
        "主表与激活页已接入新访 DAU、开口率、首次任务开始、首次任务完成、首次发起任务率、首次任务完成率; 后续只校准 task 口径。",
    ),
    (
        "用户活跃与使用分布",
        "DAU / WAU / MAU、消息数、对话段、会话时长",
        "糯米鸡, idoubi, 宋伟, 王泉澴, 吕枫, 陈作君",
        "已开发",
        "活跃规模、活跃结构与频次、Agent 使用量与深度已拆分; DAU 使用发消息用户口径, 对话数、消息数、活跃任务数、人均指标和会话时长已进入页面。",
    ),
    (
        "用户活跃与使用分布",
        "附件、IM、Workspace connector、connector、自动化任务、skill",
        "汪健, 吕枫, 程志达, 徐瑞, 糯米鸡",
        "已开发",
        "功能分布已包含附件上传、IM 连接、Workspace connector、通用 connector、自动化任务和 skill 使用。",
    ),
    (
        "Agent & 工程质量",
        "Agent 响应、失败、中断、Sandbox",
        "黄彬, 吕枫, 宋伟, 程志达",
        "已开发",
        "Agent & 工程质量页已收敛到响应体验、失败与中断、Sandbox 与执行环境; 模型资源消耗迁到业务经营推理成本下探。",
    ),
    (
        "Agent & 工程质量",
        "基础工程健康",
        "汪健, 徐瑞, 黄彬, 宋伟, 吕枫",
        "已开发",
        "工程健康不再单独成页; 页面首次加载、API 失败、实时连接、AI Gateway 和 Sandbox 已合并到 Agent & 工程质量。",
    ),
    (
        "业务经营",
        "freetrial、D+3、订阅状态、收入",
        "Zhi Tian, idoubi, 糯米鸡, 江伟忠",
        "已开发",
        "业务经营页已有 freetrial 规模、结账过程、D+3 付费结果、订阅状态和收入。",
    ),
    (
        "用户获取",
        "渠道组、单渠道、campaign、KOC / KOL",
        "许睿, 王思琦, 江伟忠, idoubi",
        "已具备待算数",
        "结构已做, 不进主表新增实体; 继续补 KOC / KOL、YouTube、campaign 和付费 / 免费渠道组的映射与数值。",
    ),
    (
        "用户获取",
        "各渠道→注册 / freetrial / 首消息 / 付费转化",
        "王思琦, 徐瑞, 许睿, 江伟忠, idoubi",
        "已具备待算数",
        "结构已做, 不进主表新增实体; 重点是把同 cohort 归因、campaign 和 KOC 映射算准。",
    ),
    (
        "用户激活与转化",
        "首任务真实事件与任务口径",
        "王泉澴, 张立鹏, 宋伟, 糯米鸡",
        "已接入待校准",
        "首次任务开始 / 完成已接入主表和激活页; 后续只处理 task 口径校准, 不再另起一个看板实体。",
    ),
    (
        "用户活跃与使用分布",
        "用户停留时长",
        "陈作君",
        "已开发",
        "当前用 30 分钟间隔切 session, 活跃页已有会话时长 P50 / P90; 不再新增停留时长实体。",
    ),
    (
        "留存与流失",
        "留存 / 流失基础",
        "idoubi, 徐瑞, 糯米鸡, 王泉澴, 陈作君",
        "已开发",
        "当前页已按订阅状态展示历史活跃用户、DAU、MAU、7+ 天未活跃和 D1 留存; 回流保留为独立段。",
    ),
    (
        "Agent & 工程质量",
        "任务成功率、任务错误率、对话稳定性、失败原因",
        "黄彬, 张立鹏, 吕枫, 程志达, 汪健, 宋伟",
        "合并处理",
        "和首任务 / 任务口径是同一件事; 不单独新增主表实体, 失败原因后续作为任务下探。",
    ),
    (
        "用户活跃与使用分布 / 业务经营",
        "工具调用、skill、connector、文件、定时任务、token / tool rate",
        "汪健, 吕枫, 程志达, 宋伟, 徐瑞, 糯米鸡",
        "作为下探",
        "不新增大实体; tool / connector / 文件 / 自动化合并到功能分布, credits / tokens / provider / model 合并到业务经营推理成本下探。",
    ),
    (
        "Agent & 工程质量",
        "用户感知质量、WS 根因、页面首屏、SLA",
        "汪健, 徐瑞, 黄彬, 宋伟, 吕枫",
        "部分开发",
        "已补页面首次加载作为性能指标; WS 根因、故障恢复时长和业务路径 SLA 先不强行展开。",
    ),
    (
        "业务经营",
        "成本、毛利、UE",
        "许睿, 张立鹏, 糯米鸡, 江伟忠, idoubi",
        "暂缓",
        "保留成本、毛利、UE 占位, 但需要成本依据后再判断是否进主表。",
    ),
    (
        "用户活跃与使用分布",
        "国家 / 地区、端类型、IM / 移动 / Web / PC",
        "张立鹏, 徐瑞, 吕枫, 程志达, 许睿, 王思琦",
        "作为下探",
        "已在活跃页收敛为 DAU地域分布与 DAU Interface 分布; 新增 / 回访作为结构派生项, 地域、Interface、渠道作为下探。",
    ),
    (
        "留存与流失",
        "付费用户 DAU / 付费用户留存",
        "陈作君",
        "待研究",
        "已在留存页新增 freetrial / 付费 / 未订阅分层占位; 需要研究 PostHog 与 Stripe join 后再填数。",
    ),
    (
        "留存与流失",
        "D7 / D30 留存与流失归因",
        "idoubi, 徐瑞, 糯米鸡, 王泉澴",
        "待研究",
        "D7 / D30 与归因需要 cohort 和取消 / 反馈分类数据, 先不作为当前实现项。",
    ),
    (
        "Agent & 工程质量",
        "trace、用户旅程、情绪曲线、满意度",
        "马为峰, 程志达",
        "跳过",
        "本轮明确跳过; 不应用弱 proxy 伪装。",
    ),
    (
        "Agent & 工程质量",
        "bug 触发比例 / bug 反馈占比 / bug 后流失",
        "徐瑞, 黄彬",
        "跳过",
        "本轮明确跳过; 需要 bug 标签、反馈分类和留存 join, 当前不做。",
    ),
    (
        "用户活跃与使用分布 / Agent & 工程质量",
        "功能曝光 / 点击 / 授权失败 / 调用成功用户",
        "汪健, 王思琦, 徐瑞",
        "作为下探",
        "不新增实体; 与当前功能分布合并, 只补授权失败、调用成功、tool 成功 / 失败等高效下探。",
    ),
    (
        "业务经营",
        "CPA / CAC / KOC 成本 / LTV / Payback",
        "许睿, 张立鹏, 糯米鸡, 江伟忠",
        "留空占位",
        "已从用户获取投放区移除 CPA / CAC / LTV / Payback 独立空行; KOC / KOL 作为广告花费下探占位, 其余待业务经营或渠道效率有依据后再接。",
    ),
    (
        "交互与自助分析",
        "搜索、下载、原始数据、角色化 dashboard / Metabase",
        "王思琦, 李林鑫",
        "外部承接",
        "这类不应全塞进表格; 需要 MaxCenter / BI 自助能力承接。",
    ),
]


def appendix_feedback_row(sheet_title: str, seq: int, values: list[str], *, kind: str = "row") -> dict:
    return {
        "id": f"{sheet_title}-feedback-{seq}".replace(" ", "_"),
        "parent": "",
        "level": 0,
        "kind": kind,
        "ratio": False,
        "value_kind": "text",
        "values": list(values),
        "comments": [],
        "definition": "员工反馈汇总。",
        "has_children": False,
        "hidden": False,
        "keep_empty": True,
        "default_open": False,
    }


def build_employee_feedback_appendix(source_sheet: dict) -> dict:
    sheet_title = source_sheet["title"]
    rows: list[dict] = []
    seq = 0

    def add(values: list[str], *, kind: str = "row") -> None:
        nonlocal seq
        seq += 1
        rows.append(appendix_feedback_row(sheet_title, seq, values, kind=kind))

    add(["反馈对象"], kind="section")
    for name, feedback in EMPLOYEE_FEEDBACK_ROWS:
        add([name, feedback])

    add(["高频诉求"], kind="section")
    for topic, people, category in FEEDBACK_THEME_ROWS:
        add([topic, people, category])

    add(["1. 已开发 / 待开发"], kind="section")
    for row in FEEDBACK_COVERAGE_ROWS:
        add(list(row))

    max_col = max((len(row["values"]) for row in rows), default=1)
    return {**source_sheet, "rows": rows, "max_col": max_col}


TERMINOLOGY_CONFLICT_ROWS = [
    (
        "UV",
        "UV / 新增 UV / 网站访问会话数",
        "页面显示沿用 UV；定义必须写明是 session 去重还是 person 去重。",
        "不同数据源不要直接相加；需要在 tooltip 或数据源列说明粒度。",
        "待校准",
    ),
    (
        "新访 DAU / 新增 DAU",
        "全量首次发消息 / 新增 UV cohort 首日发消息",
        "新用户链路用新访 DAU；活跃规模用新增 DAU。",
        "获客/激活页看新访 cohort；活跃页看全量首次发消息。",
        "已裁决",
    ),
    (
        "注册率",
        "新访注册率 / 登录入口注册率",
        "页面行名保留注册率，依靠父级分母解释。",
        "同名可保留，但 tooltip 必须说明具体分母。",
        "已裁决",
    ),
    (
        "开口率",
        "新访开口率 / 注册后开口率",
        "页面行名保留开口率，依靠父级分母解释。",
        "同名可保留，但 tooltip 必须说明具体分母。",
        "已裁决",
    ),
    (
        "新增 freetrial 用户",
        "freetrial_d0_users / fulfilled_subscription",
        "页面行名保留新增 freetrial 用户；定义层记录字段粒度差异。",
        "若 fulfilled_subscription 是事件数，不能在业务结论里当去重用户数使用。",
        "待校准",
    ),
    (
        "D1 回访 vs D1 留存",
        "次日任意行为回访 / 次日再次发消息",
        "两个口径都保留，不能合并。",
        "回访看是否回来；留存看是否再次使用核心能力。",
        "已裁决",
    ),
    (
        "付费用户平均收入",
        "MRR / 付费用户 或 MRR / 当前付费订阅数",
        "页面暂保留付费用户平均收入。",
        "当前计算若用订阅数，需确认一人一订阅假设是否成立。",
        "待校准",
    ),
    (
        "当前订阅状态",
        "active / trialing / past_due",
        "必须写订阅数，不能写用户数。",
        "Stripe subscription 状态和 customer 去重用户是两种粒度。",
        "已裁决",
    ),
    (
        "现金收入 / 净现金收入",
        "Stripe cash gross / cash net",
        "当前只有 Stripe，页面不加 Stripe 前缀。",
        "未来若接入其他收款渠道，再加来源前缀。",
        "已裁决",
    ),
]

DATA_SOURCE_ROWS = [
    (
        "PostHog 产品事件",
        "新增 UV、注册、freetrial、D0 开口、留存、活跃、文件/Connector/能力采纳、checkout funnel。",
        "generated_posthog_metrics/posthog_*.csv",
        "由 pull_operating_dashboard_generated_metrics.py 拉取和派生；用户类指标默认按 person 去重，事件类指标按 event 次数。",
    ),
    (
        "Stripe / finance 派生表",
        "现金收入、净现金收入、MRR、订阅状态、D+3 扣款、取消/结束订阅数。",
        "generated_posthog_metrics/calc_finance_2026-05-15_2026-05-26.csv",
        "取消/结束订阅数是截至当日累计 stock，不是当日新增取消 flow。",
    ),
    (
        "Google Ads",
        "广告花费、展示、点击、付费获客成本、CAC 等投放指标。",
        "generated_posthog_metrics/google_ads_daily_2026-05-15_2026-05-26.csv",
        "当前已接 Google Ads；Meta / KOC 等未完整映射时保留为空或待映射。",
    ),
    (
        "Grafana / AI Gateway",
        "AI Gateway 请求、错误、token/credits、模型/provider 分布、沙盒与入口安全指标。",
        "generated_posthog_metrics/grafana_*.csv / grafana_*.json",
        "工程质量类指标按监控源聚合，和 PostHog 用户事件不能直接相加。",
    ),
    (
        "源工作簿",
        "原始表结构、历史字段、个人看板反馈和人工补充上下文。",
        "web_restore_sources/JJHV_MoClaw_Dashboard_v31.xlsx",
        "生成器读取源工作簿后重建正式信息架构，最终以本 HTML 和脚本口径为准。",
    ),
]


def terminology_owner(canonical: str) -> str:
    if any(token in canonical for token in ["UV", "注册", "开口", "CTA", "CPA", "CAC"]):
        return "用户获取 / 激活转化"
    if any(token in canonical for token in ["DAU", "WAU", "MAU", "活跃", "频次", "使用", "附件", "Connector", "IM"]):
        return "用户活跃与使用分布"
    if any(token in canonical for token in ["D1", "D7", "D30", "留存", "回访", "未订阅"]):
        return "留存与流失"
    if any(token in canonical for token in ["freetrial", "付费", "订阅", "扣款", "现金", "收入", "MRR", "ARPU", "订单", "结账", "退款", "credits"]):
        return "业务经营"
    if any(token in canonical for token in ["Agent", "Gateway", "沙盒", "页面首次加载", "错误", "失败", "token", "响应"]):
        return "Agent & 工程质量"
    return "跨页"


def terminology_grain(canonical: str, definition: str) -> str:
    text = canonical + definition
    if "subscription 数" in definition or "订阅数" in canonical:
        return "subscription"
    if "customer" in definition:
        return "customer"
    if "flow" in definition:
        return "flow"
    if "事件" in definition or "次数" in canonical:
        return "event"
    if "比例" in definition or "率" in canonical:
        return "rate"
    if "金额" in definition or "收入" in canonical or "MRR" in canonical:
        return "money"
    if "用户" in text or "DAU" in canonical or "WAU" in canonical or "MAU" in canonical or "UV" in canonical:
        return "user"
    return "metric"


def appendix_row(sheet_title: str, seq: int, values: list[str], *, kind: str = "row") -> dict:
    return {
        "id": f"{sheet_title}-terminology-{seq}",
        "parent": "",
        "level": 0,
        "kind": kind,
        "ratio": False,
        "value_kind": "text",
        "values": list(values),
        "comments": [],
        "definition": "口径字典。",
        "has_children": False,
        "hidden": False,
        "keep_empty": True,
        "default_open": False,
    }


def build_terminology_appendix(source_sheet: dict) -> dict:
    sheet_title = source_sheet["title"]
    rows: list[dict] = []
    seq = 0

    def add(values: list[str], *, kind: str = "row") -> None:
        nonlocal seq
        seq += 1
        rows.append(appendix_row(sheet_title, seq, values, kind=kind))

    add(["附录 1 口径字典"], kind="section")
    add(["名词", "主管页面", "口径定义", "粒度"], kind="header")
    for canonical, spec in TERMINOLOGY_DICTIONARY.items():
        definition = spec["definition"]
        add(
            [
                canonical,
                terminology_owner(canonical),
                definition,
                terminology_grain(canonical, definition),
            ]
        )

    add(["2. 口径冲突项"], kind="section")
    add(["冲突项", "容易混淆的叫法 / 字段", "裁决", "说明", "状态"], kind="header")
    for row in TERMINOLOGY_CONFLICT_ROWS:
        add(list(row))

    add(["3. 数据源说明"], kind="section")
    add(["数据源", "覆盖指标", "当前文件 / 来源", "刷新方式 / 注意事项"], kind="header")
    for row in DATA_SOURCE_ROWS:
        add(list(row))

    max_col = max((len(row["values"]) for row in rows), default=1)
    return {**source_sheet, "rows": rows, "max_col": max_col}


def apply_operating_schema(sheets: list[dict]) -> list[dict]:
    lookup = row_lookup(sheets)
    by_title = {sheet["title"]: sheet for sheet in sheets}
    if "用户获取" in by_title:
        by_title["用户获取"] = build_acquisition_schema(by_title["用户获取"], lookup)
    if "用户激活与转化" in by_title:
        by_title["用户激活与转化"] = build_activation_schema(by_title["用户激活与转化"], lookup)
    if "用户留存与流失" in by_title:
        activity_sheet = {**by_title["用户留存与流失"], "title": "用户活跃与使用分布"}
        by_title["用户活跃与使用分布"] = build_activity_usage_schema(activity_sheet, lookup)
        by_title["用户留存与流失"] = build_retention_schema(by_title["用户留存与流失"], lookup)
    if "Agent 质量" in by_title:
        by_title["Agent 质量"] = build_agent_schema(by_title["Agent 质量"], lookup)
    if "财务" in by_title:
        by_title["财务"] = build_finance_schema(by_title["财务"], lookup)
    if "Dashboard" in by_title:
        schema_lookup = row_lookup([by_title[title] for title in ORDER if title in by_title])
        by_title["Dashboard"] = build_dashboard_home_schema(by_title["Dashboard"], schema_lookup)
    if "附录1_口径裁决" in by_title:
        by_title["附录1_口径裁决"] = build_terminology_appendix(by_title["附录1_口径裁决"])
    if "附录2 个人看板反馈" in by_title:
        by_title["附录2 个人看板反馈"] = build_employee_feedback_appendix(by_title["附录2 个人看板反馈"])
    return [by_title[title] for title in ORDER if title in by_title]


def source_values(lookup: list[tuple[str, dict, str]], label: list[str] | str, *, titles: list[str] | None = None, parent: str | None = None) -> list[str]:
    return find_source_values(lookup, label, titles=titles, parent=parent)


def channel_children(parent_label: str, *, include_campaign: bool = False) -> list[dict]:
    children = [
        {"label": "Direct", "source": "直接访问", "source_parent": parent_label},
        {"label": "SEO / Organic Search", "source": "SEO / 自然搜索", "source_parent": parent_label},
        {"label": "Organic Social", "source": "自然社交", "source_parent": parent_label},
        {"label": "Referral", "source": "外链推荐", "source_parent": parent_label},
        {
            "label": "Google Ads",
            "source": "Google Ads",
            "source_parent": parent_label,
            "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 2}] if include_campaign else [],
        },
        {"label": "Meta Ads", "source": "Meta Ads", "source_parent": parent_label},
        {"label": "KOC / KOL（待映射）", "values": blank_values()},
    ]
    return children


def value_by_channel(lookup: list[tuple[str, dict, str]], parent_label: str, channel_label: str) -> list[str]:
    source_label = {
        "Direct": "直接访问",
        "SEO / Organic Search": "SEO / 自然搜索",
        "Organic Social": "自然社交",
        "Referral": "外链推荐",
    }.get(channel_label, channel_label)
    return source_values(lookup, source_label, titles=["用户获取"], parent=parent_label)


def channel_ratio_children(
    lookup: list[tuple[str, dict, str]],
    numerator_parent: str,
    denominator_parent: str,
    *,
    include_campaign: bool = True,
) -> list[dict]:
    # Do not divide separate attribution rows. Channel-level funnel rates need
    # same-user cohort joins; until that exists, keep the drilldown structure
    # but leave values empty.
    return blank_channel_children(include_campaign=include_campaign, keep_empty=True)


def blank_channel_children(*, include_campaign: bool = True, keep_empty: bool = False) -> list[dict]:
    output: list[dict] = []
    for label in ["Direct", "SEO / Organic Search", "Organic Social", "Referral", "Google Ads", "Meta Ads", "KOC / KOL（待映射）"]:
        child = {"label": label, "values": blank_values(), "keep_empty": keep_empty}
        if label == "Google Ads" and include_campaign:
            child["children"] = [{"label": "campaign（待数据）", "values": blank_values(), "level": 3, "keep_empty": keep_empty}]
        output.append(child)
    return output


def new_uv_channel_children(
    filename: str,
    column: str,
    parent_values: list[str],
    *,
    include_campaign: bool = True,
) -> list[dict]:
    output: list[dict] = []
    for label in ["Direct", "SEO / Organic Search", "Organic Social", "Referral", "Google Ads", "Meta Ads"]:
        values = extra_group_series(filename, "channel", label, column, decimals=0)
        child = {"label": label, "values": values_with_share(values, parent_values)}
        if label == "Google Ads" and include_campaign:
            child["children"] = [{"label": "campaign（待数据）", "values": blank_values(), "level": 3}]
        output.append(child)
    output.append({"label": "KOC / KOL（待映射）", "values": blank_values()})
    return output


def new_uv_channel_ratio_children(
    filename: str,
    numerator_column: str,
    denominator_column: str,
    *,
    include_campaign: bool = True,
) -> list[dict]:
    output: list[dict] = []
    for label in ["Direct", "SEO / Organic Search", "Organic Social", "Referral", "Google Ads", "Meta Ads"]:
        numerator = extra_group_series(filename, "channel", label, numerator_column, decimals=0)
        denominator = extra_group_series(filename, "channel", label, denominator_column, decimals=0)
        child = {"label": label, "values": ratio_values(numerator, denominator)}
        if label == "Google Ads" and include_campaign:
            child["children"] = [{"label": "campaign（待数据）", "values": blank_values(), "level": 3, "keep_empty": True}]
        output.append(child)
    output.append({"label": "KOC / KOL（待映射）", "values": blank_values(), "keep_empty": True})
    return output


def onboarding_path_children() -> list[dict]:
    return [
        {
            "label": "onboarding 页面",
            "values": blank_values(),
            "keep_empty": True,
            "keep_empty_children": True,
            "children": [
                {"label": "/start", "values": blank_values(), "level": 3, "keep_empty": True},
                {"label": "/usecase", "values": blank_values(), "level": 3, "keep_empty": True},
            ],
        },
        {"label": "首页", "values": blank_values(), "keep_empty": True},
    ]


def paid_efficiency_children(label: str, google_values: list[str] | None = None) -> list[dict]:
    google_child: dict = {
        "label": "Google Ads",
        "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 3}],
    }
    if google_values is None:
        google_child.update(
            {
                "source": f"Google Ads {normalize_label_for_definition(label)}",
                "source_parent": normalize_label_for_definition(label),
            }
        )
    else:
        google_child["values"] = google_values
    return [
        google_child,
        {"label": "Meta Ads（待数据）", "values": blank_values()},
        {"label": "KOC / KOL（待数据）", "values": blank_values()},
    ]


def add_series(*series: list[str], kind: str = "number", decimals: int = 0) -> list[str]:
    return sum_value_series(*series, kind=kind, decimals=decimals)


def subtract_series(left: list[str], right: list[str], *, kind: str = "number", decimals: int = 0) -> list[str]:
    output: list[str] = []
    for left_value, right_value in zip(left, right):
        left_number = parse_display_number(left_value)
        right_number = parse_display_number(right_value)
        if left_number is None and right_number is None:
            output.append("—")
            continue
        output.append(fmt_extra_value((left_number or 0) - (right_number or 0), kind=kind, decimals=decimals))
    return output


def currency_division(numerator: list[str], denominator: list[str], *, factor: float = 1) -> list[str]:
    output: list[str] = []
    for numerator_value, denominator_value in zip(numerator, denominator):
        numerator_number = parse_display_number(numerator_value)
        denominator_number = parse_display_number(denominator_value)
        if numerator_number is None or denominator_number in (None, 0):
            output.append("—")
            continue
        output.append(fmt_extra_value(numerator_number / denominator_number * factor, kind="currency", decimals=1))
    return output


def values_with_share(child: list[str], parent: list[str]) -> list[str]:
    output: list[str] = []
    for child_value, parent_value in zip(child, parent):
        child_number = parse_display_number(child_value)
        parent_number = parse_display_number(parent_value)
        output.append(append_child_share(strip_share_annotation(child_value), child_number, parent_number))
    return output


def build_dashboard_home_schema(source_sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    task_csv = "posthog_task_proxy_2026-05-15_2026-05-26.csv"
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    engineering_csv = "posthog_engineering_quality_more_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    entry_csv = "grafana_entry_security_daily_2026-05-15_2026-05-26.csv"
    capability_csv = "posthog_capability_adoption_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"
    new_uv_cohort_csv = "posthog_new_uv_cohort_2026-05-15_2026-05-26.csv"
    nextday_cohort_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    google_ads_csv = "google_ads_daily_2026-05-15_2026-05-26.csv"
    new_uv = extra_series(new_uv_cohort_csv, "new_uv", decimals=0)
    google_ads_uv = source_values(lookup, "Google Ads", titles=["用户获取"], parent="新增 UV")
    meta_ads_uv = source_values(lookup, "Meta Ads", titles=["用户获取"], parent="新增 UV")
    koc_uv = source_values(lookup, "KOC / KOL（待映射）", titles=["用户获取"], parent="新增 UV")
    paid_uv = add_series(google_ads_uv, meta_ads_uv, koc_uv)
    organic_uv = subtract_series(new_uv, paid_uv)

    ad_spend = extra_series(google_ads_csv, "spend", kind="currency", decimals=1)
    paid_clicks = extra_series(google_ads_csv, "clicks", decimals=0)
    paid_uv_cost = currency_division(ad_spend, paid_uv)
    paid_trial_users = add_series(
        source_values(lookup, "Google Ads", titles=["用户获取"], parent="新增试订阅用户"),
        source_values(lookup, "Meta Ads", titles=["用户获取"], parent="新增试订阅用户"),
        source_values(lookup, "KOC / KOL（待映射）", titles=["用户获取"], parent="新增试订阅用户"),
    )

    login_cta_users = extra_series(web_core_csv, "new_uv_landing_cta_users", decimals=0)
    new_reg = extra_series(new_uv_cohort_csv, "registered_d0_users", decimals=0)
    new_trial = extra_series(new_uv_cohort_csv, "freetrial_d0_users", decimals=0)
    new_visit_dau = extra_series(new_uv_cohort_csv, "freetrial_message_d0_users", decimals=0)
    d0_messages = extra_series(new_uv_cohort_csv, "d0_messages", decimals=0)
    first_task_start = extra_series(new_uv_cohort_csv, "first_task_start_d0_users", decimals=0)
    first_task_done = extra_series(new_uv_cohort_csv, "first_task_done_d0_users", decimals=0)
    new_paid_users = extra_series(finance_csv, "new_paid_users", decimals=0)

    dau = extra_series(retention_csv, "dau", decimals=0)
    full_product_new_dau = extra_series(retention_csv, "new_dau", decimals=0)
    returning_dau = subtract_series(dau, full_product_new_dau)
    wau = extra_series(retention_csv, "wau", decimals=0)
    mau = extra_series(retention_csv, "mau", decimals=0)
    new_dau_cohort = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    dau_d1_retained = extra_series(nextday_csv, "next_chat_users", decimals=0)
    dau_d7_retained = extra_series(nextday_csv, "d7_chat_users", decimals=0)
    im_users = category_series(capability_csv, "im", "users")
    connector_users = category_series(capability_csv, "connector", "users")
    automation_users = category_series(capability_csv, "automation", "users")
    attachment_users = category_series(capability_csv, "attachment", "users")

    messages = extra_series(agent_csv, "messages", decimals=0)
    tool_use_requests = extra_series(agent_csv, "requests_with_tool_use", decimals=0)
    tool_use_count = first_present_series(
        extra_series(grafana_csv, "tool_use_count", decimals=0),
        extra_series(agent_csv, "tool_use_count", decimals=0),
    )
    llm_total = extra_series(agent_csv, "llm_total", decimals=0)
    message_failed = extra_series(agent_csv, "message_failed", decimals=0)
    active_tasks = extra_series(task_csv, "active_tasks", decimals=0)
    active_duration_p50 = source_values(lookup, "对话时长 P50 (分钟)", titles=["Dashboard"])
    active_duration_p90 = source_values(lookup, "对话时长 P90 (分钟)", titles=["Dashboard"])
    ttfb_p50 = extra_series(grafana_csv, "ttfb_p50_s", decimals=1)
    response_p50 = seconds_from_ms_series(agent_csv, "llm_p50_ms", decimals=1)
    page_load_p50 = seconds_from_ms_series(engineering_csv, "page_load_p50_ms", decimals=1)
    env_init_p50 = seconds_from_ms_series(engineering_csv, "env_init_p50_ms", decimals=1)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    checkout_started_users = extra_series(checkout_csv, "checkout_started_users", decimals=0)
    checkout_failed = add_series(
        extra_series(checkout_csv, "checkout_start_failed", decimals=0),
        extra_series(checkout_csv, "checkout_verify_failed", decimals=0),
        extra_series(checkout_csv, "checkout_fulfillment_failed", decimals=0),
    )
    inference_credits = extra_series(grafana_csv, "credits_consumed", decimals=0)
    llm_cost_dashboard_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Daily Credits Consumption",
        series_label="llm:request_completed",
        decimals=0,
    )
    cash_gross = extra_series(finance_csv, "cash_gross", kind="currency", decimals=0)
    cash_net = extra_series(finance_csv, "cash_net", kind="currency", decimals=0)
    refund = extra_series(finance_csv, "cash_refund", kind="currency", decimals=0)
    mrr = extra_series(finance_csv, "mrr_active", kind="currency", decimals=0)
    active_paid_subs = extra_series(finance_csv, "active_paid_subs", decimals=0)
    past_due_subs = extra_series(finance_csv, "status_past_due", decimals=0)

    sections = [
        {
            "title": "新用户链路",
            "items": [
                {
                    "label": "新增 UV",
                    "values": new_uv,
                    "default_open": True,
                    "children": [
                        {
                            "label": "付费渠道新增 UV",
                            "values": values_with_share(paid_uv, new_uv),
                            "children": [
                                {"label": "广告花费", "values": ad_spend, "level": 2},
                                {"label": "付费新增 UV 成本", "kind": "derived", "values": paid_uv_cost, "level": 2},
                            ],
                            "default_open": True,
                        },
                    ],
                },
                {
                    "label": "新增注册用户",
                    "values": new_reg,
                    "default_open": True,
                    "children": [
                        {"label": "注册率 (百分比)", "kind": "derived", "values": ratio_values(new_reg, new_uv)},
                    ],
                },
                {
                    "label": "新增 freetrial 用户",
                    "values": new_trial,
                    "default_open": True,
                    "children": [
                        {"label": "freetrial 率 (百分比)", "kind": "derived", "values": ratio_values(new_trial, new_reg)},
                    ],
                },
                {
                    "label": "新访 DAU",
                    "values": new_visit_dau,
                    "default_open": True,
                    "children": [
                        {"label": "开口率 (百分比)", "kind": "derived", "values": ratio_values(new_visit_dau, new_trial)},
                    ],
                },
                {"label": "新增付费用户", "values": new_paid_users},
            ],
        },
        {
            "title": "用户活跃与留存",
            "items": [
                {
                    "label": "DAU",
                    "values": dau,
                    "default_open": True,
                    "children": [
                        {"label": "新增 DAU", "values": append_share_series(full_product_new_dau, dau)},
                        {"label": "回访 DAU", "values": append_share_series(returning_dau, dau)},
                    ],
                },
                {"label": "WAU", "values": wau},
                {
                    "label": "MAU",
                    "values": mau,
                    "children": [
                        {"label": "DAU / MAU", "values": ratio_values(dau, mau), "ratio": True},
                    ],
                },
                {"label": "D1 留存率", "values": ratio_values(dau_d1_retained, new_dau_cohort), "ratio": True},
                {"label": "D7 留存率", "values": ratio_values(dau_d7_retained, new_dau_cohort), "ratio": True},
            ],
        },
        {
            "title": "用户使用",
            "items": [
                {"label": "活跃时长 P50（分钟）", "values": active_duration_p50},
                {"label": "活跃时长 P90（分钟）", "values": active_duration_p90},
                {"label": "人均日消息数", "values": division_values(messages, dau)},
                {"label": "人均日任务数", "values": division_values(active_tasks, dau)},
                {"label": "人均任务完成率", "values": blank_values(), "keep_empty": True, "ratio": True},
                {"label": "消息 tool 渗透率", "kind": "derived", "values": ratio_values(tool_use_requests, llm_total)},
                {
                    "label": "含 tool 消息平均 tool 调用数",
                    "kind": "derived",
                    "values": division_values_zero_when_no_numerator(tool_use_count, tool_use_requests),
                },
                {"label": "功能-IM 连接比例", "values": ratio_values(im_users, dau), "ratio": True},
                {"label": "功能-Connector 连接比例", "values": ratio_values(connector_users, dau), "ratio": True},
                {"label": "功能-定时任务比例", "values": ratio_values(automation_users, dau), "ratio": True},
                {"label": "功能-附件上传比例", "values": ratio_values(attachment_users, dau), "ratio": True},
            ],
        },
        {
            "title": "Agent & 工程质量",
            "items": [
                {"label": "消息失败率 (百分比)", "values": ratio_values(message_failed, messages)},
                {"label": "Chat 页面首次加载 P50（秒）", "values": page_load_p50, "keep_empty": True},
                {"label": "沙盒启动 P50（秒）", "values": env_init_p50, "keep_empty": True},
                {"label": "LLM 首 token P50（秒）", "values": ttfb_p50},
                {"label": "完整响应 P50（秒）", "values": response_p50},
            ],
        },
        {
            "title": "业务经营",
            "items": [
                {"label": "现金收入", "values": cash_gross},
                {"label": "当前有效付费订阅数", "values": active_paid_subs},
                {
                    "label": "总服务成本（待数据）",
                    "values": blank_values(),
                    "keep_empty": True,
                    "default_open": True,
                    "children": [
                        {
                            "label": "推理成本 credits",
                            "values": inference_credits,
                            "children": [
                            ],
                        },
                        {"label": "沙盒成本（待数据）", "values": blank_values(), "keep_empty": True},
                        {"label": "服务器成本（待数据）", "values": blank_values(), "keep_empty": True},
                    ],
                },
                {"label": "广告成本", "values": ad_spend},
                {
                    "label": "毛利（待数据）",
                    "values": blank_values(),
                    "keep_empty": True,
                    "default_open": True,
                    "children": [
                        {"label": "毛利率（待数据）", "kind": "derived", "values": blank_values(), "keep_empty": True},
                    ],
                },
            ],
        },
    ]
    return build_schema_sheet(source_sheet, lookup, sections)


def build_acquisition_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    new_uv_cohort_csv = "posthog_new_uv_cohort_2026-05-15_2026-05-26.csv"
    new_uv_channel_csv = "posthog_new_uv_channel_cohort_2026-05-15_2026-05-26.csv"
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    google_ads_csv = "google_ads_daily_2026-05-15_2026-05-26.csv"
    titles = ["用户获取"]
    new_uv = extra_series(new_uv_cohort_csv, "new_uv", decimals=0)
    new_reg = extra_series(new_uv_cohort_csv, "registered_d0_users", decimals=0)
    new_trial = extra_series(new_uv_cohort_csv, "freetrial_d0_users", decimals=0)
    new_visit_dau = extra_series(new_uv_cohort_csv, "freetrial_message_d0_users", decimals=0)
    new_uv_message_users = extra_series(new_uv_cohort_csv, "first_message_d0_users", decimals=0)
    new_paid = extra_series(finance_csv, "new_paid_users", decimals=0)
    ad_spend = extra_series(google_ads_csv, "spend", kind="currency", decimals=1)
    paid_impressions = extra_series(google_ads_csv, "impressions", decimals=0)
    paid_clicks = extra_series(google_ads_csv, "clicks", decimals=0)
    cpm = currency_division(ad_spend, paid_impressions, factor=1000)
    ctr = ratio_values(paid_clicks, paid_impressions)
    cpc = currency_division(ad_spend, paid_clicks)
    freetrial_rate = ratio_values(new_trial, new_reg)
    sections = [
        {
            "title": "投放",
            "items": [
                {
                    "label": "广告花费 (美元)",
                    "values": ad_spend,
                    "children": [
                        {
                            "label": "Google Ads",
                            "values": ad_spend,
                            "children": [{"label": "campaign（待数据）", "values": blank_values(), "level": 2}],
                        },
                        {"label": "Meta Ads（待数据）", "values": blank_values(), "keep_empty": True},
                        {"label": "KOC / KOL（待成本与映射）", "values": blank_values(), "keep_empty": True},
                    ],
                },
                {"label": "付费曝光 (个)", "values": paid_impressions},
                {"label": "CPM (美元)", "kind": "derived", "values": cpm, "children": paid_efficiency_children("CPM", cpm)},
                {"label": "付费点击 (个)", "values": paid_clicks},
                {"label": "CTR (百分比)", "kind": "derived", "values": ctr, "children": paid_efficiency_children("CTR", ctr)},
                {"label": "CPC (美元)", "kind": "derived", "values": cpc, "children": paid_efficiency_children("CPC", cpc)},
            ],
        },
        {
            "title": "新访数量",
            "items": [
                {
                    "label": "新增 UV",
                    "values": new_uv,
                    "children": new_uv_channel_children(new_uv_channel_csv, "new_uv", new_uv, include_campaign=True),
                    "default_open": True,
                },
            ],
        },
        {
            "title": "新访质量",
            "items": [
                {
                    "label": "新增注册",
                    "values": new_reg,
                    "children": new_uv_channel_children(new_uv_channel_csv, "registered_d0_users", new_reg, include_campaign=True),
                },
                {
                    "label": "注册率 (百分比)",
                    "kind": "derived",
                    "values": ratio_values(new_reg, new_uv),
                    "children": channel_ratio_children(lookup, "新增注册用户", "新增 UV"),
                },
                {
                    "label": "新增 freetrial 用户",
                    "values": new_trial,
                    "children": new_uv_channel_children(new_uv_channel_csv, "freetrial_d0_users", new_trial, include_campaign=True),
                },
                {
                    "label": "freetrial 率 (百分比)",
                    "kind": "derived",
                    "values": freetrial_rate,
                    "keep_empty": True,
                    "children": channel_ratio_children(lookup, "新增试用用户 (个)", "新增注册用户"),
                },
                {
                    "label": "新访 DAU",
                    "values": new_visit_dau,
                    "children": [
                        {
                            "label": "开口率 (百分比)",
                            "kind": "derived",
                            "values": ratio_values(new_visit_dau, new_trial),
                            "children": new_uv_channel_ratio_children(
                                new_uv_channel_csv,
                                "freetrial_message_d0_users",
                                "freetrial_d0_users",
                                include_campaign=True,
                            ),
                        }
                    ],
                },
                {
                    "label": "首日人均消息数",
                    "kind": "derived",
                    "values": division_values(extra_series(new_uv_cohort_csv, "d0_messages", decimals=0), new_uv_message_users),
                    "children": blank_channel_children(keep_empty=True),
                },
                {
                    "label": "前三日人均消息数",
                    "kind": "derived",
                    "values": division_values(extra_series(new_uv_cohort_csv, "d0_d2_messages", decimals=0), new_uv_message_users),
                    "children": blank_channel_children(keep_empty=True),
                },
                {"label": "当日首次付费用户", "values": new_paid},
                {
                    "label": "付费率 (百分比)",
                    "kind": "derived",
                    "values": blank_values(),
                    "keep_empty": True,
                    "children": channel_ratio_children(lookup, "新增付费用户", "新增 UV"),
                },
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_activation_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    new_uv_cohort_csv = "posthog_new_uv_cohort_2026-05-15_2026-05-26.csv"
    web_core_csv = "posthog_web_activation_core_2026-05-15_2026-05-26.csv"
    new_uv = extra_series(new_uv_cohort_csv, "new_uv", decimals=0)
    login_cta_users = extra_series(web_core_csv, "new_uv_landing_cta_users", decimals=0)
    reg_done = extra_series(new_uv_cohort_csv, "registered_d0_users", decimals=0)
    trial = extra_series(new_uv_cohort_csv, "freetrial_d0_users", decimals=0)
    new_visit_dau = extra_series(new_uv_cohort_csv, "freetrial_message_d0_users", decimals=0)
    new_uv_message_users = extra_series(new_uv_cohort_csv, "first_message_d0_users", decimals=0)
    first_task_start = extra_series(new_uv_cohort_csv, "first_task_start_d0_users", decimals=0)
    first_task_done = extra_series(new_uv_cohort_csv, "first_task_done_d0_users", decimals=0)
    next_return = extra_series(new_uv_cohort_csv, "next_return_d1_users", decimals=0)
    next_session = extra_series(new_uv_cohort_csv, "next_chat_d1_users", decimals=0)
    next_task_start = extra_series(new_uv_cohort_csv, "next_task_start_d1_users", decimals=0)
    next_task_done = extra_series(new_uv_cohort_csv, "next_task_done_d1_users", decimals=0)
    freetrial_rate = ratio_values(trial, reg_done)
    sections = [
        {
            "title": "访问到注册",
            "items": [
                {"label": "新增 UV", "values": new_uv},
                {
                    "label": "登录入口点击用户数",
                    "values": login_cta_users,
                    "default_open": True,
                    "children": [
                        {"label": "登录入口点击率", "kind": "derived", "values": ratio_values(login_cta_users, new_uv)},
                    ],
                },
                {
                    "label": "注册完成用户数",
                    "values": reg_done,
                    "default_open": True,
                    "children": [
                        {"label": "注册率 (百分比)", "kind": "derived", "values": ratio_values(reg_done, login_cta_users)},
                    ],
                },
            ],
        },
        {
            "title": "注册到 freetrial",
            "items": [
                {
                    "label": "新增 freetrial 用户",
                    "source": ["新增试用用户", "新增试用用户 (个)"],
                    "values": trial,
                    "default_open": True,
                    "children": [
                        {"label": "freetrial 率 (百分比)", "kind": "derived", "values": freetrial_rate, "keep_empty": True},
                    ],
                },
            ],
        },
        {
            "title": "freetrial 到使用",
            "items": [
                {
                    "label": "新访 DAU",
                    "values": new_visit_dau,
                    "default_open": True,
                    "children": [
                        {"label": "开口率 (百分比)", "kind": "derived", "values": ratio_values(new_visit_dau, trial)}
                    ],
                },
                {"label": "首日人均消息数", "kind": "derived", "values": division_values(extra_series(new_uv_cohort_csv, "d0_messages", decimals=0), new_uv_message_users)},
                {"label": "首次任务开始用户数", "values": first_task_start, "default_open": True, "children": [{"label": "首次发起任务率 (百分比)", "kind": "derived", "values": ratio_values(first_task_start, new_uv_message_users)}]},
                {"label": "首次任务完成用户数", "values": first_task_done, "default_open": True, "children": [{"label": "首次任务完成率 (百分比)", "kind": "derived", "values": ratio_values(first_task_done, first_task_start)}]},
            ],
        },
        {
            "title": "使用到回访",
            "items": [
                {"label": "次日回访用户数", "values": next_return, "default_open": True, "children": [{"label": "次日回访率 (百分比)", "kind": "derived", "values": ratio_values(next_return, new_uv_message_users)}]},
                {"label": "次日发起会话用户数", "values": next_session, "default_open": True, "children": [{"label": "次日回访开口率 (百分比)", "kind": "derived", "values": ratio_values(next_session, next_return)}]},
                {"label": "次日发起任务用户数", "values": next_task_start, "default_open": True, "children": [{"label": "次日回访发起任务率 (百分比)", "kind": "derived", "values": ratio_values(next_task_start, next_return)}]},
                {"label": "次日完成任务用户数", "values": next_task_done, "default_open": True, "children": [{"label": "次日任务完成率 (百分比)", "kind": "derived", "values": ratio_values(next_task_done, next_task_start)}]},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_activity_usage_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    task_csv = "posthog_task_proxy_2026-05-15_2026-05-26.csv"
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    distribution_detail_csv = "posthog_activity_distribution_detail_2026-05-15_2026-05-26.csv"
    capability_csv = "posthog_capability_adoption_2026-05-15_2026-05-26.csv"
    capability_detail_csv = "posthog_capability_adoption_detail_2026-05-15_2026-05-26.csv"

    dau = extra_series(retention_csv, "dau", decimals=0)
    wau = extra_series(retention_csv, "wau", decimals=0)
    mau = extra_series(retention_csv, "mau", decimals=0)
    new_dau = first_present_series(extra_series(retention_csv, "new_dau", decimals=0), extra_series(nextday_csv, "new_dau_cohort", decimals=0))
    returning_dau = first_present_series(extra_series(retention_csv, "returning_dau", decimals=0), subtract_series(dau, new_dau))
    new_wau = extra_series(retention_csv, "new_wau", decimals=0)
    returning_wau = first_present_series(extra_series(retention_csv, "returning_wau", decimals=0), subtract_series(wau, new_wau))
    new_mau = extra_series(retention_csv, "new_mau", decimals=0)
    returning_mau = first_present_series(extra_series(retention_csv, "returning_mau", decimals=0), subtract_series(mau, new_mau))
    freq_heavy = extra_series(retention_csv, "mau_freq_heavy_daily", decimals=0)
    freq_medium = extra_series(retention_csv, "mau_freq_medium", decimals=0)
    freq_light = extra_series(retention_csv, "mau_freq_light", decimals=0)
    freq_low = extra_series(retention_csv, "mau_freq_low", decimals=0)
    freq_cooling = extra_series(retention_csv, "mau_freq_cooling", decimals=0)
    sessions = extra_series(agent_csv, "chat_sessions", decimals=0)
    messages = extra_series(agent_csv, "messages", decimals=0)
    tool_use_requests = extra_series(agent_csv, "requests_with_tool_use", decimals=0)
    tool_use_count = first_present_series(
        extra_series(grafana_csv, "tool_use_count", decimals=0),
        extra_series(agent_csv, "tool_use_count", decimals=0),
    )
    llm_total = extra_series(agent_csv, "llm_total", decimals=0)
    single_message_sessions = source_values(lookup, "单消息对话数", titles=["Dashboard"])
    active_tasks = extra_series(task_csv, "active_tasks", decimals=0)
    new_region_children = distribution_children(distribution_detail_csv, "new", "region", limit=10)
    returning_region_children = distribution_children(distribution_detail_csv, "returning", "region", limit=10)
    im_users = category_series(capability_csv, "im", "users")
    im_new_users = category_series(capability_csv, "im", "new_users")
    im_existing_users = category_series(capability_csv, "im", "existing_users")
    im_events = category_series(capability_csv, "im", "events")
    connector_users = category_series(capability_csv, "connector", "users")
    connector_new_users = category_series(capability_csv, "connector", "new_users")
    connector_existing_users = category_series(capability_csv, "connector", "existing_users")
    connector_events = category_series(capability_csv, "connector", "events")
    automation_users = category_series(capability_csv, "automation", "users")
    automation_new_users = category_series(capability_csv, "automation", "new_users")
    automation_existing_users = category_series(capability_csv, "automation", "existing_users")
    automation_events = category_series(capability_csv, "automation", "events")
    attachment_users = category_series(capability_csv, "attachment", "users")
    attachment_new_users = category_series(capability_csv, "attachment", "new_users")
    attachment_existing_users = category_series(capability_csv, "attachment", "existing_users")
    im_new_detail_children = capability_detail_children(capability_detail_csv, "im", cohort="new", level=2)
    im_existing_detail_children = capability_detail_children(capability_detail_csv, "im", cohort="existing", level=2)
    connector_new_detail_children = capability_detail_children(capability_detail_csv, "connector", cohort="new", level=2)
    connector_existing_detail_children = capability_detail_children(capability_detail_csv, "connector", cohort="existing", level=2)
    new_web_interface_users = distribution_value_series(distribution_detail_csv, "new", "interface", "Web App")
    returning_web_interface_users = distribution_value_series(distribution_detail_csv, "returning", "interface", "Web App")
    new_interface_children = [
        {"label": "Web App", "values": append_share_series(new_web_interface_users, new_dau), "level": 2},
        {
            "label": "IM",
            "values": append_share_series(im_new_users, new_dau),
            "level": 2,
            "children": capability_detail_children(capability_detail_csv, "im", cohort="new", level=3),
        },
    ]
    returning_interface_children = [
        {"label": "Web App", "values": append_share_series(returning_web_interface_users, returning_dau), "level": 2},
        {
            "label": "IM",
            "values": append_share_series(im_existing_users, returning_dau),
            "level": 2,
            "children": capability_detail_children(capability_detail_csv, "im", cohort="existing", level=3),
        },
    ]

    upload_users = extra_series(file_csv, "upload_started_users", decimals=0)
    upload_count = extra_series(file_csv, "upload_started", decimals=0)

    sections = [
        {
            "title": "活跃规模",
            "items": [
                {
                    "label": "DAU",
                    "values": dau,
                    "default_open": True,
                    "children": [
                        {"label": "新增 DAU", "values": append_share_series(new_dau, dau)},
                        {"label": "回访 DAU", "values": append_share_series(returning_dau, dau)},
                    ],
                },
                {
                    "label": "WAU",
                    "values": wau,
                    "children": [
                        {"label": "新增 WAU", "values": append_share_series(new_wau, wau)},
                        {"label": "回访 WAU", "values": append_share_series(returning_wau, wau)},
                    ],
                },
                {
                    "label": "MAU",
                    "values": mau,
                    "children": [
                        {"label": "新增 MAU", "values": append_share_series(new_mau, mau)},
                        {"label": "回访 MAU", "values": append_share_series(returning_mau, mau)},
                    ],
                },
                {"label": "DAU / WAU", "kind": "derived", "values": ratio_values(dau, wau)},
                {"label": "DAU / MAU", "kind": "derived", "values": ratio_values(dau, mau)},
            ],
        },
        {
            "title": "活跃频次",
            "items": [
                {
                    "label": "活跃频次分层",
                    "values": mau,
                    "default_open": True,
                    "children": [
                        {"label": "重度日用", "values": append_share_series(freq_heavy, mau)},
                        {"label": "中度", "values": append_share_series(freq_medium, mau)},
                        {"label": "轻度", "values": append_share_series(freq_light, mau)},
                        {"label": "低频", "values": append_share_series(freq_low, mau)},
                        {"label": "冷却", "values": append_share_series(freq_cooling, mau)},
                    ],
                },
            ],
        },
        {
            "title": "Agent 使用量与深度",
            "items": [
                {"label": "会话时长 P50（分钟）", "source": "对话时长 P50 (分钟)", "titles": ["Dashboard"]},
                {"label": "会话时长 P90（分钟）", "source": "对话时长 P90 (分钟)", "titles": ["Dashboard"]},
                {"label": "对话数", "values": sessions},
                {"label": "人均对话数", "kind": "derived", "values": division_values(sessions, dau)},
                {"label": "消息数", "values": messages},
                {"label": "人均消息数", "kind": "derived", "values": division_values(messages, dau)},
                {"label": "tool调用数", "values": tool_use_count},
                {"label": "每条消息平均 tool 调用数", "kind": "derived", "values": division_values(tool_use_count, messages)},
                {"label": "消息 tool 渗透率", "kind": "derived", "values": ratio_values(tool_use_requests, llm_total)},
                {
                    "label": "含 tool 消息平均 tool 调用数",
                    "kind": "derived",
                    "values": division_values_zero_when_no_numerator(tool_use_count, tool_use_requests),
                },
                {"label": "活跃任务数", "values": active_tasks},
                {"label": "人均日任务数", "kind": "derived", "values": division_values(active_tasks, dau)},
            ],
        },
        {
            "title": "功能使用量与深度",
            "items": [
                {
                    "label": "IM 连接",
                    "values": append_share_series(im_users, dau),
                    "children": [
                        {
                            "label": "新增使用用户",
                            "values": append_share_series(im_new_users, dau),
                            "children": im_new_detail_children,
                        },
                        {
                            "label": "存量使用用户",
                            "values": append_share_series(im_existing_users, dau),
                            "children": im_existing_detail_children,
                        },
                        {"label": "使用次数", "values": im_events},
                    ],
                },
                {
                    "label": "Connector 连接",
                    "values": append_share_series(connector_users, dau),
                    "children": [
                        {
                            "label": "新增使用用户",
                            "values": append_share_series(connector_new_users, dau),
                            "children": connector_new_detail_children,
                        },
                        {
                            "label": "存量使用用户",
                            "values": append_share_series(connector_existing_users, dau),
                            "children": connector_existing_detail_children,
                        },
                        {"label": "使用次数", "values": connector_events},
                    ],
                },
                {
                    "label": "定时任务",
                    "values": append_share_series(automation_users, dau),
                    "children": [
                        {"label": "新增使用用户", "values": append_share_series(automation_new_users, dau)},
                        {"label": "存量使用用户", "values": append_share_series(automation_existing_users, dau)},
                        {"label": "使用次数", "values": automation_events},
                    ],
                },
                {
                    "label": "附件上传",
                    "values": append_share_series(attachment_users, dau),
                    "children": [
                        {"label": "新增使用用户", "values": append_share_series(attachment_new_users, dau)},
                        {"label": "存量使用用户", "values": append_share_series(attachment_existing_users, dau)},
                        {"label": "使用次数", "values": upload_count},
                    ],
                },
            ],
        },
        {
            "title": "用户分布",
            "items": [
                {
                    "label": "DAU地域分布",
                    "values": dau,
                    "keep_empty_children": True,
                    "children": [
                        {
                            "label": "新增",
                            "values": append_share_series(new_dau, dau),
                            "children": new_region_children,
                        },
                        {
                            "label": "回访",
                            "values": append_share_series(returning_dau, dau),
                            "children": returning_region_children,
                        },
                    ],
                },
                {
                    "label": "DAU Interface 分布",
                    "values": dau,
                    "keep_empty_children": True,
                    "children": [
                        {
                            "label": "新增",
                            "values": append_share_series(new_dau, dau),
                            "children": new_interface_children,
                        },
                        {
                            "label": "回访",
                            "values": append_share_series(returning_dau, dau),
                            "children": returning_interface_children,
                        },
                    ],
                },
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_retention_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    nextday_csv = "posthog_new_dau_nextday_2026-05-15_2026-05-26.csv"
    subscription_csv = "posthog_retention_subscription_segments_2026-05-15_2026-05-26.csv"
    segments = [
        ("unsubscribed", "未订阅用户"),
        ("freetrial", "freetrial 用户"),
        ("paid", "付费用户"),
        ("past_due", "扣款失败用户"),
        ("canceled_subscription", "订阅过期用户"),
    ]
    segment_series: dict[str, dict[str, list[str]]] = {}
    for key, _label in segments:
        segment_series[key] = {
            "users": extra_series(subscription_csv, f"{key}_users", decimals=0),
            "dau": extra_series(subscription_csv, f"{key}_dau", decimals=0),
            "mau": extra_series(subscription_csv, f"{key}_mau", decimals=0),
            "inactive_7d": extra_series(subscription_csv, f"{key}_inactive_7d", decimals=0),
            "d1_cohort": extra_series(subscription_csv, f"{key}_d1_cohort", decimals=0),
            "d1_retained": extra_series(subscription_csv, f"{key}_d1_retained", decimals=0),
            "return_users": extra_series(subscription_csv, f"{key}_return_users", decimals=0),
            "return_7d_active": extra_series(subscription_csv, f"{key}_return_7d_active", decimals=0),
        }
    total_users = sum_value_series(*(segment_series[key]["users"] for key, _label in segments))
    total_return = sum_value_series(*(segment_series[key]["return_users"] for key, _label in segments))
    total_return_active = sum_value_series(*(segment_series[key]["return_7d_active"] for key, _label in segments))
    new_dau_cohort = extra_series(nextday_csv, "new_dau_cohort", decimals=0)
    d1_retained = extra_series(nextday_csv, "next_return_users", decimals=0)
    pool_rows = [
        {"label": label, "values": append_share_series(segment_series[key]["users"], total_users)}
        for key, label in segments
    ]
    risk_rows = [
        {"label": "未订阅 7+ 天未活跃", "values": append_share_series(segment_series["unsubscribed"]["inactive_7d"], segment_series["unsubscribed"]["users"])},
        {"label": "freetrial 7+ 天未活跃", "values": append_share_series(segment_series["freetrial"]["inactive_7d"], segment_series["freetrial"]["users"])},
        {"label": "付费 7+ 天未活跃", "values": append_share_series(segment_series["paid"]["inactive_7d"], segment_series["paid"]["users"])},
        {"label": "扣款失败 7+ 天未活跃", "values": append_share_series(segment_series["past_due"]["inactive_7d"], segment_series["past_due"]["users"])},
        {"label": "订阅过期 7+ 天未活跃", "values": append_share_series(segment_series["canceled_subscription"]["inactive_7d"], segment_series["canceled_subscription"]["users"])},
    ]
    return_children = [
        {"label": label.replace("用户", "回流"), "values": append_share_series(segment_series[key]["return_users"], total_return)}
        for key, label in segments
    ]
    sections = [
        {
            "title": "历史活跃用户池（按订阅状态）",
            "items": pool_rows,
        },
        {
            "title": "一周未活跃",
            "items": risk_rows,
        },
        {
            "title": "回流",
            "items": [
                {"label": "回流用户", "values": total_return, "default_open": True, "children": return_children},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_agent_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    engineering_csv = "posthog_engineering_quality_more_2026-05-15_2026-05-26.csv"
    retention_csv = "posthog_retention_activity_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    grafana_errors_csv = "grafana_ai_gateway_errors_by_type_2026-05-15_2026-05-26.csv"

    dau = extra_series(retention_csv, "dau", decimals=0)
    messages = extra_series(agent_csv, "messages", decimals=0)
    message_failed = extra_series(agent_csv, "message_failed", decimals=0)
    stream_errors = extra_series(agent_csv, "stream_errors", decimals=0)
    chat_stream_errored = extra_series(engineering_csv, "chat_stream_errored", decimals=0)
    ws_error = extra_series(engineering_csv, "ws_error", decimals=0)
    ws_error_users = extra_series(engineering_csv, "ws_error_users", decimals=0)
    api_calls = extra_series(engineering_csv, "api_calls", decimals=0)
    api_fail = extra_series(engineering_csv, "api_fail", decimals=0)
    page_load_p50 = seconds_from_ms_series(engineering_csv, "page_load_p50_ms", decimals=1)
    page_load_p95 = seconds_from_ms_series(engineering_csv, "page_load_p95_ms", decimals=1)
    env_init_p50 = seconds_from_ms_series(engineering_csv, "env_init_p50_ms", decimals=1)
    env_init_p95 = seconds_from_ms_series(engineering_csv, "env_init_p95_ms", decimals=1)
    env_init_first_daily_p50 = seconds_from_ms_series(engineering_csv, "env_init_first_daily_p50_ms", decimals=1)
    env_init_first_daily_p95 = seconds_from_ms_series(engineering_csv, "env_init_first_daily_p95_ms", decimals=1)
    llm_ttfb_p50 = first_present_series(seconds_from_ms_series(agent_csv, "llm_ttfb_p50_ms", decimals=1), extra_series(grafana_csv, "ttfb_p50_s", decimals=1))
    llm_ttfb_p95 = first_present_series(seconds_from_ms_series(agent_csv, "llm_ttfb_p95_ms", decimals=1), extra_series(grafana_csv, "ttfb_p95_s", decimals=1))
    first_message_ttfb_p50 = seconds_from_ms_series(agent_csv, "first_message_ttfb_p50_ms", decimals=1)
    first_message_ttfb_p95 = seconds_from_ms_series(agent_csv, "first_message_ttfb_p95_ms", decimals=1)
    llm_response_p50 = seconds_from_ms_series(agent_csv, "llm_p50_ms", decimals=1)
    llm_response_p95 = first_present_series(seconds_from_ms_series(agent_csv, "llm_p95_ms", decimals=1), extra_series(grafana_csv, "request_p95_s", decimals=1))
    first_message_response_p50 = seconds_from_ms_series(agent_csv, "first_message_llm_p50_ms", decimals=1)
    first_message_response_p95 = seconds_from_ms_series(agent_csv, "first_message_llm_p95_ms", decimals=1)
    env_init_count = extra_series(engineering_csv, "env_init_count", decimals=0)
    env_init_failed = extra_series(engineering_csv, "env_init_failed", decimals=0)
    env_init_success = subtract_series(env_init_count, env_init_failed)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    gateway_error_children = [
        {
            "label": "流式中断",
            "values": append_share_series(extra_group_series(grafana_errors_csv, "error_type", "stream_abort", decimals=0), gateway_errors),
            "keep_empty": True,
        },
        {
            "label": "超时",
            "values": append_share_series(extra_group_series(grafana_errors_csv, "error_type", "timeout", decimals=0), gateway_errors),
            "keep_empty": True,
        },
        {
            "label": "上游错误",
            "values": append_share_series(extra_group_series(grafana_errors_csv, "error_type", "upstream_error", decimals=0), gateway_errors),
            "keep_empty": True,
        },
        {
            "label": "Bedrock 错误",
            "values": append_share_series(extra_group_series(grafana_errors_csv, "error_type", "bedrock_error", decimals=0), gateway_errors),
            "keep_empty": True,
        },
    ]

    sections = [
        {
            "title": "Agent 响应体验",
            "items": [
                {"label": "Chat 页面首次加载 P50（秒）", "values": page_load_p50, "keep_empty": True},
                {"label": "Chat 页面首次加载 P95（秒）", "values": page_load_p95, "keep_empty": True},
                {
                    "label": "沙盒启动 P50（秒）",
                    "values": env_init_p50,
                    "default_open": True,
                    "children": [
                        {"label": "用户每日首次沙盒启动 P50（秒）", "values": env_init_first_daily_p50, "keep_empty": True},
                    ],
                },
                {
                    "label": "沙盒启动 P95（秒）",
                    "values": env_init_p95,
                    "default_open": True,
                    "children": [
                        {"label": "用户每日首次沙盒启动 P95（秒）", "values": env_init_first_daily_p95, "keep_empty": True},
                    ],
                },
                {
                    "label": "首 token P50（秒）",
                    "values": llm_ttfb_p50,
                    "default_open": True,
                    "children": [
                        {"label": "用户首条消息 首 token P50（秒）", "values": first_message_ttfb_p50, "keep_empty": True},
                    ],
                },
                {
                    "label": "LLM 首 token P95（秒）",
                    "values": llm_ttfb_p95,
                    "default_open": True,
                    "children": [
                        {"label": "用户首条消息 首 token P95（秒）", "values": first_message_ttfb_p95, "keep_empty": True},
                    ],
                },
                {
                    "label": "完整响应 P50（秒）",
                    "values": llm_response_p50,
                    "default_open": True,
                    "children": [
                        {"label": "用户首条消息 完整响应 P50（秒）", "values": first_message_response_p50, "keep_empty": True},
                    ],
                },
                {
                    "label": "完整响应 P95（秒）",
                    "values": llm_response_p95,
                    "default_open": True,
                    "children": [
                        {"label": "用户首条消息 完整响应 P95（秒）", "values": first_message_response_p95, "keep_empty": True},
                    ],
                },
            ],
        },
        {
            "title": "Agent 失败与中断",
            "items": [
                {
                    "label": "消息失败数",
                    "values": message_failed,
                    "default_open": True,
                    "children": [
                        {"label": "消息失败率 (百分比)", "kind": "derived", "values": ratio_values(message_failed, messages)},
                    ],
                },
                {
                    "label": "流式错误数",
                    "source": "stream_errors",
                    "values": stream_errors,
                    "default_open": True,
                    "children": [
                        {"label": "流式错误率 (百分比)", "kind": "derived", "values": ratio_values(stream_errors, messages)},
                    ],
                },
                {
                    "label": "对话流中断",
                    "values": chat_stream_errored,
                    "default_open": True,
                    "children": [
                        {"label": "对话流中断率 (百分比)", "kind": "derived", "values": ratio_values(chat_stream_errored, messages)},
                    ],
                },
                {
                    "label": "实时连接错误",
                    "values": ws_error,
                    "default_open": True,
                    "children": [
                        {"label": "实时连接错误率 (百分比)", "kind": "derived", "values": ratio_values(ws_error_users, dau)},
                    ],
                },
                {
                    "label": "AI Gateway 错误",
                    "values": gateway_errors,
                    "default_open": True,
                    "children": [
                        {"label": "AI Gateway 错误率 (百分比)", "kind": "derived", "values": ratio_values(gateway_errors, gateway_requests)},
                        *gateway_error_children,
                    ],
                },
                {
                    "label": "API 失败",
                    "values": api_fail,
                    "default_open": True,
                    "children": [
                        {"label": "API 失败率 (百分比)", "kind": "derived", "values": ratio_values(api_fail, api_calls)},
                    ],
                },
                {
                    "label": "沙盒启动次数",
                    "values": env_init_count,
                    "default_open": True,
                    "children": [
                        {"label": "沙盒启动失败", "values": env_init_failed},
                        {"label": "沙盒启动失败率 (百分比)", "kind": "derived", "values": ratio_values(env_init_failed, env_init_count)},
                    ],
                },
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_engineering_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    engineering_csv = "posthog_engineering_quality_more_2026-05-15_2026-05-26.csv"
    file_csv = "posthog_file_usage_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    grafana_errors_csv = "grafana_ai_gateway_errors_by_type_2026-05-15_2026-05-26.csv"
    entry_csv = "grafana_entry_security_daily_2026-05-15_2026-05-26.csv"
    sandbox_csv = "grafana_sandbox_fleet_daily_2026-05-15_2026-05-26.csv"
    checkout_csv = "posthog_finance_checkout_funnel_2026-05-15_2026-05-26.csv"

    api_calls = extra_series(engineering_csv, "api_calls", decimals=0)
    api_success = extra_series(engineering_csv, "api_success", decimals=0)
    api_fail = extra_series(engineering_csv, "api_fail", decimals=0)
    env_init_count = extra_series(engineering_csv, "env_init_count", decimals=0)
    env_init_failed = extra_series(engineering_csv, "env_init_failed", decimals=0)
    upload_started = extra_series(file_csv, "upload_started", decimals=0)
    upload_completed = extra_series(file_csv, "upload_completed", decimals=0)
    upload_failed = extra_series(file_csv, "upload_failed", decimals=0)
    entry_requests = extra_series(entry_csv, "entry_requests", decimals=0)
    entry_5xx = extra_series(entry_csv, "entry_5xx", decimals=0)
    target_5xx = extra_series(entry_csv, "target_5xx", decimals=0)
    elb_5xx = extra_series(entry_csv, "elb_5xx", decimals=0)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    gateway_errors = zero_missing_by_denominator(extra_series(grafana_csv, "gateway_errors", decimals=0), gateway_requests)
    sandbox_checked = extra_series(sandbox_csv, "sandbox_checked", decimals=0)
    sandbox_unreachable = extra_series(sandbox_csv, "sandbox_unreachable", decimals=0)
    checkout_started_users = extra_series(checkout_csv, "checkout_started_users", decimals=0)
    checkout_failed = sum_value_series(
        extra_series(checkout_csv, "checkout_start_failed", decimals=0),
        extra_series(checkout_csv, "checkout_verify_failed", decimals=0),
        extra_series(checkout_csv, "checkout_fulfillment_failed", decimals=0),
    )
    page_loads = extra_series(engineering_csv, "page_loads", decimals=0)
    page_load_p50 = seconds_from_ms_series(engineering_csv, "page_load_p50_ms", decimals=1)
    page_load_p95 = seconds_from_ms_series(engineering_csv, "page_load_p95_ms", decimals=1)
    gateway_error_children = [
        {"label": "流式中断", "values": extra_group_series(grafana_errors_csv, "error_type", "stream_abort", decimals=0)},
        {"label": "超时", "values": extra_group_series(grafana_errors_csv, "error_type", "timeout", decimals=0)},
        {"label": "上游错误", "values": extra_group_series(grafana_errors_csv, "error_type", "upstream_error", decimals=0)},
        {"label": "Bedrock 错误", "values": extra_group_series(grafana_errors_csv, "error_type", "bedrock_error", decimals=0)},
    ]

    sections = [
        {
            "title": "入口",
            "items": [
                {"label": "入口请求", "values": entry_requests},
                {
                    "label": "入口 5xx",
                    "values": entry_5xx,
                    "children": [
                        {"label": "Target 5xx", "values": target_5xx},
                        {"label": "ELB 5xx", "values": elb_5xx},
                    ],
                },
                {"label": "入口服务 5xx 率 (百分比)", "kind": "derived", "values": ratio_values(entry_5xx, entry_requests)},
            ],
        },
        {
            "title": "页面体验",
            "items": [
                {
                    "label": "页面首次加载次数",
                    "values": page_loads,
                    "children": [
                        {"label": "页面首次加载 P50（秒）", "values": page_load_p50, "keep_empty": True},
                        {"label": "页面首次加载 P95（秒）", "values": page_load_p95, "keep_empty": True},
                    ],
                },
            ],
        },
        {
            "title": "HTTP 服务",
            "items": [
                {"label": "API 请求", "values": api_calls},
                {"label": "API 成功", "values": api_success},
                {"label": "API 失败", "values": api_fail},
                {"label": "API 成功率 (百分比)", "kind": "derived", "values": ratio_values(api_success, api_calls)},
                {
                    "label": "API 耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "API 耗时 P50（秒）", "values": seconds_from_ms_series(engineering_csv, "api_p50_ms", decimals=1)},
                        {"label": "API 耗时 P95（秒）", "values": seconds_from_ms_series(engineering_csv, "api_p95_ms", decimals=1)},
                        {"label": "API 耗时 P99（秒）", "values": seconds_from_ms_series(engineering_csv, "api_p99_ms", decimals=1)},
                    ],
                },
            ],
        },
        {
            "title": "实时连接",
            "items": [
                {"label": "实时连接错误", "values": extra_series(engineering_csv, "ws_error", decimals=0)},
                {"label": "消息发送失败", "values": extra_series(engineering_csv, "message_failed", decimals=0)},
                {"label": "对话流中断", "values": extra_series(engineering_csv, "chat_stream_errored", decimals=0)},
            ],
        },
        {
            "title": "AI Gateway",
            "items": [
                {"label": "AI Gateway 请求", "values": gateway_requests},
                {
                    "label": "AI Gateway 错误",
                    "values": gateway_errors,
                    "children": gateway_error_children,
                },
                {"label": "AI Gateway 错误率 (百分比)", "kind": "derived", "values": ratio_values(gateway_errors, gateway_requests)},
                {
                    "label": "LLM 首 token",
                    "values": blank_values(),
                    "children": [
                        {"label": "首 token P50（秒）", "values": extra_series(grafana_csv, "ttfb_p50_s", decimals=1)},
                        {"label": "首 token P95（秒）", "values": extra_series(grafana_csv, "ttfb_p95_s", decimals=1)},
                    ],
                },
                {
                    "label": "完整响应耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "完整响应 P95（秒）", "values": extra_series(grafana_csv, "request_p95_s", decimals=1)},
                        {"label": "完整响应 P99（秒）", "values": extra_series(grafana_csv, "request_p99_s", decimals=1)},
                    ],
                },
            ],
        },
        {
            "title": "沙盒",
            "items": [
                {"label": "沙盒检查次数", "values": sandbox_checked},
                {"label": "沙盒不可达", "values": sandbox_unreachable},
                {"label": "沙盒不可达率 (百分比)", "kind": "derived", "values": ratio_values(sandbox_unreachable, sandbox_checked)},
                {
                    "label": "沙盒启动耗时",
                    "values": blank_values(),
                    "children": [
                        {"label": "沙盒启动 P50（秒）", "values": seconds_from_ms_series(engineering_csv, "env_init_p50_ms", decimals=1)},
                        {"label": "沙盒启动 P95（秒）", "values": seconds_from_ms_series(engineering_csv, "env_init_p95_ms", decimals=1)},
                    ],
                },
                {"label": "沙盒重启失败", "values": extra_series(engineering_csv, "sandbox_restart_failed", decimals=0)},
            ],
        },
        {
            "title": "附件与支付",
            "items": [
                {"label": "附件上传开始", "values": upload_started},
                {"label": "附件上传成功", "values": upload_completed},
                {"label": "附件上传成功率 (百分比)", "kind": "derived", "values": ratio_values(upload_completed, upload_started)},
                {"label": "附件上传失败", "values": upload_failed},
                {"label": "附件上传失败率 (百分比)", "kind": "derived", "values": ratio_values(upload_failed, upload_started)},
                {"label": "发起结账用户", "values": checkout_started_users},
                {"label": "结账失败", "values": checkout_failed},
                {"label": "结账失败率 (百分比)", "kind": "derived", "values": ratio_values(checkout_failed, checkout_started_users)},
                {"label": "结账发起失败", "values": extra_series(engineering_csv, "checkout_start_failed", decimals=0)},
                {"label": "结账验证失败", "values": extra_series(engineering_csv, "checkout_verify_failed", decimals=0)},
                {"label": "结账履约失败", "values": extra_series(engineering_csv, "checkout_fulfillment_failed", decimals=0)},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def build_finance_schema(sheet: dict, lookup: list[tuple[str, dict, str]]) -> dict:
    finance_csv = "calc_finance_2026-05-15_2026-05-26.csv"
    google_ads_csv = "google_ads_daily_2026-05-15_2026-05-26.csv"
    new_uv_channel_csv = "posthog_new_uv_channel_cohort_2026-05-15_2026-05-26.csv"
    grafana_csv = "grafana_ai_gateway_daily_2026-05-15_2026-05-26.csv"
    agent_csv = "posthog_agent_quality_more_2026-05-15_2026-05-26.csv"
    task_csv = "posthog_task_proxy_2026-05-15_2026-05-26.csv"

    active_paid_subs = extra_series(finance_csv, "active_paid_subs", decimals=0)
    mrr = extra_series(finance_csv, "mrr_active", kind="currency", decimals=0)
    new_paid = extra_series(finance_csv, "new_paid_users", decimals=0)
    cash_gross = extra_series(finance_csv, "cash_gross", kind="currency", decimals=0)
    refund = extra_series(finance_csv, "cash_refund", kind="currency", decimals=0)
    cash_net = extra_series(finance_csv, "cash_net", kind="currency", decimals=0)
    ad_spend = extra_series(google_ads_csv, "spend", kind="currency", decimals=1)
    google_ads_reg = extra_group_series(new_uv_channel_csv, "channel", "Google Ads", "registered_d0_users", decimals=0)
    messages = extra_series(agent_csv, "messages", decimals=0)
    active_tasks = extra_series(task_csv, "active_tasks", decimals=0)
    gateway_requests = extra_series(grafana_csv, "gateway_requests", decimals=0)
    credits = extra_series(grafana_csv, "credits_consumed", decimals=0)
    llm_cost_dashboard_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Daily Credits Consumption",
        series_label="llm:request_completed",
        decimals=0,
    )
    anthropic_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Credits by Provider",
        breakdown_value="anthropic",
        decimals=0,
    )
    fal_run_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Credits by Provider",
        breakdown_value="fal-run",
        decimals=0,
    )
    deepseek_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Credits by Provider",
        breakdown_value="deepseek",
        decimals=0,
    )
    fal_queue_credits = dashboard_trend_series(
        "1412422_llm_cost.json",
        "Credits by Provider",
        breakdown_value="fal-queue",
        decimals=0,
    )

    sections = [
        {
            "title": "收入",
            "items": [
                {"label": "Stripe 现金收入", "values": cash_gross},
                {"label": "退款", "values": refund},
                {"label": "净收入", "values": cash_net},
                {
                    "label": "月经常性收入 MRR",
                    "values": mrr,
                    "children": [
                        {"label": "当前有效付费订阅数", "values": active_paid_subs},
                        {"label": "ARPU", "kind": "derived", "values": currency_division(mrr, active_paid_subs)},
                    ],
                },
                {"label": "新增付费用户", "values": new_paid},
                {"label": "截至当日已取消/结束订阅数", "values": extra_series(finance_csv, "paid_churn_subs", decimals=0)},
            ],
        },
        {
            "title": "直接服务成本",
            "items": [
                {
                    "label": "推理成本 credits",
                    "values": credits,
                    "children": [
                        {"label": "LLM Cost Dashboard credits", "values": llm_cost_dashboard_credits},
                        {"label": "anthropic credits", "values": anthropic_credits},
                        {"label": "fal-run credits", "values": fal_run_credits},
                        {"label": "deepseek credits", "values": deepseek_credits},
                        {"label": "fal-queue credits", "values": fal_queue_credits},
                        {"label": "AI Gateway 请求", "values": gateway_requests},
                        {"label": "credits / 消息", "values": division_values(credits, messages)},
                        {"label": "credits / 任务", "values": division_values(credits, active_tasks)},
                    ],
                },
                {"label": "云服务成本（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "支付手续费（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "总服务成本（待数据）", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "毛利",
            "items": [
                {"label": "毛利（待数据）", "values": blank_values(), "keep_empty": True},
                {"label": "毛利率（待数据）", "values": blank_values(), "keep_empty": True},
            ],
        },
        {
            "title": "获客支出",
            "items": [
                {
                    "label": "广告支出",
                    "values": ad_spend,
                    "children": [
                        {"label": "Google Ads 花费", "values": ad_spend},
                        {"label": "KOC / KOL 花费（待数据）", "values": blank_values(), "keep_empty": True},
                        {"label": "其他渠道花费（待数据）", "values": blank_values(), "keep_empty": True},
                    ],
                },
                {"label": "注册 CPA", "values": currency_division(ad_spend, google_ads_reg), "ratio": True},
                {"label": "CAC", "values": currency_division(ad_spend, new_paid), "ratio": True},
            ],
        },
        {
            "title": "扣获客后的经营结果",
            "items": [
                {"label": "贡献利润（待数据）", "values": blank_values(), "keep_empty": True},
            ],
        },
    ]
    return build_schema_sheet(sheet, lookup, sections)


def cell_html(row: dict, value: str, first: bool, *, show_info: bool = True) -> str:
    cls = ["l"] if first else ["num"]
    if row.get("ratio"):
        cls.append("ratio")
    attrs = ""
    if row["kind"] == "section":
        content = section_title_html(value) if first else ""
    elif first:
        toggle = '<span class="tg"></span>' if row["has_children"] else '<span class="tg empty"></span>'
        info = ""
        if show_info and row.get("definition") and row["kind"] not in {"header", "section"}:
            info = f'<span class="info-dot" title="{html.escape(row["definition"], quote=True)}">i</span>'
        content = toggle + label_html(value) + info
    else:
        content = value_html(value)
    return f'<td class="{" ".join(cls)}"{attrs}>{content}</td>'


def label_html(value: str) -> str:
    escaped = html.escape(display_label_text(value))
    return re.sub(r"(\d+(?:[.,:：/-]\d+)*(?:\.\d+)?%?)", r'<span class="inline-num">\1</span>', escaped)


def value_html(value: str) -> str:
    value = display_label_text(value)
    match = re.match(r"^(.+?)\s+\((-?\d+(?:\.\d+)?)%\)$", value)
    if not match:
        return html.escape(value)
    return f'{html.escape(match.group(1))} <span class="share-pct">({html.escape(match.group(2))}%)</span>'


def section_title_html(value: str) -> str:
    value = display_label_text(value)
    match = re.match(r"^\s*(\d+)[\.\、]\s*(.+?)\s*$", value)
    if match:
        return f"「{html.escape(match.group(1))}」{html.escape(match.group(2))}"
    return f"「{html.escape(value)}」"


def section_plain_title(value: str) -> str:
    match = re.match(r"^\s*\d+[\.\、]\s*(.+?)\s*$", value)
    return match.group(1).strip() if match else value.strip("「」 ")


def make_section_row(sheet_title: str, number: int, title: str) -> dict:
    row_id = f"{sheet_title}-section-{number}".replace(" ", "_")
    return {
        "id": row_id,
        "parent": "",
        "level": 0,
        "kind": "section",
        "ratio": False,
        "value_kind": "number",
        "values": [f"{number}. {title}"] + [""] * METRIC_VALUE_COUNT,
        "comments": [],
        "definition": f"分区标题：{title}。",
        "has_children": False,
        "hidden": False,
    }


def row_has_data(row: dict) -> bool:
    if row["kind"] in {"header", "section"}:
        return True
    return any(has_present_value(value) for value in row["values"][1:])


def remove_or_promote_empty_rows(rows: list[dict], *, keep_empty_parents: bool = False) -> list[dict]:
    rows_by_id = {row["id"]: row for row in rows}
    parent_children: dict[str, list[dict]] = {}
    for row in rows:
        if row.get("parent"):
            parent_children.setdefault(row["parent"], []).append(row)

    removed: set[str] = set()
    for row in rows:
        if row["kind"] in {"header", "section"}:
            continue
        if row_has_data(row):
            continue
        if keep_empty_parents and parent_children.get(row["id"]):
            continue
        removed.add(row["id"])

    for row in rows:
        while row.get("parent") in removed:
            old_parent = rows_by_id[row["parent"]]
            row["parent"] = old_parent.get("parent", "")
            row["level"] = max(0, row.get("level", 0) - 1)
            if row["kind"] == "child" and not row["parent"]:
                row["kind"] = "derived" if row.get("ratio") else "row"
                row["hidden"] = False

    kept = [row for row in rows if row["id"] not in removed]
    child_ids = {row["parent"] for row in kept if row.get("parent")}
    for row in kept:
        row["has_children"] = row["id"] in child_ids
        row["hidden"] = row["kind"] == "child"
    return kept


def normalize_sections(sheet_title: str, rows: list[dict]) -> list[dict]:
    if sheet_title.startswith("附录"):
        return rows

    output: list[dict] = []
    inserted_leading = False
    seen_section = False
    section_number = 0

    for row in rows:
        if row["kind"] == "header":
            output.append(row)
            continue

        if row["kind"] != "section" and not seen_section and sheet_title in LEADING_SECTION_TITLES and not inserted_leading:
            title = section_plain_title(LEADING_SECTION_TITLES[sheet_title])
            section_number += 1
            output.append(make_section_row(sheet_title, section_number, title))
            seen_section = True
            inserted_leading = True

        if row["kind"] == "section":
            section_number += 1
            title = section_plain_title(row["values"][0])
            row = row.copy()
            row.update(
                {
                    "parent": "",
                    "level": 0,
                    "has_children": False,
                    "hidden": False,
                    "values": [f"{section_number}. {title}"] + [""] * (len(row["values"]) - 1),
                }
            )
            seen_section = True
        output.append(row)

    section_ids = {row["id"] for row in output if row["kind"] == "section"}
    for row in output:
        if row.get("parent") in section_ids:
            row["parent"] = ""
            row["hidden"] = False
            if row["kind"] == "child":
                row["kind"] = "derived" if row.get("ratio") else "row"
            row["level"] = 0

    child_ids = {row["parent"] for row in output if row.get("parent")}
    for row in output:
        if row["kind"] == "section":
            row["has_children"] = False
        else:
            row["has_children"] = row["id"] in child_ids
            row["hidden"] = row["kind"] == "child"
    return output


def activation_group_for_label(label: str) -> str:
    normalized = normalize_label_for_definition(label)
    if normalized in {"UV", "新增 UV", "回访 UV"} or normalized in CHANNEL_LABELS:
        return "网站访问"
    if "CTA" in normalized or "官网主页" in normalized or "引导页" in normalized:
        return "CTA 点击"
    if normalized in {"新增注册用户"} or "登录完成" in normalized:
        return "登录完成"
    if (
        "发起绑卡" in normalized
        or "新增试用" in normalized
        or "绑卡失败" in normalized
        or "扣款" in normalized
        or normalized == "失败率"
    ):
        return "freetrial"
    if "登录失败" in normalized or normalized.startswith("失败：") or "登录页连续点击" in normalized:
        return "登录失败原因"
    return "首次使用"


def engineering_group_for_label(label: str) -> str:
    normalized = normalize_label_for_definition(label)
    if normalized in {"总请求", "受影响用户", "客户端异常", "客户端连续点击"}:
        return "用户感知"
    if any(token in normalized for token in ["端到端", "页面", "输入响应", "首字节"]):
        return "响应速度"
    if any(token in normalized for token in ["断连", "连接时长", "环境启动", "冷启动"]):
        return "连接与环境"
    return "完成与中止"


def restructure_engineering(rows: list[dict]) -> list[dict]:
    header = [row for row in rows if row["kind"] == "header"]
    body = [row for row in rows if row["kind"] != "header" and row["kind"] != "section"]
    groups = ["用户感知", "响应速度", "连接与环境", "完成与中止"]
    buckets = {group: [] for group in groups}
    row_group: dict[str, str] = {}

    for row in body:
        if row.get("parent") and row["parent"] in row_group:
            group = row_group[row["parent"]]
        else:
            group = engineering_group_for_label(row["values"][0])
        row_group[row["id"]] = group
        buckets.setdefault(group, []).append(row)

    output = header[:]
    number = 0
    for group in groups:
        group_rows = buckets.get(group, [])
        if not group_rows:
            continue
        number += 1
        output.append(make_section_row("工程质量", number, group))
        output.extend(group_rows)
    return normalize_sections("工程质量", output)


def restructure_activation(rows: list[dict]) -> list[dict]:
    header = [row for row in rows if row["kind"] == "header"]
    body = [row for row in rows if row["kind"] != "header" and row["kind"] != "section"]
    groups = ["网站访问", "CTA 点击", "登录完成", "freetrial", "登录失败原因", "首次使用"]
    buckets = {group: [] for group in groups}
    row_group: dict[str, str] = {}

    for row in body:
        if row.get("parent") and row["parent"] in row_group:
            group = row_group[row["parent"]]
        else:
            group = activation_group_for_label(row["values"][0])
        row_group[row["id"]] = group
        buckets.setdefault(group, []).append(row)

    output = header[:]
    number = 0
    for group in groups:
        group_rows = buckets.get(group, [])
        if not group_rows:
            continue
        number += 1
        output.append(make_section_row("用户激活与转化", number, group))
        output.extend(group_rows)
    return normalize_sections("用户激活与转化", output)


def row_classes(row: dict, *, interactive: bool = True) -> list[str]:
    classes = ["grid-row", row["kind"]]
    if interactive and row["has_children"]:
        classes.append("has-children")
    if row.get("default_open"):
        classes.append("open")
    if row["hidden"]:
        classes.append("hidden-row")
    if row["ratio"]:
        classes.append("ratio-row")
    if row["level"]:
        classes.append(f"level-{min(row['level'], 3)}")
    return classes


def row_attrs(row: dict, *, interactive: bool = True) -> str:
    attrs = [f'class="{" ".join(row_classes(row, interactive=interactive))}"', f'data-id="{html.escape(row["id"], quote=True)}"']
    if row["parent"]:
        attrs.append(f'data-parent="{html.escape(row["parent"], quote=True)}"')
    if row["hidden"]:
        attrs.append("hidden")
    return " ".join(attrs)


def left_cell_html(
    row: dict,
    *,
    sheet_title: str,
    section_row: dict | None = None,
    show_info: bool = True,
) -> str:
    value = row["values"][0]
    attrs = ""
    if row["kind"] == "section":
        attrs = anchor_attrs(
            "section",
            sheet_title=sheet_title,
            row=row,
            section_row=row,
            section_label=value,
        )
        content = section_title_html(value)
    else:
        if row["kind"] == "header":
            attrs = anchor_attrs(
                "row",
                sheet_title=sheet_title,
                row=row,
                row_key="header",
                section_key="header",
                row_label=value,
                section_label=value,
            )
        else:
            attrs = anchor_attrs(
                "row",
                sheet_title=sheet_title,
                row=row,
                section_row=section_row,
                row_label=value,
            )
        toggle = '<span class="tg"></span>' if row["has_children"] else '<span class="tg empty"></span>'
        info = ""
        if show_info and row.get("definition") and row["kind"] not in {"header", "section"}:
            info = f'<span class="info-dot" title="{html.escape(row["definition"], quote=True)}">i</span>'
        content = toggle + label_html(value) + info
    return f'<div class="cell left-cell"{attrs}>{content}</div>'


def right_cells_html(row: dict, *, sheet_title: str, column_labels: list[str], section_row: dict | None = None) -> str:
    cells = []
    for index, value in enumerate(row["values"][1:], start=1):
        attrs = ""
        if row["kind"] != "section":
            label = column_labels[index] if index < len(column_labels) else ""
            header_kwargs = {"row_key": "header", "section_key": "header"} if row["kind"] == "header" else {}
            attrs = anchor_attrs(
                "cell",
                sheet_title=sheet_title,
                row=row,
                column=column_key(label, index),
                section_row=section_row,
                row_label=row["values"][0],
                column_label=label,
                **header_kwargs,
            )
        cells.append(f'<div class="cell num"{attrs}>{value_html(value)}</div>')
    return "".join(cells)


def render_main_block(rows: list[dict], kind: str, *, sheet_title: str, column_labels: list[str]) -> str:
    left = []
    right = []
    current_section: dict | None = None
    for row in rows:
        if row["kind"] == "section":
            current_section = row
        left.append(f'<div {row_attrs(row, interactive=True)}>{left_cell_html(row, sheet_title=sheet_title, section_row=current_section)}</div>')
        right.append(f'<div {row_attrs(row, interactive=False)}>{right_cells_html(row, sheet_title=sheet_title, column_labels=column_labels, section_row=current_section)}</div>')
    return (
        f'<div class="table-block {kind}-block">'
        '<div class="freeze-grid">'
        f'<div class="left-pane">{"".join(left)}</div>'
        f'<div class="right-pane"><div class="right-inner">{"".join(right)}</div></div>'
        '</div></div>'
    )


APPENDIX2_SEGMENTS = [
    ("反馈对象", "1. 个人反馈", ["反馈对象", "诉求"]),
    ("高频诉求", "2. 高频诉求归属", ["主题", "涉及的人", "归属分类"]),
    ("1. 已开发 / 待开发", "3. 已开发 / 待开发", ["维度", "指标", "提出人", "状态", "备注"]),
]


def _is_blank_appendix(v: str) -> bool:
    s = (v or "").strip()
    return s in {"", "—", "-", "–", "None"}


def _segment_appendix2_rows(rows: list[dict]) -> list[dict]:
    """Split 附录2 rows into labeled segments at known marker labels."""
    starts = {label: (title, cols) for label, title, cols in APPENDIX2_SEGMENTS}
    skip_labels = {"维度", "研发组", "人名", "共识诉求"}  # source-row column/placeholder labels we replace or remove
    segments: list[dict] = []
    current: dict | None = None
    for row in rows:
        first = (row["values"][0] if row["values"] else "").strip()
        if first in starts:
            title, cols = starts[first]
            current = {"title": title, "columns": cols, "rows": []}
            segments.append(current)
            continue
        if first in skip_labels:
            continue
        if current is None:
            current = {"title": None, "columns": None, "rows": []}
            segments.append(current)
        current["rows"].append(row)
    return segments


def _synthetic_header_row(sheet_title: str, segment_title: str, columns: list[str]) -> dict:
    rid = f"{sheet_title}-seg-{segment_title}-hdr".replace(" ", "_")
    return {
        "id": rid,
        "parent": "",
        "level": 0,
        "kind": "header",
        "ratio": False,
        "value_kind": "text",
        "values": list(columns),
        "comments": [],
        "definition": "段落表头。",
        "has_children": False,
        "hidden": False,
    }


def _synthetic_section_row(sheet_title: str, title: str, columns: int) -> dict:
    rid = f"{sheet_title}-seg-{title}-sec".replace(" ", "_")
    return {
        "id": rid,
        "parent": "",
        "level": 0,
        "kind": "section",
        "ratio": False,
        "value_kind": "text",
        "values": [title] + [""] * max(columns - 1, 0),
        "comments": [],
        "definition": f"分区：{title}。",
        "has_children": False,
        "hidden": False,
    }


def render_appendix_sheet(sheet: dict, idx: int) -> str:
    sheet_classes = ["sheet"]
    is_appendix = sheet["title"].startswith("附录")
    if is_appendix:
        sheet_classes.append("appendix")
    out = [f'<section class="{" ".join(sheet_classes)}" data-sheet="{idx}"><div class="wrap">']

    if sheet["title"] == "附录2 个人看板反馈":
        segments = _segment_appendix2_rows(sheet["rows"])
    else:
        segments = [{"title": None, "columns": None, "rows": sheet["rows"]}]

    for seg in segments:
        seg_rows: list[dict] = []
        if seg["title"]:
            seg_rows.append(_synthetic_section_row(sheet["title"], seg["title"], len(seg["columns"]) + 1 if seg["columns"] else 1))
        if seg["columns"]:
            seg_rows.append(_synthetic_header_row(sheet["title"], seg["title"], seg["columns"]))
        seg_rows.extend(seg["rows"])

        if seg["columns"]:
            max_used = len(seg["columns"]) - 1
        else:
            max_used = 0
            for row in seg_rows:
                if row["kind"] == "header":
                    continue
                for i in range(len(row["values"]) - 1, 0, -1):
                    if not _is_blank_appendix(row["values"][i]):
                        if i > max_used:
                            max_used = i
                        break
            if max_used == 0:
                max_used = 1
        keep_cols = max_used + 1

        for row in seg_rows:
            trimmed = row["values"][:keep_cols]
            if len(trimmed) < keep_cols:
                trimmed = trimmed + [""] * (keep_cols - len(trimmed))
            if row["kind"] != "header":
                trimmed = [trimmed[0]] + ["" if _is_blank_appendix(v) else v for v in trimmed[1:]]
            row["values"] = trimmed

        colgroup = (
            "<colgroup>"
            '<col class="c-appx-label">'
            + "".join('<col class="c-appx-text">' for _ in range(max_used))
            + "</colgroup>"
        )

        table_open = False

        def close_table() -> None:
            nonlocal table_open
            if table_open:
                out.append("</tbody></table></div>")
                table_open = False

        def open_table(kind: str) -> None:
            nonlocal table_open
            close_table()
            out.append(f'<div class="table-block {kind}-block">')
            out.append(f'<table class="t">{colgroup}<tbody>')
            table_open = True

        for row in seg_rows:
            if row["kind"] == "header" and not table_open:
                open_table("header")
            elif row["kind"] == "section":
                open_table("section")
            elif not table_open:
                open_table("plain")

            classes = [row["kind"]]
            if row["has_children"]:
                classes.append("has-children")
            if row["hidden"]:
                classes.append("hidden-row")
            if row["ratio"]:
                classes.append("ratio-row")
            if row["level"]:
                classes.append(f"level-{min(row['level'], 3)}")
            attrs = [f'class="{" ".join(classes)}"', f'data-id="{html.escape(row["id"], quote=True)}"']
            if row["parent"]:
                attrs.append(f'data-parent="{html.escape(row["parent"], quote=True)}"')
            if row["hidden"]:
                attrs.append("hidden")
            cells = "".join(cell_html(row, v, i == 0, show_info=not is_appendix) for i, v in enumerate(row["values"]))
            out.append(f"<tr {' '.join(attrs)}>{cells}</tr>")
        close_table()

    out.append("</div></section>")
    return "\n".join(out)


def render_sheet(sheet: dict, idx: int) -> str:
    if sheet["title"].startswith("附录"):
        return render_appendix_sheet(sheet, idx)

    out = [f'<section class="sheet" data-sheet="{idx}"><div class="wrap">']
    current_rows: list[dict] = []
    current_kind = "plain"
    column_labels = sheet["rows"][0]["values"] if sheet["rows"] else []

    def flush() -> None:
        nonlocal current_rows
        if current_rows:
            out.append(render_main_block(current_rows, current_kind, sheet_title=sheet["title"], column_labels=column_labels))
            current_rows = []

    for row in sheet["rows"]:
        if row["kind"] == "header":
            flush()
            current_kind = "header"
            current_rows = [row]
            flush()
            current_kind = "plain"
            continue
        if row["kind"] == "section":
            flush()
            current_kind = "section"
            current_rows = [row]
            continue
        current_rows.append(row)
    flush()
    out.append("</div></section>")
    return "\n".join(out)


CSS = r"""
:root {
  --navy:#252b33;
  --navy-2:#343b45;
  --derived-bg:#f2f3f5;
  --ratio-bg:#f7f8f9;
  --paper:#ffffff;
  --bg:#f4f5f6;
  --line:#d8dde3;
  --line-strong:#aeb8c3;
  --ink:#0b0f14;
  --ink-2:#26313d;
  --muted:#7c8794;
  --child:#fafbfc;
  --hover:#f6f9fb;
  --today:#252b33;
  --main-table-width:__MAIN_TABLE_WIDTH__px;
}
* { box-sizing:border-box; }
html, body { margin:0; background:var(--bg); color:var(--ink-2); }
body {
  font-family:"KaiTi SC","楷体-简","Kaiti SC","STKaiti",serif;
  font-size:16px;
  line-height:1.16;
  -webkit-font-smoothing:antialiased;
  text-rendering:optimizeLegibility;
}
.num, .tab, .inline-num, .tg, .info-dot {
  font-family:"KaiTi SC","楷体-简","Kaiti SC","STKaiti",serif;
  font-variant-numeric:tabular-nums lining-nums;
  font-feature-settings:"tnum" 1, "lnum" 1;
}
.topbar {
  position:sticky;
  top:0;
  z-index:20;
  background:rgba(255,255,255,.96);
  border-bottom:1px solid var(--line);
  backdrop-filter:saturate(140%) blur(8px);
}
.tabs { display:flex; gap:18px; overflow-x:auto; padding:0 20px; max-width:1840px; margin:0 auto; scrollbar-width:none; }
.tabs::-webkit-scrollbar { display:none; }
.tab {
  border:0;
  background:transparent;
  color:#697482;
  padding:9px 0 8px;
  border-bottom:2px solid transparent;
  font-size:16px;
  white-space:nowrap;
  cursor:pointer;
  letter-spacing:0;
}
.tab.on { color:var(--navy); border-bottom-color:var(--navy); font-weight:600; }
.tab.group-start { margin-left:18px; }
.tab.appendix-start { margin-left:auto; }
main { max-width:1840px; margin:0 auto; padding:10px 20px 22px; }
.sheet { display:none; }
.sheet.on { display:block; }
.wrap { overflow:visible; background:transparent; }
.table-block {
  border:1px solid var(--line-strong);
  background:var(--paper);
  box-shadow:none;
  margin:0 0 12px;
  overflow:hidden;
}
.sheet:not(.appendix) .table-block { width:100%; max-width:calc(var(--main-table-width) + 2px); }
.sheet.appendix .table-block { width:100%; }
.header-block {
  position:sticky;
  top:40px;
  z-index:18;
  margin-bottom:12px;
}
.section-block { margin-bottom:14px; }
.freeze-grid {
  display:grid;
  grid-template-columns:__LEFT_COL_WIDTH__px minmax(0, 1fr);
  width:100%;
  background:var(--paper);
}
.left-pane {
  position:relative;
  z-index:4;
  background:var(--paper);
}
.right-pane {
  width:100%;
  overflow-x:auto;
  overflow-y:hidden;
  scrollbar-width:none;
  background:var(--paper);
}
.right-pane::-webkit-scrollbar { display:none; }
.right-inner { width:__RIGHT_INNER_WIDTH__px; min-width:__RIGHT_INNER_WIDTH__px; }
.grid-row {
  display:grid;
  height:24px;
  min-height:24px;
}
.grid-row[hidden],
tr[hidden] {
  display:none !important;
}
.left-pane .grid-row { grid-template-columns:__LEFT_COL_WIDTH__px; }
.right-pane .grid-row {
  grid-template-columns:__RIGHT_GRID_TEMPLATE__;
}
.cell {
  min-width:0;
  height:24px;
  padding:3px 8px;
  border-bottom:1px solid var(--line);
  color:var(--ink);
  font-size:16px;
  font-weight:400;
  line-height:18px;
  overflow:hidden;
  white-space:nowrap;
  text-overflow:clip;
}
.left-cell {
  text-align:left;
  background:inherit;
  border-right:1px solid var(--line-strong);
}
.num {
  text-align:right;
  color:var(--ink);
  font-size:16px;
  font-weight:400;
  font-variant-numeric:tabular-nums lining-nums;
}
.header-block .grid-row,
.header-block .cell {
  height:28px;
  min-height:28px;
  line-height:21px;
}
.header-block .cell {
  background:#eef1f4;
  color:#111820;
  font-weight:600;
  border-bottom:1px solid #8d99a6;
}
.header-block .right-pane .grid-row > .cell:nth-child(__TODAY_GRID_NTH__),
.right-pane .grid-row:not(.section) > .cell:nth-child(__TODAY_GRID_NTH__) {
  background:var(--today);
  color:#f7f8f9;
  font-weight:700;
  border-bottom-color:#4b5563;
}
.section-block {
  border-color:var(--line-strong);
}
.section-block .grid-row.section,
.section-block .grid-row.section .cell {
  height:28px;
  min-height:28px;
  line-height:21px;
}
.grid-row.section .cell {
  background:#fff;
  color:#252b33;
  font-weight:600;
  border-bottom-color:#b6c0cb;
}
.grid-row.section .left-cell {
  padding-left:8px;
}
.grid-row.row .left-cell,
.grid-row.row.has-children .left-cell {
  font-weight:600;
}
.grid-row.derived .cell {
  background:var(--derived-bg);
  color:#2f3740;
  font-size:16px;
  font-style:normal;
  font-weight:400;
}
.grid-row.derived .left-cell {
  color:#252b33;
  font-weight:500;
}
.grid-row.child .cell {
  background:var(--child);
  color:#252b33;
  font-size:16px;
  font-style:normal;
  font-weight:400;
}
.grid-row.child .left-cell {
  color:#4b5563;
  font-weight:400;
}
.grid-row.ratio-row .cell { background:var(--ratio-bg); }
.grid-row.derived.ratio-row .cell { background:#f1f2f4; }
.right-pane .grid-row:not(.section) > .cell:nth-child(__TODAY_GRID_NTH__) .share-pct { color:#d7dce2; }
.grid-row.level-1 .left-cell { padding-left:28px; }
.grid-row.level-2 .left-cell { padding-left:44px; }
.grid-row.level-3 .left-cell { padding-left:60px; }
.grid-row.has-children { cursor:pointer; }
.grid-row.has-children:hover .cell { background:var(--hover); }
.right-pane .grid-row.has-children:hover > .cell:nth-child(__TODAY_GRID_NTH__) { background:#303740; color:#fff; }
.grid-row.open .left-cell .tg::before { content:"−"; }
.grid-row:last-child .cell { border-bottom:0; }
table.t { width:var(--main-table-width); min-width:var(--main-table-width); table-layout:fixed; border-collapse:separate; border-spacing:0; background:var(--paper); }
.sheet.appendix table.t { width:100%; min-width:0; table-layout:auto; }
col.c-appx-label { width:220px; }
col.c-appx-text { width:auto; }
.sheet.appendix td { vertical-align:top; line-height:1.55; }
.sheet.appendix td:first-child {
  position:static !important;
  border-right:1px solid var(--line);
  width:auto;
  min-width:0;
  max-width:none;
  background:inherit !important;
}
.sheet.appendix td,
.sheet.appendix td.num {
  text-align:left;
  white-space:normal;
}
col.c-label { width:__LEFT_COL_WIDTH__px; }
col.c-date { width:76px; }
col.c-today { width:108px; }
col.c-comp { width:128px; }
col.c-wow { width:78px; }
col.c-mom { width:78px; }
td { border-bottom:1px solid var(--line); padding:3px 8px; height:24px; vertical-align:middle; font-size:16px; font-weight:400; }
td:first-child {
  position:sticky;
  left:0;
  z-index:10;
  border-right:1px solid var(--line-strong);
  width:__LEFT_COL_WIDTH__px;
  min-width:__LEFT_COL_WIDTH__px;
  max-width:320px;
  color:var(--ink);
  text-align:left;
  background:#fff;
}
td.num { text-align:right; color:var(--ink); white-space:nowrap; font-size:16px; font-weight:400; letter-spacing:0; }
tr.header td {
  background:#eef1f4;
  color:#111820;
  font-weight:600;
  border-bottom:1px solid #8d99a6;
  box-shadow:none;
}
tr.header td:first-child { color:#111820; text-align:left; padding-left:8px; z-index:13; }
tr.header td:nth-child(__TODAY_TABLE_NTH__),
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__) { background:var(--today); color:#fff; border-bottom-color:#4b5563; }
tr.header td:nth-child(__TODAY_TABLE_NTH__) { font-weight:700; }
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__).num { font-weight:700; color:#f7f8f9; }
tr.section td {
  background:#fff;
  color:#252b33;
  font-weight:600;
  border-top:0;
  border-bottom-color:#b6c0cb;
  box-shadow:none;
  height:28px;
}
tr.section td:first-child {
  z-index:12;
  text-align:left;
  padding-left:6px;
}
tr:not(.section):last-child td { border-bottom:0; }
tr.row td { background:#fff; color:#0b0f14; font-weight:400; }
tr.row td:first-child { font-weight:600; }
tr.row.has-children td:first-child { font-weight:600; }
tr.derived td { background:var(--derived-bg); color:#2f3740; font-size:16px; font-style:normal; font-weight:400; }
tr.derived td:first-child { color:#252b33; font-style:normal; font-weight:500; }
tr.derived td.num { color:#252b33; font-size:16px; font-weight:400; }
tr.child td { background:var(--child); color:#3b434d; font-size:16px; font-style:normal; font-weight:400; }
tr.child td:first-child { color:#4b5563; font-style:normal; font-weight:400; }
tr.child td.num { color:#252b33; font-size:16px; font-style:normal; font-weight:400; }
.share-pct {
  color:#5f6975;
  font-size:12px;
  font-style:italic;
  font-weight:400;
}
tr.ratio-row td { background:var(--ratio-bg); }
tr.derived.ratio-row td { background:#f1f2f4; }
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__),
tr.derived td:nth-child(__TODAY_TABLE_NTH__).num,
tr.child td:nth-child(__TODAY_TABLE_NTH__).num,
tr.ratio-row td:nth-child(__TODAY_TABLE_NTH__),
tr.derived.ratio-row td:nth-child(__TODAY_TABLE_NTH__).num {
  background:var(--today);
  color:#f7f8f9;
  font-weight:700;
}
tr.section td:nth-child(__TODAY_TABLE_NTH__),
tr.section td:nth-child(__TODAY_TABLE_NTH__).num {
  background:#fff;
  color:#252b33;
  font-weight:600;
  border-bottom-color:#b6c0cb;
}
tr:not(.header) td:nth-child(__TODAY_TABLE_NTH__) .share-pct { color:#d7dce2; }
tr.level-1 td:first-child { padding-left:28px; }
tr.level-2 td:first-child { padding-left:44px; font-style:normal; }
tr.level-3 td:first-child { padding-left:60px; font-style:normal; }
tr.has-children { cursor:pointer; }
tr.has-children:hover td { background:var(--hover); }
tr.has-children:hover td:nth-child(__TODAY_TABLE_NTH__) { background:#303740; color:#fff; }
tr.open > td:first-child .tg::before { content:"−"; }
.tg {
  display:inline-flex;
  width:12px;
  height:10px;
  margin-right:6px;
  align-items:center;
  justify-content:center;
  color:#a5aeb8;
  font-size:12px;
  font-weight:400;
  line-height:1;
  vertical-align:1px;
}
.tg::before { content:"+"; }
.tg.empty { visibility:hidden; }
.inline-num { color:inherit; opacity:.72; }
.info-dot {
  display:inline-flex;
  margin-left:7px;
  width:12px;
  height:12px;
  border-radius:50%;
  align-items:center;
  justify-content:center;
  border:1px solid #b9c1ca;
  background:#f7f8f9;
  color:#6f7a86;
  font-size:8px;
  font-weight:500;
  line-height:1;
  vertical-align:2px;
  cursor:help;
  opacity:.82;
}
.info-dot:hover {
  background:#e9edf1;
  color:#252b33;
  border-color:#8f9aa7;
}
@media (max-width:900px) {
  body { font-size:16px; }
  main { padding:10px; }
  .tabs { padding-left:10px; padding-right:10px; }
  td:first-child { width:240px; min-width:240px; }
  td { padding:3px 7px; height:24px; }
}
"""


JS = r"""
(function initDashboardBase() {
const sheets = __SHEETS__;
const DEFAULT_DATE_SCROLL_LEFT = __DEFAULT_DATE_SCROLL_LEFT__;
const baseState = window.dashboardBaseState || { currentSheet: 0, initializedScroll: new Set(), wired: false };
if (!(baseState.initializedScroll instanceof Set)) {
  baseState.initializedScroll = new Set(baseState.initializedScroll || []);
}
window.dashboardBaseState = baseState;

function refreshCommentOverlay(options = {}) {
  const overlay = window.dashboardCommentOverlay;
  if (!overlay) return;
  if (options.closeFloating) {
    if (typeof overlay.closeComposer === 'function') overlay.closeComposer();
    if (typeof overlay.closePopover === 'function') overlay.closePopover();
  }
  if (typeof overlay.scheduleRenderPins === 'function') overlay.scheduleRenderPins();
  else if (typeof overlay.renderPins === 'function') overlay.renderPins();
}

function applyDefaultDateScroll(sheetIndex = baseState.currentSheet) {
  if (baseState.initializedScroll.has(sheetIndex)) return;
  baseState.initializedScroll.add(sheetIndex);
  document.querySelectorAll(`.sheet[data-sheet="${sheetIndex}"] .right-pane`).forEach(pane => {
    pane.scrollLeft = DEFAULT_DATE_SCROLL_LEFT;
  });
}

function sync() {
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('on', Number(t.dataset.sheet) === baseState.currentSheet));
  document.querySelectorAll('.sheet').forEach(s => s.classList.toggle('on', Number(s.dataset.sheet) === baseState.currentSheet));
  applyDefaultDateScroll(baseState.currentSheet);
  window.scrollTo({ top: 0, behavior: 'instant' });
  refreshCommentOverlay({ closeFloating: true });
}

function closeChildren(id) {
  document.querySelectorAll(`.grid-row[data-parent="${CSS.escape(id)}"]`).forEach(child => {
    child.hidden = true;
    child.classList.remove('open');
    closeChildren(child.dataset.id);
  });
}

function wire() {
  if (baseState.wired) return;
  baseState.wired = true;
  document.querySelectorAll('.tab').forEach(btn => {
    btn.addEventListener('click', () => {
      baseState.currentSheet = Number(btn.dataset.sheet);
      sync();
    });
  });
  let syncing = false;
  document.querySelectorAll('.right-pane').forEach(pane => {
    pane.addEventListener('scroll', () => {
      if (syncing) return;
      syncing = true;
      const left = pane.scrollLeft;
      document.querySelectorAll('.sheet.on .right-pane').forEach(other => {
        if (other !== pane) other.scrollLeft = left;
      });
      syncing = false;
      refreshCommentOverlay();
    });
  });
  document.querySelectorAll('.left-pane .grid-row.has-children').forEach(row => {
    row.addEventListener('click', event => {
      event.stopPropagation();
      const id = row.dataset.id;
      const children = Array.from(document.querySelectorAll(`.grid-row[data-parent="${CSS.escape(id)}"]`));
      const leftChildren = children.filter(c => c.closest('.left-pane'));
      const anyVisible = leftChildren.some(c => !c.hidden);
      document.querySelectorAll(`.left-pane .grid-row[data-id="${CSS.escape(id)}"]`).forEach(parent => parent.classList.toggle('open', !anyVisible));
      if (anyVisible) closeChildren(id);
      else children.forEach(c => { c.hidden = false; });
      refreshCommentOverlay();
    });
  });
}

wire();
sync();

window.initDashboardCommentOverlay = async function initDashboardCommentOverlay() {
  if (!window.DashboardCommentModel || !window.DashboardCommentStore || !window.DashboardCommentOverlay) return;
  const existing = window.dashboardCommentOverlay;
  if (existing && existing.initialized) return existing;
  if (existing && typeof existing.init === 'function') {
    await existing.init();
    return existing;
  }
  if (existing && typeof existing.destroy === 'function') existing.destroy();
  const model = window.DashboardCommentModel;
  const store = window.DashboardCommentStore.createDefaultStore();
  const overlay = new window.DashboardCommentOverlay.CommentOverlay({
    model,
    store,
    pageKey: "moclaw_operating_dashboard",
    pageVersion: "v1"
  });
  window.dashboardCommentOverlay = overlay;
  await overlay.init();
  return overlay;
};

function scheduleDashboardCommentOverlayInit() {
  if (document.readyState !== 'loading') {
    window.initDashboardCommentOverlay();
    return;
  }
  if (window.dashboardCommentOverlayInitScheduled) return;
  window.dashboardCommentOverlayInitScheduled = true;
  window.addEventListener('DOMContentLoaded', () => {
    window.dashboardCommentOverlayInitScheduled = false;
    window.initDashboardCommentOverlay();
  }, { once: true });
}

scheduleDashboardCommentOverlayInit();
})();
"""


def render_css() -> str:
    return (
        CSS.replace("__MAIN_TABLE_WIDTH__", str(MAIN_TABLE_WIDTH))
        .replace("__LEFT_COL_WIDTH__", str(LEFT_COL_WIDTH))
        .replace("__RIGHT_INNER_WIDTH__", str(RIGHT_INNER_WIDTH))
        .replace("__RIGHT_GRID_TEMPLATE__", RIGHT_GRID_TEMPLATE)
        .replace("__TODAY_GRID_NTH__", str(TODAY_GRID_NTH))
        .replace("__TODAY_TABLE_NTH__", str(TODAY_TABLE_NTH))
    )


def render_js(sheets: list[dict]) -> str:
    payload = json.dumps([{"title": display_sheet_title(s["title"])} for s in sheets], ensure_ascii=False)
    return JS.replace("__SHEETS__", payload).replace("__DEFAULT_DATE_SCROLL_LEFT__", str(DEFAULT_DATE_SCROLL_LEFT))


def main() -> None:
    MISSING_DEFINITIONS.clear()
    sheets = workbook_model()
    comments_css = COMMENTS_CSS_PATH.read_text(encoding="utf-8")
    comment_scripts = [path.read_text(encoding="utf-8") for path in COMMENT_SCRIPT_PATHS]
    parts = [
        "<!doctype html>",
        '<html lang="zh-CN"><head><meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>MoClaw Operating Dashboard · Final</title>",
        f"<style>{render_css()}</style>",
        f"<style>{comments_css}</style></head><body>",
        '<nav class="topbar"><div class="tabs">',
    ]
    for i, sheet in enumerate(sheets):
        group_class = " group-start" if sheet["title"] in {"用户获取", "Agent 质量", "财务"} else ""
        appendix_class = " appendix-start" if sheet["title"].startswith("附录1") else ""
        parts.append(f'<button class="tab{group_class}{appendix_class}" data-sheet="{i}">{html.escape(display_sheet_title(sheet["title"]))}</button>')
    parts.extend(['</div></nav>', "<main>"])
    for i, sheet in enumerate(sheets):
        parts.append(render_sheet(sheet, i))
    parts.extend(
        [
            "</main>",
            *[f"<script>{script}</script>" for script in comment_scripts],
            f"<script>{render_js(sheets)}</script>",
            "</body></html>",
        ]
    )
    OUT.write_text("\n".join(parts), encoding="utf-8")
    write_missing_definitions()
    print(OUT)
    print("sheets", len(sheets), "rows", sum(len(s["rows"]) for s in sheets))
    print("missing_definitions", len(set(MISSING_DEFINITIONS)), MISSING_DEFS_OUT)


def write_missing_definitions() -> None:
    by_sheet: dict[str, list[str]] = {}
    for sheet, label in sorted(set(MISSING_DEFINITIONS)):
        by_sheet.setdefault(sheet, []).append(label)
    lines = [
        "MoClaw Operating Dashboard missing field definitions",
        f"Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Count: {sum(len(v) for v in by_sheet.values())}",
        "",
    ]
    for sheet in ORDER:
        labels = by_sheet.get(sheet)
        if not labels:
            continue
        lines.append(f"[{sheet}]")
        lines.extend(f"- {label}" for label in labels)
        lines.append("")
    MISSING_DEFS_OUT.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
