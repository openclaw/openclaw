#!/usr/bin/env python3
"""
Conversion Utilities - Handle file format conversion for large file analyzer
"""

import os
import sys
import subprocess
import logging
from pathlib import Path

class ConversionUtils:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def detect_file_format(self, file_path: str) -> str:
        """Detect file format based on extension and content"""
        path = Path(file_path)
        ext = path.suffix.lower()
        
        if ext in ['.csv', '.txt']:
            return 'csv'
        elif ext in ['.xlsx', '.xls']:
            return 'excel'
        elif ext == '.json':
            return 'json'
        else:
            # Try to detect by content
            try:
                with open(file_path, 'rb') as f:
                    header = f.read(8)
                    if header.startswith(b'PK'):
                        return 'excel'  # Excel files are ZIP archives
                    elif b',' in header[:100]:
                        return 'csv'
                    elif b'{' in header[:100] or b'[' in header[:100]:
                        return 'json'
            except Exception as e:
                self.logger.warning(f"Could not detect format by content: {e}")
            
            return 'unknown'
    
    def convert_excel_to_csv(self, excel_path: str, csv_path: str, sheet_name: str = None) -> bool:
        """
        Convert Excel file to CSV using available methods
        Returns True if successful, False otherwise
        """
        # Method 1: Try using system tools (if available)
        if self._try_system_conversion(excel_path, csv_path, sheet_name):
            return True
        
        # Method 2: Try using Python libraries (if available)
        if self._try_python_conversion(excel_path, csv_path, sheet_name):
            return True
        
        # Method 3: Fallback - provide instructions for manual conversion
        self._provide_manual_conversion_instructions(excel_path, csv_path)
        return False
    
    def _try_system_conversion(self, excel_path: str, csv_path: str, sheet_name: str = None) -> bool:
        """Try system-level conversion tools"""
        # Check for xlsx2csv
        try:
            cmd = ['xlsx2csv']
            if sheet_name:
                cmd.extend(['-s', sheet_name])
            cmd.extend([excel_path, csv_path])
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
            if result.returncode == 0:
                self.logger.info(f"Converted {excel_path} to {csv_path} using xlsx2csv")
                return True
            else:
                self.logger.warning(f"xlsx2csv failed: {result.stderr}")
        except FileNotFoundError:
            self.logger.info("xlsx2csv not available")
        except subprocess.TimeoutExpired:
            self.logger.warning("xlsx2csv conversion timed out")
        except Exception as e:
            self.logger.warning(f"xlsx2csv error: {e}")
        
        return False
    
    def _try_python_conversion(self, excel_path: str, csv_path: str, sheet_name: str = None) -> bool:
        """Try Python library conversion"""
        try:
            # Try pandas first (most reliable)
            import pandas as pd
            
            # Read Excel file
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(excel_path)
            
            # Write to CSV
            df.to_csv(csv_path, index=False, encoding='utf-8')
            self.logger.info(f"Converted {excel_path} to {csv_path} using pandas")
            return True
            
        except ImportError:
            self.logger.info("pandas not available, trying alternative methods")
        except Exception as e:
            self.logger.warning(f"pandas conversion failed: {e}")
        
        try:
            # Try openpyxl for .xlsx files
            from openpyxl import load_workbook
            import csv
            
            wb = load_workbook(excel_path, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            
            wb.close()
            self.logger.info(f"Converted {excel_path} to {csv_path} using openpyxl")
            return True
            
        except ImportError:
            self.logger.info("openpyxl not available")
        except Exception as e:
            self.logger.warning(f"openpyxl conversion failed: {e}")
        
        try:
            # Try xlrd for .xls files
            import xlrd
            
            workbook = xlrd.open_workbook(excel_path)
            worksheet = workbook.sheet_by_index(0) if not sheet_name else workbook.sheet_by_name(sheet_name)
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row_idx in range(worksheet.nrows):
                    writer.writerow(worksheet.row_values(row_idx))
            
            self.logger.info(f"Converted {excel_path} to {csv_path} using xlrd")
            return True
            
        except ImportError:
            self.logger.info("xlrd not available")
        except Exception as e:
            self.logger.warning(f"xlrd conversion failed: {e}")
        
        return False
    
    def _provide_manual_conversion_instructions(self, excel_path: str, csv_path: str):
        """Provide instructions for manual conversion"""
        instructions = f"""
MANUAL CONVERSION REQUIRED:

Your Excel file '{excel_path}' needs to be converted to CSV format.

Please follow these steps:

1. Open the Excel file in Microsoft Excel, Google Sheets, or LibreOffice Calc
2. Go to File → Save As (or Export)
3. Choose "CSV (Comma delimited) (*.csv)" as the file type
4. Save it as '{os.path.basename(csv_path)}'
5. Upload the CSV file back to this chat

Alternative command-line method (if you have access):
- Install xlsx2csv: pip install xlsx2csv
- Run: xlsx2csv "{excel_path}" "{csv_path}"

Once converted, the analysis will proceed automatically.
        """
        print(instructions)
        self.logger.info("Manual conversion instructions provided")
    
    def get_conversion_status(self, original_file: str) -> dict:
        """Get status of conversion attempt"""
        csv_file = original_file.rsplit('.', 1)[0] + '.csv'
        temp_csv = original_file + '.temp_converted.csv'
        
        status = {
            'original_file': original_file,
            'csv_file_exists': os.path.exists(csv_file),
            'temp_csv_exists': os.path.exists(temp_csv),
            'conversion_needed': self.detect_file_format(original_file) == 'excel'
        }
        
        if status['csv_file_exists']:
            status['target_file'] = csv_file
        elif status['temp_csv_exists']:
            status['target_file'] = temp_csv
        else:
            status['target_file'] = None
            
        return status