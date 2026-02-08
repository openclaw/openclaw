#!/usr/bin/env python3
"""
Excel to CSV Converter - Lightweight converter for large Excel files
Uses system tools when available, falls back to basic parsing
"""

import os
import sys
import subprocess
import logging
from pathlib import Path

class ExcelConverter:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.available_tools = self._check_available_tools()
    
    def _check_available_tools(self):
        """Check what conversion tools are available"""
        tools = {
            'xlsx2csv': False,
            'pandas': False,
            'xlrd': False,
            'openpyxl': False
        }
        
        # Check xlsx2csv command line tool
        try:
            result = subprocess.run(['xlsx2csv', '--version'], 
                                  capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                tools['xlsx2csv'] = True
                self.logger.info("xlsx2csv tool available")
        except (subprocess.TimeoutExpired, FileNotFoundError, subprocess.SubprocessError):
            pass
        
        # Check Python libraries
        try:
            import pandas as pd
            tools['pandas'] = True
            self.logger.info("pandas available")
        except ImportError:
            pass
            
        try:
            import xlrd
            tools['xlrd'] = True
            self.logger.info("xlrd available")
        except ImportError:
            pass
            
        try:
            import openpyxl
            tools['openpyxl'] = True
            self.logger.info("openpyxl available")
        except ImportError:
            pass
            
        return tools
    
    def convert_excel_to_csv(self, excel_path: str, csv_path: str = None, sheet_name: str = None) -> str:
        """
        Convert Excel file to CSV with available tools
        Returns path to converted CSV file
        """
        if csv_path is None:
            csv_path = str(Path(excel_path).with_suffix('.csv'))
        
        self.logger.info(f"Converting {excel_path} to {csv_path}")
        
        # Try tools in order of preference
        if self.available_tools['xlsx2csv']:
            return self._convert_with_xlsx2csv(excel_path, csv_path, sheet_name)
        elif self.available_tools['pandas']:
            return self._convert_with_pandas(excel_path, csv_path, sheet_name)
        elif self.available_tools['openpyxl']:
            return self._convert_with_openpyxl(excel_path, csv_path, sheet_name)
        elif self.available_tools['xlrd']:
            return self._convert_with_xlrd(excel_path, csv_path, sheet_name)
        else:
            # Fallback: try to install minimal dependencies or use system approach
            return self._convert_with_fallback(excel_path, csv_path, sheet_name)
    
    def _convert_with_xlsx2csv(self, excel_path: str, csv_path: str, sheet_name: str = None) -> str:
        """Convert using xlsx2csv command line tool"""
        try:
            cmd = ['xlsx2csv']
            if sheet_name:
                cmd.extend(['-s', sheet_name])
            cmd.extend([excel_path, csv_path])
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
            if result.returncode == 0:
                self.logger.info(f"Successfully converted with xlsx2csv: {csv_path}")
                return csv_path
            else:
                self.logger.error(f"xlsx2csv failed: {result.stderr}")
                raise Exception(f"xlsx2csv conversion failed: {result.stderr}")
        except Exception as e:
            self.logger.error(f"xlsx2csv conversion error: {e}")
            raise
    
    def _convert_with_pandas(self, excel_path: str, csv_path: str, sheet_name: str = None) -> str:
        """Convert using pandas"""
        try:
            import pandas as pd
            
            # Read Excel file
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(excel_path)
            
            # Write to CSV
            df.to_csv(csv_path, index=False, encoding='utf-8')
            self.logger.info(f"Successfully converted with pandas: {csv_path}")
            return csv_path
        except Exception as e:
            self.logger.error(f"pandas conversion error: {e}")
            raise
    
    def _convert_with_openpyxl(self, excel_path: str, csv_path: str, sheet_name: str = None) -> str:
        """Convert using openpyxl (for .xlsx files only)"""
        try:
            from openpyxl import load_workbook
            import csv
            
            wb = load_workbook(excel_path, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            
            wb.close()
            self.logger.info(f"Successfully converted with openpyxl: {csv_path}")
            return csv_path
        except Exception as e:
            self.logger.error(f"openpyxl conversion error: {e}")
            raise
    
    def _convert_with_xlrd(self, excel_path: str, csv_path: str, sheet_name: str = None) -> str:
        """Convert using xlrd (for .xls files)"""
        try:
            import xlrd
            import csv
            
            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for row_idx in range(ws.nrows):
                    writer.writerow(ws.row_values(row_idx))
            
            self.logger.info(f"Successfully converted with xlrd: {csv_path}")
            return csv_path
        except Exception as e:
            self.logger.error(f"xlrd conversion error: {e}")
            raise
    
    def _convert_with_fallback(self, excel_path: str, csv_path: str, sheet_name: str = None) -> str:
        """
        Fallback conversion method
        Attempts to install minimal dependencies or provide alternative solutions
        """
        self.logger.warning("No conversion tools available, attempting fallback...")
        
        # Try to install minimal dependencies
        try:
            self.logger.info("Attempting to install xlsx2csv...")
            subprocess.run([sys.executable, '-m', 'pip', 'install', 'xlsx2csv'], 
                          capture_output=True, timeout=60)
            # Re-check available tools
            self.available_tools = self._check_available_tools()
            if self.available_tools['xlsx2csv']:
                return self._convert_with_xlsx2csv(excel_path, csv_path, sheet_name)
        except Exception as e:
            self.logger.warning(f"Failed to install xlsx2csv: {e}")
        
        # Final fallback: provide instructions
        error_msg = f"""
        Cannot convert Excel file automatically. Please try one of these options:
        
        1. Install conversion tools:
           pip install xlsx2csv pandas openpyxl
        
        2. Convert manually in Excel:
           - Open the file in Excel
           - Save As -> CSV (UTF-8)
           - Upload the CSV file instead
        
        3. Use online converter (for non-sensitive data):
           - Upload to a trusted online Excel-to-CSV converter
           - Download and re-upload the CSV
        
        Original file: {excel_path}
        Expected CSV output: {csv_path}
        """
        self.logger.error(error_msg)
        raise Exception("No suitable conversion method available. " + error_msg)

def main():
    """Command line interface for testing"""
    if len(sys.argv) < 2:
        print("Usage: python excel_converter.py <excel_file> [csv_output]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    csv_output = sys.argv[2] if len(sys.argv) > 2 else None
    
    converter = ExcelConverter()
    try:
        result = converter.convert_excel_to_csv(excel_file, csv_output)
        print(f"Successfully converted: {result}")
    except Exception as e:
        print(f"Conversion failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()